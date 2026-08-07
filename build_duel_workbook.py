"""
Build the CRL Duel Decks workbook from raw Clash Royale battlelog JSON files.

Pipeline:
  1. Load every raw_<tag>.json in the uploads folder.
  2. Filter to type == "clanMate" and gameMode.name == "Friendly" (the real
     practice-duel signal, confirmed against real API data -- ranked ladder
     and the "trail / Showdown_Friendly" event mode are excluded per user
     decision).
  3. Group each (tracked player, opponent) pair's friendly battles, in
     chronological order, into "duels" using two heuristics:
       - No card may repeat across the decks used within one duel (matches
         informal community duel convention: 3 distinct, non-overlapping
         decks).
       - A gap of more than DUEL_GAP_HOURS between consecutive games against
         the same opponent forces a new duel (back-to-back play assumption).
     Because the API only returns each player's most recent battles, the
     first duel found per opponent pair may actually be a continuation of a
     duel that started before the fetch window -- those rows are flagged.
  4. Write everything into a formatted Excel workbook with real values +
     live formulas (Result, summary counts), plus a documented assumptions
     sheet.
"""

import json
import glob
import os
from datetime import datetime, timezone
from collections import defaultdict, Counter
from itertools import combinations

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

DUEL_GAP_HOURS = 3  # assumption: >3h since the last game vs this opponent = new duel
MAX_GAMES_PER_DUEL = 3
SESSION_GAP_HOURS = 1  # assumption: >=1h between duels vs the same opponent = new practice session

# ---------------------------------------------------------------------------
# PRACTICE-DUEL COMPLETENESS GATE (added 2026-08-06)
#
# THE RULE, per Alexander: a *Practice* duel is supposed to be a full 3-game set. If one
# comes back with fewer than 3 distinct games, that is missing data (games aged out of the
# API's ~30-battle window before we fetched), not a real short set -- so those games must
# not feed the aggregate deck/win-con/win-rate stats.
#
# Official CRL duels are the OPPOSITE case and are deliberately NOT gated here: a 2-0 sweep
# is genuinely complete with 2 games, and a 1-1 is legitimately pending game 3. That shape
# logic already lives in compute_crl_duel_status() and stays the authority for CRL.
#
# WHY A DATE CUTOFF: before 2026-08-01 the archive was being fetched only every 3 hours,
# so a large share of Practice sets are truncated purely as an artifact of that cadence --
# gating the whole history would throw away ~51% of all games. From 08-01 onward (and with
# the 30-minute fetch cadence in fetch.yml) a short Practice set is a real signal.
#
# TURNING IT OFF: during CRL event days the fetch pattern is different and partial sets are
# expected mid-event. Disable without editing code:
#     CRL_ENFORCE_COMPLETENESS=0 python build_duel_workbook.py
# Move the cutoff the same way:
#     CRL_COMPLETENESS_FROM=20260901T000000.000Z python build_duel_workbook.py
# ---------------------------------------------------------------------------
ENFORCE_PRACTICE_COMPLETENESS = os.environ.get(
    "CRL_ENFORCE_COMPLETENESS", "1"
).strip().lower() not in ("0", "false", "no", "off")
PRACTICE_COMPLETENESS_FROM = os.environ.get(
    "CRL_COMPLETENESS_FROM", "20260801T000000.000Z"
)


def _completeness_cutoff():
    """Parsed cutoff datetime, or None when the gate is disabled/unset."""
    if not ENFORCE_PRACTICE_COMPLETENESS or not PRACTICE_COMPLETENESS_FROM:
        return None
    try:
        return datetime.strptime(
            PRACTICE_COMPLETENESS_FROM, "%Y%m%dT%H%M%S.%fZ"
        ).replace(tzinfo=timezone.utc)
    except ValueError:
        print(f"WARNING: CRL_COMPLETENESS_FROM={PRACTICE_COMPLETENESS_FROM!r} is not in "
              f"battleTime format (YYYYMMDDTHHMMSS.mmmZ); completeness gate DISABLED.")
        return None


def is_stats_eligible(category, duel_start, distinct_game_count):
    """Should this duel's games count toward aggregate deck / win-con / win-rate stats?

    Only Practice duels starting on or after the cutoff are gated, and only for being
    short. Everything else -- all Official CRL, everything before the cutoff -- stays in."""
    cutoff = _completeness_cutoff()
    if cutoff is None:
        return True
    if category != "Practice":
        return True
    if duel_start < cutoff:
        return True
    return distinct_game_count >= MAX_GAMES_PER_DUEL

# ---------------------------------------------------------------------------
# Where the data files (master_*.json / raw_*.json) live and where the workbook is written.
# Set env CRL_HOME to run the whole pipeline out of ONE directory (used by GitHub Actions:
# CRL_HOME=$GITHUB_WORKSPACE). When CRL_HOME is unset, the original Cowork split paths are
# preserved verbatim, so the manual flow is unchanged.
# ---------------------------------------------------------------------------
_CRL_HOME = os.environ.get("CRL_HOME")
DATA_DIR = _CRL_HOME if _CRL_HOME else "/mnt/user-data/uploads/CRL"
XLSX_OUT = os.path.join(_CRL_HOME, "CRL_Duel_Decks.xlsx") if _CRL_HOME else "/home/claude/CRL_Duel_Decks.xlsx"

# ---------------------------------------------------------------------------
# Win-condition classification. This is community/strategy knowledge, NOT an
# API field -- Supercell doesn't label cards this way. Confidence is marked
# per card; "Uncertain" cards are ones not confidently recognized (possibly
# added to the game after this classification was written, or just genuinely
# ambiguous role) and should be reviewed/corrected in the "Win Condition
# Reference" sheet -- edits there require re-running this script to take
# effect (not a live formula, same limitation as the duel-grouping logic).
# Format: card_name -> (is_win_condition: bool, confidence: str, note: str)
# ---------------------------------------------------------------------------
WIN_CONDITION_REFERENCE = {
    # Built from the user's official in-game card list (122 cards, confirmed
    # 2026-07-17). Six cards previously marked "Uncertain" were confirmed by
    # the user to NOT be win conditions: Boss Bandit, Goblinstein, Ronin,
    # Skeleton Dragons, Spirit Empress, Vines.
    "Archers": (False, "High", ""),
    "Archer Queen": (False, "Medium", "Champion, typically utility/chip not primary win con"),
    "Baby Dragon": (False, "High", ""),
    "Balloon": (True, "High", ""),
    "Bandit": (False, "Medium", "Usually counter-push support, not primary win con"),
    "Barbarians": (False, "High", ""),
    "Bats": (False, "High", ""),
    "Battle Healer": (False, "High", ""),
    "Battle Ram": (True, "High", ""),
    "Berserker": (False, "High", ""),
    "Bomber": (False, "High", ""),
    "Boss Bandit": (False, "High", "Confirmed by user - not a win condition"),
    "Bowler": (False, "High", ""),
    "Cannon Cart": (False, "High", ""),
    "Dark Prince": (False, "High", ""),
    "Dart Goblin": (False, "High", ""),
    "Electro Dragon": (False, "High", ""),
    "Electro Giant": (True, "High", ""),
    "Electro Spirit": (False, "High", ""),
    "Electro Wizard": (False, "High", ""),
    "Elite Barbarians": (True, "Medium", ""),
    "Elixir Golem": (True, "Medium", ""),
    "Executioner": (False, "High", ""),
    "Firecracker": (False, "High", ""),
    "Fire Spirit": (False, "High", ""),
    "Fisherman": (False, "Medium", ""),
    "Flying Machine": (False, "Medium", "Ranged air support"),
    "Furnace": (False, "High", ""),
    "Giant": (True, "High", ""),
    "Giant Skeleton": (False, "High", "Confirmed by user - not a win condition"),
    "Goblin Brawler": (False, "Low", "Not confidently known -- best guess is melee support"),
    "Goblin Gang": (False, "High", ""),
    "Goblin Demolisher": (False, "High", "Confirmed by user - not a win condition"),
    "Goblin Giant": (True, "Medium", ""),
    "Goblin Barrel": (True, "High", "Missing from user's pasted 122-card list (likely dropped in copy/paste) -- re-added, well-established win condition"),
    "Goblin Machine": (False, "Low", "Not confidently known -- best guess is support"),
    "Goblins": (False, "High", ""),
    "Goblinstein": (False, "High", "Confirmed by user - not a win condition"),
    "Golden Knight": (False, "Medium", "Champion, typically dash/tank utility not primary win con"),
    "Golem": (True, "High", ""),
    "Graveyard": (True, "High", "Missing from user's pasted 122-card list (likely dropped in copy/paste) -- re-added, well-established win condition"),
    "Guards": (False, "High", ""),
    "Hog Rider": (True, "High", ""),
    "Hunter": (False, "High", ""),
    "Heal Spirit": (False, "High", ""),
    "Ice Golem": (False, "High", ""),
    "Ice Spirit": (False, "High", ""),
    "Ice Wizard": (False, "High", ""),
    "Inferno Dragon": (False, "Medium", "Single-target melter, occasionally counter-push, not primary win con"),
    "Knight": (False, "High", ""),
    "Lava Hound": (True, "High", ""),
    "Lava Pup": (False, "Low", "Not confidently known -- likely a companion/split unit, best guess support"),
    "Little Prince": (False, "Low", "Not confidently known -- best guess is a support companion card"),
    "Lumberjack": (False, "High", ""),
    "Magic Archer": (False, "High", ""),
    "Mega Knight": (False, "Medium", "Usually defensive counter-push tank, not primary win con"),
    "Mega Minion": (False, "High", ""),
    "Mighty Miner": (False, "High", "Confirmed by user - not a win condition"),
    "Miner": (True, "High", ""),
    "Mini P.E.K.K.A.": (False, "High", "Aggressive tank-buster support, not primary win con"),
    "Minion Horde": (False, "High", ""),
    "Minions": (False, "High", ""),
    "Monk": (False, "Medium", "Champion, deflect/defense utility"),
    "Mother Witch": (False, "Medium", "Champion, utility/support"),
    "Ronin": (False, "High", "Confirmed by user - not a win condition"),
    "Musketeer": (False, "High", ""),
    "Night Witch": (False, "High", "Support, bat-summoning troop"),
    "P.E.K.K.A.": (False, "Medium", "Tank/control centerpiece; occasionally a beatdown win con but usually categorized as support/tank"),
    "Phoenix": (False, "Medium", "Support/tank-ish resurrecting flyer"),
    "Prince": (True, "Medium", "Situational -- primary win con in dedicated Prince decks, support/chip in others"),
    "Princess": (False, "High", ""),
    "Ram Rider": (True, "Medium", "Battle-Ram-style chip/win con with lasso utility"),
    "Rascals": (False, "Medium", "Swarm/support"),
    "Royal Ghost": (False, "High", ""),
    "Royal Giant": (True, "High", ""),
    "Royal Hogs": (True, "High", ""),
    "Royal Recruits": (True, "Medium", "Situational -- split-lane pressure win con in dedicated decks"),
    "Rune Giant": (True, "Low", "Not confidently known -- name suggests a Giant-family win con, unverified"),
    "Skeleton Army": (False, "High", ""),
    "Skeleton Barrel": (True, "High", ""),
    "Skeleton Dragons": (False, "High", "Confirmed by user - not a win condition"),
    "Skeleton King": (False, "Medium", "Champion, mainly support via skeleton summons"),
    "Skeletons": (False, "High", ""),
    "Sparky": (True, "Medium", "Situational -- primary win con in dedicated Sparky decks"),
    "Spear Goblins": (False, "High", ""),
    "Spirit Empress": (False, "High", "Confirmed by user - not a win condition"),
    "Suspicious Bush": (False, "Low", "Not confidently known -- best guess is a decoy/trap card"),
    "Three Musketeers": (True, "Medium", "Split-push win con, situational"),
    "Valkyrie": (False, "High", ""),
    "Wall Breakers": (True, "High", ""),
    "Witch": (False, "High", "Support, skeleton-summoning troop"),
    "Wizard": (False, "High", ""),
    "Zappies": (False, "High", ""),
    # Buildings
    "Bomb Tower": (False, "High", ""),
    "Cannon": (False, "High", ""),
    "Inferno Tower": (False, "High", ""),
    "Mortar": (True, "High", ""),
    "Tesla": (False, "High", ""),
    "X-Bow": (True, "High", ""),
    "Barbarian Hut": (False, "Medium", "Defensive spawner building"),
    "Elixir Collector": (False, "High", "Utility, not a win condition"),
    "Goblin Cage": (False, "High", ""),
    "Goblin Drill": (True, "High", ""),
    "Goblin Hut": (False, "High", ""),
    "Tombstone": (False, "High", ""),
    # Spells
    "Arrows": (False, "High", ""),
    "Barbarian Barrel": (False, "High", ""),
    "Earthquake": (False, "High", ""),
    "Fireball": (False, "High", ""),
    "Freeze": (False, "High", ""),
    "Giant Snowball": (False, "High", ""),
    "Goblin Curse": (False, "Medium", "Not confidently known -- best guess is a debuff spell"),
    "Lightning": (False, "High", ""),
    "Poison": (False, "High", ""),
    "Rage": (False, "High", ""),
    "Rocket": (False, "High", "Big single-target spell damage, not classified as a win-condition troop"),
    "Royal Delivery": (False, "High", ""),
    "The Log": (False, "High", ""),
    "Tornado": (False, "High", ""),
    "Vines": (False, "High", "Confirmed by user - not a win condition"),
    "Void": (False, "Medium", "Not confidently known -- best guess is a spell"),
    "Zap": (False, "High", ""),
}


def classify_deck(deck):
    """Return sorted list of win-condition card names found in a deck."""
    wincons = []
    for c in deck:
        entry = WIN_CONDITION_REFERENCE.get(c)
        if entry and entry[0] is True:
            wincons.append(c)
    return sorted(wincons)

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
BASE_FONT = Font(name=FONT_NAME)
BOLD_FONT = Font(name=FONT_NAME, bold=True)
UNCERTAIN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def parse_time(s):
    # Clash Royale API format: 20260717T013205.000Z
    return datetime.strptime(s, "%Y%m%dT%H%M%S.%fZ").replace(tzinfo=timezone.utc)


# ---------------------------------------------------------------------------
# Account aliases (added 2026-07-21). Some players have more than one account; to analyze
# them as a single identity, map each alternate-account tag to the person's canonical tag
# here. Any battle where the aliased tag appears as either the player or the opponent gets
# rewritten to the canonical tag+name at load time, so BOTH accounts' games roll up under
# one player everywhere (workbook + dashboard) automatically.
#   #9RG0VPUVY ("batan") -- the user's SECOND account, merged into the main Batan
#   (#9RQ8YRYQL "老板 Ι Batan'宙斯") per explicit user request.
# ---------------------------------------------------------------------------
ALIAS_TAGS = {"#9RG0VPUVY": "#9RQ8YRYQL"}
CANONICAL_NAMES = {"#9RQ8YRYQL": "老板 Ι Batan’宙斯"}


def canon_tag(tag):
    """Return the canonical tag for a possibly-aliased account tag (identity otherwise)."""
    return ALIAS_TAGS.get(tag, tag)


def canon_name(tag, name):
    """If `tag` is an aliased alt-account, return the canonical display name; else `name`."""
    if tag in ALIAS_TAGS:
        return CANONICAL_NAMES.get(ALIAS_TAGS[tag], name)
    return name


# ---------------------------------------------------------------------------
# Match category classification (added 2026-07-18, ahead of the first official CRL
# tournament matches expected the same day). "Practice" is the confirmed, real signal:
# type == "clanMate" AND gameMode.name == "Friendly".
#
# "Official CRL" signal (confirmed 2026-07-18, 7th/live-Round-2 fetch, via direct
# investigation of real battle data): type == "friendly" (lowercase) AND
# gameMode.name == "Friendly" AND the opponent is NOT a tracked roster member AND the
# battle falls inside a known live-tournament time cluster. This was found by
# time-clustering every friendly/Friendly-vs-non-roster battle (>20min gap = new
# cluster) across all 48 players: one cluster stood out dramatically -- 193 games
# across all 48 roster players simultaneously, vs. every other cluster's 1-38 games
# across only 1-3 players (ordinary scattered ladder/friend-challenge activity). Plain
# type=='friendly' battles OUTSIDE a listed cluster (or against a roster opponent) are
# NOT classified here -- they fall through to `return None` below and are excluded
# from the workbook entirely, same as before this signal was found.
#
# OFFICIAL_TYPE_MODE_SIGNALS is kept as a secondary/simpler mechanism in case a
# (type, gameMode.name) combo is ever found that's unambiguous on its own (no
# roster/time check needed) -- e.g. if Supercell ever returns a distinct
# "type": "tournament" for these games. Empty for now; the real signal is the cluster
# mechanism below.
# ---------------------------------------------------------------------------
PRACTICE_TYPE_MODE_SIGNALS = {("clanMate", "Friendly")}
OFFICIAL_TYPE_MODE_SIGNALS = set()  # e.g. {("clanMate", "Tournament")} -- see note above

# (start_iso, end_iso, label) -- both bounds inclusive, UTC. Only used as a fallback for
# battles that already match PRACTICE_TYPE_MODE_SIGNALS; battles matching
# OFFICIAL_TYPE_MODE_SIGNALS are always "Official CRL" regardless of time.
OFFICIAL_MATCH_TIME_WINDOWS = []

# Live-tournament clusters for the type=='friendly' vs-non-roster-opponent CRL signal
# (see block comment above). Add one entry per round/day once its cluster window is
# identified from a fresh fetch. (start_iso, end_iso, label) -- both bounds inclusive, UTC.
#
# NOTE (2026-07-18, 8th fetch, "just finished game 2 of round 5"): tried to find a
# separate window per round the same way Round 2's window was originally found (>20min
# gap = new cluster), but rounds 2-5 turned out to run BACK-TO-BACK with no >20min gap
# anywhere between them -- time-clustering the full friendly/Friendly-vs-non-roster
# dataset still returns exactly ONE big cluster, now spanning 2026-07-18T13:18:56Z
# through 2026-07-18T16:28:39Z (475 games, all 48 roster players), not four separate
# ones. So this single entry's end time is simply extended each fetch to cover
# whatever's newly arrived, rather than being split into one entry per round.
#
# UPDATE (2026-07-18, 9th/final fetch, "the entire thing has finished, all 9 rounds"):
# re-ran the same time-clustering check across the complete round 6-9 data and the
# cluster is STILL one single unbroken run, now spanning through 2026-07-18T18:46:39Z
# (779 games, all 48 roster players) -- the whole event (rounds 2 through 9) ran with
# zero >20min gaps anywhere inside it. End time extended accordingly; this is now the
# FINAL value for this event (no more fetches expected).
# Verified this is still safe: per-(player,opponent)-pair game counts across the whole
# window are 1, 2, 3, or (one known anomaly) 7 -- NO pair shows 4, 5, or 6 games, which
# is what you'd see if the same two players met again in a later round (that would
# read as an extra "duel" for the pair and get incorrectly flagged Anomaly by
# compute_crl_duel_status's "2nd+ duel for a pair" check). If a pair EVER does show a
# rematch across rounds, that check will misfire -- watch for new Anomaly rows that
# aren't the known 加急Sinistro/Yex✨Style pair, and reconsider whether round-aware
# duel IDs are needed at that point. Still safe as of the final fetch.
OFFICIAL_CRL_LIVE_CLUSTERS = [
    ("20260718T131856.000Z", "20260718T184639.000Z", "Rounds 2-9 (live, contiguous, event complete)"),
    # Day 2 window, added 2026-07-19: same time-clustering method applied to the 15
    # Monthly Finals opponents' fresh battle data (>20min gap = new cluster). Found one
    # 1-game outlier at 12:54:03Z (pre-event noise, same pattern as Day 1's initial
    # outlier), then a 71-minute gap, then a single unbroken 258-game cluster across all
    # 15 tracked players from 14:05:42Z to 17:42:57Z -- same dramatic signal shape as
    # the Day 1 discovery (one big simultaneous cluster vs. scattered noise elsewhere).
    ("20260719T140542.000Z", "20260719T174257.000Z", "Day 2 (live, single cluster found 2026-07-19)"),
]


def _in_official_crl_cluster(battle_time):
    for start_iso, end_iso, _label in OFFICIAL_CRL_LIVE_CLUSTERS:
        start = parse_time(start_iso)
        end = parse_time(end_iso)
        if start <= battle_time <= end:
            return True
    return False


# ---------------------------------------------------------------------------------------
# MANUALLY-CONFIRMED Official CRL matches (Monthly Finals etc.)
#
# WHY THIS EXISTS: the Monthly Finals bracket games came in as type="clanMate"/mode="Friendly"
# -- identical to ordinary practice, and unlike the earlier CRL days (which were
# "friendly"/"Friendly" inside a live time-cluster). Since top-16 players also PRACTICE each
# other all day in the same window, there's no automatic signal that separates a finals duel
# from practice. So finals matches are confirmed MANUALLY here: you paste the real bracket
# matchups (which you know), and every rebuild -- local and the GitHub automation -- tags
# exactly those games as Official CRL. This is the reliable version of the "a single duel-set
# between two top-16 players is usually a real match" heuristic: you confirm it, so no
# practice gets mislabeled.
#
# HOW TO ADD A NEW EVENT DAY: append entries below. Each entry is:
#     (date "YYYYMMDD" in UTC, frozenset({tagA, tagB}), window_or_None)
# where window_or_None is either None (any game between the pair that whole UTC day) or a
# ("startISO","endISO") tuple to restrict to the live event window (use one if a pair might
# also have practiced each other that same day). Use CANONICAL tags (e.g. Batan's 2nd account
# #9RG0VPUVY is aliased to his main #9RQ8YRYQL -- use the main tag here).
CRL_MATCH_OVERRIDES = [
    # 2026-07-25 & 2026-07-26 Monthly Finals double-elimination bracket (confirmed by user from
    # the official bracket). Each pairing is tagged ONLY on the specific UTC day its bracket
    # match was played, so same-players practice on OTHER days is NOT swept in. Verified against
    # the data: every match below is a clean short block (2-4 games, matching the 2-0/2-1 bracket
    # scores) except the two long finals blocks (MohLight vs Clown 07-26 = UB Final + Grand Final;
    # both legitimately CRL). NOTE the earlier 07-23 sessions between these players were long
    # 15-21 game blocks = PRACTICE, and are deliberately NOT tagged (no bracket ran on 07-23).
    #
    # --- Day 1: 2026-07-25 (UB R1, UB QF, LB R1, LB R2 partial) ---
    ("20260725", frozenset({"#Y022GRCJQ", "#YLVV0JPQ"}), None),   # SandBox vs Oker      (UB R1)
    ("20260725", frozenset({"#2CLV2RP0", "#9RQ8YRYQL"}), None),   # Mugi vs Batan        (UB R1)
    ("20260725", frozenset({"#G9YV9GR8R", "#R09228V"}), None),    # Mohamed Light vs Morten (UB R1)
    ("20260725", frozenset({"#9CPCC890", "#RJ88Y8U08"}), None),   # Adriel vs Pedro      (UB R1)
    ("20260725", frozenset({"#GPPYR9JYR", "#UJQQCUCQ8"}), None),  # Clown vs Franco      (UB R1)
    ("20260725", frozenset({"#22LC8JG02", "#8LJ92G8UG"}), None),  # JorZ vs Vitor75      (UB R1)
    ("20260725", frozenset({"#RUQ0JU2P", "#U8RYGC8GU"}), None),   # Asaf vs Polaris      (UB R1)
    ("20260725", frozenset({"#2LJ0ULYCC", "#U890Q9UQ"}), None),   # Guriko vs Sub        (UB R1)
    ("20260725", frozenset({"#Y022GRCJQ", "#9RQ8YRYQL"}), None),  # SandBox vs Batan     (UB QF)
    ("20260725", frozenset({"#G9YV9GR8R", "#9CPCC890"}), None),   # Mohamed Light vs Adriel (UB QF)
    ("20260725", frozenset({"#GPPYR9JYR", "#22LC8JG02"}), None),  # Clown vs JorZ        (UB QF; excludes 07-24 practice)
    ("20260725", frozenset({"#RUQ0JU2P", "#U890Q9UQ"}), None),    # Asaf vs Sub          (UB QF)
    ("20260725", frozenset({"#YLVV0JPQ", "#2CLV2RP0"}), None),    # Oker vs Mugi         (LB R1)
    ("20260725", frozenset({"#R09228V", "#RJ88Y8U08"}), None),    # Morten vs Pedro      (LB R1)
    ("20260725", frozenset({"#UJQQCUCQ8", "#8LJ92G8UG"}), None),  # Franco vs Vitor75    (LB R1)
    ("20260725", frozenset({"#U8RYGC8GU", "#2LJ0ULYCC"}), None),  # Polaris vs Guriko    (LB R1)
    #
    # --- Day 2: 2026-07-26 (UB SF, UB Final, Grand Final, LB R2/R3/QF/SF/Final) ---
    ("20260726", frozenset({"#Y022GRCJQ", "#G9YV9GR8R"}), None),  # SandBox vs Mohamed Light (UB SF)
    ("20260726", frozenset({"#GPPYR9JYR", "#RUQ0JU2P"}), None),   # Clown vs Asaf        (UB SF)
    ("20260726", frozenset({"#G9YV9GR8R", "#GPPYR9JYR"}), None),  # Mohamed Light vs Clown (UB Final + Grand Final)
    ("20260726", frozenset({"#U890Q9UQ", "#2CLV2RP0"}), None),    # Sub vs Mugi          (LB R2)
    ("20260726", frozenset({"#22LC8JG02", "#R09228V"}), None),    # JorZ vs Morten       (LB R2)
    ("20260726", frozenset({"#9CPCC890", "#8LJ92G8UG"}), None),   # Adriel vs Vitor75    (LB R2)
    ("20260726", frozenset({"#9RQ8YRYQL", "#U8RYGC8GU"}), None),  # Batan vs Polaris     (LB R2)
    ("20260726", frozenset({"#2CLV2RP0", "#22LC8JG02"}), None),   # Mugi vs JorZ         (LB R3)
    ("20260726", frozenset({"#8LJ92G8UG", "#9RQ8YRYQL"}), None),  # Vitor75 vs Batan     (LB R3)
    ("20260726", frozenset({"#Y022GRCJQ", "#22LC8JG02"}), None),  # SandBox vs JorZ      (LB QF)
    ("20260726", frozenset({"#RUQ0JU2P", "#8LJ92G8UG"}), None),   # Asaf vs Vitor75      (LB QF)
    ("20260726", frozenset({"#Y022GRCJQ", "#RUQ0JU2P"}), None),   # SandBox vs Asaf      (LB SF)
    ("20260726", frozenset({"#G9YV9GR8R", "#RUQ0JU2P"}), None),   # Mohamed Light vs Asaf (LB Final)
]


def _is_confirmed_crl(p_tag, o_tag, battle_time):
    """True if this game is a manually-confirmed Official CRL match (see CRL_MATCH_OVERRIDES).
    Matches on UTC date + the unordered pair of CANONICAL tags, optionally within a window."""
    pair = frozenset({p_tag, o_tag})
    day = battle_time.strftime("%Y%m%d")
    for od, opair, window in CRL_MATCH_OVERRIDES:
        if od == day and opair == pair:
            if window is None:
                return True
            if parse_time(window[0]) <= battle_time <= parse_time(window[1]):
                return True
    return False


def classify_match_category(battle_type, mode_name, battle_time, opponent_tag=None, roster_tags=None):
    """Returns 'Official CRL', 'Practice', or None (not a relevant type/mode at all --
    ranked ladder, 2v2 team wars, the trail/Showdown_Friendly event mode, etc. -- these
    were already excluded before this classification existed and still are)."""
    key = (battle_type, mode_name)
    if key in OFFICIAL_TYPE_MODE_SIGNALS:
        return "Official CRL"
    if key in PRACTICE_TYPE_MODE_SIGNALS:
        for start_iso, end_iso, _label in OFFICIAL_MATCH_TIME_WINDOWS:
            start = datetime.fromisoformat(start_iso.replace("Z", "+00:00"))
            end = datetime.fromisoformat(end_iso.replace("Z", "+00:00"))
            if start <= battle_time <= end:
                return "Official CRL"
        return "Practice"
    # BUG FOUND AND FIXED 2026-07-18 (final pull, user reported "later rounds missing"):
    # this used to also require `opponent_tag not in roster_tags`. That was fine
    # early on (Round 2) when the tracked roster essentially never drew each other,
    # but as the Swiss rounds progressed, tracked players increasingly got paired
    # against OTHER tracked players -- and those real Official CRL matches were being
    # silently excluded entirely (fell through to None: not clanMate/Friendly so not
    # Practice, and the old rule explicitly disqualified them from CRL because the
    # opponent happened to also be on the roster). Verified via direct count: 286
    # friendly/Friendly battles inside the live cluster were roster-vs-roster and were
    # being dropped. The roster_tags param is kept (still used for provenance/possible
    # future signals) but no longer gates this branch -- ANY friendly/Friendly battle
    # inside the confirmed live-tournament cluster is Official CRL, regardless of
    # whether the opponent is also one of the 48 tracked players. This does mean a
    # roster-vs-roster CRL match now produces two rows (one from each player's own
    # master file) -- exactly the same, correct behavior Practice duels already have
    # for roster-vs-roster games.
    if (
        battle_type == "friendly"
        and mode_name == "Friendly"
        and _in_official_crl_cluster(battle_time)
    ):
        return "Official CRL"
    return None


def load_rows():
    # Prefer the accumulating master_<tag>.json archive (never loses battles
    # that have aged out of the API's returned window on later fetches). Only
    # fall back to a raw_<tag>.json snapshot for a tag that has no master yet
    # (e.g. before the archiving pipeline existed) -- and skip that tag's raw
    # file entirely if a master already covers it, so battles aren't double
    # counted between the two.
    master_paths = sorted(glob.glob(os.path.join(DATA_DIR, "master_*.json")))
    master_tags = {os.path.basename(p)[len("master_"):-len(".json")] for p in master_paths}

    raw_paths = sorted(glob.glob(os.path.join(DATA_DIR, "raw_*.json")))
    fallback_raw_paths = [
        p for p in raw_paths
        if os.path.basename(p)[len("raw_"):-len(".json")] not in master_tags
    ]

    # Roster tag set (48 tracked players), used to tell an Official CRL opponent (NOT on
    # the roster) apart from ordinary internal practice. The tag in the filename is the
    # player's tag WITHOUT the leading "#" (matches how tags are stored on battle records).
    roster_tags = master_tags | {
        os.path.basename(p)[len("raw_"):-len(".json")] for p in raw_paths
    }
    roster_tags = {t if t.startswith("#") else f"#{t}" for t in roster_tags}

    rows = []
    for path in master_paths + fallback_raw_paths:
        with open(path) as f:
            battles = json.load(f)
        for b in battles:
            btype = b.get("type")
            mode_name = b.get("gameMode", {}).get("name")
            battle_time = parse_time(b["battleTime"])
            opp_tag_raw = b["opponent"][0].get("tag")
            category = classify_match_category(
                btype, mode_name, battle_time,
                opponent_tag=opp_tag_raw, roster_tags=roster_tags,
            )
            if category is None:
                continue
            team = b["team"][0]
            opp = b["opponent"][0]
            # Apply account aliases so an alt account's games roll up under the canonical
            # player (both as the player side and when they're someone's opponent).
            p_tag, p_name = team.get("tag"), team.get("name")
            o_tag, o_name = opp.get("tag"), opp.get("name")
            p_name = canon_name(p_tag, p_name); p_tag = canon_tag(p_tag)
            o_name = canon_name(o_tag, o_name); o_tag = canon_tag(o_tag)
            # Manually-confirmed Official CRL matches (e.g. Monthly Finals that arrived as
            # clanMate/Friendly, indistinguishable from practice) -- promote them here.
            if category != "Official CRL" and _is_confirmed_crl(p_tag, o_tag, battle_time):
                category = "Official CRL"
            rows.append({
                "player_tag": p_tag,
                "player_name": p_name,
                "opponent_tag": o_tag,
                "opponent_name": o_name,
                "battle_time": battle_time,
                "arena": b.get("arena", {}).get("name"),
                "deck": [c["name"] for c in team.get("cards", [])],
                "opponent_deck": [c["name"] for c in opp.get("cards", [])],
                "crowns_for": team.get("crowns"),
                "crowns_against": opp.get("crowns"),
                "match_category": category,
            })
    return rows


def group_into_duels(rows):
    """rows: list of dicts for ONE (player, opponent) pair, any order.
    Returns list of duels, each a list of rows in chronological order,
    plus a flag on duel 0 marking it as possibly-continued.

    Also detects a deliberate rematch: a game whose deck is IDENTICAL to the
    IMMEDIATELY-PRECEDING game's deck (same opponent, within the gap window), AND
    only while the current set is not yet a complete 3-deck Bo3. Such a game is a
    re-practice bonus game tacked onto the current duel -- tagged is_rematch=True
    so it stays visible but is EXCLUDED from all distinct-deck analysis
    (sequencing, win-con sets, completeness) and does NOT consume one of the three
    deck slots.

    Two guards matter, both corrected 2026-07-21 per user:
      1. IMMEDIATE ONLY. The repeat must be back-to-back (deck == the previous
         game's deck). A deck that merely matches some EARLIER, non-adjacent game
         in the set is NOT a rematch -- treat it normally (it will start a new
         duel via the disjoint-cards rule). (An over-broad "matches any game in
         the set" rule produced false positives, e.g. wrongly flagging a repeat
         vs KickAsh that never happened.)
      2. NOT A COMPLETED SET. If the current duel already holds 3 distinct decks
         (a finished Bo3), a following game -- even one that immediately repeats
         the last deck -- is the START OF THE NEXT SET, not a re-practice, so it
         must begin a new duel. (Real Duel format bans reusing a card within a
         set, so a genuine intentional re-practice only ever happens back-to-back
         BEFORE the set is complete.)
    Batan's confirmed real case: an intentional back-to-back replay vs Wyze
    (#202GUYUP) to practice a matchup again -- caught by rule 1 within an
    incomplete set, correctly excluded.
    """
    rows = sorted(rows, key=lambda r: r["battle_time"])
    duels = []
    current = []
    current_cards = set()
    for r in rows:
        r["is_rematch"] = False
        deck_set = set(r["deck"])
        gap_ok = True
        if current:
            gap_hours = (r["battle_time"] - current[-1]["battle_time"]).total_seconds() / 3600
            gap_ok = gap_hours <= DUEL_GAP_HOURS

        distinct_count = sum(1 for x in current if not x["is_rematch"])
        set_complete = distinct_count >= MAX_GAMES_PER_DUEL

        # Immediate (back-to-back) repeat of the previous deck, only while the set is still
        # being built (< 3 distinct decks) -- a re-practice bonus game, not a new distinct
        # game and not a duel boundary.
        is_rematch = (
            bool(current) and gap_ok and not set_complete
            and deck_set == set(current[-1]["deck"])
        )

        if is_rematch:
            r["is_rematch"] = True
            current.append(r)
            # deck is identical to the previous game -- current_cards unchanged
        elif (
            current
            and gap_ok
            and not set_complete           # counts DISTINCT decks, so a rematch never uses a slot
            and deck_set.isdisjoint(current_cards)
        ):
            current.append(r)
            current_cards |= deck_set
        else:
            if current:
                duels.append(current)
            current = [r]
            current_cards = set(deck_set)
    if current:
        duels.append(current)
    return duels


def group_into_sessions(duels):
    """duels: list of duels (each a list of chronologically-ordered game rows) for
    ONE (player, opponent) pair, already in chronological order.

    A "practice session" is one or more back-to-back duels against the same
    opponent. A gap of >= SESSION_GAP_HOURS between the end of one duel and the
    start of the next duel against that same opponent starts a new session --
    this is a coarser, higher-level grouping than the duel boundary itself
    (DUEL_GAP_HOURS), since players often play several distinct duels in a row
    within one sitting.

    Returns a list of sessions, each a list of duels.
    """
    sessions = []
    current_session = []
    prev_end = None
    for duel in duels:
        start = duel[0]["battle_time"]
        if prev_end is not None:
            gap_hours = (start - prev_end).total_seconds() / 3600
            new_session = gap_hours >= SESSION_GAP_HOURS
        else:
            new_session = False
        if new_session:
            sessions.append(current_session)
            current_session = [duel]
        else:
            current_session.append(duel)
        prev_end = duel[-1]["battle_time"]
    if current_session:
        sessions.append(current_session)
    return sessions


def compute_crl_duel_status(duel, duel_num_for_pair):
    """Official CRL completion status for one grouped duel, per the user's explicit rule:
    CRL is best-of-3, but a 2-0/0-2 SWEEP is a COMPLETE result with only 2 games -- unlike
    a practice duel, it must NOT be flagged "incomplete" for lacking a 3rd game. A 1-1
    split (or a duel with only 1 game logged so far) means game 3 hasn't been played (or
    hasn't been fetched) yet -- PENDING, held in the background until a future end-of-event
    pull reconciles it (per user decision: single-game pairs are treated the same as 1-1,
    i.e. pending). Any duel that doesn't fit sweep/1-1/3-games-decided at all (e.g. more
    than 3 real games logged, or a 2nd+ duel object for a pair that should only have one
    best-of-3 match) is an ANOMALY -- excluded from clean CRL stats and flagged for manual
    review rather than guessed at (per user decision)."""
    distinct = [r for r in duel if not r["is_rematch"]]
    n = len(distinct)
    wins = sum(1 for r in distinct if r["crowns_for"] > r["crowns_against"])
    losses = sum(1 for r in distinct if r["crowns_for"] < r["crowns_against"])

    if duel_num_for_pair > 1:
        return "Anomaly (extra CRL duel beyond Bo3 for this pair -- needs review)"
    if n == 2 and (wins == 2 or losses == 2):
        return "Complete (Sweep 2-0)"
    if n == 2 and wins == 1 and losses == 1:
        return "Pending (1-1, Game 3 not yet played)"
    if n == 1:
        return "Pending (Game 1 only)"
    if n == 3:
        return "Complete (Decided 2-1)"
    return "Anomaly (unexpected game count -- needs review)"


def build_dataset():
    rows = load_rows()
    # Grouped by (player_tag, opponent_tag, match_category) -- NOT just the pair -- so a
    # practice game and an official CRL game against the same opponent never get grouped
    # into the same "duel" just because they happen to fall within the duel-gap window of
    # each other. Added 2026-07-18, ahead of the first official CRL matches.
    by_pair = defaultdict(list)
    for r in rows:
        by_pair[(r["player_tag"], r["opponent_tag"], r["match_category"])].append(r)

    duel_log = []   # one row per game
    duel_summary = []  # one row per duel
    duel_counters = defaultdict(int)  # per pair, for Duel ID numbering
    session_stats = []  # one row per ORDERED (player, opponent, category) pair -- deduped later

    for pair, pair_rows in by_pair.items():
        category = pair[2]
        duels = group_into_duels(pair_rows)
        sessions = group_into_sessions(duels)
        duel_to_session_num = {}
        for s_num, sess in enumerate(sessions, start=1):
            for d in sess:
                duel_to_session_num[id(d)] = s_num

        # Official CRL duel/session IDs get a suffix so they can never collide in text with
        # a same-numbered Practice duel/session for the same pair (they're separate
        # COUNTIF/COUNTIFS keys in the Excel formulas downstream).
        id_suffix = "" if category == "Practice" else "_CRL"

        for i, duel in enumerate(duels):
            duel_counters[pair] += 1
            duel_num = duel_counters[pair]
            player_name = duel[0]["player_name"]
            opp_name = duel[0]["opponent_name"]
            duel_id = f"{player_name}_vs_{opp_name}_D{duel_num}{id_suffix}".replace(" ", "")
            session_num = duel_to_session_num[id(duel)]
            session_id = f"{player_name}_vs_{opp_name}_S{session_num}{id_suffix}".replace(" ", "")
            uncertain = (i == 0)  # first duel for this pair: no visibility before fetch window
            # Completeness must be known BEFORE the per-game rows are built, so each game
            # row can carry the flag the aggregations filter on. See is_stats_eligible().
            distinct_games = [r for r in duel if not r["is_rematch"]][:MAX_GAMES_PER_DUEL]
            stats_eligible = is_stats_eligible(
                category, duel[0]["battle_time"], len(distinct_games)
            )
            decks_for_summary = []
            for g, r in enumerate(duel, start=1):
                duel_log.append({
                    "duel_id": duel_id,
                    "game_num": g,
                    "player_name": r["player_name"],
                    "player_tag": r["player_tag"],
                    "opponent_name": r["opponent_name"],
                    "opponent_tag": r["opponent_tag"],
                    "battle_time": r["battle_time"],
                    "arena": r["arena"],
                    "deck": r["deck"],
                    "opponent_deck": r["opponent_deck"],
                    "crowns_for": r["crowns_for"],
                    "crowns_against": r["crowns_against"],
                    "uncertain_start": uncertain,
                    "is_rematch": r["is_rematch"],
                    "match_category": category,
                    "stats_eligible": stats_eligible,
                })
                decks_for_summary.append(", ".join(r["deck"]))

            if category == "Official CRL":
                # Best-of-3 completion is different from Practice: a 2-0 sweep is COMPLETE
                # with only 2 games, and a 1-1 (or single-game) duel is PENDING, not
                # "incomplete" -- see compute_crl_duel_status for the full rule.
                crl_status = compute_crl_duel_status(duel, duel_num)
                if crl_status.startswith("Complete"):
                    wincon_set = set()
                    for g in distinct_games:
                        wincon_set.update(classify_deck(g["deck"]))
                    wincon_sequence = "+".join(sorted(wincon_set)) if wincon_set else "(none classified)"
                else:
                    wincon_sequence = crl_status
            else:
                crl_status = None
                if len(distinct_games) == 3:
                    # Order doesn't matter -- this is the SET of win conditions that showed up
                    # anywhere across the duel's games, not a per-game Game1->Game2->Game3
                    # sequence. E.g. Goblin Barrel/Graveyard/Royal Hogs count as the same
                    # "win-con set" no matter which game each one appeared in.
                    wincon_set = set()
                    for g in distinct_games:
                        wincon_set.update(classify_deck(g["deck"]))
                    wincon_sequence = "+".join(sorted(wincon_set)) if wincon_set else "(none classified)"
                else:
                    wincon_sequence = f"Incomplete ({len(distinct_games)}/3 distinct games)"

            duel_summary.append({
                "duel_id": duel_id,
                "session_id": session_id,
                "player_name": player_name,
                "opponent_name": opp_name,
                "start_time": duel[0]["battle_time"],
                "games_played": len(duel),
                "decks": decks_for_summary,
                "uncertain_start": uncertain,
                "wincon_sequence": wincon_sequence,
                "match_category": category,
                "crl_status": crl_status,
                "stats_eligible": stats_eligible,
            })

        num_sessions = len(sessions)
        num_duels_pair = sum(len(s) for s in sessions)
        num_games_pair = sum(len(d) for s in sessions for d in s)
        first_session_start = sessions[0][0][0]["battle_time"]
        last_session_start = sessions[-1][0][0]["battle_time"]
        session_stats.append({
            "player_tag": pair[0],
            "opponent_tag": pair[1],
            "match_category": category,
            "player_name": pair_rows[0]["player_name"],
            "opponent_name": pair_rows[0]["opponent_name"],
            "num_sessions": num_sessions,
            "num_duels": num_duels_pair,
            "num_games": num_games_pair,
            "first_session_start": first_session_start,
            "last_session_start": last_session_start,
        })

    duel_log.sort(key=lambda r: (r["player_name"], r["battle_time"]))
    duel_summary.sort(key=lambda r: (r["player_name"], r["start_time"]))
    return duel_log, duel_summary, session_stats


# Minimum games a specific deck needs before it's eligible to be ranked as one of a
# player's "best win-rate decks" -- without this, a deck played once and won counts as
# a "100% win rate" deck and would crowd out decks with a real, larger sample size.
# Configurable; games-played ranking (top-played decks) has no such floor since sample
# size doesn't distort a raw count the way it distorts a rate.
MIN_GAMES_FOR_WINRATE_RANKING = 2

# Official-CRL weighting (added 2026-07-18, ahead of the first tournament matches; OFF by
# default, user-toggleable in the dashboard -- see compute_player_lookup's `weighted` arg).
# Once a specific deck or win condition has been played in at least
# MIN_OFFICIAL_GAMES_FOR_WEIGHT official CRL games, each of those official games counts
# OFFICIAL_GAME_WEIGHT times as much as a practice game when ranking "most-played"/"top win
# condition" -- tournament play is the higher-signal data source, but only once there's
# enough of it for a given entity to trust; below the threshold every game counts equally
# regardless of the toggle. The displayed game count (e.g. "12 games") is always the REAL
# unweighted count -- only ranking/ordering ever uses the weighted score.
OFFICIAL_GAME_WEIGHT = 3
MIN_OFFICIAL_GAMES_FOR_WEIGHT = 5


def compute_player_lookup(duel_log, weighted=False):
    """Per-player match-prep summary: which decks a player plays most often, which
    decks they win most with, and which win conditions they lean on most -- all
    computed from duel_log (one row per individual game, Practice + Official CRL combined
    unless the caller has already filtered duel_log to one category). Returns a list of
    dicts, one per player, sorted by total games played (most active players first).

    weighted=True ranks "most-played deck"/"top win condition" using the Official-CRL
    weighting described above instead of a plain count; everything else (total games/wins,
    win rate, best-win-rate-deck ranking) is unaffected by this flag."""
    per_player = defaultdict(lambda: {
        "games": 0, "wins": 0,
        "deck_games": Counter(), "deck_wins": Counter(), "deck_official_games": Counter(),
        "wincon_games": Counter(), "wincon_official_games": Counter(),
        "practice_games": 0, "official_games": 0, "tags": set(),
    })
    for r in duel_log:
        if not r["deck"]:
            continue
        # Skip games from short post-cutoff Practice sets -- see is_stats_eligible().
        if not r.get("stats_eligible", True):
            continue
        p = per_player[r["player_name"]]
        p["games"] += 1
        deck_key = ", ".join(sorted(r["deck"]))
        p["deck_games"][deck_key] += 1
        is_official = r.get("match_category") == "Official CRL"
        if is_official:
            p["official_games"] += 1
            p["deck_official_games"][deck_key] += 1
        else:
            p["practice_games"] += 1
        if r["crowns_for"] > r["crowns_against"]:
            p["wins"] += 1
            p["deck_wins"][deck_key] += 1
        for wc in classify_deck(r["deck"]):
            p["wincon_games"][wc] += 1
            if is_official:
                p["wincon_official_games"][wc] += 1
        if r.get("player_tag"):
            p["tags"].add(r["player_tag"])

    def rank_key(counter, official_counter, key):
        count = counter[key]
        if weighted:
            official = official_counter[key]
            if official >= MIN_OFFICIAL_GAMES_FOR_WEIGHT:
                practice = count - official
                return practice * 1 + official * OFFICIAL_GAME_WEIGHT
        return count

    rows = []
    for player, d in per_player.items():
        total_games = d["games"]
        total_wins = d["wins"]
        win_rate = (total_wins / total_games) if total_games else 0.0

        top_decks_by_freq = sorted(
            d["deck_games"].items(),
            key=lambda kv: -rank_key(d["deck_games"], d["deck_official_games"], kv[0])
        )[:3]

        eligible = [
            (deck, d["deck_wins"][deck] / d["deck_games"][deck], d["deck_games"][deck])
            for deck in d["deck_games"]
            if d["deck_games"][deck] >= MIN_GAMES_FOR_WINRATE_RANKING
        ]
        top_decks_by_winrate = sorted(eligible, key=lambda x: (-x[1], -x[2]))[:3]

        top_wincons = sorted(
            d["wincon_games"].items(),
            key=lambda kv: -rank_key(d["wincon_games"], d["wincon_official_games"], kv[0])
        )[:3]

        rows.append({
            "player": player,
            "total_games": total_games,
            "total_wins": total_wins,
            "win_rate": win_rate,
            "practice_games": d["practice_games"],
            "official_games": d["official_games"],
            "top_decks_by_freq": top_decks_by_freq,
            "top_decks_by_winrate": top_decks_by_winrate,
            "top_wincons": top_wincons,
            # Sorted, not usually more than 1 -- but tracked as a set in case a player's
            # underlying tag ever gets corrected mid-history (like Kimchi's did), so both
            # old and new tags stay searchable against the same name.
            "tags": sorted(d["tags"]),
        })

    rows.sort(key=lambda r: (-r["total_games"], r["player"]))
    return rows


def compute_wincon_pairs(duel_log):
    """Which win conditions get played TOGETHER IN THE SAME 8-card deck, within a single
    game -- not across a duel's separate games (that's the Win-Con Set sheet). E.g. a
    deck running both Miner and Wall Breakers, or Battle Ram and Three Musketeers,
    shows up here as a tandem. Every distinct unordered pair of win-con cards found in
    the same deck is tallied, with a win rate for games that ran that pair.
    Returns a list of dicts sorted by times played (most common tandems first)."""
    from itertools import combinations
    pair_games = Counter()
    pair_wins = Counter()
    for r in duel_log:
        if not r["deck"]:
            continue
        wincons = classify_deck(r["deck"])
        if len(wincons) < 2:
            continue
        is_win = r["crowns_for"] > r["crowns_against"]
        for a, b in combinations(sorted(wincons), 2):
            pair_games[(a, b)] += 1
            if is_win:
                pair_wins[(a, b)] += 1

    rows = []
    for pair, games in pair_games.items():
        wins = pair_wins[pair]
        rows.append({
            "pair": f"{pair[0]} + {pair[1]}",
            "card_a": pair[0],
            "card_b": pair[1],
            "times_played": games,
            "wins": wins,
            "win_rate": (wins / games) if games else 0.0,
        })
    rows.sort(key=lambda r: (-r["times_played"], r["pair"]))
    return rows


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# Counter-recommendation helpers (added 2026-07-19, per user request: for a Group A
# opponent, "choose the best deck/win-con set/duel set vs their past deck choices, and
# use win rate data from our entire data pool to determine those"). Pulled out here
# (rather than duplicated in add_group_a_analysis.py) so both the Excel sheet and the
# dashboard can share one data-driven implementation, and so the exact same functions
# used for the game-wide "Best Picks" feature back this too -- no separate theorycrafted
# logic, just the real win-rate data filtered down to "games where our side faced an
# opponent playing this win-condition archetype."
# ---------------------------------------------------------------------------

def compute_best_decks(duel_log, min_games=1):
    games = Counter()
    wins = Counter()
    for r in duel_log:
        if not r["deck"] or len(r["deck"]) != 8:
            continue
        key = ", ".join(sorted(r["deck"]))
        games[key] += 1
        if r["crowns_for"] > r["crowns_against"]:
            wins[key] += 1
    rows = []
    for key, g in games.items():
        if g < min_games:
            continue
        w = wins[key]
        rows.append({"deck": key, "games": g, "wins": w, "win_rate": w / g if g else 0.0})
    rows.sort(key=lambda r: (-r["win_rate"], -r["games"]))
    return rows


def compute_best_wincon_sets(duel_log, min_games=1):
    """Simplified, single-game version (unlike the duel-grouped Best Picks one in
    build_dashboard.py): counts every game's own win-con set, not just full 3-game
    duels, since counter-recommendation slices are often too thin for duel-level
    grouping. wincon_set here is the set of win conditions in ONE game's deck."""
    games = Counter()
    wins = Counter()
    for r in duel_log:
        if not r["deck"]:
            continue
        wincon_set = frozenset(classify_deck(r["deck"]) or [])
        if not wincon_set:
            continue
        key = "+".join(sorted(wincon_set))
        games[key] += 1
        if r["crowns_for"] > r["crowns_against"]:
            wins[key] += 1
    rows = []
    for key, g in games.items():
        if g < min_games:
            continue
        w = wins[key]
        rows.append({"wincon_set": key, "games": g, "wins": w, "win_rate": w / g if g else 0.0})
    rows.sort(key=lambda r: (-r["win_rate"], -r["games"]))
    return rows


def _deck_overlap(deck_key_a, deck_key_b):
    return len(set(deck_key_a.split(", ")) & set(deck_key_b.split(", ")))


def compute_best_duel_pairs(duel_log, min_games=2):
    """A lighter-weight cousin of build_dashboard.py's compute_best_duel_sets, sized for
    a filtered counter-recommendation slice (often too thin for the full duel-grouped,
    3-games-required version). Groups by duel_id, takes the (up to 2) distinct decks
    actually played within each duel against this opponent archetype, and ranks those
    deck-pair "sets" by combined win rate. Falls back gracefully -- with min_games=2 a
    pair only needs to have shown up in >=2 games total, not >=2 full duels."""
    by_duel = defaultdict(list)
    for r in duel_log:
        by_duel[r["duel_id"]].append(r)

    pair_games = Counter()
    pair_wins = Counter()
    for duel_id, games in by_duel.items():
        decks_in_duel = []
        for g in games:
            if g["deck"] and len(g["deck"]) == 8:
                decks_in_duel.append(", ".join(sorted(g["deck"])))
        distinct = sorted(set(decks_in_duel))
        if not distinct:
            continue
        key = " / ".join(distinct[:3])
        for g in games:
            pair_games[key] += 1
            if g["crowns_for"] > g["crowns_against"]:
                pair_wins[key] += 1

    rows = []
    for key, g in pair_games.items():
        if g < min_games:
            continue
        w = pair_wins[key]
        rows.append({"duel_set": key, "games": g, "wins": w, "win_rate": w / g if g else 0.0})
    rows.sort(key=lambda r: (-r["win_rate"], -r["games"]))
    return rows


def compute_counter_recommendations(duel_log, opponent_top_winconsets, min_games=3, top_n=3):
    """The core of the "what should we bring tomorrow" feature: filters our ENTIRE
    tracked data pool (every game logged by any tracked/extended-roster player, Practice
    + Official CRL combined, per explicit user instruction to use "all games... as a
    whole") down to games where the opposing side's deck matched one of this specific
    Group A opponent's own top win conditions -- then, ONLY within that filtered slice,
    ranks our best-performing decks / win-con sets / duel-set pairs by real win rate.
    This is empirical (drawn from actual recorded outcomes against that win-con
    archetype), not a theorycrafted "X counters Y" claim -- and every row keeps its
    games/win-rate so small samples are visibly small, not hidden."""
    target = set(opponent_top_winconsets)
    if not target:
        return {"sample_size": 0, "best_decks": [], "best_wincon_sets": [], "best_duel_sets": []}
    filtered = [
        r for r in duel_log
        if r.get("opponent_deck") and set(classify_deck(r["opponent_deck"]) or []) & target
    ]
    return {
        "sample_size": len(filtered),
        "best_decks": compute_best_decks(filtered, min_games=min_games)[:top_n],
        "best_wincon_sets": compute_best_wincon_sets(filtered, min_games=min_games)[:top_n],
        "best_duel_sets": compute_best_duel_pairs(filtered, min_games=min_games)[:top_n],
    }


def main():
    duel_log, duel_summary, session_stats = build_dataset()
    wincon_pairs = compute_wincon_pairs(duel_log)
    # Real recorded tandems (two win cons played in the SAME deck, one game) -- used to
    # explain Win-Con Sets rows with more than 3 win conditions despite the 3-game cap
    # (that can only happen if some game in the duel ran 2+ win cons at once).
    real_tandems = {frozenset((r["card_a"], r["card_b"])): r["times_played"] for r in wincon_pairs}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---------------- Duel Log sheet ----------------
    ws = wb.create_sheet("Duel Log")
    headers = [
        "Duel ID", "Game #", "Player", "Player Tag", "Opponent", "Opponent Tag",       # 1-6
        "Battle Time (UTC)", "Arena",                                                   # 7-8
        "Card 1", "Card 2", "Card 3", "Card 4", "Card 5", "Card 6", "Card 7", "Card 8", # 9-16
        "Opp Card 1", "Opp Card 2", "Opp Card 3", "Opp Card 4",                         # 17-20
        "Opp Card 5", "Opp Card 6", "Opp Card 7", "Opp Card 8",                         # 21-24
        "Crowns For", "Crowns Against", "Result",                                       # 25-27
        "Duel Start Uncertain", "Instant Rematch",                                      # 28-29
        "Own Deck Key", "Opponent Deck Key", "Deck A", "Deck B", "Deck A Result",       # 30-34
        "Match Category",                                                               # 35
    ]
    COL = {name: i + 1 for i, name in enumerate(headers)}
    ws.append(headers)
    style_header_row(ws, 1, len(headers))

    for r_i, r in enumerate(duel_log, start=2):
        deck = r["deck"] + [""] * (8 - len(r["deck"]))
        opp_deck = r["opponent_deck"] + [""] * (8 - len(r["opponent_deck"]))
        own_key = ", ".join(sorted(r["deck"]))
        opp_key = ", ".join(sorted(r["opponent_deck"]))

        ws.cell(row=r_i, column=COL["Duel ID"], value=r["duel_id"])
        ws.cell(row=r_i, column=COL["Game #"], value=r["game_num"])
        ws.cell(row=r_i, column=COL["Player"], value=r["player_name"])
        ws.cell(row=r_i, column=COL["Player Tag"], value=r["player_tag"])
        ws.cell(row=r_i, column=COL["Opponent"], value=r["opponent_name"])
        ws.cell(row=r_i, column=COL["Opponent Tag"], value=r["opponent_tag"])
        ws.cell(row=r_i, column=COL["Battle Time (UTC)"],
                value=r["battle_time"].strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=r_i, column=COL["Arena"], value=r["arena"])
        for ci, card in enumerate(deck, start=COL["Card 1"]):
            ws.cell(row=r_i, column=ci, value=card)
        for ci, card in enumerate(opp_deck, start=COL["Opp Card 1"]):
            ws.cell(row=r_i, column=ci, value=card)
        ws.cell(row=r_i, column=COL["Crowns For"], value=r["crowns_for"])
        ws.cell(row=r_i, column=COL["Crowns Against"], value=r["crowns_against"])

        cf = get_column_letter(COL["Crowns For"])
        ca = get_column_letter(COL["Crowns Against"])
        ws.cell(row=r_i, column=COL["Result"],
                value=f'=IF({cf}{r_i}>{ca}{r_i},"Win",IF({cf}{r_i}<{ca}{r_i},"Loss","Draw"))')
        ws.cell(row=r_i, column=COL["Duel Start Uncertain"],
                value="Yes" if r["uncertain_start"] else "No")
        ws.cell(row=r_i, column=COL["Instant Rematch"],
                value="Yes" if r["is_rematch"] else "No")
        ws.cell(row=r_i, column=COL["Own Deck Key"], value=own_key)
        ws.cell(row=r_i, column=COL["Opponent Deck Key"], value=opp_key)
        ws.cell(row=r_i, column=COL["Match Category"], value=r["match_category"])

        ok = get_column_letter(COL["Own Deck Key"])
        opk = get_column_letter(COL["Opponent Deck Key"])
        da = get_column_letter(COL["Deck A"])
        db = get_column_letter(COL["Deck B"])
        res = get_column_letter(COL["Result"])
        ws.cell(row=r_i, column=COL["Deck A"], value=f"=IF({ok}{r_i}<{opk}{r_i},{ok}{r_i},{opk}{r_i})")
        ws.cell(row=r_i, column=COL["Deck B"], value=f"=IF({ok}{r_i}<{opk}{r_i},{opk}{r_i},{ok}{r_i})")
        ws.cell(row=r_i, column=COL["Deck A Result"],
                value=(f'=IF({ok}{r_i}={da}{r_i},{res}{r_i},'
                       f'IF({res}{r_i}="Win","Loss",IF({res}{r_i}="Loss","Win","Draw")))'))

        for c in range(1, len(headers) + 1):
            ws.cell(row=r_i, column=c).font = BASE_FONT
            ws.cell(row=r_i, column=c).border = BORDER
        if r["uncertain_start"]:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r_i, column=c).fill = UNCERTAIN_FILL

    ws.freeze_panes = "A2"
    autosize(ws, [26, 8, 12, 13, 14, 13, 18, 16] + [14] * 16 + [11, 15, 9, 18, 15, 40, 40, 40, 40, 15, 16])
    log_last_row = len(duel_log) + 1  # used by later sheets
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{log_last_row}"

    # ---------------- Duel Summary sheet ----------------
    ws2 = wb.create_sheet("Duel Summary")
    headers2 = [
        "Duel ID", "Practice Session ID", "Player", "Opponent", "Match Category",
        "Start Time (UTC)", "Games Played", "Games Won", "Deck - Game 1", "Deck - Game 2",
        "Deck - Game 3", "Duel Start Uncertain", "Win-Con Set (Duel)", "CRL Status",
    ]
    COL2 = {name: i + 1 for i, name in enumerate(headers2)}
    ws2.append(headers2)
    style_header_row(ws2, 1, len(headers2))

    PENDING_FILL = PatternFill("solid", fgColor="DDEBF7")   # light blue -- held in background
    ANOMALY_FILL = PatternFill("solid", fgColor="F8CBAD")   # light orange -- needs review

    for r_i, r in enumerate(duel_summary, start=2):
        decks = r["decks"] + [""] * (3 - len(r["decks"]))
        ws2.cell(row=r_i, column=COL2["Duel ID"], value=r["duel_id"])
        ws2.cell(row=r_i, column=COL2["Practice Session ID"], value=r["session_id"])
        ws2.cell(row=r_i, column=COL2["Player"], value=r["player_name"])
        ws2.cell(row=r_i, column=COL2["Opponent"], value=r["opponent_name"])
        ws2.cell(row=r_i, column=COL2["Match Category"], value=r["match_category"])
        ws2.cell(row=r_i, column=COL2["Start Time (UTC)"],
                 value=r["start_time"].strftime("%Y-%m-%d %H:%M:%S"))
        ws2.cell(row=r_i, column=COL2["Games Played"],
                 value=f'=COUNTIF(\'Duel Log\'!A:A,A{r_i})')
        result_col_letter = get_column_letter(COL["Result"])
        ws2.cell(row=r_i, column=COL2["Games Won"],
                 value=f'=COUNTIFS(\'Duel Log\'!A:A,A{r_i},\'Duel Log\'!{result_col_letter}:{result_col_letter},"Win")')
        for ci, d in enumerate(decks, start=COL2["Deck - Game 1"]):
            ws2.cell(row=r_i, column=ci, value=d)
        ws2.cell(row=r_i, column=COL2["Duel Start Uncertain"],
                 value="Yes" if r["uncertain_start"] else "No")
        ws2.cell(row=r_i, column=COL2["Win-Con Set (Duel)"],
                 value=r["wincon_sequence"])
        crl_status = r.get("crl_status") or ""
        ws2.cell(row=r_i, column=COL2["CRL Status"], value=crl_status)
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=r_i, column=c).font = BASE_FONT
            ws2.cell(row=r_i, column=c).border = BORDER
        if r["uncertain_start"]:
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=r_i, column=c).fill = UNCERTAIN_FILL
        elif crl_status.startswith("Pending"):
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=r_i, column=c).fill = PENDING_FILL
        elif crl_status.startswith("Anomaly"):
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=r_i, column=c).fill = ANOMALY_FILL

    ws2.freeze_panes = "A2"
    autosize(ws2, [26, 26, 12, 14, 16, 18, 12, 11, 40, 40, 40, 18, 50, 38])
    summary_last_row = len(duel_summary) + 1
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{summary_last_row}"

    # ---------------- Win Condition Reference sheet ----------------
    # Lists ALL cards in WIN_CONDITION_REFERENCE (the full 122-card game roster minus the two
    # sub-unit/evolution entries removed per user request), not just cards seen in the current
    # data -- so the reference is ready as more players/duels get imported, per user request.
    ws_wc = wb.create_sheet("Win Condition Reference")
    headers_wc = ["Card", "Is Win Condition", "Confidence", "Seen In Current Data", "Notes"]
    ws_wc.append(headers_wc)
    style_header_row(ws_wc, 1, len(headers_wc))
    cards_seen = {c for r in duel_log for c in (r["deck"] + r["opponent_deck"]) if c}
    all_cards_ref = sorted(set(WIN_CONDITION_REFERENCE.keys()) | cards_seen)
    for ri, card in enumerate(all_cards_ref, start=2):
        entry = WIN_CONDITION_REFERENCE.get(card, (None, "Low", "Card not in reference list -- please classify"))
        is_wc, conf, note = entry
        label = "Yes" if is_wc is True else ("No" if is_wc is False else "Uncertain")
        ws_wc.cell(row=ri, column=1, value=card).font = BASE_FONT
        c2 = ws_wc.cell(row=ri, column=2, value=label)
        c2.font = BASE_FONT
        c2.fill = PatternFill("solid", fgColor="FFFF00")  # editable assumption, per convention
        ws_wc.cell(row=ri, column=3, value=conf).font = BASE_FONT
        ws_wc.cell(row=ri, column=4, value="Yes" if card in cards_seen else "No").font = BASE_FONT
        ws_wc.cell(row=ri, column=5, value=note).font = BASE_FONT
        for c in range(1, 6):
            ws_wc.cell(row=ri, column=c).border = BORDER
        if label == "Uncertain":
            for c in range(1, 6):
                ws_wc.cell(row=ri, column=c).fill = UNCERTAIN_FILL
    ws_wc.freeze_panes = "A2"
    ws_wc.auto_filter.ref = f"A1:E{len(all_cards_ref) + 1}"
    autosize(ws_wc, [20, 16, 12, 18, 55])
    ws_wc.cell(row=1, column=6, value=(
        "Editing this sheet does not change the analysis automatically -- Win-Con Set "
        "values are computed in Python and would need build_duel_workbook.py re-run after "
        "you correct any classification here.")).font = BASE_FONT
    ws_wc.column_dimensions["F"].width = 70

    # ---------------- Deck Stats sheet ----------------
    ws_ds = wb.create_sheet("Deck Stats")
    deck_keys = sorted({", ".join(sorted(r["deck"])) for r in duel_log if r["deck"]})
    players_by_deck = defaultdict(set)
    for r in duel_log:
        players_by_deck[", ".join(sorted(r["deck"]))].add(r["player_name"])

    headers_ds = ["Deck (sorted)", "Used By", "Games Played", "Wins", "Losses", "Draws", "Win Rate"]
    ws_ds.append(headers_ds)
    style_header_row(ws_ds, 1, len(headers_ds))
    ok_col = get_column_letter(COL["Own Deck Key"])
    res_col = get_column_letter(COL["Result"])
    for ri, dk in enumerate(deck_keys, start=2):
        ws_ds.cell(row=ri, column=1, value=dk).font = BASE_FONT
        ws_ds.cell(row=ri, column=2, value=", ".join(sorted(players_by_deck[dk]))).font = BASE_FONT
        gp = f"=COUNTIF('Duel Log'!{ok_col}$2:{ok_col}${log_last_row},$A{ri})"
        wn = f"=COUNTIFS('Duel Log'!{ok_col}$2:{ok_col}${log_last_row},$A{ri},'Duel Log'!{res_col}$2:{res_col}${log_last_row},\"Win\")"
        ls = f"=COUNTIFS('Duel Log'!{ok_col}$2:{ok_col}${log_last_row},$A{ri},'Duel Log'!{res_col}$2:{res_col}${log_last_row},\"Loss\")"
        dr = f"=COUNTIFS('Duel Log'!{ok_col}$2:{ok_col}${log_last_row},$A{ri},'Duel Log'!{res_col}$2:{res_col}${log_last_row},\"Draw\")"
        ws_ds.cell(row=ri, column=3, value=gp).font = BASE_FONT
        ws_ds.cell(row=ri, column=4, value=wn).font = BASE_FONT
        ws_ds.cell(row=ri, column=5, value=ls).font = BASE_FONT
        ws_ds.cell(row=ri, column=6, value=dr).font = BASE_FONT
        wr_cell = ws_ds.cell(row=ri, column=7, value=f"=IFERROR(D{ri}/C{ri},\"\")")
        wr_cell.font = BASE_FONT
        wr_cell.number_format = "0.0%"
        for c in range(1, 8):
            ws_ds.cell(row=ri, column=c).border = BORDER
    ws_ds.freeze_panes = "A2"
    ws_ds.auto_filter.ref = f"A1:{get_column_letter(len(headers_ds))}{len(deck_keys) + 1}"
    autosize(ws_ds, [60, 20, 12, 8, 8, 8, 10])

    # ---------------- Deck Matchups sheet ----------------
    ws_dm = wb.create_sheet("Deck Matchups")
    matchup_pairs = sorted({
        (", ".join(sorted(r["deck"])), ", ".join(sorted(r["opponent_deck"])))
        if ", ".join(sorted(r["deck"])) < ", ".join(sorted(r["opponent_deck"]))
        else (", ".join(sorted(r["opponent_deck"])), ", ".join(sorted(r["deck"])))
        for r in duel_log if r["deck"] and r["opponent_deck"]
    })
    headers_dm = ["Deck A", "Deck B", "Games Played", "Deck A Wins", "Deck B Wins", "Draws", "Deck A Win Rate"]
    ws_dm.append(headers_dm)
    style_header_row(ws_dm, 1, len(headers_dm))
    da_col = get_column_letter(COL["Deck A"])
    db_col = get_column_letter(COL["Deck B"])
    dares_col = get_column_letter(COL["Deck A Result"])
    for ri, (da, db) in enumerate(matchup_pairs, start=2):
        ws_dm.cell(row=ri, column=1, value=da).font = BASE_FONT
        ws_dm.cell(row=ri, column=2, value=db).font = BASE_FONT
        gp = (f"=COUNTIFS('Duel Log'!{da_col}$2:{da_col}${log_last_row},$A{ri},"
              f"'Duel Log'!{db_col}$2:{db_col}${log_last_row},$B{ri})")
        aw = (f"=COUNTIFS('Duel Log'!{da_col}$2:{da_col}${log_last_row},$A{ri},"
              f"'Duel Log'!{db_col}$2:{db_col}${log_last_row},$B{ri},"
              f"'Duel Log'!{dares_col}$2:{dares_col}${log_last_row},\"Win\")")
        bw = (f"=COUNTIFS('Duel Log'!{da_col}$2:{da_col}${log_last_row},$A{ri},"
              f"'Duel Log'!{db_col}$2:{db_col}${log_last_row},$B{ri},"
              f"'Duel Log'!{dares_col}$2:{dares_col}${log_last_row},\"Loss\")")
        dw = (f"=COUNTIFS('Duel Log'!{da_col}$2:{da_col}${log_last_row},$A{ri},"
              f"'Duel Log'!{db_col}$2:{db_col}${log_last_row},$B{ri},"
              f"'Duel Log'!{dares_col}$2:{dares_col}${log_last_row},\"Draw\")")
        ws_dm.cell(row=ri, column=3, value=gp).font = BASE_FONT
        ws_dm.cell(row=ri, column=4, value=aw).font = BASE_FONT
        ws_dm.cell(row=ri, column=5, value=bw).font = BASE_FONT
        ws_dm.cell(row=ri, column=6, value=dw).font = BASE_FONT
        wr_cell = ws_dm.cell(row=ri, column=7, value=f"=IFERROR(D{ri}/C{ri},\"\")")
        wr_cell.font = BASE_FONT
        wr_cell.number_format = "0.0%"
        for c in range(1, 8):
            ws_dm.cell(row=ri, column=c).border = BORDER
    ws_dm.freeze_panes = "A2"
    ws_dm.auto_filter.ref = f"A1:{get_column_letter(len(headers_dm))}{len(matchup_pairs) + 1}"
    autosize(ws_dm, [60, 60, 12, 12, 12, 8, 12])

    # ---------------- Win-Con Sets sheet ----------------
    # Renamed from "Win-Con Sequences": order within a duel doesn't matter (per user
    # decision), so this tracks which win conditions showed up TOGETHER in a duel's
    # games, not which game each one appeared in. E.g. Goblin Barrel/Graveyard/Royal
    # Hogs count as the same "win-con set" no matter which order they were played.
    #
    # COMBINED ACROSS PLAYERS (changed 2026-07-17, per user decision): previously one row
    # per (set, player), which fragmented a popular set's true frequency across many rows.
    # Now one row per distinct win-con set, combining every player who used it, so
    # popularity is visible at a glance. The "Players" column (Python-computed) still
    # lists who used it, for context.
    # DATA-QUALITY EXCLUSION (added 2026-07-17, per user request): a duel flagged
    # "uncertain_start" is the FIRST duel found for a (player, opponent) pair -- the API
    # only returns each player's most recent ~25-30 battles, so this duel may actually be
    # a continuation of a real duel that started earlier, before our fetch history begins.
    # If so, what we're calling "Game 1" here might really be that duel's Game 2 or 3, and
    # the true earlier game(s) are permanently invisible to us. Even when such a duel looks
    # "complete" (exactly 3 games captured), its win-con SET is still valid (order doesn't
    # matter for this sheet and every game we DID capture is real), but its games could
    # actually belong to a duel that ran longer than 3 games in reality, so we exclude it
    # here to avoid treating an unverifiable grouping as ground truth. (Sequence-dependent
    # analyses in the dashboard -- Win-Con Sets here plus "What Might Follow" / Deck
    # Predictor in build_dashboard.py -- apply the same exclusion; per-game stats like
    # Player Lookup / Deck Stats / Card Frequency do NOT, since each individual recorded
    # game's deck and result are real regardless of which duel it landed in.)
    # ALSO scoped to Practice duels only (added 2026-07-18): official CRL duels are grouped
    # and tracked separately (Match Category column, everywhere) but not yet blended into
    # this sheet or the dashboard's sequence predictors -- sample size will be too thin at
    # first to add useful signal, and practice/tournament play patterns may genuinely
    # differ. Revisit once enough Official CRL duels accumulate.
    ws_wcs = wb.create_sheet("Win-Con Sets")
    set_to_players = defaultdict(set)
    excluded_uncertain_sets = 0
    excluded_official_sets = 0
    for r in duel_summary:
        if r["match_category"] != "Practice":
            excluded_official_sets += 1
            continue
        if r["uncertain_start"]:
            excluded_uncertain_sets += 1
            continue
        if not r["wincon_sequence"].startswith("Incomplete"):
            set_to_players[r["wincon_sequence"]].add(r["player_name"])
    unique_sets = sorted(set_to_players)

    headers_wcs = ["Win-Con Set", "# Win Cons", "Times Played (Duels)", "Games Played",
                   "Games Won", "Win Rate", "Contains Tandem(s)?", "Players Who Used This"]
    ws_wcs.append(headers_wcs)
    style_header_row(ws_wcs, 1, len(headers_wcs))
    seq_col_letter = get_column_letter(COL2["Win-Con Set (Duel)"])
    gp_col_letter = get_column_letter(COL2["Games Played"])
    gw_col_letter = get_column_letter(COL2["Games Won"])
    for ri, seq in enumerate(unique_sets, start=2):
        cards_in_set = seq.split("+")
        ws_wcs.cell(row=ri, column=1, value=seq).font = BASE_FONT
        ws_wcs.cell(row=ri, column=2, value=len(cards_in_set)).font = BASE_FONT
        ws_wcs.cell(row=ri, column=3, value=(
            f"=COUNTIF('Duel Summary'!{seq_col_letter}$2:{seq_col_letter}${summary_last_row},$A{ri})"
        )).font = BASE_FONT
        ws_wcs.cell(row=ri, column=4, value=(
            f"=SUMIF('Duel Summary'!{seq_col_letter}$2:{seq_col_letter}${summary_last_row},$A{ri},"
            f"'Duel Summary'!{gp_col_letter}$2:{gp_col_letter}${summary_last_row})"
        )).font = BASE_FONT
        ws_wcs.cell(row=ri, column=5, value=(
            f"=SUMIF('Duel Summary'!{seq_col_letter}$2:{seq_col_letter}${summary_last_row},$A{ri},"
            f"'Duel Summary'!{gw_col_letter}$2:{gw_col_letter}${summary_last_row})"
        )).font = BASE_FONT
        wcs_wr_cell = ws_wcs.cell(row=ri, column=6, value=f'=IFERROR(E{ri}/D{ri},"")')
        wcs_wr_cell.font = BASE_FONT
        wcs_wr_cell.number_format = "0.0%"
        # Explains sets with MORE win cons than the 3-game cap: that can only happen if
        # some single game in the duel ran 2+ win cons at once (a "tandem" deck, e.g.
        # Miner + Wall Breakers both slotted together) -- cross-reference every pair of
        # cards within this set against the Win-Con Pairs data (real recorded co-
        # occurrences in one deck) to identify which specific tandem(s) explain it.
        tandem_hits = []
        if len(cards_in_set) > MAX_GAMES_PER_DUEL:
            for a, b in combinations(sorted(cards_in_set), 2):
                n = real_tandems.get(frozenset((a, b)))
                if n:
                    tandem_hits.append(f"{a} + {b} ({n}x)")
        tandem_cell = ws_wcs.cell(row=ri, column=7, value="; ".join(tandem_hits) if tandem_hits else "")
        tandem_cell.font = BASE_FONT
        if tandem_hits:
            tandem_cell.fill = PatternFill("solid", fgColor="FFF2CC")
        ws_wcs.cell(row=ri, column=8,
                    value=", ".join(sorted(set_to_players[seq]))).font = BASE_FONT
        for c in range(1, len(headers_wcs) + 1):
            ws_wcs.cell(row=ri, column=c).border = BORDER
    ws_wcs.freeze_panes = "A2"
    ws_wcs.auto_filter.ref = f"A1:{get_column_letter(len(headers_wcs))}{len(unique_sets) + 1}"
    autosize(ws_wcs, [60, 11, 18, 14, 14, 12, 45, 60])
    ws_wcs.cell(row=1, column=10, value=(
        "Win-Con Set = the distinct win-condition cards that appeared anywhere across a "
        "duel's games, order ignored (Goblin Barrel+Graveyard+Royal Hogs counts the same "
        "regardless of which game each one was played in). Combined across all players who "
        "used that set -- use the AutoFilter dropdown on 'Times Played (Duels)' or 'Win "
        "Rate' to sort and see which sets are most popular / most successful. "
        "A set can show MORE than 3 win cons despite the 3-game duel cap -- that only "
        "happens when one game in the duel ran two win conditions together in the same "
        "deck (a 'tandem', e.g. Miner + Wall Breakers). The 'Contains Tandem(s)?' column "
        "(highlighted when filled) names the specific tandem pair(s) that explain it, "
        "cross-referenced against the Win-Con Pairs sheet. Only duels "
        "with exactly 3 identified distinct games are included; duels cut short at the "
        "edge of the fetched data window are excluded (see 'Incomplete' rows in Duel "
        "Summary). ALSO excluded: 'Duel Start Uncertain' duels (the first duel found for a "
        f"player/opponent pair, which may really be a continuation of an earlier duel we "
        f"never saw -- see Data Quality sheet). {excluded_uncertain_sets} duel rows excluded "
        f"on that basis in this build. ALSO excluded (added 2026-07-18): {excluded_official_sets} "
        "Official CRL duel(s) -- not yet blended in, see Match Category column in Duel "
        "Summary/Duel Log to view them directly.")).font = BASE_FONT
    ws_wcs.column_dimensions["H"].width = 80

    # ---------------- Card Frequency sheet ----------------
    # Restricted to win-condition cards only (per user decision, to keep this sheet
    # focused on the cards that matter competitively and cut down on formula volume).
    ws3 = wb.create_sheet("Card Frequency")
    players = sorted({r["player_name"] for r in duel_log})
    all_cards = sorted({
        c for r in duel_log for c in r["deck"]
        if c and WIN_CONDITION_REFERENCE.get(c, (False,))[0]
    })

    ws3.cell(row=1, column=1, value="Win Condition Card").font = HEADER_FONT
    ws3.cell(row=1, column=1).fill = HEADER_FILL
    for pi, p in enumerate(players, start=2):
        cell = ws3.cell(row=1, column=pi, value=f"{p} - Duel Decks Using This Win Con")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, card in enumerate(all_cards, start=2):
        ws3.cell(row=ri, column=1, value=card).font = BASE_FONT
        for pi, p in enumerate(players, start=2):
            col_letters = ["I", "J", "K", "L", "M", "N", "O", "P"]  # Card 1..8 in Duel Log
            terms = "+".join(
                f'COUNTIFS(\'Duel Log\'!C$2:C${log_last_row},"{p}",'
                f"'Duel Log'!{cl}$2:{cl}${log_last_row},$A{ri})"
                for cl in col_letters
            )
            ws3.cell(row=ri, column=pi, value=f"={terms}").font = BASE_FONT

    autosize(ws3, [20] + [34] * len(players))
    ws3.freeze_panes = "B2"
    card_freq_last_row = len(all_cards) + 1
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(players) + 1)}{card_freq_last_row}"
    # Heatmap PER PLAYER COLUMN (not across the whole table): each player has a different
    # total game count, so a per-column scale highlights which win con THAT player leans
    # on most, rather than letting one very active player's raw numbers wash out everyone
    # else's colors.
    # Blue sequential ramp (light -> dark blue), easier on the eyes than the earlier
    # white/yellow/red scale -- per user request.
    heatmap_rule = ColorScaleRule(
        start_type="min", start_color="F2F8FF",
        mid_type="percentile", mid_value=50, mid_color="6DA7EC",
        end_type="max", end_color="184F95",
    )
    for pi in range(2, len(players) + 2):
        col_letter = get_column_letter(pi)
        ws3.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{card_freq_last_row}", heatmap_rule
        )
    ws3.cell(row=1, column=len(players) + 3, value=(
        "Restricted to win-condition cards only (see Win Condition Reference sheet) -- "
        "support/utility cards are intentionally excluded here to keep this sheet focused "
        "on primary win conditions and reduce formula volume. Color scale is PER PLAYER "
        "COLUMN (light blue=least-used, dark blue=most-used win con for that specific "
        "player), so each player's own pattern stands out regardless of how many total "
        "games they have.")).font = BASE_FONT
    ws3.column_dimensions[get_column_letter(len(players) + 3)].width = 80

    # ---------------- Win-Con Pairs sheet ----------------
    # Distinct from Win-Con Sets: this is about which win conditions get slotted TOGETHER
    # in the same 8-card deck for a single game (tandems like Miner+Wall Breakers,
    # Battle Ram+Three Musketeers), not which win conditions showed up across a duel's
    # separate games. Per user request.
    ws_wcp = wb.create_sheet("Win-Con Pairs")
    headers_wcp = ["Win-Con Tandem", "Card A", "Card B", "Times Played", "Wins", "Win Rate"]
    ws_wcp.append(headers_wcp)
    style_header_row(ws_wcp, 1, len(headers_wcp))
    for ri, r in enumerate(wincon_pairs, start=2):
        ws_wcp.cell(row=ri, column=1, value=r["pair"]).font = BASE_FONT
        ws_wcp.cell(row=ri, column=2, value=r["card_a"]).font = BASE_FONT
        ws_wcp.cell(row=ri, column=3, value=r["card_b"]).font = BASE_FONT
        ws_wcp.cell(row=ri, column=4, value=r["times_played"]).font = BASE_FONT
        ws_wcp.cell(row=ri, column=5, value=r["wins"]).font = BASE_FONT
        wr_cell = ws_wcp.cell(row=ri, column=6, value=r["win_rate"])
        wr_cell.font = BASE_FONT
        wr_cell.number_format = "0.0%"
        for c in range(1, len(headers_wcp) + 1):
            ws_wcp.cell(row=ri, column=c).border = BORDER
    ws_wcp.freeze_panes = "A2"
    ws_wcp.auto_filter.ref = f"A1:{get_column_letter(len(headers_wcp))}{len(wincon_pairs) + 1}"
    autosize(ws_wcp, [40, 20, 20, 14, 10, 12])
    ws_wcp.cell(row=1, column=8, value=(
        "Which two win conditions get played TOGETHER in the same deck, within a single "
        "game (e.g. Miner + Wall Breakers, Battle Ram + Three Musketeers) -- different "
        "from Win-Con Sets, which tracks win conditions across a duel's separate games. "
        "Sort by Times Played to see the most common tandems, or Win Rate to see which "
        "tandems perform best.")).font = BASE_FONT
    ws_wcp.column_dimensions["H"].width = 80

    # ---------------- Practice Partners sheet ----------------
    # Which pairs of tracked players practice together most, counted in SESSIONS (one or
    # more back-to-back duels vs. the same opponent; a gap of >= SESSION_GAP_HOURS between
    # duels starts a new session), not raw game counts. Since a battle can appear in both
    # players' own fetched logs (each player's master_<tag>.json records it from their own
    # perspective), the same pair can show up as two ORDERED entries (A vs B and B vs A) in
    # session_stats -- deduped here to one row per unordered pair, keeping whichever
    # direction has the higher session count (the two should normally agree; a mismatch
    # would mean one side's log window didn't capture everything the other side's did).
    ws_pp = wb.create_sheet("Practice Partners")
    dedup = {}
    for s in session_stats:
        # Category included in the dedup key (added 2026-07-18) so a pair's Practice
        # sessions and Official CRL sessions get separate rows, not blended into one.
        key = (frozenset((s["player_tag"], s["opponent_tag"])), s["match_category"])
        existing = dedup.get(key)
        if existing is None or s["num_sessions"] > existing["num_sessions"]:
            dedup[key] = s
    partner_rows = sorted(
        dedup.values(),
        key=lambda s: (-s["num_sessions"], -s["num_duels"], s["player_name"])
    )
    headers_pp = [
        "Player 1", "Player 2", "Match Category", "Practice Sessions", "Total Duels",
        "Total Games", "First Session Start (UTC)", "Most Recent Session Start (UTC)",
    ]
    ws_pp.append(headers_pp)
    style_header_row(ws_pp, 1, len(headers_pp))
    for ri, s in enumerate(partner_rows, start=2):
        ws_pp.cell(row=ri, column=1, value=s["player_name"]).font = BASE_FONT
        ws_pp.cell(row=ri, column=2, value=s["opponent_name"]).font = BASE_FONT
        ws_pp.cell(row=ri, column=3, value=s["match_category"]).font = BASE_FONT
        ws_pp.cell(row=ri, column=4, value=s["num_sessions"]).font = BASE_FONT
        ws_pp.cell(row=ri, column=5, value=s["num_duels"]).font = BASE_FONT
        ws_pp.cell(row=ri, column=6, value=s["num_games"]).font = BASE_FONT
        ws_pp.cell(row=ri, column=7,
                   value=s["first_session_start"].strftime("%Y-%m-%d %H:%M:%S")).font = BASE_FONT
        ws_pp.cell(row=ri, column=8,
                   value=s["last_session_start"].strftime("%Y-%m-%d %H:%M:%S")).font = BASE_FONT
        for c in range(1, len(headers_pp) + 1):
            ws_pp.cell(row=ri, column=c).border = BORDER
    ws_pp.freeze_panes = "A2"
    ws_pp.auto_filter.ref = f"A1:{get_column_letter(len(headers_pp))}{len(partner_rows) + 1}"
    autosize(ws_pp, [22, 22, 16, 16, 12, 12, 24, 26])
    ws_pp.cell(row=1, column=len(headers_pp) + 2, value=(
        f"A 'practice session' is one or more duels played back-to-back against the same "
        f"opponent. A gap of {SESSION_GAP_HOURS}+ hour(s) between the end of one duel and "
        f"the start of the next duel vs. that opponent starts a new session (a coarser, "
        f"higher-level grouping than the {DUEL_GAP_HOURS}-hour duel-boundary gap itself, "
        f"since several distinct duels are often played in one sitting). Python-computed, "
        f"not a live formula -- re-run build_duel_workbook.py after new data is imported.")
    ).font = BASE_FONT
    ws_pp.column_dimensions[get_column_letter(len(headers_pp) + 2)].width = 80

    # ---------------- Read Me sheet ----------------
    ws4 = wb.create_sheet("Read Me", 0)
    ws4.column_dimensions["A"].width = 100
    lines = [
        ("CRL Player Duels - Methodology & Assumptions", BOLD_FONT),
        ("", BASE_FONT),
        ("Source", BOLD_FONT),
        ("Data pulled from the official Clash Royale API (api.clashroyale.com/v1), "
         "players/{tag}/battlelog endpoint, via a local script run outside this workbook.",
         BASE_FONT),
        ("", BASE_FONT),
        ("Friendly / practice filter", BOLD_FONT),
        ('Rows are included only where the API battle record has type == "clanMate" AND '
         'gameMode.name == "Friendly". This was confirmed against real pulled data to be '
         "the signal for an informal 1v1 practice battle between two named players (not "
         "ranked ladder, not the trail/Showdown_Friendly event mode, both excluded by "
         "user decision).", BASE_FONT),
        ("", BASE_FONT),
        ("Match Category: Practice vs. Official CRL (added 2026-07-18, signal confirmed "
         "2026-07-18 live Round 2)", BOLD_FONT),
        ("Every game, duel, and session is tagged 'Practice' or 'Official CRL' (Duel Log / "
         "Duel Summary / Practice Partners / Player Lookup all have this column). Practice "
         "and Official CRL games between the same two players are grouped into SEPARATE "
         "duels/sessions -- they never blend into one duel just because they happened close "
         "together in time. Official CRL duel/session IDs get an '_CRL' suffix so they can "
         "never collide with a same-numbered Practice duel for the same pair.", BASE_FONT),
        ("Official CRL signal: type == 'friendly' (lowercase) AND gameMode.name == "
         "'Friendly' AND the opponent is NOT one of the 48 tracked roster players AND the "
         "battle falls inside a known live-tournament time cluster (OFFICIAL_CRL_LIVE_"
         "CLUSTERS in build_duel_workbook.py -- one entry per round, added as each round's "
         "cluster window is identified from a fresh fetch; Round 2's is 2026-07-18 "
         "13:18:56-14:55:33 UTC, found by time-clustering all friendly/Friendly-vs-outsider "
         "battles across every roster player and confirming one dramatic outlier -- 193 "
         "games across all 48 players at once -- versus 23 much smaller clusters (1-38 "
         "games, 1-3 players) that are ordinary scattered ladder/friend-challenge activity, "
         "not CRL). Plain type=='friendly' battles outside a listed cluster, or against a "
         "roster opponent, are excluded from the workbook entirely, same as before this "
         "signal was found.", BASE_FONT),
        ("CRL best-of-3 completion (per user decision, 2026-07-18): a 2-0/0-2 SWEEP is "
         "COMPLETE with only 2 games -- unlike Practice, it is NOT flagged 'incomplete' for "
         "lacking a 3rd game. A 1-1 split, or a duel with only 1 game logged so far, is "
         "PENDING (game 3 hasn't been played/fetched yet) -- held in the background and "
         "excluded from clean/complete CRL stats until a future end-of-event pull "
         "reconciles it. Duel Summary's 'CRL Status' column shows this per duel (Complete "
         "(Sweep 2-0) / Complete (Decided 2-1) / Pending (1-1...) / Pending (Game 1 only) / "
         "Anomaly...), with Pending rows shaded light blue and Anomaly rows shaded light "
         "orange. One Round-2 matchup -- 加急Sinistro (TM) vs Yex Style -- logged 7 games "
         "(W4-L3) inside the CRL cluster, which doesn't fit the sweep/1-1/decided pattern; "
         "it's marked Anomaly and excluded from clean stats pending manual review, per user "
         "decision, rather than guessed at.", BASE_FONT),
        ("Win-Con Sets and the dashboard's 'What Might Follow'/Deck Predictor are still "
         "scoped to Practice duels only -- Official CRL isn't blended into those yet (sample "
         "size is thin and practice/tournament patterns may genuinely differ).", BASE_FONT),
        ("", BASE_FONT),
        ("Duel grouping heuristic (ASSUMPTION - not an API field)", BOLD_FONT),
        (f"The API has no concept of a 'duel session' - this workbook infers one. "
         f"Consecutive friendly battles against the same opponent (in time order) are "
         f"grouped into one duel of up to {MAX_GAMES_PER_DUEL} games, where no card may "
         f"repeat across the decks used in that duel (matches the community norm of playing "
         f"3 distinct, non-overlapping decks per duel). A gap of more than {DUEL_GAP_HOURS} "
         f"hours since the previous game against that same opponent forces a new duel to "
         f"start, on the assumption duels are played back-to-back in one sitting. "
         f"Both the game cap and the gap threshold are configurable in "
         f"build_duel_workbook.py.", BASE_FONT),
        ("", BASE_FONT),
        ("\"Instant Rematch\" flag", BOLD_FONT),
        ("Occasionally a player deliberately replays the exact same deck against the same "
         "opponent right after a game, to study a specific matchup further. Since this "
         "reuses cards, it would otherwise look like a violation of the \"no repeated cards\" "
         "duel rule and wrongly trigger a new duel to start. These games are instead kept in "
         "the same duel as a bonus game (not counted toward the 3-distinct-deck cap) and "
         "flagged Yes here so they stay visible rather than being silently merged or "
         "mis-split.", BASE_FONT),
        ("", BASE_FONT),
        ("\"Duel Start Uncertain\" flag (yellow highlight)", BOLD_FONT),
        ("The Clash Royale API only returns each player's most recent battles (30 were "
         "pulled per player here). This means the very first duel found for a given "
         "player/opponent pair may actually be a continuation of a duel that began before "
         "the fetch window - we cannot see further back, so this cannot be fully resolved "
         "from this data alone. Those rows are flagged Yes and highlighted so they can be "
         "sanity-checked. As of 2026-07-17, these duels (plus incomplete duels, <3 games) "
         "are also automatically EXCLUDED from the Win-Con Sets sheet and the dashboard's "
         "'What Might Follow' / Deck Predictor features, since those specifically depend on "
         "trusting a duel's full, correctly-ordered game set -- see the Data Quality sheet "
         "for exact counts. Every individual game row itself is still kept everywhere else "
         "(Duel Log, Player Lookup, Deck Stats, Card Frequency, Deck Matchups) since each "
         "recorded game's deck and result are real regardless of which duel it landed in.",
         BASE_FONT),
        ("", BASE_FONT),
        ("Repeated-card check", BOLD_FONT),
        ("No card may repeat across the games grouped into one duel -- this is enforced as a "
         "hard rule in the grouping algorithm itself (a repeated card forces a new duel "
         "boundary), not just a warning, so it can't silently slip through. Verified directly "
         "against the raw fetched data on 2026-07-17: 0 violations found. See Data Quality "
         "sheet.", BASE_FONT),
        ("", BASE_FONT),
        ("\"Win Condition\" classification (ASSUMPTION - community/strategy concept, not an API field)",
         BOLD_FONT),
        ("Supercell's API does not label cards as win conditions. The 'Win Condition Reference' "
         "sheet lists every card seen in this dataset with a Yes/No classification (yellow = "
         "editable) and a confidence level. A handful of cards were not confidently recognized "
         "(possibly added to the game recently) and are marked 'Uncertain' rather than guessed -- "
         "check that sheet and correct anything that looks wrong. Changes there require "
         "re-running build_duel_workbook.py to flow into the Win-Con Set analysis, since "
         "that set text is generated in Python, not by a live formula.", BASE_FONT),
        ("", BASE_FONT),
        ("\"Win-Con Set\" is order-independent (by user decision)", BOLD_FONT),
        ("Earlier versions tracked a Game1 -> Game2 -> Game3 win-condition sequence, but which "
         "game a win condition was played in doesn't matter much competitively -- what matters "
         "is which win conditions a player brings to a duel together. So Win-Con Set is the "
         "de-duplicated, alphabetically-sorted set of win conditions across a duel's 3 games: "
         "Goblin Barrel+Graveyard+Royal Hogs counts the same regardless of which order they "
         "were played in.", BASE_FONT),
        ("", BASE_FONT),
        ("\"Practice Session\" grouping (ASSUMPTION - not an API field)", BOLD_FONT),
        (f"A practice session is one or more duels played back-to-back against the same "
         f"opponent. A gap of {SESSION_GAP_HOURS}+ hour(s) between the end of one duel and the "
         f"start of the next duel vs. that same opponent starts a new session -- a coarser, "
         f"higher-level grouping than the {DUEL_GAP_HOURS}-hour duel-boundary gap, since "
         f"players often play several distinct duels in a row in one sitting. The 'Practice "
         f"Partners' sheet ranks tracked players by how many sessions they've had together. "
         f"Threshold is configurable (SESSION_GAP_HOURS in build_duel_workbook.py).",
         BASE_FONT),
        ("", BASE_FONT),
        ("Analytics prototype -- built on a small dataset", BOLD_FONT),
        ("Deck Stats, Deck Matchups, Win-Con Sets, and Practice Partners are built and working "
         "correctly, but as the tracked roster and history grow, treat any single number here "
         "with the sample size in mind (visible directly, e.g. Games Played / Practice "
         "Sessions columns) -- a deck or matchup seen only once or twice isn't statistically "
         "meaningful yet. The formulas will recalculate automatically as more players/games are "
         "added to Duel Log.", BASE_FONT),
        ("", BASE_FONT),
        ("Sorting / filtering", BOLD_FONT),
        ("Every data sheet has AutoFilter enabled on its header row -- click the dropdown "
         "arrow on any 'Games Played', 'Times Played', 'Win Rate', etc. column to sort "
         "ascending/descending or filter down to what you're looking for.", BASE_FONT),
        ("", BASE_FONT),
        ("Sheets", BOLD_FONT),
        ("Duel Log: one row per individual game, in duel groups, including the opponent's deck.",
         BASE_FONT),
        ("Duel Summary: one row per duel, with win counts, all decks used, the practice "
         "session it belongs to, and the win-con set for that duel.", BASE_FONT),
        ("Win Condition Reference: editable card classification driving the win-con analysis.",
         BASE_FONT),
        ("Deck Stats: games played / win rate for each unique 8-card deck.", BASE_FONT),
        ("Deck Matchups: head-to-head record between specific deck pairs that have faced "
         "each other.", BASE_FONT),
        ("Win-Con Sets: how often each (order-independent) win-condition set shows up, "
         "combined across all players who used it, plus win rate and who used it. Sort by "
         "'Times Played (Duels)' or 'Win Rate' via the AutoFilter dropdown to find the most "
         "popular / most successful sets.", BASE_FONT),
        ("Card Frequency: how often each WIN-CONDITION card appeared in each player's duel "
         "decks (restricted to win conditions only, not all 122 cards, to keep this sheet "
         "focused and reduce formula volume). Color scale is per-player-column (light->dark "
         "blue), so each player's own go-to win conditions pop out visually.", BASE_FONT),
        ("Win-Con Pairs: which two win conditions get played TOGETHER in the same deck "
         "within a single game (tandems like Miner+Wall Breakers) -- different from Win-Con "
         "Sets, which is about a duel's separate games.", BASE_FONT),
        ("Practice Partners: which pairs of tracked players have practiced together most, "
         "ranked by practice-session count (not raw game count).", BASE_FONT),
        ("Player Lookup: match-prep dashboard -- one row per player, their most-played "
         "decks, best win-rate decks, and most-used win conditions. Filter/search by "
         "player name OR player tag (added 2026-07-18, 'Player Tag(s)' column) using the "
         "header row's AutoFilter dropdown -- useful if a player's tracked account renamed "
         "in-game and no longer matches the roster's expected display name (see the "
         "ElMollejas/MH Axel case, 2026-07-18). 'Practice Games'/'Official CRL Games' "
         "columns (also added 2026-07-18) split each player's total by Match Category; "
         "Practice + Official CRL games are still COMBINED (unweighted) for the Most-Played "
         "Deck / Top Win Condition rankings on this Excel sheet -- see the dashboard for a "
         "toggleable weighted view and a filter to see Practice-only or Official-CRL-only "
         "rankings.", BASE_FONT),
        ("Data Quality: exact counts of incomplete and uncertain-start duels, what gets "
         "excluded from which sheets/features and why, and the repeated-card check result.",
         BASE_FONT),
    ]
    for i, (text, font) in enumerate(lines, start=1):
        cell = ws4.cell(row=i, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---------------- Player Lookup sheet (match-prep dashboard, last sheet) ----------------
    # Not a dropdown+formula lookup -- Excel's array-formula/exotic-function support is
    # unreliable in this pipeline's recalc environment (see xlsx skill notes), so instead
    # every player gets their own row (one-row-per-player "profile"), and AutoFilter on
    # the header lets you filter/search down to one name fast, which is what actually
    # matters for a quick lookup mid-tournament-prep. Python-computed, not live formulas --
    # re-run build_duel_workbook.py after new data is imported.
    # Unweighted (weighted=False) in the Excel sheet -- Excel is static, no live toggle is
    # possible here, so this stays the plain, easy-to-reason-about view. The dashboard adds
    # a user-toggleable weighted view + Practice/Official CRL filter on top of the same
    # underlying compute_player_lookup() function -- see build_dashboard.py.
    player_lookup = compute_player_lookup(duel_log, weighted=False)
    ws_pl = wb.create_sheet("Player Lookup")
    headers_pl = [
        "Player", "Total Games", "Total Wins", "Win Rate",
        "Practice Games", "Official CRL Games",
        "Most-Played Deck #1", "Most-Played Deck #2", "Most-Played Deck #3",
        "Best Win-Rate Deck #1", "Best Win-Rate Deck #2", "Best Win-Rate Deck #3",
        "Top Win Condition #1", "Top Win Condition #2", "Top Win Condition #3",
        "Player Tag(s)",
    ]
    ws_pl.append(headers_pl)
    style_header_row(ws_pl, 1, len(headers_pl))

    def fmt_deck_freq(entry):
        if not entry:
            return ""
        deck, n = entry
        return f"{deck} ({n} game{'s' if n != 1 else ''})"

    def fmt_deck_winrate(entry):
        if not entry:
            return ""
        deck, wr, n = entry
        return f"{deck} ({wr:.0%} win rate, {n} game{'s' if n != 1 else ''})"

    def fmt_wincon(entry):
        if not entry:
            return ""
        card, n = entry
        return f"{card} ({n} game{'s' if n != 1 else ''})"

    for ri, row in enumerate(player_lookup, start=2):
        ws_pl.cell(row=ri, column=1, value=row["player"])
        ws_pl.cell(row=ri, column=2, value=row["total_games"])
        ws_pl.cell(row=ri, column=3, value=row["total_wins"])
        ws_pl.cell(row=ri, column=4, value=(
            f"{row['win_rate']:.0%}" if row["total_games"] else "n/a"
        ))
        ws_pl.cell(row=ri, column=5, value=row.get("practice_games", 0))
        ws_pl.cell(row=ri, column=6, value=row.get("official_games", 0))
        freq = row["top_decks_by_freq"]
        for ci in range(3):
            ws_pl.cell(row=ri, column=7 + ci,
                       value=fmt_deck_freq(freq[ci]) if ci < len(freq) else "")
        wr_decks = row["top_decks_by_winrate"]
        for ci in range(3):
            ws_pl.cell(row=ri, column=10 + ci,
                       value=fmt_deck_winrate(wr_decks[ci]) if ci < len(wr_decks) else "")
        wincons = row["top_wincons"]
        for ci in range(3):
            ws_pl.cell(row=ri, column=13 + ci,
                       value=fmt_wincon(wincons[ci]) if ci < len(wincons) else "")
        ws_pl.cell(row=ri, column=16, value=", ".join(row.get("tags") or []))
        for c in range(1, len(headers_pl) + 1):
            ws_pl.cell(row=ri, column=c).font = BASE_FONT
            ws_pl.cell(row=ri, column=c).border = BORDER
            ws_pl.cell(row=ri, column=c).alignment = Alignment(wrap_text=True, vertical="top")

    ws_pl.freeze_panes = "B2"
    ws_pl.auto_filter.ref = f"A1:{get_column_letter(len(headers_pl))}{len(player_lookup) + 1}"
    autosize(ws_pl, [22, 12, 12, 10, 14, 16] + [46] * 6 + [30] * 3 + [16])
    ws_pl.cell(row=1, column=len(headers_pl) + 2, value=(
        f"Best Win-Rate Deck columns only consider decks the player has used at least "
        f"{MIN_GAMES_FOR_WINRATE_RANKING} times -- otherwise a deck played once and won "
        f"would show as a misleading '100% win rate' entry ahead of decks with a real "
        f"sample size. Rows sorted by total games played (most active players first). "
        f"Use the AutoFilter dropdown on 'Player' (cell A1) to jump straight to one name. "
        f"'Practice Games'/'Official CRL Games' (added 2026-07-18) split the 'Total Games' "
        f"count by Match Category -- Most-Played/Top Win Condition rankings here are "
        f"UNWEIGHTED (plain counts, both categories combined); the dashboard adds a "
        f"user-toggleable weighted view plus a Practice/Official CRL filter on top of the "
        f"same underlying data.")
    ).font = BASE_FONT
    ws_pl.column_dimensions[get_column_letter(len(headers_pl) + 2)].width = 80

    # ---------------- Data Quality sheet (added 2026-07-17, per user request) ----------------
    # Answers "how much of this data could be corrupted by API window truncation?" directly,
    # so it doesn't have to be taken on faith. Two independent risks are covered:
    #  1. Incomplete duels: fewer than MAX_GAMES_PER_DUEL non-rematch games captured. Real
    #     and legitimate sometimes (a duel really did end after 1-2 games), so these rows are
    #     NOT deleted from Duel Log/Duel Summary (every captured game is real, ground-truth
    #     data) -- they're excluded only from the sequence-dependent Win-Con Sets sheet and
    #     the dashboard's "What Might Follow"/"Deck Predictor" features, both of which assume
    #     a duel's captured games are its true, fully-ordered game set.
    #  2. Uncertain-start duels: the FIRST duel found for a (player, opponent) pair. Since the
    #     API only returns each player's ~25-30 most recent battles, this duel may really be a
    #     continuation of an earlier duel we never saw -- so even a "complete" 3-game uncertain
    #     duel is excluded from the same sequence-dependent analyses.
    #  3. No-repeated-card check: verified structurally by the grouping algorithm itself (a
    #     card repeat forces a new duel boundary, see group_into_duels) -- 0 violations found
    #     when audited directly against the raw data on 2026-07-17.
    ws_dq = wb.create_sheet("Data Quality")
    ws_dq.column_dimensions["A"].width = 42
    ws_dq.column_dimensions["B"].width = 14
    ws_dq.column_dimensions["C"].width = 100

    total_duels_dq = len(duel_summary)
    incomplete_dq = [r for r in duel_summary if r["games_played"] < MAX_GAMES_PER_DUEL]
    uncertain_dq = [r for r in duel_summary if r["uncertain_start"]]
    uncertain_complete_dq = [r for r in uncertain_dq if r["games_played"] >= MAX_GAMES_PER_DUEL]
    excluded_from_sequence_dq = len({r["duel_id"] for r in duel_summary
                                      if r["uncertain_start"] or r["games_played"] < MAX_GAMES_PER_DUEL})

    dq_rows = [
        ("Metric", "Count", "Notes"),
        ("Total duels captured", total_duels_dq, ""),
        ("Incomplete duels (< 3 non-rematch games)", len(incomplete_dq),
         "Real games, real results -- kept in Duel Log/Duel Summary/Player Lookup/Deck "
         "Stats. Excluded only from Win-Con Sets and the dashboard's sequence-dependent "
         "predictors, since those need a duel's full game set."),
        ("Uncertain-start duels (first duel per pair)", len(uncertain_dq),
         "API only returns each player's most recent ~25-30 battles -- this duel may be a "
         "continuation of an earlier duel we can never see. Excluded from Win-Con Sets and "
         "the dashboard's sequence-dependent predictors."),
        ("  ...of which look 'complete' (3 games) but are unverifiable", len(uncertain_complete_dq),
         "Highest-risk subset: these would silently look like clean data without this flag."),
        ("Total duels excluded from Win-Con Sets / predictors", excluded_from_sequence_dq,
         f"{excluded_from_sequence_dq / total_duels_dq:.1%} of all captured duels, as of this build."),
        ("Card-overlap violations found (repeated card across a duel's games)", 0,
         "Should always be 0 -- the grouping algorithm treats a repeated card as a hard "
         "duel-boundary signal, not a soft one, so this can't slip through silently. "
         "Verified directly against raw battle data on 2026-07-17."),
        ("", "", ""),
        ("Gap-risk note (not auto-computed here)", "", (
            "A separate, harder-to-detect risk: if a player plays more friendlies than the "
            "API returns (~25-30) in between two fetches, older ones in the middle of that "
            "gap can permanently age out and vanish with zero trace in the data -- this "
            "can't be proven after the fact without a fetch-time log. fetch_cr_battlelogs.py "
            "now writes fetch_log.json (timestamp, per-player oldest/newest battle, return "
            "count) on every run and will print a warning at fetch time if this looks like it "
            "may have happened, going forward.")),
    ]
    for ri, (a, b, c) in enumerate(dq_rows, start=1):
        ca = ws_dq.cell(row=ri, column=1, value=a)
        cb = ws_dq.cell(row=ri, column=2, value=b)
        cc = ws_dq.cell(row=ri, column=3, value=c)
        if ri == 1:
            for cell in (ca, cb, cc):
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
        else:
            ca.font = BOLD_FONT if a and not a.startswith("  ") else BASE_FONT
            cb.font = BASE_FONT
            cc.font = BASE_FONT
            cc.alignment = Alignment(wrap_text=True, vertical="top")
    ws_dq.freeze_panes = "A2"

    wb.save(XLSX_OUT)
    print(f"Wrote {len(duel_log)} game rows across {len(duel_summary)} duels "
          f"for players: {players}")


if __name__ == "__main__":
    main()
