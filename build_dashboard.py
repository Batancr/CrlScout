"""
Build crl_opponent_scout.html -- a self-contained, no-dependency HTML dashboard for
fast opponent lookup during CRL play, from the data already computed in
CRL_Duel_Decks.xlsx (Player Lookup, Win-Con Sets, Win-Con Pairs, Deck Stats sheets).
Card icons are pulled from the Clash Royale API's own battle data. By default they're
hotlinked (need internet to display) -- for guaranteed offline rendering, run
download_card_icons.py somewhere with real internet access first (this sandbox's
network is blocked from api-assets.clashroyale.com), which saves the PNGs into a
card_icons/ folder next to this script; this script auto-detects that folder and
embeds the icons as base64 instead.

USAGE: run this AFTER build_duel_workbook.py and recalc.py (the workbook must have
cached formula VALUES, not just formula text, for openpyxl's data_only=True read to
work -- LibreOffice's recalc.py writes those cached values back into the file).

    python3 build_duel_workbook.py
    python3 /mnt/skills/public/xlsx/scripts/recalc.py CRL_Duel_Decks.xlsx   # or wherever recalc.py lives
    python3 build_dashboard.py

This is a STATIC SNAPSHOT -- it has no live connection back to the Excel file. Re-run
it (after the two steps above) any time the underlying data changes to refresh it.
"""
import glob
import json
import os
import re
from collections import Counter, defaultdict
import openpyxl

XLSX_PATH = os.path.join(os.environ["CRL_HOME"], "CRL_Duel_Decks.xlsx") if os.environ.get("CRL_HOME") else "CRL_Duel_Decks.xlsx"


def rows_of(wb, sheet, headers_row=1):
    """Read a sheet into a list of dicts, stopping at the first blank header
    (that's where the free-text assumption/notes column starts on these sheets)."""
    ws = wb[sheet]
    raw_headers = [c.value for c in ws[headers_row]]
    headers = []
    for h in raw_headers:
        if h is None:
            break
        headers.append(h)
    ncols = len(headers)
    out = []
    for row in ws.iter_rows(min_row=headers_row + 1, max_col=ncols, values_only=True):
        if row[0] is None:
            continue
        out.append(dict(zip(headers, row)))
    return out


def find_master_paths():
    search_globs = [
        "/mnt/user-data/uploads/CRL/master_*.json",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "master_*.json"),
        "master_*.json",
    ]
    paths, seen = [], set()
    for pattern in search_globs:
        for p in sorted(glob.glob(pattern)):
            if p not in seen:
                seen.add(p)
                paths.append(p)
    return paths


def build_card_meta():
    """Card icon URLs, champion (hero) status, and evolution-eligibility come straight
    from the Clash Royale API's own battle-log responses -- not guessed or pulled from a
    third-party source. Every card object in a real fetched battle already includes
    iconUrls.medium, rarity (champion cards are the "hero" cards -- Golden Knight,
    Skeleton King, Archer Queen, Monk, Mighty Miner, etc.), and evolutionLevel/
    maxEvolutionLevel (whether this specific card slot is an Evolution in that battle).
    Scans every master_<tag>.json this script can find. Returns:
      icons: {card_name: icon_url}
      champions: set of card names that are Champion/Hero cards
      evolution_capable: set of card names ever seen evolved in the data (i.e. cards
        this roster has slotted as an Evolution at least once -- a simplification,
        since Deck Stats/Player Lookup aggregate multiple deck instances together and
        don't track evolution choice per specific instance)."""
    icons, icons_evo, icons_hero, champions, evolution_capable = {}, {}, {}, set(), set()
    for path in find_master_paths():
        try:
            with open(path) as f:
                battles = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue
        for b in battles:
            for side in ("team", "opponent"):
                for p in b.get(side, []):
                    for c in p.get("cards", []) + p.get("supportCards", []):
                        name = c.get("name")
                        if not name:
                            continue
                        icon_urls = c.get("iconUrls", {})
                        url = icon_urls.get("medium")
                        if url and name not in icons:
                            icons[name] = url
                        evo_url = icon_urls.get("evolutionMedium")
                        if evo_url and name not in icons_evo:
                            icons_evo[name] = evo_url
                        hero_url = icon_urls.get("heroMedium")
                        if hero_url and name not in icons_hero:
                            icons_hero[name] = hero_url
                        if c.get("rarity") == "champion":
                            champions.add(name)
                        if c.get("evolutionLevel") is not None:
                            evolution_capable.add(name)
    return icons, icons_evo, icons_hero, champions, evolution_capable


def build_card_elixir():
    """card_name -> elixirCost, straight from the Clash Royale API's own card objects
    (every card in a real battle log includes elixirCost) -- not guessed or looked up
    from a third-party source. Used to compute a player's deck elixir-cost tendency."""
    elixir = {}
    for path in find_master_paths():
        try:
            with open(path) as f:
                battles = json.load(f)
        except (json.JSONDecodeError, OSError):
            continue
        for b in battles:
            for side in ("team", "opponent"):
                for p in b.get(side, []):
                    for c in p.get("cards", []) + p.get("supportCards", []):
                        name = c.get("name")
                        cost = c.get("elixirCost")
                        if name and cost is not None and name not in elixir:
                            elixir[name] = cost
    return elixir


def build_player_briefs(duel_log, card_elixir, min_games_for_winrate):
    """Per-player 'opponent brief' data: best AND worst win-rate decks (the workbook's
    Player Lookup sheet only has best; worst is useful for a quick 'what beats them'
    read), plus an average deck elixir-cost tendency computed from real elixirCost
    values in the battle data. Same MIN_GAMES_FOR_WINRATE_RANKING threshold as the
    workbook so numbers stay consistent between the two. Returns {player: {...}}."""
    per_player = defaultdict(lambda: {
        "deck_games": Counter(), "deck_wins": Counter(), "elixir_samples": [],
    })
    for r in duel_log:
        if not r["deck"]:
            continue
        p = per_player[r["player_name"]]
        deck_key = ", ".join(sorted(r["deck"]))
        p["deck_games"][deck_key] += 1
        if r["crowns_for"] > r["crowns_against"]:
            p["deck_wins"][deck_key] += 1
        costs = [card_elixir[c] for c in r["deck"] if c in card_elixir]
        if costs:
            p["elixir_samples"].append(sum(costs) / len(costs))

    briefs = {}
    for player, d in per_player.items():
        eligible = [
            (deck, d["deck_wins"][deck] / d["deck_games"][deck], d["deck_games"][deck])
            for deck in d["deck_games"]
            if d["deck_games"][deck] >= min_games_for_winrate
        ]
        best = sorted(eligible, key=lambda x: (-x[1], -x[2]))[:3]
        worst = sorted(eligible, key=lambda x: (x[1], -x[2]))[:3]
        avg_elixir = (
            sum(d["elixir_samples"]) / len(d["elixir_samples"])
            if d["elixir_samples"] else None
        )
        briefs[player] = {
            "best_decks": [{"deck": deck, "win_rate": wr, "games": n} for deck, wr, n in best],
            "worst_decks": [{"deck": deck, "win_rate": wr, "games": n} for deck, wr, n in worst],
            "avg_elixir": round(avg_elixir, 2) if avg_elixir is not None else None,
        }
    return briefs


def _safe_card_filename_base(name):
    """Mirrors download_card_icons.py's safe_filename() base-name logic, so lookups
    here find exactly the files that script wrote (e.g. "Mini P.E.K.K.A" -> "Mini_P_E_K_K_A")."""
    return re.sub(r"[^A-Za-z0-9]+", "_", name).strip("_")


def embed_local_icons(icons, icons_evo, icons_hero):
    """If a card_icons/ folder (from download_card_icons.py, run somewhere with real
    internet access -- this sandbox's egress proxy blocks api-assets.clashroyale.com)
    exists next to this script, replace hotlinked URLs with embedded base64 data URLs
    so the dashboard works fully offline -- for the base card art AND, where present,
    the real evolution-form and hero-form art (not just a generic badge icon).

    Scans the folder directly by filename convention rather than trusting
    manifest.json's schema, since that file's format has changed across versions of
    download_card_icons.py (flat name->file vs. nested name->{variant:file}) and a
    stale manifest shouldn't hide icons that are actually present on disk.
    Falls back silently to the hotlinked URL (or no image) for anything missing.
    Returns (icons, icons_evo, icons_hero, embedded_count)."""
    import base64
    local_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "card_icons")
    if not os.path.isdir(local_dir):
        return icons, icons_evo, icons_hero, 0

    on_disk = set(os.listdir(local_dir))

    def _load(fname):
        fpath = os.path.join(local_dir, fname)
        if not os.path.getsize(fpath):
            return None
        with open(fpath, "rb") as f:
            return f"data:image/png;base64,{base64.b64encode(f.read()).decode('ascii')}"

    embedded = dict(icons)
    embedded_evo = dict(icons_evo)
    embedded_hero = dict(icons_hero)
    count = 0
    all_names = set(icons) | set(icons_evo) | set(icons_hero)
    for name in all_names:
        base_name = _safe_card_filename_base(name)
        base_fname = f"{base_name}.png"
        evo_fname = f"{base_name}_evolution.png"
        hero_fname = f"{base_name}_hero.png"
        if base_fname in on_disk:
            data_url = _load(base_fname)
            if data_url:
                embedded[name] = data_url
                count += 1
        if evo_fname in on_disk:
            data_url = _load(evo_fname)
            if data_url:
                embedded_evo[name] = data_url
                count += 1
        if hero_fname in on_disk:
            data_url = _load(hero_fname)
            if data_url:
                embedded_hero[name] = data_url
                count += 1
    return embedded, embedded_evo, embedded_hero, count


# ---------------------------------------------------------------------------
# "Best Picks" quick-reference feature (added 2026-07-18, per user request): three
# game-day-quick-reference rankings -- top win-rate single decks, top win-rate win-con
# SETS (reuses the same order-independent-set concept as the workbook's Win-Con Sets
# sheet), and top win-rate DUEL sets (which 3-deck combination performed best together).
# Each is computed separately per Match Category (the caller passes an already-filtered
# duel_log slice) so results never blend Practice and Official CRL.
# ---------------------------------------------------------------------------

# Baked-in floors are now just 1 (i.e. "has at least one game/duel") -- the real
# small-sample filtering happens live in the dashboard via the "Min times played"
# dropdown (added 2026-07-19) so the user can raise/lower it on the fly instead of
# rebuilding. BEST_PICKS_MIN_DUELSET_DUELS/BEST_PICKS_MIN_CLUSTERS_TARGET still control
# how loose the deck-overlap clustering threshold gets, since that's a build-time
# decision -- the full (unfiltered) cluster rows are returned either way so the client
# can filter tighter or looser without needing a different clustering threshold.
BEST_PICKS_MIN_DECK_GAMES = 1        # a deck needs at least this many games to be ranked
BEST_PICKS_MIN_WINCON_SET_DUELS = 1  # a win-con set needs at least this many duels
BEST_PICKS_MIN_DUELSET_DUELS = 3    # a duel-set cluster needs at least this many duels to
                                     # count toward the clustering-threshold decision below
BEST_PICKS_MIN_CLUSTERS_TARGET = 5  # keep loosening the deck-overlap threshold until at
                                     # least this many duel-set clusters qualify


def compute_best_decks(duel_log, min_games=BEST_PICKS_MIN_DECK_GAMES):
    games = Counter()
    wins = Counter()
    for r in duel_log:
        if not r["deck"] or len(r["deck"]) != 8:
            continue
        key = ", ".join(sorted(r["deck"]))
        games[key] += 1
        if r["crowns_for"] > r["crowns_against"]:
            wins[key] += 1
    rows = []
    for key, g in games.items():
        if g < min_games:
            continue
        w = wins[key]
        rows.append({"deck": key, "games": g, "wins": w, "win_rate": w / g if g else 0.0})
    rows.sort(key=lambda r: (-r["win_rate"], -r["games"]))
    return rows


def compute_best_wincon_sets(duel_log, min_duels=BEST_PICKS_MIN_WINCON_SET_DUELS):
    by_duel = defaultdict(list)
    for r in duel_log:
        by_duel[r["duel_id"]].append(r)

    set_duels = Counter()
    set_games = Counter()
    set_wins = Counter()
    for duel_id, games in by_duel.items():
        if games[0]["uncertain_start"]:
            continue
        non_rematch = [g for g in games if not g["is_rematch"]][:MAX_GAMES_PER_DUEL]
        if len(non_rematch) < MAX_GAMES_PER_DUEL:
            continue
        wincon_set = set()
        for g in non_rematch:
            wincon_set.update(classify_deck(g["deck"]))
        key = "+".join(sorted(wincon_set)) if wincon_set else "(none classified)"
        set_duels[key] += 1
        for g in games:
            set_games[key] += 1
            if g["crowns_for"] > g["crowns_against"]:
                set_wins[key] += 1

    rows = []
    for key, d in set_duels.items():
        if d < min_duels:
            continue
        g = set_games[key]
        w = set_wins[key]
        rows.append({
            "wincon_set": key, "duels": d, "games": g, "wins": w,
            "win_rate": w / g if g else 0.0,
        })
    rows.sort(key=lambda r: (-r["win_rate"], -r["duels"]))
    return rows


def _deck_overlap(deck_key_a, deck_key_b):
    return len(set(deck_key_a.split(", ")) & set(deck_key_b.split(", ")))


def _cluster_decks_by_overlap(deck_keys, threshold):
    """Union-find over the deck-key set: two decks are merged into the same family if they
    share at least `threshold` of their 8 cards. Returns {family_root: [deck_key, ...]}."""
    parent = {d: d for d in deck_keys}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    deck_list = list(deck_keys)
    for i in range(len(deck_list)):
        for j in range(i + 1, len(deck_list)):
            if _deck_overlap(deck_list[i], deck_list[j]) >= threshold:
                union(deck_list[i], deck_list[j])

    families = defaultdict(list)
    for d in deck_list:
        families[find(d)].append(d)
    return families


def compute_best_duel_sets(duel_log, min_duels=BEST_PICKS_MIN_DUELSET_DUELS):
    """Which combination of (up to) 3 decks used together in a duel has the best win rate.
    Exact full-deck repeats are rare, so instead of requiring identical decks, decks are
    clustered into "families" of near-duplicates (sharing at least N of 8 cards), and a
    duel-set is the combination of families its games belonged to. Starts requiring an
    exact 8/8-card match; if that (and 7/8) doesn't turn up enough distinct qualifying
    clusters, progressively loosens to 6, 5, then 4 shared cards -- the user's own
    suggested approach for finding "enough matches" despite decks rarely repeating
    exactly. The threshold actually used is always returned alongside the results, so
    it's clear how loose the comparison got."""
    by_duel = defaultdict(list)
    for r in duel_log:
        by_duel[r["duel_id"]].append(r)

    duel_decksets = {}
    for duel_id, games in by_duel.items():
        if games[0]["uncertain_start"]:
            continue
        non_rematch = [g for g in games if not g["is_rematch"]][:MAX_GAMES_PER_DUEL]
        if len(non_rematch) < MAX_GAMES_PER_DUEL:
            continue
        deck_keys = []
        valid = True
        for g in non_rematch:
            if len(g["deck"]) != 8:
                valid = False
                break
            deck_keys.append(", ".join(sorted(g["deck"])))
        if not valid:
            continue
        duel_decksets[duel_id] = deck_keys

    if not duel_decksets:
        return {"threshold_used": None, "rows": []}

    all_deck_keys = {dk for decks in duel_decksets.values() for dk in decks}

    chosen = None
    for threshold in (8, 7, 6, 5, 4):
        families = _cluster_decks_by_overlap(all_deck_keys, threshold)
        deck_to_family = {m: root for root, members in families.items() for m in members}

        cluster_duels = Counter()
        cluster_games = Counter()
        cluster_wins = Counter()
        cluster_example = {}
        for duel_id, deck_keys in duel_decksets.items():
            fam_ids = tuple(sorted(deck_to_family[dk] for dk in deck_keys))
            cluster_duels[fam_ids] += 1
            cluster_example.setdefault(fam_ids, deck_keys)
            for g in by_duel[duel_id]:
                cluster_games[fam_ids] += 1
                if g["crowns_for"] > g["crowns_against"]:
                    cluster_wins[fam_ids] += 1

        # `rows` holds EVERY cluster regardless of duel count -- the client-side "Min
        # times played" dropdown does the small-sample filtering live. `qualifying` (at
        # the BEST_PICKS_MIN_DUELSET_DUELS floor) is what decides whether this threshold
        # produced enough real clusters to stop loosening further.
        rows = []
        qualifying = 0
        for fam_ids, d in cluster_duels.items():
            g = cluster_games[fam_ids]
            w = cluster_wins[fam_ids]
            rows.append({
                "example_decks": cluster_example[fam_ids],
                "family_sizes": [len(families[fid]) for fid in fam_ids],
                "duels": d, "games": g, "wins": w,
                "win_rate": w / g if g else 0.0,
            })
            if d >= min_duels:
                qualifying += 1
        if qualifying >= BEST_PICKS_MIN_CLUSTERS_TARGET or threshold == 4:
            chosen = {"threshold_used": threshold, "rows": rows}
            if qualifying:
                break
            # threshold==4 and still nothing -- keep `chosen` as the emptiest result so
            # the caller can report "no reliable duel-set found at any tolerance" cleanly

    chosen["rows"].sort(key=lambda r: (-r["win_rate"], -r["duels"]))
    return chosen


def build_wincon_transitions(duel_log):
    """'What typically follows' data: within each duel, looking only at the first up to
    MAX_GAMES_PER_DUEL non-rematch games in order, every win condition seen in an
    EARLIER game is paired with every win condition (and every full deck) seen in a
    LATER game of that same duel. Aggregated across every tracked player's duels (not
    opponent-specific -- there usually isn't enough history against one specific
    opponent alone for this to be reliable), so this reflects general roster-wide
    sequencing tendencies. A player-specific breakdown is also built for a "just this
    player" toggle in the dashboard."""
    from collections import defaultdict as dd
    from itertools import combinations

    # DATA-QUALITY EXCLUSION (added 2026-07-17, per user request): skip uncertain-start
    # duels here too -- the first duel found for a (player, opponent) pair may really be a
    # continuation of an earlier duel we never saw (API only returns each player's most
    # recent ~25-30 battles), so its "Game 1"/"Game 2"/"Game 3" labeling can't be trusted
    # for sequence predictions. See the workbook's Data Quality sheet for exact counts.
    # ALSO scoped to Practice duels only (added 2026-07-18): Official CRL duels are tracked
    # separately (Match Category, everywhere in the workbook) but not yet blended into this
    # predictor -- sample size will be too thin at first, and practice/tournament play may
    # genuinely differ. Revisit once enough Official CRL duels accumulate.
    excluded_uncertain_duel_ids = set()
    by_duel = dd(list)
    for r in duel_log:
        if r.get("match_category") not in (None, "Practice"):
            continue
        if r["uncertain_start"]:
            excluded_uncertain_duel_ids.add(r["duel_id"])
            continue
        if not r["is_rematch"]:
            by_duel[r["duel_id"]].append(r)

    wincon_trans = dd(lambda: dd(int))          # global: from_wincon (single card) -> to_wincon -> count
    deck_trans = dd(lambda: dd(int))             # global: from_deck_key -> to_deck_key -> count
    wincon_trans_by_player = dd(lambda: dd(lambda: dd(int)))  # player -> from -> to -> count

    # Multi-select version: a deck can run more than one win condition at once (e.g.
    # Goblin Barrel + Wall Breakers + Miner in the same deck) -- keyed on the EXACT SET
    # of win cons seen in the earlier game (sorted, joined with SET_SEP), not decomposed
    # per-card. Falls back to the per-card wincon_trans above (aggregated across the
    # selected cards) when there's no history of that exact combo yet.
    SET_SEP = "::"
    wincon_set_trans = dd(lambda: dd(int))
    wincon_set_trans_by_player = dd(lambda: dd(lambda: dd(int)))

    # Game-3 predictor: same idea, but keyed on the PAIR of win conditions seen in two
    # earlier games (game 1 and game 2 of the duel, in order) -- predicts what shows up
    # in a later (game 3+) game. Needs duels with at least 3 non-rematch games tracked.
    TRIPLE_SEP = "|||"
    triple_trans = dd(lambda: dd(int))           # global: "wincon1|||wincon2" -> to_wincon -> count
    triple_trans_by_player = dd(lambda: dd(lambda: dd(int)))  # player -> "w1|||w2" -> to -> count
    triple_set_trans = dd(lambda: dd(int))        # global: "A::B|||C" (multi-wincon G1/G2) -> to -> count
    triple_set_trans_by_player = dd(lambda: dd(lambda: dd(int)))
    # Game-3 FULL DECK predictor: keyed on the exact (deck1, deck2) pair -> deck3. Exact
    # hits will be rare (huge deck space), so the dashboard falls back to a card-overlap
    # search against the plain deck_trans data (blended from both selected decks) when
    # this has no history for a given pair -- same fallback pattern as the game-2 predictor.
    deck3_trans = dd(lambda: dd(int))             # global: "deck1|||deck2" -> deck3 -> count
    all_wincons_seen = set()

    for duel_id, games in by_duel.items():
        games = sorted(games, key=lambda r: r["game_num"])[:MAX_GAMES_PER_DUEL]
        # This block is already Practice-only (Official CRL excluded above). Practice duels
        # are now always best-of-3 (3 unique decks, even after a 2-0), so a practice duel
        # with fewer than 3 distinct decks is a TRUNCATED set -- a game was lost to the
        # API's sliding window -- and its game-to-game transitions would be wrong (its
        # "game 1" may really be a later game of the set). Require a complete Bo3 here so
        # only clean 3-deck sets feed the transition/predictor models. (Added 2026-07-20
        # per user: "duel sets entered into the data [should be] complete Bo3, with 3
        # unique decks, not incomplete 2 game duels.")
        if len(games) < 3:
            continue
        player = games[0]["player_name"]
        per_game = []
        for g in games:
            wincons = classify_deck(g["deck"])
            deck_key = ", ".join(sorted(g["deck"])) if g["deck"] else None
            per_game.append((wincons, deck_key))
            all_wincons_seen.update(wincons)
        for i, j in combinations(range(len(per_game)), 2):
            from_wincons, from_deck = per_game[i]
            to_wincons, to_deck = per_game[j]
            if from_deck and to_deck:
                deck_trans[from_deck][to_deck] += 1
            from_set_key = SET_SEP.join(sorted(set(from_wincons)))
            for fw in from_wincons:
                for tw in to_wincons:
                    wincon_trans[fw][tw] += 1
                    wincon_trans_by_player[player][fw][tw] += 1
            if from_set_key:
                for tw in to_wincons:
                    wincon_set_trans[from_set_key][tw] += 1
                    wincon_set_trans_by_player[player][from_set_key][tw] += 1

        if len(per_game) >= 3:
            for i, j, k in combinations(range(len(per_game)), 3):
                g1_wincons, _ = per_game[i]
                g2_wincons, _ = per_game[j]
                g3_wincons, _ = per_game[k]
                for w1 in g1_wincons:
                    for w2 in g2_wincons:
                        key = f"{w1}{TRIPLE_SEP}{w2}"
                        for w3 in g3_wincons:
                            triple_trans[key][w3] += 1
                            triple_trans_by_player[player][key][w3] += 1
                g1_set_key = SET_SEP.join(sorted(set(g1_wincons)))
                g2_set_key = SET_SEP.join(sorted(set(g2_wincons)))
                if g1_set_key and g2_set_key:
                    combo_key = f"{g1_set_key}{TRIPLE_SEP}{g2_set_key}"
                    for w3 in g3_wincons:
                        triple_set_trans[combo_key][w3] += 1
                        triple_set_trans_by_player[player][combo_key][w3] += 1
                d1_deck = per_game[i][1]
                d2_deck = per_game[j][1]
                d3_deck = per_game[k][1]
                if d1_deck and d2_deck and d3_deck:
                    deck3_trans[f"{d1_deck}{TRIPLE_SEP}{d2_deck}"][d3_deck] += 1

    def top_n(d, n=8):
        return {k: sorted(v.items(), key=lambda kv: -kv[1])[:n] for k, v in d.items()}

    return {
        "wincon": top_n(wincon_trans),
        "deck": top_n(deck_trans, n=6),
        "wincon_by_player": {p: top_n(v) for p, v in wincon_trans_by_player.items()},
        "wincon_set": top_n(wincon_set_trans),
        "wincon_set_by_player": {p: top_n(v) for p, v in wincon_set_trans_by_player.items()},
        "wincon3": top_n(triple_trans),
        "wincon3_by_player": {p: top_n(v) for p, v in triple_trans_by_player.items()},
        "wincon3_set": top_n(triple_set_trans),
        "wincon3_set_by_player": {p: top_n(v) for p, v in triple_set_trans_by_player.items()},
        "deck3": top_n(deck3_trans, n=6),
        "all_wincons": sorted(all_wincons_seen),
        "set_sep": SET_SEP,
        "triple_sep": TRIPLE_SEP,
        "excluded_uncertain_duels": len(excluded_uncertain_duel_ids),
    }


wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# player_lookup is (re)built below, directly from compute_player_lookup() rather than the
# Excel sheet -- see the "Four variants" block after combined_duel_log, added 2026-07-18
# for the Match Category filter + weighting toggle.
# Full lists exported (no top-80 cap) -- the dashboard now has a live "Min times
# played" filter plus a "Show top N" display cap (added 2026-07-19), so trimming here
# would just throw away rows the user might want to raise the minimum past 80 to see.
wincon_sets = sorted(
    rows_of(wb, "Win-Con Sets"),
    key=lambda r: -(r["Times Played (Duels)"] or 0),
)
deck_stats = sorted(
    rows_of(wb, "Deck Stats"),
    key=lambda r: -(r["Games Played"] or 0),
)
card_icons, card_icons_evo, card_icons_hero, card_champions, card_evolution_capable = build_card_meta()
card_icons, card_icons_evo, card_icons_hero, embedded_count = embed_local_icons(
    card_icons, card_icons_evo, card_icons_hero)

# For a handful of commonly-played cards, the user wants the MAIN card art itself
# swapped to the evolution/hero form (not just a small corner badge) -- these are
# meta-defining picks where the evolved/hero look is what's instantly recognizable.
# No badge is shown for these (the main image already communicates it); every other
# card keeps the base art + small corner badge behavior as before.
FORCE_EVO_ART = {
    "Wall Breakers", "Bats", "Royal Ghost", "Lumberjack", "Goblin Cage",
    "Royal Hogs", "Goblin Barrel", "Battle Ram", "Mortar", "Skeleton Barrel",
    "Royal Giant", "Goblin Drill", "Tesla", "Archers", "Baby Dragon",
    "Inferno Dragon",
}
FORCE_HERO_ART = {"Magic Archer", "Dark Prince", "Balloon"}

forced_art_cards = set()
missing_forced_art = []
for name in FORCE_EVO_ART:
    if name in card_icons_evo:
        card_icons[name] = card_icons_evo[name]
        forced_art_cards.add(name)
    else:
        missing_forced_art.append(f"{name} (evolution)")
for name in FORCE_HERO_ART:
    if name in card_icons_hero:
        card_icons[name] = card_icons_hero[name]
        forced_art_cards.add(name)
    else:
        missing_forced_art.append(f"{name} (hero)")
if missing_forced_art:
    print("WARNING: no local art found for these forced-art cards, main icon left as "
          f"base art: {', '.join(missing_forced_art)}")

# Reuse the build script's own duel-grouping/classification logic (same source of
# truth as the workbook) to compute win-con "what follows" transitions -- avoids
# re-deriving duel/game order by re-parsing Excel text.
import sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_duel_workbook import (  # noqa: E402
    build_dataset, MAX_GAMES_PER_DUEL, classify_deck, MIN_GAMES_FOR_WINRATE_RANKING, parse_time,
    classify_match_category, compute_counter_recommendations, OFFICIAL_CRL_LIVE_CLUSTERS,
    canon_tag, canon_name,
)


# ---- Event-day labeling for Official CRL games (added 2026-07-19, per user request: "those
# day 2, top 64 games should have like a little note or flag indicating that they're day 2
# games") ----
# OFFICIAL_CRL_LIVE_CLUSTERS is an ordered list of (start_iso, end_iso, label) windows, one
# per event day discovered so far (see build_duel_workbook.py for the discovery method).
# Number them Day 1, Day 2, ... in the order they appear in that list (chronological, since
# new days are always appended, never inserted). A battle_time that doesn't fall in ANY
# known window shouldn't happen for a row already classified "Official CRL" (that
# classification itself depends on falling inside one of these windows), but fall back to
# "Day ?" rather than crashing if the list and the classifier ever drift apart.
def _event_day_label(battle_time):
    for i, (start_iso, end_iso, _label) in enumerate(OFFICIAL_CRL_LIVE_CLUSTERS):
        start = parse_time(start_iso)
        end = parse_time(end_iso)
        if start <= battle_time <= end:
            return f"Day {i + 1}"
    return "Day X"

duel_log, _duel_summary, _session_stats = build_dataset()

# ---- Shadow player profiles ----
# The ~50 tracked players' own battle logs also reveal the DECKS of everyone THEY
# played against in practice -- including opponents who aren't part of the tracked
# roster themselves (no master_<tag>.json of their own, since we only have an API
# token whitelisted for the ~50 tracked tags). Those opponents are still real players
# who could show up in an actual CRL match, so build lightweight "shadow" profiles for
# them from the opponent side of tracked players' games and fold them into the same
# searchable dashboard pool -- clearly marked as lower-confidence, one-sided data
# (only what a tracked player happened to see, never their own full practice history).
tracked_names = {r["player_name"] for r in duel_log}
mirrored_rows = []
for r in duel_log:
    if r["opponent_name"] in tracked_names:
        continue  # already has a first-hand profile from their own master file
    if not r["opponent_deck"]:
        continue
    mirrored = dict(r)
    mirrored["duel_id"] = r["duel_id"] + "_shadow"  # keep shadow-side duel grouping separate
    mirrored["player_name"] = r["opponent_name"]
    mirrored["player_tag"] = r["opponent_tag"]
    mirrored["opponent_name"] = r["player_name"]
    mirrored["opponent_tag"] = r["player_tag"]
    mirrored["deck"] = r["opponent_deck"]
    mirrored["opponent_deck"] = r["deck"]
    mirrored["crowns_for"] = r["crowns_against"]
    mirrored["crowns_against"] = r["crowns_for"]
    mirrored_rows.append(mirrored)

combined_duel_log = duel_log + mirrored_rows

# Battles we ALREADY have a first-hand or mirrored copy of, from the tracked roster's own
# accumulated master_<tag>.json archives -- keyed by (unordered tag pair, battle_time). Used
# below to de-duplicate a scouted player's own battle log against games we already know about
# from the roster's side (e.g. a scouted opponent who has also played official CRL games
# against a tracked player -- those games are already correctly classified "Official CRL" via
# the roster's own data; re-importing them from the scouted player's side as generic "Scouted"
# rows would both double-count the game AND downgrade its classification).
_known_battle_keys = {
    (frozenset((r["player_tag"], r["opponent_tag"])), r["battle_time"])
    for r in combined_duel_log
}

# ---- Scouted players (added 2026-07-18) ----
# A "scouted" player is someone NOT on the tracked roster -- no master_<tag>.json archive,
# fetched as a one-off via fetch_scout_player.py (e.g. an upcoming opponent). Their own
# battle log (scout_<tag>.json, a raw API snapshot, ~25-30 most recent battles, never
# archived/accumulated) is parsed directly into duel_log-style rows so they get a normal
# searchable dashboard profile -- deck frequency, win rate, top win conditions -- without
# needing to be part of the paired duel/session grouping used for tracked-vs-tracked
# practice. Tagged match_category="Scouted" (not "Practice"/"Official CRL") so they never
# leak into the roster's Practice-only or Official-CRL-only pools/predictors -- they only
# show up in the "All Games" pool. ALL battle types in their log are included here (not
# just clanMate/Friendly), since the point is a quick one-off look at whatever this
# specific person has been playing, not a rigorous duel-pairing analysis. Any battle already
# covered by the roster's own data (see _known_battle_keys above) is skipped here, so a
# scouted player's total game count never double-counts games we already had a copy of.
scout_paths = sorted(glob.glob("scout_*.json"))
scout_rows = []
scout_rows_deduped = 0
scouted_names = set()
for path in scout_paths:
    with open(path) as f:
        battles = json.load(f)
    for b in battles:
        team = b.get("team", [{}])[0]
        opp = b.get("opponent", [{}])[0]
        deck = [c["name"] for c in team.get("cards", [])]
        opp_deck = [c["name"] for c in opp.get("cards", [])]
        if not deck:
            continue
        battle_time = parse_time(b["battleTime"])
        key = (frozenset((team.get("tag"), opp.get("tag"))), battle_time)
        if key in _known_battle_keys:
            scout_rows_deduped += 1
            continue
        _pt = canon_tag(team.get("tag")); _pn = canon_name(team.get("tag"), team.get("name"))
        _ot = canon_tag(opp.get("tag")); _on = canon_name(opp.get("tag"), opp.get("name"))
        scouted_names.add(_pn)
        scout_rows.append({
            "duel_id": f"scout_{_pt}",
            "game_num": 1,
            "player_name": _pn,
            "player_tag": _pt,
            "opponent_name": _on,
            "opponent_tag": _ot,
            "battle_time": battle_time,
            "deck": deck,
            "opponent_deck": opp_deck,
            "crowns_for": team.get("crowns", 0),
            "crowns_against": opp.get("crowns", 0),
            "is_rematch": False,
            "uncertain_start": False,
            "match_category": "Scouted",
        })
if scout_rows or scout_rows_deduped:
    print(f"Scouted players loaded: {sorted(scouted_names)} ({len(scout_rows)} battles total "
          f"from {len(scout_paths)} scout_*.json file(s)); {scout_rows_deduped} battle(s) "
          f"skipped as already covered by the roster's own data.")
combined_duel_log = combined_duel_log + scout_rows

# ---- Extended Roster (added 2026-07-19) ----
# Permanently-tracked opponents who are NOT part of the original 48-player roster.
# Unlike a one-off "scouted" player, these accumulate an archive over time
# (extended_<tag>.json, merged/deduped by battleTime -- same shape as master_<tag>.json)
# via fetch_extended_roster.py, and are fetched on every future pickup alongside the
# original 48. Deliberately kept OUT of roster_tags used by classify_match_category /
# the Official CRL cluster signal -- if they were counted as "roster", games between them
# and the original 48 would stop being detected as Official CRL. So: tracked
# permanently, but NOT roster for classification purposes (per explicit user decision,
# 2026-07-19). Their own battles are classified with the SAME rules as the roster
# (Practice / Official CRL / excluded-type), not dumped as generic "Scouted" -- this lets
# them feed the same Best Picks / player-lookup pools their classification allows. Any
# battle already covered by the roster's own data (shared games vs. the original 48) is
# deduped the same way scouted players are, via _known_battle_keys (extended below to
# also include scout_rows first, so a player who's BOTH scouted and extended-roster --
# shouldn't normally happen, but just in case -- doesn't double count either way).
# uncertain_start is always set True for these rows (single flat battles pulled from a
# ~25-30-game recent snapshot per player, no reliable duel-boundary detection possible),
# which automatically keeps them out of Win-Con Sets / the sequence-dependent predictors
# (both already require certain-start, grouped multi-game duels) while still letting them
# appear in Player Lookup / Best Picks deck & win-rate aggregates.
_known_battle_keys |= {
    (frozenset((r["player_tag"], r["opponent_tag"])), r["battle_time"])
    for r in scout_rows
}
_roster_tags_only = {
    f"#{os.path.basename(p)[len('master_'):-len('.json')]}" for p in find_master_paths()
}
extended_paths = sorted(
    p for p in glob.glob("extended_*.json")
    if os.path.basename(p) not in ("extended_fetch_log.json", "extended_roster_tags.json")
)
extended_rows = []
extended_rows_deduped = 0
extended_rows_excluded_type = 0
extended_names = set()
for path in extended_paths:
    with open(path) as f:
        battles = json.load(f)
    for b in battles:
        team = b.get("team", [{}])[0]
        opp = b.get("opponent", [{}])[0]
        deck = [c["name"] for c in team.get("cards", [])]
        opp_deck = [c["name"] for c in opp.get("cards", [])]
        if not deck:
            continue
        battle_time = parse_time(b["battleTime"])
        btype = b.get("type")
        mode_name = b.get("gameMode", {}).get("name")
        opp_tag_raw = opp.get("tag")
        category = classify_match_category(
            btype, mode_name, battle_time,
            opponent_tag=opp_tag_raw, roster_tags=_roster_tags_only,
        )
        if category is None:
            extended_rows_excluded_type += 1
            continue
        key = (frozenset((team.get("tag"), opp.get("tag"))), battle_time)
        if key in _known_battle_keys:
            extended_rows_deduped += 1
            continue
        _pt = canon_tag(team.get("tag")); _pn = canon_name(team.get("tag"), team.get("name"))
        _ot = canon_tag(opp.get("tag")); _on = canon_name(opp.get("tag"), opp.get("name"))
        extended_names.add(_pn)
        extended_rows.append({
            "duel_id": f"extended_{_pt}_{b.get('battleTime')}",
            "game_num": 1,
            "player_name": _pn,
            "player_tag": _pt,
            "opponent_name": _on,
            "opponent_tag": _ot,
            "battle_time": battle_time,
            "deck": deck,
            "opponent_deck": opp_deck,
            "crowns_for": team.get("crowns", 0),
            "crowns_against": opp.get("crowns", 0),
            "is_rematch": False,
            "uncertain_start": True,
            "match_category": category,
            "_extended_roster": True,
        })
if extended_rows or extended_rows_deduped:
    print(f"Extended-roster players loaded: {len(extended_names)} players "
          f"({len(extended_rows)} battles total from {len(extended_paths)} "
          f"extended_*.json file(s)); {extended_rows_deduped} battle(s) skipped as "
          f"already covered by the roster's own data, {extended_rows_excluded_type} "
          f"skipped as an excluded battle type (2v2/ranked ladder/etc.).")
combined_duel_log = combined_duel_log + extended_rows
extended_names_all = extended_names

# ---- Deck Explorer data export (added 2026-07-19) ----
# Per-player flat game list (deck + match category + win/loss), used by the dashboard's
# interactive "Deck Explorer" -- a filterable most-played-decks view with a live card-
# overlap threshold (so 1-2 card swaps of the same deck cluster together), plus a
# Practice-vs-Official-CRL toggle. Originally requested scoped to the 7 Group A players,
# then the user asked to make it generally available for any player on the dashboard
# ("add that as an available feature for all players in general in my dashboard as well,
# it will be useful in the future") -- so this covers every tracked/extended-roster player
# who's ever been the "player" side of a logged game, not just Group A. Deck cards are
# stored as a single sorted comma-joined string (same format used elsewhere) to keep the
# exported JSON smaller; the client splits it back into a card list for rendering. Kept
# deliberately raw/ungrouped here -- the archetype clustering itself happens client-side
# in JS (clusterDecksByOverlap), since the overlap threshold is a live control, not a
# build-time decision.
player_decks = defaultdict(list)
for r in combined_duel_log:
    if not r.get('deck') or len(r['deck']) != 8:
        continue
    if r.get('crowns_for') is None or r.get('crowns_against') is None:
        continue
    player_decks[r['player_tag']].append({
        'd': ', '.join(sorted(r['deck'])),
        'c': 'crl' if r.get('match_category') == 'Official CRL' else 'other',
        'w': r['crowns_for'] > r['crowns_against'],
    })
player_decks = dict(player_decks)
print(f"Deck Explorer data: {len(player_decks)} players with at least one full-deck game "
      f"({sum(len(v) for v in player_decks.values())} total games indexed).")

from build_duel_workbook import compute_player_lookup  # noqa: E402


def _fmt_deck_freq(entry):
    if not entry:
        return ""
    deck, n = entry
    return f"{deck} ({n} game{'s' if n != 1 else ''})"


def _fmt_deck_winrate(entry):
    if not entry:
        return ""
    deck, wr, n = entry
    return f"{deck} ({wr:.0%} win rate, {n} game{'s' if n != 1 else ''})"


def _fmt_wincon(entry):
    if not entry:
        return ""
    card, n = entry
    return f"{card} ({n} game{'s' if n != 1 else ''})"


def _format_player_lookup(raw_rows):
    """compute_player_lookup() output -> the same JSON-friendly row shape the dashboard
    has always used (formatted deck/win-con strings), for a given duel_log slice + weight
    setting. Used to build all four dashboard variants (blended, blended-weighted,
    Practice-only, Official-CRL-only) from the same underlying data/logic."""
    out = []
    for row in raw_rows:
        freq = row["top_decks_by_freq"]
        wr_decks = row["top_decks_by_winrate"]
        wincons = row["top_wincons"]
        out.append({
            "Player": row["player"],
            "Total Games": row["total_games"],
            "Total Wins": row["total_wins"],
            "Win Rate": f"{row['win_rate']:.0%}" if row["total_games"] else "n/a",
            "Practice Games": row.get("practice_games", 0),
            "Official CRL Games": row.get("official_games", 0),
            "Most-Played Deck #1": _fmt_deck_freq(freq[0]) if len(freq) > 0 else "",
            "Most-Played Deck #2": _fmt_deck_freq(freq[1]) if len(freq) > 1 else "",
            "Most-Played Deck #3": _fmt_deck_freq(freq[2]) if len(freq) > 2 else "",
            "Best Win-Rate Deck #1": _fmt_deck_winrate(wr_decks[0]) if len(wr_decks) > 0 else "",
            "Best Win-Rate Deck #2": _fmt_deck_winrate(wr_decks[1]) if len(wr_decks) > 1 else "",
            "Best Win-Rate Deck #3": _fmt_deck_winrate(wr_decks[2]) if len(wr_decks) > 2 else "",
            "Top Win Condition #1": _fmt_wincon(wincons[0]) if len(wincons) > 0 else "",
            "Top Win Condition #2": _fmt_wincon(wincons[1]) if len(wincons) > 1 else "",
            "Top Win Condition #3": _fmt_wincon(wincons[2]) if len(wincons) > 2 else "",
            "Player Tag(s)": ", ".join(row.get("tags") or []),
            "_is_extended": row["player"] in extended_names_all and row["player"] not in scouted_names,
            "_is_shadow": (
                row["player"] not in tracked_names
                and row["player"] not in scouted_names
                and row["player"] not in extended_names_all
            ),
            "_is_scouted": row["player"] in scouted_names,
        })
    out.sort(key=lambda r: -(r.get("Total Games") or 0))
    return out


# Four variants of the same underlying data (added 2026-07-18, alongside the Match
# Category split): "All" games blended, either unweighted (plain counts -- the default,
# matches the static Excel Player Lookup sheet exactly) or weighted (Official CRL games
# count more once a deck/win-con has enough of them -- see OFFICIAL_GAME_WEIGHT /
# MIN_OFFICIAL_GAMES_FOR_WEIGHT in build_duel_workbook.py), OR filtered down to ONLY
# Practice or ONLY Official CRL games. The dashboard lets the user pick category + toggle
# weighting live -- weighting only has an effect in "All" mode.
practice_only_log = [r for r in combined_duel_log if r.get("match_category") == "Practice"]
official_only_log = [r for r in combined_duel_log if r.get("match_category") == "Official CRL"]

player_lookup = _format_player_lookup(compute_player_lookup(combined_duel_log, weighted=False))
player_lookup_weighted = _format_player_lookup(compute_player_lookup(combined_duel_log, weighted=True))
player_lookup_practice = _format_player_lookup(compute_player_lookup(practice_only_log, weighted=False))
player_lookup_official = _format_player_lookup(compute_player_lookup(official_only_log, weighted=False))

print(f"Player pool built: {len(player_lookup)} total profiles "
      f"({len(player_lookup) - len(tracked_names) - len(scouted_names) - len(extended_names_all)} "
      f"shadow/opponent-only, {len(extended_names_all)} extended-roster, {len(scouted_names)} "
      f"scouted), {len(player_lookup_official)} with any Official CRL games so far.")

# ---- Group A Scouting quick-access panel (added 2026-07-19) ----
# User's Day-2 group-stage group (snake-seeded, seeds 1/16/17/32/33/48/49/64 -- rulebook
# 4.1.3.8.3). Surfaced as its own visible dashboard section (not just searchable), since
# "where is this dashboard" feedback showed the earlier version only made these players
# findable via search, without a dedicated panel. GROUP_A_ROSTER is the single source of
# truth here -- edit this list (name, tag) to change who's shown; a player with no
# matching profile in player_lookup yet (i.e. not fetched) is flagged "pending" rather
# than silently omitted.
# status: "confirmed" (actual Group A opponent), "on_deck" (not yet confirmed, scouted
# ahead of a possible disqualification reshuffle), or "reference" (explicitly NOT in the
# user's group, kept only because the user wants to see this player's decks).
#
# Corrected 2026-07-19 per user follow-up: "remove wyze, fullly potatoe, ink from the day
# 2 analysis feature, and add sandbox, once sandbox is there that's the entirety of my
# group except viiper... keep viiper there since i want to see his decks for my
# reference." Removed Wyze❤️Ultimo / fluffypotato99 / LF丨张✨Ink❤️llb (not actually in the
# group); added SandBox (confirmed); JL Viiper's status changed confirmed -> reference.
# REPLACED 2026-07-19: per explicit user request ("replace the current day 2 opponents
# with the 15 opponents who I am up against in the monthly final, give me the same
# stats/info and dashboard features on those players as you did for the day 2 players"),
# the entire prior Day-2 Group A roster (7 opponents + 6 reference-only players) is
# replaced with the 15 possible Monthly Finals (Day 3) opponents. "status" is reused as:
#   "confirmed" -- one of the user's 5 "projected" (most-likely) Day 3 opponents
#   "on_deck"   -- one of the 10 other possible Day 3 opponents
# (no more "reference" category this round -- fully replaced, not appended to)
GROUP_A_ROSTER = [
    ("老板 Ι Batan’宙斯", "#9RQ8YRYQL", True, "confirmed"),   # you
    ("Mugi", "#2CLV2RP0", False, "confirmed"),
    ("SandBox", "#Y022GRCJQ", False, "confirmed"),
    ("40k Oker", "#YLVV0JPQ", False, "confirmed"),
    ("Mohamed Light", "#G9YV9GR8R", False, "confirmed"),
    ("Adriel", "#9CPCC890", False, "confirmed"),
    ("Pedro™️", "#RJ88Y8U08", False, "on_deck"),
    ("Asaf", "#RUQ0JU2P", False, "on_deck"),
    ("Clown (KickAsh)", "#GPPYR9JYR", False, "on_deck"),
    ("Vitor75", "#8LJ92G8UG", False, "on_deck"),
    ("Sub", "#U890Q9UQ", False, "on_deck"),
    ("SK Morten", "#R09228V", False, "on_deck"),
    ("Guriko", "#2LJ0ULYCC", False, "on_deck"),
    ("Polaris", "#U8RYGC8GU", False, "on_deck"),
    ("JorZ", "#22LC8JG02", False, "on_deck"),
    ("FrancoMedinaSL", "#UJQQCUCQ8", False, "on_deck"),
]
_lookup_by_tag = {}
for row in player_lookup:
    for t in (row.get("Player Tag(s)") or "").split(","):
        t = t.strip()
        if t:
            _lookup_by_tag[t] = row["Player"]
group_a = []
for name, tag, is_you, status in GROUP_A_ROSTER:
    matched_name = _lookup_by_tag.get(tag)
    group_a.append({
        "name": matched_name or name,
        "tag": tag,
        "is_you": is_you,
        "status": status,
        "is_confirmed": status == "confirmed",  # kept for any older code path checking this
        "has_data": matched_name is not None,
    })
print(f"Group A panel: {sum(1 for g in group_a if g['has_data'])}/{len(group_a)} members have data "
      f"({', '.join(g['name'] for g in group_a if not g['has_data']) or 'none'} pending); "
      f"{sum(1 for g in group_a if g['status']=='on_deck')} on-deck, "
      f"{sum(1 for g in group_a if g['status']=='reference')} reference-only.")

# ---- Group A CRL duel history (added 2026-07-19, per user request: "for each player
# show me all their crl game duel history ... I can click a tab or something and see all
# the matchups they played, with decks visually shown") ----
# For each Group A opponent (excluding "you"/Batan), every row from combined_duel_log
# where they're the player side AND the match is classified Official CRL -- their full
# recorded CRL duel history in this archive, deck vs. deck, most recent first. This is
# ALL of their tracked Official CRL games (not narrowed to just games vs. Batan), since
# scouting value comes from seeing their full CRL deck/matchup pattern, not only the one
# side we happen to have already played. Rendered client-side with real deck-icon strips
# (same cardIconStrip helper used everywhere else on the dashboard), not just aggregated
# stats like the rest of the profile view.
# Two parallel histories per player: Official CRL and Practice. The dashboard's history
# modal has a CRL/Practice toggle (added 2026-07-21 per user -- "practice duels are
# important to see as well"). Batan (you) is now INCLUDED here too: he's a top-16 player,
# and his second account (#9RG0VPUVY, aliased in via ALIAS_TAGS) means he now has a real
# practice history worth viewing. Practice rows carry no event-day badge (day badges are a
# CRL-tournament concept).
def _history_rows(tag, category):
    games = [
        r for r in combined_duel_log
        if r["player_tag"] == tag and r.get("match_category") == category
    ]
    games.sort(key=lambda r: r["battle_time"], reverse=True)
    out = []
    for r in games:
        row = {
            "battle_time": r["battle_time"].strftime("%Y-%m-%d %H:%M") if hasattr(r["battle_time"], "strftime") else str(r["battle_time"]),
            "opponent_name": r["opponent_name"],
            "opponent_tag": r["opponent_tag"],
            "player_deck": r["deck"],
            "opponent_deck": r["opponent_deck"],
            "crowns_for": r["crowns_for"],
            "crowns_against": r["crowns_against"],
            "result": "W" if (r["crowns_for"] or 0) > (r["crowns_against"] or 0)
                      else ("L" if (r["crowns_for"] or 0) < (r["crowns_against"] or 0) else "T"),
        }
        if category == "Official CRL":
            row["event_day"] = _event_day_label(r["battle_time"]) if hasattr(r["battle_time"], "strftime") else "Day X"
        out.append(row)
    return out

# Manual per-player watch notes (added 2026-07-21). Keyed by tag; rendered as a highlighted
# call-out at the top of that player's history modal. Add a line here per player as you learn
# things to watch for. (Kept here, not in the data files, so notes survive every data rebuild.)
PLAYER_NOTES = {
    "#2CLV2RP0": [
        "Watch out for his Mega Knight / Magic Archer / Hog Rider / Lightning cycle deck.",
        "Watch out for his Goblin Drill / Mighty Miner / Barbarian Barrel triple-spell control deck.",
        "The MK-Lightning cycle and the Drill triple-spell decks are his likely broad Game-1 blind picks (he + his coach aiming to cover a wide deck variety with one pick).",
    ],
}

group_a_history = {}           # Official CRL
group_a_practice_history = {}  # Practice
for g in group_a:
    if not g["has_data"]:
        continue
    tag = g["tag"]
    group_a_history[tag] = _history_rows(tag, "Official CRL")
    group_a_practice_history[tag] = _history_rows(tag, "Practice")
# ---- Recent Practice Trends per player (added 2026-07-21 per user) ----
# For each Group A player: their PRACTICE games since the end of the most recent Official CRL
# day (dynamic -- the last window in OFFICIAL_CRL_LIVE_CLUSTERS, so this auto-advances when a
# new event day is added). Position-agnostic (all games in a set count, since players hide
# their real Game-1 picks as Game 2/3 in practice) and rematch-excluded. Two views:
#   1. Most-run DECKS, fuzzy-grouped so near-identical lists (<=2 card variance, e.g. a bait
#      deck swapping 1-2 tech cards) count as one deck.
#   2. Most-frequent WIN-CONDITION COMBINATIONS -- the FULL win-con set of each deck, so
#      "Goblin Barrel + Wall Breakers", "Miner + Wall Breakers", and "Balloon + Miner" are
#      shown as distinct pairings rather than collapsed into a bare "Wall Breakers".
_RECENT_CUTOFF = max(parse_time(end) for _s, end, _l in OFFICIAL_CRL_LIVE_CLUSTERS) \
    if OFFICIAL_CRL_LIVE_CLUSTERS else None

def _cluster_decks(decks):
    # union-find: decks sharing >=6/8 cards (<=2 card variance) merge into one cluster
    n = len(decks); parent = list(range(n)); sets = [set(d) for d in decks]
    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]; i = parent[i]
        return i
    for i in range(n):
        for j in range(i + 1, n):
            if len(sets[i] & sets[j]) >= 6:
                parent[find(i)] = find(j)
    groups = defaultdict(list)
    for i in range(n):
        groups[find(i)].append(i)
    return list(groups.values())

group_a_recent_practice = {}
for g in group_a:
    if not g["has_data"] or _RECENT_CUTOFF is None:
        continue
    tag = g["tag"]
    games = [r for r in combined_duel_log
             if r["player_tag"] == tag and r.get("match_category") == "Practice"
             and not r.get("is_rematch") and r["deck"] and len(r["deck"]) == 8
             and hasattr(r["battle_time"], "strftime") and r["battle_time"] > _RECENT_CUTOFF]
    if not games:
        group_a_recent_practice[tag] = {"n_games": 0, "clusters": [], "wincons": []}
        continue
    decks = [list(r["deck"]) for r in games]
    clusters = _cluster_decks(decks)
    clusters.sort(key=lambda idxs: -len(idxs))
    cl_out = []
    for idxs in clusters[:5]:
        exact = Counter(tuple(decks[i]) for i in idxs)
        rep = list(exact.most_common(1)[0][0])
        wc = classify_deck(rep)
        cl_out.append({
            "count": len(idxs),
            "deck": rep,
            "wincon": " + ".join(wc) if wc else "(no win-con)",
        })
    # win-con COMBINATIONS (full set per deck), not decomposed
    combo = Counter()
    for r in games:
        wc = classify_deck(r["deck"])
        combo[" + ".join(wc) if wc else "(no win-con)"] += 1
    wincons = [{"combo": k, "count": v} for k, v in combo.most_common(6)]
    group_a_recent_practice[tag] = {
        "n_games": len(games),
        "cutoff": _RECENT_CUTOFF.strftime("%b %d"),
        "clusters": cl_out,
        "wincons": wincons,
    }
print("Recent practice trends computed for "
      f"{sum(1 for v in group_a_recent_practice.values() if v['n_games'])} players "
      f"(cutoff {_RECENT_CUTOFF.strftime('%Y-%m-%d %H:%M') if _RECENT_CUTOFF else 'n/a'}).")

print("Group A duel history (CRL / Practice): " + ", ".join(
    f"{g['name']} {len(group_a_history.get(g['tag'], []))}c/{len(group_a_practice_history.get(g['tag'], []))}p"
    for g in group_a if g["has_data"]
))

# ---- Group A "Recommended for Tomorrow" counter-recommendations (added 2026-07-19,
# per explicit user request) ----
# For each Group A / on-deck opponent: pull their own top win-condition sets from every
# game of theirs in combined_duel_log (Practice + Official CRL, all their games we have --
# not narrowed to Official CRL only, unlike the history view above, since we want their
# broadest real deck-choice signal). Then filter OUR ENTIRE tracked data pool
# (combined_duel_log -- every roster + extended-roster player's Practice + Official CRL
# games combined, per the user's explicit "all games, practice and crl as a whole"
# instruction) down to games where we faced an opponent playing one of those same win
# conditions, and rank our own best-performing deck / win-con set / duel-set pair within
# that filtered slice by real win rate. Purely empirical -- no theorycrafted "X counters
# Y" claim, and every recommendation keeps its games/win-rate so a thin sample is visibly
# thin. Mirrors the same "Recommended for Tomorrow" section in add_group_a_analysis.py's
# Excel sheet (via the shared compute_counter_recommendations helper), though the data
# pool here is combined_duel_log (roster + extended-roster) rather than that script's
# narrower main-roster-only pool -- intentionally the broadest pool available.
group_a_recommendations = {}
for g in group_a:
    if g["is_you"] or not g["has_data"]:
        continue
    tag = g["tag"]
    their_games = [r for r in combined_duel_log if r["player_tag"] == tag]
    wincon_games = Counter()
    for r in their_games:
        for wc in (classify_deck(r["deck"]) or []):
            wincon_games[wc] += 1
    top_wincons = [wc for wc, _n in wincon_games.most_common(3) if wc != "(none classified)"][:2]
    if top_wincons:
        reco = compute_counter_recommendations(combined_duel_log, top_wincons, min_games=3, top_n=3)
    else:
        reco = {"sample_size": 0, "best_decks": [], "best_wincon_sets": [], "best_duel_sets": []}
    group_a_recommendations[tag] = {"top_wincons": top_wincons, **reco}
print("Group A recommendations: " + ", ".join(
    f"{g['name']} (pool {group_a_recommendations[g['tag']]['sample_size']}g)"
    for g in group_a if g["has_data"] and not g["is_you"]
))

# ---- Group A Matchup Prep (added 2026-07-19, per explicit user request: "For the my
# matchup analysis page, can you make such a page in the excel, and a feature as well in
# the day 2, for all my day 2 opponents") ----
# The flip side of "Recommended for Tomorrow" above (which only uses OUR own results to
# find what WE should bring): this answers "what's most likely to beat THEM, and what
# already has" -- the same structure as the Excel "My Matchup Analysis" sheet built for the
# user, generalized to each of the 7 actual Day 2 opponents (confirmed + on-deck; reference
# players and "you" excluded, same scope as Group A Sequencing below).
# 1) Their own top win-cons (from their own games).
# 2) EMPIRICAL -- decks that have actually beaten them personally (min 3 games faced).
# 3) PREDICTIVE -- across the ENTIRE tracked pool, what beats players using their top
#    win-cons (much larger sample than their own history alone).
# 4) Double-confirmed cross-reference of 2 and 3.
MATCHUP_PREP_MIN_GAMES_DECK = 3
MATCHUP_PREP_MIN_GAMES_WINCON = 5


def compute_matchup_prep(duel_log, target_tag):
    target_games = [r for r in duel_log if r["player_tag"] == target_tag and r["deck"] and len(r["deck"]) == 8]
    wincon_games = Counter()
    wincon_wins = Counter()
    for r in target_games:
        won = r["crowns_for"] > r["crowns_against"]
        for wc in (classify_deck(r["deck"]) or []):
            wincon_games[wc] += 1
            if won:
                wincon_wins[wc] += 1
    top_wincons = [wc for wc, _g in wincon_games.most_common(3)]

    opp_deck_games = Counter()
    opp_deck_wins = Counter()
    for r in target_games:
        if not r.get("opponent_deck") or len(r["opponent_deck"]) != 8:
            continue
        dk = ", ".join(sorted(r["opponent_deck"]))
        opp_deck_games[dk] += 1
        if r["crowns_for"] < r["crowns_against"]:
            opp_deck_wins[dk] += 1
    empirical = sorted(
        [{"deck": dk, "games": g, "wins": opp_deck_wins[dk], "win_rate": opp_deck_wins[dk] / g}
         for dk, g in opp_deck_games.items() if g >= MATCHUP_PREP_MIN_GAMES_DECK],
        key=lambda x: (-x["win_rate"], -x["games"]),
    )[:8]

    target_set = set(top_wincons)
    predicted_decks, predicted_wincons, sample_size = [], [], 0
    if target_set:
        filtered = [r for r in duel_log if r.get("deck") and set(classify_deck(r["deck"]) or []) & target_set]
        sample_size = len(filtered)
        pdg, pdw, pwg, pww = Counter(), Counter(), Counter(), Counter()
        for r in filtered:
            if not r.get("opponent_deck") or len(r["opponent_deck"]) != 8:
                continue
            dk = ", ".join(sorted(r["opponent_deck"]))
            pdg[dk] += 1
            opp_won = r["crowns_for"] < r["crowns_against"]
            if opp_won:
                pdw[dk] += 1
            for wc in (classify_deck(r["opponent_deck"]) or []):
                pwg[wc] += 1
                if opp_won:
                    pww[wc] += 1
        predicted_decks = sorted(
            [{"deck": dk, "games": g, "wins": pdw[dk], "win_rate": pdw[dk] / g}
             for dk, g in pdg.items() if g >= MATCHUP_PREP_MIN_GAMES_DECK],
            key=lambda x: (-x["win_rate"], -x["games"]),
        )[:6]
        predicted_wincons = sorted(
            [{"wincon": wc, "games": g, "wins": pww[wc], "win_rate": pww[wc] / g}
             for wc, g in pwg.items() if g >= MATCHUP_PREP_MIN_GAMES_WINCON],
            key=lambda x: (-x["win_rate"], -x["games"]),
        )[:6]

    personal_threats = {dk for dk, g in opp_deck_games.items() if g >= 2 and opp_deck_wins[dk] / g >= 0.5}
    predicted_keys = {row["deck"] for row in predicted_decks}
    double_confirmed = sorted(personal_threats & predicted_keys)

    return {
        "total_games": len(target_games),
        "top_wincons": top_wincons,
        "wincon_games": [{"wincon": wc, "games": g, "wins": wincon_wins[wc], "win_rate": wincon_wins[wc] / g}
                          for wc, g in wincon_games.most_common(6)],
        "empirical": empirical,
        "predicted_sample_size": sample_size,
        "predicted_decks": predicted_decks,
        "predicted_wincons": predicted_wincons,
        "double_confirmed": double_confirmed,
    }


group_a_matchup_prep = {}
for g in group_a:
    if g["is_you"] or not g["has_data"] or g["status"] not in ("confirmed", "on_deck"):
        continue
    group_a_matchup_prep[g["tag"]] = compute_matchup_prep(combined_duel_log, g["tag"])
print("Group A Matchup Prep: " + ", ".join(
    f"{g['name']} (pool {group_a_matchup_prep[g['tag']]['predicted_sample_size']}g)"
    for g in group_a if g["tag"] in group_a_matchup_prep
))

# ---- Group A Sequencing (added 2026-07-19, per explicit user request, same feature as
# add_group_a_sequencing_analysis.py's Excel sheet, ported to the dashboard) ----
# Scoped to the same 7 confirmed+on_deck players as Matchup Prep above. Two angles:
# 1) Spell + win-condition combos across all their logged decks.
# 2) B03 positional tendencies -- which win-con/spell they tend to slot into which
#    position (Game 1/2/3) of a duel set. NOTE: a naive "does their win-con/spell change
#    game-to-game" stat was tried and dropped -- real Clash Royale Duel format bans reusing
#    any card across a duel's games, so that comes out ~100% for everyone by the game's own
#    rules, not a behavioral signal (see add_group_a_sequencing_analysis.py for the full
#    writeup). The position-based "usual slot" view is the real signal.
SEQ_SPELL_CARDS = {
    "Arrows", "Barbarian Barrel", "Clone", "Earthquake", "Fireball", "Freeze",
    "Giant Snowball", "Goblin Curse", "Lightning", "Mirror", "Poison", "Rage",
    "Rocket", "Royal Delivery", "The Log", "Tornado", "Void", "Zap",
}


def compute_group_a_sequencing(duel_log, target_tag):
    player_games = [r for r in duel_log if r["player_tag"] == target_tag and r["deck"] and len(r["deck"]) == 8]

    combo_games, combo_wins = Counter(), Counter()
    for r in player_games:
        wc = tuple(classify_deck(r["deck"]) or [])
        sp = tuple(sorted(set(r["deck"]) & SEQ_SPELL_CARDS))
        key = (wc, sp)
        combo_games[key] += 1
        if r["crowns_for"] > r["crowns_against"]:
            combo_wins[key] += 1
    combo_rows = sorted(
        [{"wincons": list(k[0]), "spells": list(k[1]), "games": g, "wins": combo_wins[k], "win_rate": combo_wins[k] / g}
         for k, g in combo_games.items()],
        key=lambda x: (-x["games"], -x["win_rate"]),
    )[:8]

    sequencing_games = [r for r in player_games if not r.get("is_rematch")]
    duels = defaultdict(list)
    for r in sequencing_games:
        duels[r["duel_id"]].append(r)

    position_wincons = {1: Counter(), 2: Counter(), 3: Counter()}
    position_spells = {1: Counter(), 2: Counter(), 3: Counter()}
    wincon_position_counts = defaultdict(Counter)
    spell_position_counts = defaultdict(Counter)
    multi_game_duels = 0
    for duel_id, games in duels.items():
        games = sorted(games, key=lambda r: r["battle_time"])[:3]
        # Completeness gate (added 2026-07-20 per user): practice duels are now always
        # best-of-3, so a practice duel with <3 distinct decks is a truncated set (a game
        # aged out of the API window) and its Game-1/2/3 position labels are unreliable --
        # exclude it. Official CRL is different: a 2-0 sweep is a complete 2-game result,
        # so >=2 is valid there.
        cat = games[0].get("match_category")
        if cat == "Practice":
            if len(games) < 3:
                continue
        elif len(games) < 2:
            continue
        multi_game_duels += 1
        for i, gm in enumerate(games, start=1):
            wc = classify_deck(gm["deck"]) or []
            sp = sorted(set(gm["deck"]) & SEQ_SPELL_CARDS)
            position_wincons[i].update(wc)
            position_spells[i].update(sp)
            for w in wc:
                wincon_position_counts[w][i] += 1
            for s in sp:
                spell_position_counts[s][i] += 1

    def modal_rows(position_counts, min_games=2):
        rows = []
        for name, counts in position_counts.items():
            total = sum(counts.values())
            if total < min_games:
                continue
            modal_pos, modal_n = counts.most_common(1)[0]
            rows.append({
                "name": name, "total": total, "modal_pos": modal_pos, "modal_n": modal_n,
                "breakdown": ", ".join(f"G{p}:{counts.get(p, 0)}" for p in (1, 2, 3) if counts.get(p, 0)),
            })
        rows.sort(key=lambda x: (-x["total"], x["modal_pos"]))
        return rows[:6]

    return {
        "total_games": len(player_games),
        "combo_rows": combo_rows,
        "multi_game_duels": multi_game_duels,
        "position_wincons": {p: c.most_common(3) for p, c in position_wincons.items()},
        "position_spells": {p: c.most_common(3) for p, c in position_spells.items()},
        "wincon_position_rows": modal_rows(wincon_position_counts),
        "spell_position_rows": modal_rows(spell_position_counts),
    }


group_a_sequencing = {}
for g in group_a:
    if g["is_you"] or not g["has_data"] or g["status"] not in ("confirmed", "on_deck"):
        continue
    group_a_sequencing[g["tag"]] = compute_group_a_sequencing(combined_duel_log, g["tag"])
print("Group A Sequencing: " + ", ".join(
    f"{g['name']} ({group_a_sequencing[g['tag']]['multi_game_duels']} multi-duels)"
    for g in group_a if g["tag"] in group_a_sequencing
))


# ---- Duel-Set Record per player (added 2026-07-22 per user) ----
# For each Group A player (INCLUDING Batan -- the user asked for "each top 16 player"):
# their duel-SET outcome distribution and set win rate, split by Official CRL vs Practice.
#
# The two categories are scored differently BY DESIGN, because the formats differ:
#   * Official CRL is first-to-2 and STOPS: a set is 2 games (a 2-0/0-2 sweep) or 3 (2-1/1-2).
#     A 2-game 1-1 is not a result -- it's a pending set whose game 3 isn't fetched yet -- so
#     it's excluded (mirrors compute_crl_duel_status). Set win = reached 2 game-wins.
#   * Practice is now ALWAYS a full best-of-three -- the 3rd game is played even after someone
#     goes 2-0, to explore deck variety. So a practice set is decided at 2 wins (first-to-2)
#     but a "dead rubber" game 3 still gets played. That means the interesting question the
#     user asked -- "do they go 2-0 then lose the last game?" -- lives ONLY in practice, and
#     needs the first-two-games sequence, not just the final tally (W-W-L and W-L-W are both
#     "2-1" on the scoreboard but mean completely different things). So practice sets are
#     bucketed by how the first two games went, then what happened in game 3:
#        up 2-0  (W W .) -> closed it out 3-0   |  dropped the dead rubber (2-0 -> 2-1 final)
#        even 1-1 (split) -> won the decider 2-1 |  lost the decider 1-2
#        down 0-2 (L L .) -> got swept 0-3       |  won the dead rubber (0-2 -> 1-2 final)
#     Set win (first-to-2) = went up 2-0, OR was even 1-1 and won the decider.
#
# Same completeness gate as sequencing: 8-card decks only, rematch (re-practice) games
# dropped, practice needs 3 distinct games to count as a complete set, CRL needs the sweep/
# decided shapes above.

def _win_of(r):
    return r["crowns_for"] > r["crowns_against"]

def compute_duel_set_record(duel_log, tag):
    by_duel = defaultdict(list)
    for r in duel_log:
        if r["player_tag"] != tag:
            continue
        if r.get("match_category") not in ("Official CRL", "Practice"):
            continue
        if not r["deck"] or len(r["deck"]) != 8:
            continue
        by_duel[r["duel_id"]].append(r)

    crl = {"sets": 0, "wins": 0, "2-0": 0, "2-1": 0, "1-2": 0, "0-2": 0, "anomaly": 0}
    prac = {
        "sets": 0, "wins": 0,
        # by full game record (for a quick scoreboard view)
        "3-0": 0, "2-1": 0, "1-2": 0, "0-3": 0,
        # by first-two-games sequence, then game 3 -- the meaningful practice buckets
        "up20": 0, "up20_closed": 0, "up20_dropped": 0,       # went 2-0: closed 3-0 vs dropped g3
        "even": 0, "even_won": 0, "even_lost": 0,             # split 1-1: won vs lost the decider
        "down02": 0, "down02_swept": 0, "down02_rubber": 0,   # went 0-2: swept 0-3 vs won dead rubber
    }

    for games in by_duel.values():
        distinct = [r for r in sorted(games, key=lambda x: x["battle_time"]) if not r.get("is_rematch")][:3]
        cat = distinct[0].get("match_category") if distinct else None
        res = [_win_of(r) for r in distinct]
        w = sum(res)
        n = len(res)

        if cat == "Official CRL":
            # Only count decided sets: a 2-game sweep, or a 3-game decided Bo3.
            if n == 2 and w == 2:
                crl["2-0"] += 1; crl["wins"] += 1; crl["sets"] += 1
            elif n == 2 and w == 0:
                crl["0-2"] += 1; crl["sets"] += 1
            elif n == 3 and w == 2:
                crl["2-1"] += 1; crl["wins"] += 1; crl["sets"] += 1
            elif n == 3 and w == 1:
                crl["1-2"] += 1; crl["sets"] += 1
            else:
                # 2-game 1-1 (pending), 1-game (pending), or a >3 / 3-0 / 0-3 anomaly
                if n >= 2:
                    crl["anomaly"] += 1
        elif cat == "Practice":
            if n < 3:
                continue  # truncated set -- a game aged out of the API window
            prac["sets"] += 1
            # scoreboard record
            prac[f"{w}-{n - w}"] = prac.get(f"{w}-{n - w}", 0) + 1
            first_two = res[:2]
            g3 = res[2]
            if first_two == [True, True]:          # up 2-0
                prac["up20"] += 1; prac["wins"] += 1
                if g3: prac["up20_closed"] += 1
                else:  prac["up20_dropped"] += 1
            elif first_two == [False, False]:      # down 0-2
                prac["down02"] += 1
                if g3: prac["down02_rubber"] += 1
                else:  prac["down02_swept"] += 1
            else:                                  # even 1-1 after two
                prac["even"] += 1
                if g3: prac["even_won"] += 1; prac["wins"] += 1
                else:  prac["even_lost"] += 1

    crl["win_rate"] = (crl["wins"] / crl["sets"]) if crl["sets"] else None
    prac["win_rate"] = (prac["wins"] / prac["sets"]) if prac["sets"] else None
    return {"crl": crl, "practice": prac}


group_a_duel_set_record = {}
for g in group_a:
    if not g["has_data"]:
        continue
    group_a_duel_set_record[g["tag"]] = compute_duel_set_record(combined_duel_log, g["tag"])
print("Duel-Set Record: " + ", ".join(
    f"{g['name']} (CRL {group_a_duel_set_record[g['tag']]['crl']['sets']}s"
    + (f"/{group_a_duel_set_record[g['tag']]['crl']['win_rate']*100:.0f}%" if group_a_duel_set_record[g['tag']]['crl']['win_rate'] is not None else "")
    + f", Prac {group_a_duel_set_record[g['tag']]['practice']['sets']}s"
    + (f"/{group_a_duel_set_record[g['tag']]['practice']['win_rate']*100:.0f}%" if group_a_duel_set_record[g['tag']]['practice']['win_rate'] is not None else "")
    + ")"
    for g in group_a if g["tag"] in group_a_duel_set_record
))

# "Best Picks" quick-reference feature (added 2026-07-18) -- top win-rate decks, win-con
# sets, and duel (deck) sets, computed separately per Match Category so the game-day view
# never blends Practice and Official CRL results.
best_picks = {}
for label, log_slice in (("all", combined_duel_log), ("practice", practice_only_log), ("official", official_only_log)):
    duel_set_result = compute_best_duel_sets(log_slice)
    best_picks[label] = {
        "decks": compute_best_decks(log_slice),
        "wincon_sets": compute_best_wincon_sets(log_slice),
        "duel_sets": duel_set_result["rows"],
        "duel_sets_threshold": duel_set_result["threshold_used"],
    }
print(f"Best Picks computed -- All: {len(best_picks['all']['decks'])} ranked decks, "
      f"{len(best_picks['all']['wincon_sets'])} ranked win-con sets, "
      f"{len(best_picks['all']['duel_sets'])} ranked duel sets "
      f"(deck-overlap threshold {best_picks['all']['duel_sets_threshold']}/8 cards). "
      f"Official CRL: {len(best_picks['official']['decks'])} ranked decks so far.")

transitions = build_wincon_transitions(combined_duel_log)
card_elixir = build_card_elixir()
player_briefs = build_player_briefs(combined_duel_log, card_elixir, MIN_GAMES_FOR_WINRATE_RANKING)
all_cards = sorted(card_icons.keys())

data = {
    "player_lookup": player_lookup,
    "player_lookup_weighted": player_lookup_weighted,
    "player_lookup_practice": player_lookup_practice,
    "player_lookup_official": player_lookup_official,
    "group_a": group_a,
    "group_a_history": group_a_history,
    "group_a_practice_history": group_a_practice_history,
    "group_a_recent_practice": group_a_recent_practice,
    # normalize each player's notes to a list (a single string still works as one note)
    "player_notes": {k: (v if isinstance(v, list) else [v]) for k, v in PLAYER_NOTES.items()},
    "group_a_recommendations": group_a_recommendations,
    "group_a_matchup_prep": group_a_matchup_prep,
    "group_a_sequencing": group_a_sequencing,
    "group_a_duel_set_record": group_a_duel_set_record,
    "player_decks": player_decks,
    "best_picks": best_picks,
    "all_cards": all_cards,
    "wincon_sets_top": wincon_sets,
    "deck_stats_top": deck_stats,
    "card_icons": card_icons,
    "card_icons_evo": card_icons_evo,
    "card_icons_hero": card_icons_hero,
    "card_champions": sorted(card_champions),
    "card_evolution_capable": sorted(card_evolution_capable),
    "forced_art_cards": sorted(forced_art_cards),
    "transitions": transitions,
    "player_briefs": player_briefs,
}
data_json = json.dumps(data, ensure_ascii=False)
print(f"Card icons found for {len(card_icons)} distinct cards "
      f"({len(card_champions)} champion cards, {len(card_evolution_capable)} seen evolved).")
print(f"Evolution-form art available for {len(card_icons_evo)} cards, "
      f"hero-form art available for {len(card_icons_hero)} cards.")
if embedded_count:
    print(f"Embedded {embedded_count} base/evolution/hero icon files as base64 "
          f"(offline-capable) from local card_icons/ folder.")
else:
    print("No local card_icons/ folder found -- icons are hotlinked (need internet to display). "
          "Run download_card_icons.py somewhere with internet access, then re-run this script "
          "to embed them for offline use.")
print(f"Win-con transition pairs computed: {sum(len(v) for v in transitions['wincon'].values())}")
print(f"Game-3 (G1+G2 -> G3) win-con transitions computed: {len(transitions['wincon3'])} distinct G1+G2 combos, "
      f"{sum(len(v) for v in transitions['wincon3'].values())} total transitions")

html = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>CRL Opponent Scout</title>
<style>
  :root {
    color-scheme: light;
    --surface-1:      #fcfcfb;
    --page-plane:      #f9f9f7;
    --text-primary:   #0b0b0b;
    --text-secondary: #52514e;
    --text-muted:     #898781;
    --gridline:       #e1e0d9;
    --baseline:       #c3c2b7;
    --border:         rgba(11,11,11,0.10);
    --series-blue:    #2a78d6;
    --series-green:   #008300;
    --series-magenta: #e87ba4;
    --series-yellow:  #eda100;
    --series-aqua:    #1baf7a;
    --series-orange:  #eb6834;
    --series-violet:  #4a3aa7;
    --series-red:     #e34948;
    --good:           #0ca30c;
    --warning:        #fab219;
    --serious:        #ec835a;
    --critical:       #d03b3b;
    --seq-100: #cde2fb; --seq-200: #9ec5f4; --seq-300: #6da7ec;
    --seq-400: #3987e5; --seq-500: #256abf; --seq-600: #184f95;
  }
  @media (prefers-color-scheme: dark) {
    :root:where(:not([data-theme="light"])) {
      color-scheme: dark;
      --surface-1:      #1a1a19;
      --page-plane:      #0d0d0d;
      --text-primary:   #ffffff;
      --text-secondary: #c3c2b7;
      --text-muted:     #898781;
      --gridline:       #2c2c2a;
      --baseline:       #383835;
      --border:         rgba(255,255,255,0.10);
      --series-blue:    #3987e5;
      --series-green:   #008300;
      --series-magenta: #d55181;
      --series-yellow:  #c98500;
      --series-aqua:    #199e70;
      --series-orange:  #d95926;
      --series-violet:  #9085e9;
      --series-red:     #e66767;
      --good:           #0ca30c;
      --warning:        #fab219;
      --serious:        #ec835a;
      --critical:       #d03b3b;
    }
  }

  * { box-sizing: border-box; }
  html, body {
    margin: 0; padding: 0;
    background: var(--page-plane);
    color: var(--text-primary);
    font-family: system-ui, -apple-system, "Segoe UI", sans-serif;
  }
  .wrap { max-width: 1180px; margin: 0 auto; padding: 24px 20px 60px; }

  header.top { margin-bottom: 20px; }
  h1 { font-size: 22px; margin: 0 0 4px; letter-spacing: -0.01em; }
  .subtitle { color: var(--text-secondary); font-size: 13.5px; margin: 0 0 18px; }

  .search-box {
    position: relative;
    margin-bottom: 8px;
  }
  .search-box input {
    width: 100%;
    font-size: 17px;
    padding: 13px 16px;
    border-radius: 10px;
    border: 1px solid var(--border);
    background: var(--surface-1);
    color: var(--text-primary);
    outline: none;
  }
  .search-box input:focus { border-color: var(--series-blue); }
  .suggestions {
    position: absolute;
    top: calc(100% + 4px);
    left: 0; right: 0;
    background: var(--surface-1);
    border: 1px solid var(--border);
    border-radius: 10px;
    box-shadow: 0 8px 24px rgba(0,0,0,0.12);
    max-height: 260px;
    overflow-y: auto;
    z-index: 20;
    display: none;
  }
  .suggestions.show { display: block; }
  .suggestions button {
    display: block; width: 100%; text-align: left;
    padding: 10px 16px; font-size: 14.5px;
    background: none; border: none; color: var(--text-primary);
    cursor: pointer;
  }
  .suggestions button:hover, .suggestions button.active { background: var(--gridline); }

  .quick-roster { display: flex; flex-wrap: wrap; gap: 6px; margin: 10px 0 28px; }
  .chip {
    font-size: 12.5px; padding: 5px 10px; border-radius: 999px;
    border: 1px solid var(--border); background: var(--surface-1);
    color: var(--text-secondary); cursor: pointer;
  }
  .chip:hover { border-color: var(--series-blue); color: var(--series-blue); }

  .profile {
    display: none;
    background: var(--surface-1);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 22px 24px;
    margin-bottom: 30px;
  }
  .profile.show { display: block; }
  .profile-head {
    display: flex; align-items: baseline; justify-content: space-between;
    flex-wrap: wrap; gap: 10px; margin-bottom: 16px;
  }
  .profile-head h2 { margin: 0; font-size: 24px; }
  .stat-row { display: flex; gap: 28px; flex-wrap: wrap; }
  .stat-tile { min-width: 90px; }
  .stat-tile .num {
    font-size: 26px; font-weight: 600; font-variant-numeric: tabular-nums;
    line-height: 1.1;
  }
  .stat-tile .label { font-size: 12px; color: var(--text-muted); margin-top: 2px; }
  .stat-tile .num.winrate-good { color: var(--good); }
  .stat-tile .num.winrate-warn { color: var(--warning); }
  .stat-tile .num.winrate-bad { color: var(--critical); }

  .section-title {
    font-size: 12.5px; text-transform: uppercase; letter-spacing: 0.04em;
    color: var(--text-muted); margin: 22px 0 10px; font-weight: 600;
  }

  .brief-card {
    margin-top: 18px; border: 1.5px solid var(--seq-400); border-radius: 12px;
    background: var(--seq-100); padding: 14px 16px 16px;
  }
  .brief-head { display: flex; align-items: baseline; gap: 10px; flex-wrap: wrap; margin-bottom: 10px; }
  .brief-title { font-size: 13px; font-weight: 700; color: var(--seq-600); text-transform: uppercase; letter-spacing: 0.03em; }
  .brief-sub { font-size: 11.5px; color: var(--text-muted); }
  .brief-grid { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 14px; }
  @media (max-width: 820px) { .brief-grid { grid-template-columns: 1fr; } }
  .brief-col-label {
    font-size: 11px; text-transform: uppercase; letter-spacing: 0.03em;
    color: var(--text-muted); font-weight: 600; margin-bottom: 6px;
  }
  .brief-wincon-row { display: flex; align-items: center; gap: 6px; margin-bottom: 4px; font-size: 12.5px; }
  .brief-wincon-row .n { color: var(--text-muted); font-size: 11px; }
  .brief-deck-row { display: flex; align-items: center; gap: 6px; margin-bottom: 6px; }
  .brief-deck-row .meta { font-size: 11px; color: var(--text-muted); }
  .brief-elixir {
    display: inline-flex; align-items: baseline; gap: 6px; font-size: 22px; font-weight: 700;
    font-variant-numeric: tabular-nums; color: var(--seq-600);
  }
  .brief-elixir .unit { font-size: 12px; font-weight: 400; color: var(--text-muted); }
  .brief-elixir-tag {
    display: inline-block; margin-top: 4px; font-size: 11px; padding: 2px 8px;
    border-radius: 999px; background: var(--seq-200); color: var(--seq-600); font-weight: 600;
  }
  .brief-empty { font-size: 12px; color: var(--text-muted); }
  .card-grid { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; }
  @media (max-width: 820px) { .card-grid { grid-template-columns: 1fr; } }
  .deck-card {
    background: var(--page-plane);
    border: 1px solid var(--border);
    border-radius: 10px;
    padding: 12px 14px;
    font-size: 13px;
  }
  .deck-card .deck-meta {
    font-size: 11.5px; color: var(--text-muted); margin-top: 8px;
  }
  .icon-strip { display: flex; flex-wrap: wrap; gap: 4px; }
  .card-icon-wrap { position: relative; display: inline-block; }
  .card-icon {
    border-radius: 5px; border: 1px solid var(--border);
    object-fit: cover; background: var(--surface-1); display: block;
  }
  .card-icon-fallback {
    width: 32px; height: 40px; border-radius: 5px;
    border: 1px solid var(--border); background: var(--gridline);
    color: var(--text-muted); font-size: 13px; font-weight: 700;
    display: flex; align-items: center; justify-content: center;
  }
  .badge-champ, .badge-evo {
    position: absolute; top: -4px; right: -4px;
    width: 14px; height: 14px; border-radius: 50%;
    font-size: 9px; line-height: 14px; text-align: center;
    border: 1px solid var(--surface-1);
  }
  .badge-champ { background: var(--series-yellow); color: #4a3400; right: -4px; }
  .badge-evo { background: var(--series-violet); color: #fff; right: 10px; }
  .badge-champ.badge-img, .badge-evo.badge-img {
    background: var(--surface-1); padding: 0; overflow: hidden;
  }
  .badge-champ.badge-img img, .badge-evo.badge-img img {
    width: 100%; height: 100%; object-fit: cover; display: block;
  }
  .icon-legend {
    display: flex; gap: 16px; align-items: center;
    font-size: 11.5px; color: var(--text-muted); margin: 6px 0 0;
  }
  .icon-legend span.dot {
    display: inline-block; width: 12px; height: 12px; border-radius: 50%;
    margin-right: 4px; vertical-align: -1px;
  }
  .row-label { font-size: 11px; color: var(--text-muted); margin-top: 5px; max-width: 320px; }
  .wincon-badges { display: flex; flex-wrap: wrap; gap: 10px; }
  .wincon-badge {
    display: flex; align-items: center; gap: 10px;
    background: var(--seq-100); color: var(--seq-600);
    border-radius: 10px; padding: 8px 14px 8px 8px; font-size: 13px; font-weight: 600;
  }
  .wincon-badge .wincon-meta { font-size: 11px; font-weight: 400; opacity: 0.8; margin-top: 1px; }
  @media (prefers-color-scheme: dark) {
    .wincon-badge { background: rgba(57,135,229,0.18); color: #9ec5f4; }
  }
  .empty-hint {
    color: var(--text-secondary); font-size: 14px; padding: 30px 0;
    text-align: center;
  }

  h3.table-title {
    font-size: 15px; margin: 34px 0 8px;
    display: flex; align-items: center; gap: 8px;
  }
  .table-note { font-size: 12px; color: var(--text-muted); font-weight: 400; }
  .table-wrap {
    background: var(--surface-1);
    border: 1px solid var(--border);
    border-radius: 12px;
    overflow: hidden;
  }
  table { width: 100%; border-collapse: collapse; font-size: 13px; }
  thead th {
    text-align: left; padding: 10px 12px;
    font-size: 11.5px; text-transform: uppercase; letter-spacing: 0.03em;
    color: var(--text-muted); border-bottom: 1px solid var(--gridline);
    cursor: pointer; user-select: none; white-space: nowrap;
  }
  thead th:hover { color: var(--text-primary); }
  thead th .arrow { font-size: 10px; margin-left: 3px; opacity: 0.6; }
  tbody td {
    padding: 9px 12px; border-bottom: 1px solid var(--gridline);
    vertical-align: top;
  }
  tbody tr:last-child td { border-bottom: none; }
  tbody tr:hover { background: var(--page-plane); }
  td.num, th.num { text-align: right; font-variant-numeric: tabular-nums; }
  .wr-bar-wrap { display: flex; align-items: center; gap: 8px; justify-content: flex-end; }
  .wr-bar-track {
    width: 60px; height: 6px; border-radius: 3px; background: var(--gridline);
    overflow: hidden;
  }
  .wr-bar-fill { height: 100%; background: var(--series-blue); border-radius: 3px; }
  .name-link { color: var(--series-blue); cursor: pointer; text-decoration: none; }
  .name-link:hover { text-decoration: underline; }
  .table-controls {
    display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap;
    gap: 10px; margin: 34px 0 8px;
  }
  .table-controls-right { display: flex; gap: 10px; flex-wrap: wrap; align-items: center; }
  .filter-input {
    font-size: 13px; padding: 7px 10px; border-radius: 8px;
    border: 1px solid var(--border); background: var(--surface-1);
    color: var(--text-primary); width: 220px;
  }
  .mini-select-wrap {
    display: flex; align-items: center; gap: 6px; font-size: 12.5px;
    color: var(--text-secondary); white-space: nowrap;
  }
  .mini-select {
    font-size: 13px; padding: 6px 8px; border-radius: 8px;
    border: 1px solid var(--border); background: var(--surface-1);
    color: var(--text-primary); cursor: pointer;
  }

  /* Page nav (added 2026-07-19): splits the dashboard into two "pages" so Best Picks /
     Popular Win-Con Sets / Deck Stats -- long reference tables -- don't clutter the
     game-day-critical Scout / What Might Follow / Deck Predictor tools on first load. */
  .page-nav {
    display: flex; gap: 8px; margin: 18px 0 22px; border-bottom: 1px solid var(--gridline);
    padding-bottom: 0;
  }
  .page-nav-btn {
    font-size: 14px; font-weight: 600; padding: 10px 18px; cursor: pointer;
    border: none; background: none; color: var(--text-muted);
    border-bottom: 2px solid transparent; margin-bottom: -1px;
  }
  .page-nav-btn:hover { color: var(--text-primary); }
  .page-nav-btn.active { color: var(--series-blue); border-bottom-color: var(--series-blue); }
  footer.note {
    margin-top: 40px; font-size: 11.5px; color: var(--text-muted);
    border-top: 1px solid var(--gridline); padding-top: 14px;
  }

  .predictor {
    margin: 40px 0 0;
    background: var(--surface-1);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 22px 24px;
  }
  .predictor-sub { font-size: 13px; color: var(--text-secondary); margin: 4px 0 16px; max-width: 640px; }
  .predictor-controls {
    display: flex; flex-wrap: wrap; gap: 10px; align-items: center; margin-bottom: 18px;
  }
  .predictor-tabs { display: flex; gap: 6px; }
  .predictor-tab {
    font-size: 13px; padding: 7px 14px; border-radius: 999px;
    border: 1px solid var(--border); background: var(--page-plane);
    color: var(--text-secondary); cursor: pointer;
  }
  .predictor-tab.active { background: var(--series-blue); color: #fff; border-color: var(--series-blue); }
  .match-category-bar {
    display: flex; flex-wrap: wrap; gap: 12px 20px; align-items: center; margin-bottom: 12px;
  }
  .category-tabs { display: flex; gap: 6px; }
  .category-tab {
    font-size: 13px; padding: 7px 14px; border-radius: 999px;
    border: 1px solid var(--border); background: var(--page-plane);
    color: var(--text-secondary); cursor: pointer;
  }
  .category-tab.active { background: var(--series-blue); color: #fff; border-color: var(--series-blue); }
  .weight-toggle {
    display: flex; align-items: center; gap: 6px; font-size: 13px;
    color: var(--text-secondary); cursor: pointer; user-select: none;
  }
  .weight-toggle input[type="checkbox"] { cursor: pointer; }
  .weight-toggle.disabled { opacity: 0.45; pointer-events: none; }
  .best-picks { margin-bottom: 30px; }
  .view-tabs { display: flex; gap: 6px; margin: 4px 0 4px; }
  .view-tab {
    font-size: 13px; padding: 7px 14px; border-radius: 999px;
    border: 1px solid var(--border); background: var(--page-plane);
    color: var(--text-secondary); cursor: pointer;
  }
  .view-tab.active { background: var(--series-violet); color: #fff; border-color: var(--series-violet); }
  .best-picks-rank {
    display: inline-flex; align-items: center; justify-content: center;
    width: 22px; height: 22px; border-radius: 50%; background: var(--page-plane);
    border: 1px solid var(--border);
    font-size: 12px; font-weight: 600; color: var(--text-secondary); margin-right: 8px;
  }
  .duelset-decks { display: flex; flex-direction: column; gap: 6px; }
  .duelset-decks .deck-card { margin: 0; }
  .predictor-controls select {
    font-size: 13px; padding: 8px 10px; border-radius: 8px;
    border: 1px solid var(--border); background: var(--surface-1);
    color: var(--text-primary); min-width: 220px;
  }
  .predictor-scope-search.search-box {
    margin-bottom: 0; min-width: 220px; max-width: 260px;
  }
  .predictor-scope-search input {
    width: 100%; font-size: 13px; padding: 8px 10px; border-radius: 8px;
    border: 1px solid var(--border); background: var(--surface-1); color: var(--text-primary);
    outline: none;
  }
  .predictor-scope-search input:focus { border-color: var(--series-blue); }
  .predictor-results { display: flex; flex-direction: column; gap: 8px; }
  .predictor-row {
    display: flex; align-items: center; gap: 12px;
    padding: 8px 4px; border-bottom: 1px solid var(--gridline);
  }
  .predictor-row:last-child { border-bottom: none; }
  .predictor-row .predictor-rank {
    font-size: 12px; color: var(--text-muted); width: 18px; text-align: right; flex-shrink: 0;
  }
  .predictor-row .predictor-bar-track {
    flex: 1; height: 8px; border-radius: 4px; background: var(--gridline); overflow: hidden;
    min-width: 60px;
  }
  .predictor-row .predictor-bar-fill { height: 100%; background: var(--series-blue); border-radius: 4px; }
  .predictor-row .predictor-count { font-size: 12px; color: var(--text-muted); width: 90px; text-align: right; flex-shrink: 0; }

  .predictor-pickers { display: flex; flex-wrap: wrap; gap: 16px; margin-bottom: 18px; }
  .picker-col { min-width: 220px; }
  .picker-col-label { font-size: 12px; font-weight: 700; color: var(--text-muted); text-transform: uppercase; letter-spacing: .03em; margin-bottom: 6px; }
  .picker-chips { display: flex; flex-wrap: wrap; gap: 6px; max-height: 180px; overflow-y: auto; padding: 2px; }
  .picker-chip {
    font-size: 12.5px; padding: 6px 12px; border-radius: 999px;
    border: 1px solid var(--border); background: var(--page-plane);
    color: var(--text-secondary); cursor: pointer; user-select: none;
  }
  .picker-chip.active { background: var(--series-blue); color: #fff; border-color: var(--series-blue); }

  .deck-predictor {
    margin: 22px 0 0;
    background: var(--surface-1);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 22px 24px;
  }
  .deck-predictor-bulk {
    background: var(--page-plane); border: 1px solid var(--border); border-radius: 10px;
    padding: 12px 14px; margin-bottom: 14px;
  }
  .deck-predictor-bulk-label { display: block; font-size: 12px; color: var(--text-muted); margin-bottom: 6px; }
  .deck-predictor-bulk-row { display: flex; gap: 8px; align-items: flex-start; }
  .deck-predictor-bulk textarea {
    flex: 1; min-height: 38px; resize: vertical; font-size: 13px; padding: 8px 10px;
    border-radius: 8px; border: 1px solid var(--border); background: var(--surface-1);
    color: var(--text-primary); font-family: inherit; outline: none;
  }
  .deck-predictor-bulk textarea:focus { border-color: var(--series-blue); }
  .deck-predictor-bulk-status { font-size: 12px; margin-top: 6px; min-height: 16px; }
  .deck-predictor-bulk-status.ok { color: var(--series-blue); }
  .deck-predictor-bulk-status.warn { color: #b45309; }
  .deck-predictor-controls { display: flex; flex-wrap: wrap; gap: 10px; align-items: center; margin-bottom: 12px; }
  .deck-predictor-controls .filter-input { min-width: 220px; }
  .deck-predictor-count { font-size: 12px; color: var(--text-muted); }
  .deck-predictor-selected {
    display: flex; flex-wrap: wrap; gap: 6px; min-height: 40px; margin-bottom: 12px;
    padding: 8px; border: 1px dashed var(--border); border-radius: 10px;
  }
  .deck-predictor-selected .empty-hint { margin: 4px; font-size: 12.5px; }
  .deck-predictor-pool {
    display: flex; flex-wrap: wrap; gap: 6px; max-height: 240px; overflow-y: auto;
    padding: 8px; border: 1px solid var(--gridline); border-radius: 10px;
  }
  .deck-slot { margin-bottom: 18px; padding-bottom: 4px; }
  .deck-slot + .deck-slot { border-top: 1px solid var(--gridline); padding-top: 16px; }
  .deck-slot-label {
    font-size: 12px; font-weight: 700; color: var(--text-muted); text-transform: uppercase;
    letter-spacing: .03em; margin-bottom: 8px;
  }
  .picker-search-row { display: flex; gap: 8px; margin-bottom: 8px; }
  .picker-search-row input {
    flex: 1; font-size: 12.5px; padding: 7px 10px; border-radius: 8px;
    border: 1px solid var(--border); background: var(--surface-1); color: var(--text-primary);
    outline: none;
  }
  .picker-search-row input:focus { border-color: var(--series-blue); }
  .picker-search-status { font-size: 11.5px; color: var(--text-muted); min-height: 14px; margin: -4px 0 6px; }
  .picker-search-status.warn { color: #b45309; }
  .deck-tile {
    display: flex; align-items: center; gap: 6px; padding: 4px 10px 4px 4px;
    border-radius: 999px; border: 1px solid var(--border); background: var(--page-plane);
    cursor: pointer; user-select: none; font-size: 12.5px; color: var(--text-secondary);
  }
  .deck-tile img { width: 20px; height: 20px; border-radius: 4px; object-fit: cover; flex-shrink: 0; }
  .deck-tile.selected { background: var(--series-blue); color: #fff; border-color: var(--series-blue); }
  .deck-tile.disabled { opacity: .35; cursor: not-allowed; }

  .shadow-tag {
    font-size: 10.5px; font-weight: 700; text-transform: uppercase; letter-spacing: .02em;
    color: var(--text-muted); background: var(--page-plane); border: 1px solid var(--border);
    border-radius: 999px; padding: 2px 8px; margin-left: 8px; vertical-align: middle;
  }
  .scouted-tag {
    font-size: 10.5px; font-weight: 700; text-transform: uppercase; letter-spacing: .02em;
    color: #92400e; background: #fef3c7; border: 1px solid #fcd34d;
    border-radius: 999px; padding: 2px 8px; margin-left: 8px; vertical-align: middle;
  }
  .extended-tag {
    font-size: 10.5px; font-weight: 700; text-transform: uppercase; letter-spacing: .02em;
    color: #075985; background: #e0f2fe; border: 1px solid #7dd3fc;
    border-radius: 999px; padding: 2px 8px; margin-left: 8px; vertical-align: middle;
  }
  /* Event-day flag on each Duel History row (added 2026-07-19, per user request to flag
     which event day a given Official CRL game belongs to). Colors cycle per day so a
     future Day 3+ stays visually distinct without needing new CSS each time. */
  .day-badge {
    display: inline-block; font-size: 9.5px; font-weight: 800; text-transform: uppercase;
    letter-spacing: .03em; border-radius: 5px; padding: 1px 5px; margin-right: 6px;
    vertical-align: middle; line-height: 1.5;
  }
  .day-badge-day1 { color: #4338ca; background: #e0e7ff; border: 1px solid #a5b4fc; }
  .day-badge-day2 { color: #b45309; background: #fef3c7; border: 1px solid #fcd34d; }
  .day-badge-day3 { color: #15803d; background: #dcfce7; border: 1px solid #86efac; }
  .day-badge-day4, .day-badge-day5 { color: #9d174d; background: #fce7f3; border: 1px solid #f9a8d4; }
  .day-badge-dayx { color: var(--text-muted); background: var(--page-plane); border: 1px solid var(--border); }
  .group-a-panel {
    background: var(--page-plane); border: 1px solid var(--border); border-radius: 12px;
    padding: 14px 16px; margin: 0 0 16px;
  }
  .group-a-chips { display: flex; flex-wrap: wrap; gap: 8px; }
  .group-a-chip {
    display: inline-flex; align-items: center; gap: 6px; padding: 6px 12px;
    border-radius: 999px; border: 1px solid var(--series-blue); background: #fff;
    color: var(--series-blue); font-size: 13px; font-weight: 600; cursor: pointer;
  }
  .group-a-chip.is-you { border-color: var(--series-yellow); color: #92660a; background: #fffbeb; cursor: default; }
  .group-a-chip.pending {
    border-color: var(--border); color: var(--text-muted); background: var(--page-plane);
    cursor: default;
  }
  .group-a-chip .pending-note { font-size: 10.5px; font-weight: 500; opacity: .8; }
  .group-a-chip-wrap { display: inline-flex; align-items: center; gap: 4px; }
  .group-a-history-btn {
    display: inline-flex; align-items: center; justify-content: center;
    width: 24px; height: 24px; border-radius: 50%; border: 1px solid var(--border);
    background: #fff; color: var(--text-muted); font-size: 12px; cursor: pointer;
    line-height: 1;
  }
  .group-a-history-btn:hover { border-color: var(--series-blue); color: var(--series-blue); }
  .modal-overlay {
    display: none; position: fixed; inset: 0; background: rgba(15,23,42,.55);
    z-index: 100; align-items: flex-start; justify-content: center; overflow-y: auto;
    padding: 40px 16px;
  }
  .modal-overlay.show { display: flex; }
  .modal-box {
    background: var(--surface-1); border-radius: 14px; max-width: 900px; width: 100%;
    padding: 20px 22px 24px; box-shadow: 0 20px 60px rgba(0,0,0,.35);
  }
  .modal-header {
    display: flex; align-items: center; justify-content: space-between; margin-bottom: 4px;
  }
  .modal-header h2 { margin: 0; font-size: 18px; }
  .modal-close {
    border: none; background: none; font-size: 22px; line-height: 1; cursor: pointer;
    color: var(--text-muted); padding: 4px 8px;
  }
  .modal-close:hover { color: var(--text-primary); }
  .modal-subtitle { font-size: 12.5px; color: var(--text-muted); margin: 0 0 14px; }
  .history-row {
    display: flex; align-items: center; gap: 14px; padding: 10px 8px;
    border-bottom: 1px solid var(--gridline);
  }
  .history-row:last-child { border-bottom: none; }
  .history-meta { width: 132px; flex-shrink: 0; font-size: 11.5px; color: var(--text-muted); }
  .history-meta .opp { font-size: 13px; font-weight: 600; color: var(--text-primary); margin-bottom: 2px; }
  .history-decks { display: flex; align-items: center; gap: 10px; flex: 1; min-width: 0; }
  .history-decks .vs { font-size: 11px; color: var(--text-muted); font-weight: 700; flex-shrink: 0; }
  .history-result {
    width: 44px; flex-shrink: 0; text-align: center; font-weight: 800; font-size: 13px;
    border-radius: 8px; padding: 4px 0;
  }
  .history-result.W { background: #dcfce7; color: #166534; }
  .history-result.L { background: #fee2e2; color: #991b1b; }
  .history-result.T { background: var(--page-plane); color: var(--text-muted); }
  .history-empty { padding: 20px 8px; text-align: center; color: var(--text-muted); font-size: 13px; }
  /* CRL / Practice duel-history toggle (added 2026-07-21) */
  .hist-toggle { display: inline-flex; gap: 4px; margin: 6px 0 12px; background: var(--page-plane);
    border: 1px solid var(--border); border-radius: 10px; padding: 3px; }
  .hist-tab { border: none; background: none; cursor: pointer; font-size: 13px; font-weight: 600;
    color: var(--text-secondary); padding: 6px 14px; border-radius: 8px; display: inline-flex;
    align-items: center; gap: 6px; }
  .hist-tab:hover { color: var(--text-primary); }
  .hist-tab.active { background: var(--series-blue); color: #fff; }
  .hist-count { font-size: 11px; font-weight: 700; background: rgba(0,0,0,.14); border-radius: 999px;
    padding: 1px 7px; }
  .hist-tab.active .hist-count { background: rgba(255,255,255,.25); }
  /* Manual per-player watch note call-out (added 2026-07-21) */
  .player-note {
    background: #FFF3E0; border: 1px solid #FFB74D; border-left: 4px solid #E65100;
    color: #7A3B00; border-radius: 8px; padding: 10px 14px; margin: 0 0 14px;
    font-size: 13.5px; line-height: 1.45;
  }
  .player-note b { color: #E65100; }
  .player-note-list { margin: 6px 0 0; padding-left: 20px; }
  .player-note-list li { margin: 3px 0; }
  .group-a-note-flag { color: #E65100; font-size: 12px; margin-left: 2px; cursor: help; }
  /* Recent Practice Trends (added 2026-07-21) */
  .rp-cols { display: grid; grid-template-columns: 1.4fr 1fr; gap: 16px; }
  @media (max-width: 640px) { .rp-cols { grid-template-columns: 1fr; } }
  .rp-deck { background: var(--page-plane); border: 1px solid var(--border); border-radius: 8px;
    padding: 7px 9px; margin-bottom: 7px; }
  .rp-deck-head { font-size: 12.5px; margin-bottom: 3px; }
  .rp-count { font-weight: 800; color: var(--series-blue); }
  /* Duel-Set Record (added 2026-07-22) */
  :root { --dsr-win: #1f9d55; --dsr-win2: #7bc47f; --dsr-loss2: #f0a58f; --dsr-loss: #d64545; }
  .dsr-cols { display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }
  @media (max-width: 640px) { .dsr-cols { grid-template-columns: 1fr; } }
  .dsr-col { background: var(--page-plane); border: 1px solid var(--border); border-radius: 10px; padding: 11px 13px; }
  .dsr-col-head { font-size: 13px; font-weight: 800; margin-bottom: 9px; }
  .dsr-col-head span { font-weight: 500; font-size: 11px; color: var(--text-muted); margin-left: 4px; }
  .dsr-crl { color: #3730a3; } .dsr-prac { color: #92400e; }
  .dsr-line { display: flex; align-items: baseline; gap: 8px; margin-bottom: 8px; }
  .dsr-wr { font-size: 26px; font-weight: 800; line-height: 1; }
  .dsr-wr-sub { font-size: 12px; color: var(--text-muted); }
  .dsr-bar { display: flex; height: 12px; border-radius: 6px; overflow: hidden; margin-bottom: 8px; background: var(--border); }
  .dsr-seg { display: block; min-width: 2px; }
  .dsr-chips { display: flex; flex-wrap: wrap; gap: 5px 6px; margin-bottom: 8px; align-items: center; }
  .dsr-chip { font-size: 11.5px; padding: 2px 8px; border-radius: 999px; border: 1px solid var(--border); background: var(--surface-1); }
  .dsr-chip b { font-weight: 800; }
  .dsr-chip.w { border-color: #a7d8b6; background: rgba(31,157,85,.10); }
  .dsr-chip.l { border-color: #e9b0b0; background: rgba(214,69,69,.09); }
  .dsr-chip-note { font-size: 10.5px; color: var(--text-muted); }
  .dsr-decide { border-top: 1px dashed var(--border); padding-top: 8px; margin-top: 2px; }
  .dsr-decide-row { display: flex; justify-content: space-between; gap: 10px; font-size: 12px; margin: 4px 0; flex-wrap: wrap; }
  .dsr-decide-lbl { color: var(--text-muted); font-weight: 600; }
  .dsr-decide-val b { font-weight: 800; }
  .dsr-decide-val .l, .dsr-note .l { color: var(--dsr-loss); }
  .dsr-note { font-size: 11.5px; color: var(--text-muted); margin-top: 7px; line-height: 1.4; }
  .dsr-note b { color: var(--text-primary); font-weight: 800; }
  .deck-explorer {
    background: var(--page-plane); border: 1px solid var(--border); border-radius: 12px;
    padding: 14px 16px; margin: 8px 0 20px;
  }
  .deck-explorer-head {
    display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap;
    gap: 10px; margin-bottom: 6px;
  }
  .deck-explorer-controls { display: flex; gap: 14px; flex-wrap: wrap; }
  .deck-explorer-controls label {
    font-size: 12px; color: var(--text-muted); display: flex; align-items: center; gap: 6px;
  }
  .deck-explorer-controls select {
    font-size: 12.5px; padding: 4px 8px; border-radius: 8px; border: 1px solid var(--border);
    background: var(--surface-1); color: var(--text-primary);
  }
  .deck-explorer-note { font-size: 11.5px; color: var(--text-muted); margin: 0 0 10px; }
  .deck-explorer-row {
    display: flex; align-items: center; gap: 14px; padding: 8px 4px;
    border-bottom: 1px solid var(--gridline);
  }
  .deck-explorer-row:last-child { border-bottom: none; }
  .deck-explorer-row .der-meta { flex-shrink: 0; width: 130px; text-align: right; }
  .deck-explorer-row .der-games { font-size: 13px; font-weight: 700; }
  .deck-explorer-row .der-wr { font-size: 12px; color: var(--text-muted); }
  .deck-explorer-row .der-variants {
    font-size: 10.5px; color: var(--text-muted); background: var(--surface-1);
    border: 1px solid var(--border); border-radius: 999px; padding: 1px 7px; margin-top: 3px;
    display: inline-block;
  }
  .group-a-chip.on-deck { border-style: dashed; }
  .group-a-chip.reference { border-color: var(--text-muted); color: var(--text-muted); }
  .reco-section {
    border: 1px solid var(--border); border-radius: 10px; padding: 12px 14px 14px;
    margin-bottom: 6px; background: var(--page-plane);
  }
  .reco-heading { font-size: 13.5px; font-weight: 700; margin-bottom: 4px; }
  .reco-heading-note { font-size: 10.5px; font-weight: 500; color: var(--text-muted); margin-left: 6px; text-transform: none; }
  .reco-col-label {
    font-size: 11.5px; font-weight: 700; color: var(--text-muted); text-transform: uppercase;
    letter-spacing: .02em; margin: 10px 0 4px;
  }
  .reco-row {
    display: flex; align-items: center; justify-content: space-between; gap: 10px;
    padding: 4px 0;
  }
  .reco-row-text { padding: 6px 0; }
  .reco-label { font-size: 13px; }
  .reco-wr { font-size: 12.5px; font-weight: 700; color: var(--series-blue); flex-shrink: 0; }
  .reco-duelset {
    display: flex; align-items: center; gap: 8px; padding: 6px 0; flex-wrap: wrap;
  }
  .reco-duelset .vs { font-size: 11px; color: var(--text-muted); font-weight: 700; }
  .reco-row-threat {
    background: #FCE8E6; border-radius: 6px; padding: 6px 8px; flex-wrap: wrap;
  }
</style>
</head>
<body>
<div class="wrap">
  <header class="top">
    <h1>CRL Opponent Scout</h1>
    <p class="subtitle">Search a player to see their decks, win rates, and go-to win conditions. Data snapshot from CRL_Duel_Decks.xlsx.</p>
    <a href="monthly_finals_opponents.html" style="display:inline-block;margin-bottom:10px;padding:6px 12px;background:#E8F5E9;color:#1B5E20;border:1px solid #A5D6A7;border-radius:6px;font-size:13px;font-weight:700;text-decoration:none;">📋 Monthly Finals Opponents (Day 2 deck ratings + Day 3 scouting) &rarr;</a>
    <div class="icon-legend">
      <span><span class="dot" style="background:var(--series-yellow);"></span>&#9819; Champion / Hero card</span>
      <span><span class="dot" style="background:var(--series-violet);"></span>&#9733; Sometimes played as an Evolution</span>
    </div>
  </header>

  <div class="page-nav" id="pageNav">
    <button class="page-nav-btn active" data-page="pageScout">Scout Tools</button>
    <button class="page-nav-btn" data-page="pageStats">Best Picks &amp; Stats</button>
  </div>

  <div class="page" id="pageScout">
  <div class="match-category-bar">
    <div class="category-tabs" id="matchCategoryTabs">
      <button class="category-tab active" data-category="all">All Games</button>
      <button class="category-tab" data-category="practice">Practice Only</button>
      <button class="category-tab" data-category="official">Official CRL Only</button>
    </div>
    <label class="weight-toggle" id="weightToggleLabel">
      <input type="checkbox" id="weightToggle">
      Weight Official CRL games higher <span class="table-note" style="font-size:11px;">(only affects "All Games" rankings; no effect until a deck/win-con has 5+ official games)</span>
    </label>
  </div>
  <div class="search-box">
    <input type="text" id="searchInput" placeholder="Search a player name or tag (e.g. #80ULUJLYY)..." autocomplete="off">
    <div class="suggestions" id="suggestions"></div>
  </div>

  <div class="group-a-panel" id="groupAPanel">
    <div class="table-title" style="margin:0 0 6px;">Group A -- Day 2 Group Stage <span class="table-note">snake-seeded per rulebook 4.1.3.8.3, seeds 1/16/17/32/33/48/49/64</span></div>
    <div class="group-a-chips" id="groupAChips"></div>
  </div>

  <div class="modal-overlay" id="historyModal">
    <div class="modal-box">
      <div class="modal-header">
        <h2 id="historyModalTitle"></h2>
        <button class="modal-close" id="historyModalClose" aria-label="Close">&times;</button>
      </div>
      <p class="modal-subtitle" id="historyModalSubtitle"></p>
      <div id="historyModalBody"></div>
    </div>
  </div>

  <div class="quick-roster" id="quickRoster"></div>

  <div class="profile" id="profile">
    <div class="profile-head">
      <h2 id="profName"></h2>
    </div>
    <div class="stat-row" id="statRow"></div>

    <div class="brief-card" id="briefCard">
      <div class="brief-head">
        <span class="brief-title">Opponent Brief</span>
        <span class="brief-sub">Quick mid-tournament reference -- everything below in one glance</span>
      </div>
      <div class="brief-grid" id="briefGrid"></div>
    </div>

    <div class="section-title">Most-Played Decks</div>
    <div class="card-grid" id="mostPlayedGrid"></div>

    <div class="deck-explorer" id="deckExplorer">
      <div class="deck-explorer-head">
        <span class="section-title" style="margin:0;">Deck Explorer</span>
        <div class="deck-explorer-controls">
          <label>Games:
            <select id="deckExplorerCategory">
              <option value="all">Practice + Official CRL</option>
              <option value="crl">Official CRL only</option>
            </select>
          </label>
          <label>Group decks sharing:
            <select id="deckExplorerThreshold">
              <option value="8">8/8 cards (exact deck)</option>
              <option value="7">7+/8 cards</option>
              <option value="6">6+/8 cards</option>
              <option value="5">5+/8 cards</option>
              <option value="4">4+/8 cards</option>
            </select>
          </label>
        </div>
      </div>
      <p class="deck-explorer-note">Groups decks sharing at least the selected number of cards into one archetype row, so 1-2 card swaps of the same deck still show up together -- lower the threshold if you're not seeing many exact repeats.</p>
      <div id="deckExplorerResults"></div>
    </div>

    <div class="section-title">Best Win-Rate Decks</div>
    <div class="card-grid" id="bestWinRateGrid"></div>

    <div class="section-title">Top Win Conditions</div>
    <div class="wincon-badges" id="winconBadges"></div>
  </div>

  <div id="noResult" class="empty-hint" style="display:none;">No player matches that search.</div>

  <div class="predictor">
    <h3 class="table-title" style="margin:0 0 4px;">What Might Follow? <span class="table-note">predict game 2/3 from what the opponent opened with</span></h3>
    <p class="predictor-sub" id="predictorSub">Check every win condition your opponent played in an earlier game of this duel (a deck can run more than one at once -- e.g. Goblin Barrel + Wall Breakers + Miner together). Shows what most commonly showed up in a LATER game of the same duel, aggregated across the tracked roster's duel history.</p>
    <p class="predictor-sub" id="predictorDqNote" style="font-size:11px;opacity:0.7;"></p>
    <div class="predictor-controls">
      <div class="predictor-tabs">
        <button class="predictor-tab active" data-mode="wincon">Game 2 (from G1)</button>
        <button class="predictor-tab" data-mode="wincon3">Game 3 (from G1 + G2)</button>
      </div>
      <div class="predictor-scope-search search-box">
        <input type="text" id="predictorScopeSearch" placeholder="Search player to scope to...">
        <div class="suggestions" id="predictorScopeSuggestions"></div>
      </div>
      <select id="predictorScope">
        <option value="__all__">All tracked players</option>
      </select>
    </div>
    <div id="predictorPickers" class="predictor-pickers"></div>
    <div id="predictorResults" class="predictor-results"></div>
  </div>

  <div class="deck-predictor">
    <h3 class="table-title" style="margin:0 0 4px;">Deck Predictor <span class="table-note">pick 8 cards from an earlier game, get likely full decks for the next one</span></h3>
    <p class="predictor-sub" id="deckPredictorSub">Click the cards your opponent played in an earlier game (up to 8) from the full card pool below to rebuild their deck -- or type in the filter box and hit Enter to add the top match fast. Card matching is fuzzy: partial names, common abbreviations (wb, gob, skelly, barb...), acronyms (initials of each word), and minor typos are all recognized, not just exact spelling.</p>
    <p class="predictor-sub" id="deckPredictorDqNote" style="font-size:11px;opacity:0.7;"></p>
    <div class="predictor-tabs" style="margin-bottom:14px;">
      <button class="predictor-tab active" data-deckmode="deck2">Game 2 (from G1)</button>
      <button class="predictor-tab" data-deckmode="deck3">Game 3 (from G1 + G2)</button>
    </div>

    <div id="deckSlot_g1" class="deck-slot">
      <div class="deck-slot-label" id="deckSlotLabel_g1">Game 1 deck</div>
      <div class="deck-predictor-bulk">
        <label for="deckPredictorBulk_g1" class="deck-predictor-bulk-label">Fast entry (CRL time-crunch mode): type/paste all 8 card names, comma or newline separated, then hit Enter or Add All</label>
        <div class="deck-predictor-bulk-row">
          <textarea id="deckPredictorBulk_g1" rows="1" placeholder="e.g. hog, ice golem, musk, fb, skellies, is, cannon, log"></textarea>
          <button id="deckPredictorBulkAdd_g1" class="predictor-tab" type="button">Add All</button>
        </div>
        <div id="deckPredictorBulkStatus_g1" class="deck-predictor-bulk-status"></div>
      </div>
      <div class="deck-predictor-controls">
        <input class="filter-input" id="deckPredictorFilter_g1" placeholder="Filter card pool by name... (Enter = add top match)">
        <span id="deckPredictorCount_g1" class="deck-predictor-count">0 / 8 selected</span>
        <button id="deckPredictorClear_g1" class="predictor-tab" type="button">Clear</button>
      </div>
      <div class="deck-predictor-selected" id="deckPredictorSelected_g1"></div>
      <div class="deck-predictor-pool" id="deckPredictorPool_g1"></div>
    </div>

    <div id="deckSlot_g2" class="deck-slot" style="display:none;">
      <div class="deck-slot-label" id="deckSlotLabel_g2">Game 2 deck</div>
      <div class="deck-predictor-bulk">
        <label for="deckPredictorBulk_g2" class="deck-predictor-bulk-label">Fast entry (CRL time-crunch mode): type/paste all 8 card names, comma or newline separated, then hit Enter or Add All</label>
        <div class="deck-predictor-bulk-row">
          <textarea id="deckPredictorBulk_g2" rows="1" placeholder="e.g. hog, ice golem, musk, fb, skellies, is, cannon, log"></textarea>
          <button id="deckPredictorBulkAdd_g2" class="predictor-tab" type="button">Add All</button>
        </div>
        <div id="deckPredictorBulkStatus_g2" class="deck-predictor-bulk-status"></div>
      </div>
      <div class="deck-predictor-controls">
        <input class="filter-input" id="deckPredictorFilter_g2" placeholder="Filter card pool by name... (Enter = add top match)">
        <span id="deckPredictorCount_g2" class="deck-predictor-count">0 / 8 selected</span>
        <button id="deckPredictorClear_g2" class="predictor-tab" type="button">Clear</button>
      </div>
      <div class="deck-predictor-selected" id="deckPredictorSelected_g2"></div>
      <div class="deck-predictor-pool" id="deckPredictorPool_g2"></div>
    </div>

    <div id="deckPredictorResults" class="predictor-results" style="margin-top:16px;"></div>
  </div>
  </div><!-- /pageScout -->

  <div class="page" id="pageStats" style="display:none;">

  <div class="best-picks">
    <h3 class="table-title" style="margin:0 0 4px;">Best Picks <span class="table-note">quick reference for game day -- top win-rate decks, duel sets, and win-con sets</span></h3>
    <p class="predictor-sub">Ranked by win rate (best first). Use "Min times played" to hide small-sample rows -- a deck with 1 game and a lucky win otherwise shows as a "100% win rate" pick. "Top Duel Sets" compares which 3-deck combinations performed best together -- since exact full-deck repeats are rare, decks are grouped into near-duplicate "families" (sharing most of their 8 cards) so there's enough data to rank; the note under that table shows exactly how loose that comparison had to get.</p>
    <div class="match-category-bar">
      <div class="category-tabs" id="bestPicksCategoryTabs">
        <button class="category-tab active" data-category="all">All Games</button>
        <button class="category-tab" data-category="practice">Practice Only</button>
        <button class="category-tab" data-category="official">Official CRL Only</button>
      </div>
    </div>
    <div class="view-tabs" id="bestPicksViewTabs">
      <button class="view-tab active" data-view="decks">Top Decks</button>
      <button class="view-tab" data-view="duel_sets">Top Duel Sets</button>
      <button class="view-tab" data-view="wincon_sets">Top Win-Con Sets</button>
    </div>
    <div class="table-controls" style="margin:12px 0 4px;">
      <div></div>
      <div class="table-controls-right">
        <label class="mini-select-wrap">Min times played:
          <select class="mini-select" id="bestPicksMinPlayed">
            <option value="1">1 (all)</option>
            <option value="2">2</option>
            <option value="3" selected>3</option>
            <option value="5">5</option>
            <option value="10">10</option>
          </select>
        </label>
        <label class="mini-select-wrap">Show:
          <select class="mini-select" id="bestPicksShowCount">
            <option value="10">Top 10</option>
            <option value="15" selected>Top 15</option>
            <option value="25">Top 25</option>
            <option value="50">Top 50</option>
            <option value="all">All</option>
          </select>
        </label>
      </div>
    </div>
    <div id="bestPicksNote" class="predictor-sub" style="font-size:12px; margin-top:2px;"></div>
    <div id="bestPicksResults" class="table-wrap"></div>
  </div>

  <div class="table-controls">
    <h3 class="table-title" style="margin:0;">Popular Win-Con Sets <span class="table-note">combined across all tracked players</span></h3>
    <div class="table-controls-right">
      <input class="filter-input" id="winconFilter" placeholder="Filter by card name...">
      <label class="mini-select-wrap">Min games played:
        <select class="mini-select" id="winconMinGames">
          <option value="0">Any</option>
          <option value="2">2</option>
          <option value="3" selected>3</option>
          <option value="5">5</option>
          <option value="10">10</option>
        </select>
      </label>
      <label class="mini-select-wrap">Show:
        <select class="mini-select" id="winconShowCount">
          <option value="15">Top 15</option>
          <option value="30" selected>Top 30</option>
          <option value="60">Top 60</option>
          <option value="all">All</option>
        </select>
      </label>
    </div>
  </div>
  <div class="table-wrap">
    <table id="winconTable">
      <thead><tr>
        <th data-key="Win-Con Set" data-type="str">Win-Con Set</th>
        <th data-key="Times Played (Duels)" data-type="num" class="num">Times Played</th>
        <th data-key="Games Played" data-type="num" class="num">Games</th>
        <th data-key="Win Rate" data-type="num" class="num">Win Rate</th>
        <th data-key="Players Who Used This" data-type="str">Players</th>
      </tr></thead>
      <tbody></tbody>
    </table>
  </div>

  <div class="table-controls">
    <h3 class="table-title" style="margin:0;">Deck Stats <span class="table-note">every unique 8-card deck seen</span></h3>
    <div class="table-controls-right">
      <input class="filter-input" id="deckFilter" placeholder="Filter by card name...">
      <label class="mini-select-wrap">Min games played:
        <select class="mini-select" id="deckMinGames">
          <option value="0">Any</option>
          <option value="2">2</option>
          <option value="3" selected>3</option>
          <option value="5">5</option>
          <option value="10">10</option>
        </select>
      </label>
      <label class="mini-select-wrap">Show:
        <select class="mini-select" id="deckShowCount">
          <option value="15">Top 15</option>
          <option value="30" selected>Top 30</option>
          <option value="60">Top 60</option>
          <option value="all">All</option>
        </select>
      </label>
    </div>
  </div>
  <div class="table-wrap">
    <table id="deckTable">
      <thead><tr>
        <th data-key="Deck (sorted)" data-type="str">Deck</th>
        <th data-key="Used By" data-type="str">Used By</th>
        <th data-key="Games Played" data-type="num" class="num">Games</th>
        <th data-key="Win Rate" data-type="num" class="num">Win Rate</th>
      </tr></thead>
      <tbody></tbody>
    </table>
  </div>

  </div><!-- /pageStats -->

  <footer class="note">
    Static snapshot generated from CRL_Duel_Decks.xlsx. Re-run build_duel_workbook.py and regenerate this
    dashboard after fetching new battle data to refresh it. "Best Win-Rate Deck" entries only include decks
    played at least twice, to avoid one-off small-sample results looking misleadingly strong. Players marked
    "seen as opponent" aren't part of the ~50 tracked roster -- their profile is built only from games where a
    tracked player happened to face them, so coverage is one-sided and may be thin.
  </footer>
</div>

<script>
const DATA = __DATA_JSON__;

// Match Category filter + weighting toggle (added 2026-07-18): four precomputed player
// pools baked in at build time (see build_dashboard.py's "Four variants" block) -- picking
// a category or flipping the weight toggle just swaps which one is active, no recompute.
const PLAYER_POOLS = {
  all_unweighted: DATA.player_lookup,
  all_weighted: DATA.player_lookup_weighted || DATA.player_lookup,
  practice: DATA.player_lookup_practice || DATA.player_lookup,
  official: DATA.player_lookup_official || [],
};
let matchCategoryFilter = 'all';   // 'all' | 'practice' | 'official'
let weightOfficialGames = false;   // only meaningful when matchCategoryFilter === 'all'

let players = PLAYER_POOLS.all_unweighted;
let playerByName = {};
function rebuildPlayerPool() {
  if (matchCategoryFilter === 'practice') players = PLAYER_POOLS.practice;
  else if (matchCategoryFilter === 'official') players = PLAYER_POOLS.official;
  else players = weightOfficialGames ? PLAYER_POOLS.all_weighted : PLAYER_POOLS.all_unweighted;
  playerByName = {};
  players.forEach(p => playerByName[p['Player']] = p);
}
rebuildPlayerPool();
const winconSets = DATA.wincon_sets_top;
const deckStats = DATA.deck_stats_top;
const allCards = DATA.all_cards || [];
const cardIcons = DATA.card_icons;
const cardIconsEvo = DATA.card_icons_evo || {};
const cardIconsHero = DATA.card_icons_hero || {};
const playerBriefs = DATA.player_briefs || {};
const playerDecks = DATA.player_decks || {};
const championCards = new Set(DATA.card_champions);
const evoCapableCards = new Set(DATA.card_evolution_capable);
const forcedArtCards = new Set(DATA.forced_art_cards || []);
const transitions = DATA.transitions;

// Mini card-icon strip. Real icon URLs pulled from the Clash Royale API's own battle
// data (every card object includes iconUrls.medium) -- embedded as base64 when a local
// card_icons/ folder is present (see embed_local_icons()), otherwise hotlinked from
// Supercell's CDN; falls back to a plain text initial badge per-card if a name has no
// known icon at all. Champion (hero) cards get a small corner badge showing the card's
// actual hero-form art (falls back to a gold crown symbol if that art isn't available);
// cards this roster has slotted as an Evolution at least once get a small corner badge
// showing the card's actual evolution-form art (falls back to a purple star symbol).
// This is still a simplification for the evolution badge specifically -- evolution
// choice is per-battle-instance, but the decks shown here are aggregated across many
// instances, so the badge flags "can appear evolved" rather than "is evolved in this
// exact displayed deck".
function cardIconStrip(cardNames, size) {
  size = size || 32;
  return '<div class="icon-strip">' + cardNames.map(name => {
    const url = cardIcons[name];
    const heroUrl = cardIconsHero[name];
    const evoUrl = cardIconsEvo[name];
    // Forced-art cards already show the evolution/hero art as the MAIN icon (see
    // FORCE_EVO_ART/FORCE_HERO_ART in build_dashboard.py) -- skip the corner badge
    // entirely for those, it'd be redundant.
    const isForcedArt = forcedArtCards.has(name);
    const champBadge = (!isForcedArt && championCards.has(name))
      ? (heroUrl
          ? `<span class="badge-champ badge-img" title="Champion / Hero card"><img src="${heroUrl}" alt=""></span>`
          : '<span class="badge-champ" title="Champion / Hero card">&#9819;</span>')
      : '';
    const evoBadge = (!isForcedArt && evoCapableCards.has(name))
      ? (evoUrl
          ? `<span class="badge-evo badge-img" title="Sometimes played as an Evolution"><img src="${evoUrl}" alt=""></span>`
          : '<span class="badge-evo" title="Sometimes played as an Evolution">&#9733;</span>')
      : '';
    const badges = champBadge + evoBadge;
    const inner = url
      ? `<img class="card-icon" src="${url}" alt="${name}" title="${name}" loading="lazy" width="${size}" height="${Math.round(size*1.25)}" onerror="this.replaceWith(Object.assign(document.createElement('div'),{className:'card-icon-fallback',textContent:'${name.charAt(0)}',title:'${name}'}))">`
      : `<div class="card-icon-fallback" title="${name}">${name.charAt(0)}</div>`;
    return `<div class="card-icon-wrap">${inner}${badges}</div>`;
  }).join('') + '</div>';
}
// Parses "Card A, Card B, ... (N games)" -> ['Card A', 'Card B', ...]
function parseCardList(text) {
  if (!text) return [];
  const m = text.match(/^(.*)\s\([^)]+\)$/);
  const cardsPart = m ? m[1] : text;
  return cardsPart.split(',').map(s => s.trim()).filter(Boolean);
}

function winRateClass(pct) {
  if (pct >= 55) return 'winrate-good';
  if (pct >= 45) return 'winrate-warn';
  return 'winrate-bad';
}
function parsePct(s) {
  if (typeof s === 'number') return s * 100;
  if (!s) return null;
  const m = String(s).replace('%','').trim();
  const n = parseFloat(m);
  return isNaN(n) ? null : n;
}

// ---------- Quick roster chips ----------
const quickRoster = document.getElementById('quickRoster');
function renderQuickRoster() {
  quickRoster.innerHTML = '';
  players.slice(0, 16).forEach(p => {
    const b = document.createElement('button');
    b.className = 'chip';
    b.textContent = p['Player'];
    b.onclick = () => selectPlayer(p['Player']);
    quickRoster.appendChild(b);
  });
}
renderQuickRoster();

// ---------- Group A quick-access panel (added 2026-07-19) ----------
const groupAChips = document.getElementById('groupAChips');
function renderGroupA() {
  groupAChips.innerHTML = '';
  (DATA.group_a || []).forEach(g => {
    const wrap = document.createElement('span');
    wrap.className = 'group-a-chip-wrap';
    const b = document.createElement('button');
    b.className = 'group-a-chip' + (g.is_you ? ' is-you' : '') + (g.has_data ? '' : ' pending')
      + (g.status === 'on_deck' ? ' on-deck' : '')
      + (g.status === 'reference' ? ' reference' : '');
    const statusNote = g.status === 'on_deck'
      ? ' <span class="pending-note" title="Not yet a confirmed Group A member -- scouted ahead of time in case disqualifications reshuffle the group before Day 2.">(on deck)</span>'
      : g.status === 'reference'
      ? ' <span class="pending-note" title="Not actually in your group -- kept here only for your own reference.">(reference only)</span>'
      : '';
    const noteFlag = (((DATA.player_notes || {})[g.tag]) || []).length
      ? ` <span class="group-a-note-flag" title="Has a watch note -- open the history (▤) to read it.">📌</span>`
      : '';
    if (g.is_you) {
      b.innerHTML = `${g.name} <span class="pending-note">(you)</span>${noteFlag}`;
    } else if (g.has_data) {
      b.innerHTML = `${g.name}${statusNote}${noteFlag}`;
      b.onclick = () => selectPlayer(g.name);
    } else {
      b.innerHTML = `${g.name} <span class="pending-note">(not fetched yet)</span>${statusNote}`;
    }
    wrap.appendChild(b);
    if (g.has_data) {
      const crlN = (DATA.group_a_history && DATA.group_a_history[g.tag] || []).length;
      const pracN = (DATA.group_a_practice_history && DATA.group_a_practice_history[g.tag] || []).length;
      const hb = document.createElement('button');
      hb.className = 'group-a-history-btn';
      hb.title = `View duel history for ${g.name} (${crlN} Official CRL, ${pracN} practice) -- toggle inside`;
      hb.textContent = '▤'; // small grid glyph, reads as "history/log"
      hb.onclick = () => openHistoryModal(g);
      wrap.appendChild(hb);
    }
    groupAChips.appendChild(wrap);
  });
}
renderGroupA();

// ---------- Group A CRL duel history modal (added 2026-07-19) ----------
const historyModal = document.getElementById('historyModal');
const historyModalTitle = document.getElementById('historyModalTitle');
const historyModalSubtitle = document.getElementById('historyModalSubtitle');
const historyModalBody = document.getElementById('historyModalBody');
document.getElementById('historyModalClose').onclick = () => historyModal.classList.remove('show');
historyModal.addEventListener('click', (e) => { if (e.target === historyModal) historyModal.classList.remove('show'); });
document.addEventListener('keydown', (e) => { if (e.key === 'Escape') historyModal.classList.remove('show'); });

function historyRow(g) {
  const cardsFor = g.player_deck || [];
  const cardsAgainst = g.opponent_deck || [];
  const score = (g.crowns_for != null && g.crowns_against != null) ? `${g.crowns_for}-${g.crowns_against}` : '';
  const dayBadge = g.event_day ? `<span class="day-badge day-badge-${g.event_day.replace(/\s+/g, '').toLowerCase()}">${g.event_day}</span>` : '';
  return `<div class="history-row">
    <div class="history-meta"><div class="opp">${dayBadge}${g.opponent_name || 'Unknown'}</div>${g.battle_time || ''}${score ? ' &middot; ' + score + ' crowns' : ''}</div>
    <div class="history-decks">
      ${cardIconStrip(cardsFor, 26)}
      <span class="vs">VS</span>
      ${cardIconStrip(cardsAgainst, 26)}
    </div>
    <div class="history-result ${g.result || ''}">${g.result || '-'}</div>
  </div>`;
}

function recoDeckRow(row) {
  const cardNames = row.deck ? row.deck.split(', ') : [];
  return `<div class="reco-row">
    ${cardIconStrip(cardNames, 26)}
    <div class="reco-wr">${row.wins}/${row.games} (${(row.win_rate*100).toFixed(0)}%)</div>
  </div>`;
}
function recoWinconRow(row) {
  return `<div class="reco-row reco-row-text">
    <div class="reco-label">${row.wincon_set}</div>
    <div class="reco-wr">${row.wins}/${row.games} (${(row.win_rate*100).toFixed(0)}%)</div>
  </div>`;
}
function recoDuelSetRow(row) {
  const decks = row.duel_set ? row.duel_set.split(' / ') : [];
  return `<div class="reco-duelset">
    ${decks.map(d => cardIconStrip(d.split(', '), 22)).join('<span class="vs">+</span>')}
    <div class="reco-wr">${row.wins}/${row.games} (${(row.win_rate*100).toFixed(0)}%)</div>
  </div>`;
}
function renderRecommendationSection(g) {
  const reco = (DATA.group_a_recommendations && DATA.group_a_recommendations[g.tag]) || null;
  if (!reco || !reco.top_wincons || !reco.top_wincons.length) {
    return `<div class="reco-section"><div class="reco-heading">Recommended for Tomorrow</div>
      <div class="history-empty">Not enough classified win-condition data on ${g.name} yet to build a recommendation.</div></div>`;
  }
  const sampleNote = `Based on their top win-con${reco.top_wincons.length > 1 ? 's' : ''} -- <b>${reco.top_wincons.join(', ')}</b> --
    filtered across our whole tracked pool (Practice + Official CRL combined): <b>${reco.sample_size}</b> of our
    own games faced an opponent playing one of those win conditions.`;
  if (!reco.sample_size) {
    return `<div class="reco-section"><div class="reco-heading">Recommended for Tomorrow</div>
      <p class="modal-subtitle" style="margin:0 0 8px;">${sampleNote}</p>
      <div class="history-empty">No games in our pool have faced this win-con archetype yet -- no recommendation possible.</div></div>`;
  }
  return `<div class="reco-section">
    <div class="reco-heading">Recommended for Tomorrow <span class="reco-heading-note">data-driven, from our entire tracked pool -- not theorycrafted</span></div>
    <p class="modal-subtitle" style="margin:0 0 10px;">${sampleNote}</p>
    <div class="reco-col-label">Best deck to bring (min 3 games)</div>
    ${reco.best_decks && reco.best_decks.length ? reco.best_decks.map(recoDeckRow).join('') : '<div class="history-empty">No single deck reaches the 3-game minimum yet.</div>'}
    <div class="reco-col-label">Best win-con set to bring (min 3 games)</div>
    ${reco.best_wincon_sets && reco.best_wincon_sets.length ? reco.best_wincon_sets.map(recoWinconRow).join('') : '<div class="history-empty">No win-con set reaches the 3-game minimum yet.</div>'}
    <div class="reco-col-label">Best duel-set -- decks used together vs this archetype (min 3 games)</div>
    ${reco.best_duel_sets && reco.best_duel_sets.length ? reco.best_duel_sets.map(recoDuelSetRow).join('') : '<div class="history-empty">No duel-set combination reaches the 3-game minimum yet.</div>'}
  </div>`;
}

// ---- Matchup Prep (added 2026-07-19) -- the flip side of "Recommended for Tomorrow":
// what's most likely to beat THEM, and what already has, instead of what WE should bring.
// Mirrors the Excel "Group A Matchup Prep" sheet. Scoped to the same 7 confirmed+on_deck
// players (compute_matchup_prep() in build_dashboard.py's Python only computes this for
// that scope), so this section simply won't be present for reference-only chips.
function matchupThreatRow(row, isDoubleConfirmed) {
  const cardNames = row.deck ? row.deck.split(', ') : [];
  return `<div class="reco-row${isDoubleConfirmed ? ' reco-row-threat' : ''}">
    ${cardIconStrip(cardNames, 26)}
    <div class="reco-wr">${row.wins}/${row.games} (${(row.win_rate*100).toFixed(0)}%)</div>
    ${isDoubleConfirmed ? '<span class="reco-heading-note" style="color:#B71C1C;">&#9888; also beat them personally</span>' : ''}
  </div>`;
}
function matchupWinconRow(row) {
  return `<div class="reco-row reco-row-text">
    <div class="reco-label">${row.wincon}</div>
    <div class="reco-wr">${row.wins}/${row.games} (${(row.win_rate*100).toFixed(0)}%)</div>
  </div>`;
}
function renderMatchupPrepSection(g) {
  const prep = (DATA.group_a_matchup_prep && DATA.group_a_matchup_prep[g.tag]) || null;
  if (!prep || !prep.total_games) {
    return '';
  }
  const doubleConfirmedSet = new Set(prep.double_confirmed || []);
  return `<div class="reco-section">
    <div class="reco-heading">Matchup Prep <span class="reco-heading-note">what's most likely to beat them -- data-driven, not theorycrafted</span></div>
    <p class="modal-subtitle" style="margin:0 0 10px;">${prep.total_games} of their own logged games. Top win-cons: <b>${(prep.top_wincons||[]).join(', ') || 'none identified'}</b>.</p>
    <div class="reco-col-label">What has actually beaten them personally (min 3 games faced)</div>
    ${prep.empirical && prep.empirical.length ? prep.empirical.map(r => matchupThreatRow(r, doubleConfirmedSet.has(r.deck))).join('') : '<div class="history-empty">No single opponent deck has faced them 3+ times yet.</div>'}
    <div class="reco-col-label">What the wider tracked pool says beats their win-con archetype (${prep.predicted_sample_size || 0}g pool, min 3 games)</div>
    ${prep.predicted_decks && prep.predicted_decks.length ? prep.predicted_decks.map(r => matchupThreatRow(r, doubleConfirmedSet.has(r.deck))).join('') : '<div class="history-empty">No single deck reaches the 3-game minimum against this win-con pool yet.</div>'}
    <div class="reco-col-label">Win-conditions that beat their archetype pool-wide (min 5 games)</div>
    ${prep.predicted_wincons && prep.predicted_wincons.length ? prep.predicted_wincons.map(matchupWinconRow).join('') : '<div class="history-empty">No single win condition reaches the 5-game minimum yet.</div>'}
  </div>`;
}

// ---- Sequencing (added 2026-07-19) -- spell+win-con combos and B03 positional
// tendencies, ported from add_group_a_sequencing_analysis.py's Excel sheet. Deliberately
// does NOT show a "switch rate" stat -- real Clash Royale Duel format bans reusing any
// card across a duel's games, so that number is ~100% for everyone by the game's own
// rules, not a player tendency (confirmed against this data before dropping it).
function seqComboRow(row) {
  const wcText = (row.wincons && row.wincons.length) ? row.wincons.join(', ') : '(no win-con identified)';
  const spText = (row.spells && row.spells.length) ? row.spells.join(', ') : '(no spell identified)';
  return `<div class="reco-row reco-row-text">
    <div class="reco-label">${wcText}<span style="color:var(--text-muted);"> + ${spText}</span></div>
    <div class="reco-wr">${row.wins}/${row.games} (${(row.win_rate*100).toFixed(0)}%)</div>
  </div>`;
}
function seqPositionRow(row) {
  return `<div class="reco-row reco-row-text">
    <div class="reco-label">${row.name}</div>
    <div class="reco-wr">Usually Game ${row.modal_pos} (${row.modal_n}/${row.total}) <span style="color:var(--text-muted);">-- ${row.breakdown}</span></div>
  </div>`;
}
function renderSequencingSection(g) {
  const seq = (DATA.group_a_sequencing && DATA.group_a_sequencing[g.tag]) || null;
  if (!seq || !seq.total_games) {
    return '';
  }
  const posLabel = (pos) => {
    const wc = (seq.position_wincons[pos] || []).map(([n, c]) => `${n} (${c}g)`).join(', ') || '(none)';
    const sp = (seq.position_spells[pos] || []).map(([n, c]) => `${n} (${c}g)`).join(', ') || '(none)';
    return `<div class="reco-row reco-row-text"><div class="reco-label">Game ${pos}</div><div class="reco-wr" style="text-align:left;">Win-cons: ${wc}<br>Spells: ${sp}</div></div>`;
  };
  return `<div class="reco-section">
    <div class="reco-heading">Spell &amp; Win-Con Sequencing <span class="reco-heading-note">card-reuse is banned within a duel, so this shows WHICH slot they favor, not "switch rate"</span></div>
    <p class="modal-subtitle" style="margin:0 0 10px;">${seq.total_games} logged games, ${seq.multi_game_duels} multi-game duel set(s) found.</p>
    <div class="reco-col-label">Spell + win-con combos (across all decks played)</div>
    ${seq.combo_rows && seq.combo_rows.length ? seq.combo_rows.map(seqComboRow).join('') : '<div class="history-empty">No full-deck games recorded yet.</div>'}
    ${seq.multi_game_duels ? `
    <div class="reco-col-label">By duel position -- top win-cons / spells at each stage</div>
    ${[1,2,3].map(posLabel).join('')}
    <div class="reco-col-label">Win-condition / spell -- usual position (min 2 games)</div>
    ${(seq.wincon_position_rows||[]).concat(seq.spell_position_rows||[]).length
        ? (seq.wincon_position_rows||[]).concat(seq.spell_position_rows||[]).map(seqPositionRow).join('')
        : '<div class="history-empty">Nothing has appeared 2+ times in a sequenced position yet.</div>'}
    ` : '<div class="history-empty" style="margin-top:6px;">Not enough multi-game duel data yet to analyze positional sequencing.</div>'}
  </div>`;
}

// Recent Practice Trends section -- most-run decks (fuzzy-grouped) + win-con combinations
// since the last CRL day, position-agnostic.
function renderRecentPracticeSection(g) {
  const rp = (DATA.group_a_recent_practice || {})[g.tag];
  if (!rp) return '';
  if (!rp.n_games) {
    return `<div class="reco-section"><div class="reco-heading">Recent Practice Trends</div>
      <div class="history-empty">No practice games tracked since the last CRL day yet.</div></div>`;
  }
  const clusters = (rp.clusters || []).map(c => `
    <div class="rp-deck">
      <div class="rp-deck-head"><span class="rp-count">${c.count}×</span> <span class="wc">${c.wincon}</span></div>
      ${cardIconStrip(c.deck, 26)}
    </div>`).join('');
  const combos = (rp.wincons || []).map(w => `
    <div class="reco-row reco-row-text">
      <div class="reco-label">${w.combo}</div><div class="reco-wr">${w.count}</div>
    </div>`).join('');
  return `<div class="reco-section">
    <div class="reco-heading">Recent Practice Trends <span class="reco-heading-note">practice since ${rp.cutoff} (last CRL day) · ${rp.n_games} games · all set positions</span></div>
    <div class="rp-cols">
      <div class="rp-col">
        <div class="reco-col-label">Most-run decks <span style="color:var(--text-muted);font-weight:400;">(grouped, ≤2-card variance)</span></div>
        ${clusters || '<div class="history-empty">—</div>'}
      </div>
      <div class="rp-col">
        <div class="reco-col-label">Most-frequent win-con combos <span style="color:var(--text-muted);font-weight:400;">(paired as run together)</span></div>
        ${combos || '<div class="history-empty">—</div>'}
      </div>
    </div>
  </div>`;
}

// Duel-Set Record section -- set win rate + how sets are decided, split CRL vs Practice.
function renderDuelSetRecordSection(g) {
  const rec = (DATA.group_a_duel_set_record || {})[g.tag];
  if (!rec) return '';
  const c = rec.crl, p = rec.practice;
  const pct = (x) => x == null ? '—' : `${Math.round(x * 100)}%`;
  const bar = (segs) => {
    const tot = segs.reduce((s, x) => s + x.n, 0);
    if (!tot) return '';
    return `<div class="dsr-bar">${segs.filter(s => s.n).map(s =>
      `<span class="dsr-seg" style="flex:${s.n};background:${s.color}" title="${s.label}: ${s.n}"></span>`).join('')}</div>`;
  };
  const chip = (label, n, cls) => `<span class="dsr-chip ${cls || ''}"><b>${n}</b> ${label}</span>`;

  // ---- Official CRL block ----
  const crlBody = c.sets ? `
    <div class="dsr-line"><span class="dsr-wr">${pct(c.win_rate)}</span>
      <span class="dsr-wr-sub">set win rate · ${c.wins}/${c.sets} sets won</span></div>
    ${bar([
      {n: c["2-0"], color: 'var(--dsr-win)', label: 'Won 2-0'},
      {n: c["2-1"], color: 'var(--dsr-win2)', label: 'Won 2-1'},
      {n: c["1-2"], color: 'var(--dsr-loss2)', label: 'Lost 1-2'},
      {n: c["0-2"], color: 'var(--dsr-loss)', label: 'Lost 0-2'},
    ])}
    <div class="dsr-chips">
      ${chip('won 2-0', c["2-0"], 'w')}${chip('won 2-1', c["2-1"], 'w')}
      ${chip('lost 1-2', c["1-2"], 'l')}${chip('lost 0-2', c["0-2"], 'l')}
    </div>
    <div class="dsr-note">Sweeps (2-0 / 0-2): <b>${c["2-0"] + c["0-2"]}</b> of ${c.sets} · went the distance (2-1 / 1-2): <b>${c["2-1"] + c["1-2"]}</b>${c.anomaly ? ` · <span style="color:var(--text-muted)">${c.anomaly} pending/irregular excluded</span>` : ''}</div>
  ` : `<div class="history-empty">No decided Official CRL sets tracked yet.</div>`;

  // ---- Practice block (all 3 games always played) ----
  const pracBody = p.sets ? `
    <div class="dsr-line"><span class="dsr-wr">${pct(p.win_rate)}</span>
      <span class="dsr-wr-sub">set win rate · ${p.wins}/${p.sets} sets won (decided at 2)</span></div>
    ${bar([
      {n: p["3-0"], color: 'var(--dsr-win)', label: '3-0'},
      {n: p["2-1"], color: 'var(--dsr-win2)', label: '2-1'},
      {n: p["1-2"], color: 'var(--dsr-loss2)', label: '1-2'},
      {n: p["0-3"], color: 'var(--dsr-loss)', label: '0-3'},
    ])}
    <div class="dsr-chips">
      ${chip('3-0', p["3-0"], 'w')}${chip('2-1', p["2-1"], 'w')}
      ${chip('1-2', p["1-2"], 'l')}${chip('0-3', p["0-3"], 'l')}
      <span class="dsr-chip-note">full game record (all 3 played)</span>
    </div>
    <div class="dsr-decide">
      <div class="dsr-decide-row"><span class="dsr-decide-lbl">Went up 2-0 (${p.up20})</span>
        <span class="dsr-decide-val">closed out 3-0 <b>${p.up20_closed}</b> · <span class="l">dropped the last game <b>${p.up20_dropped}</b></span></span></div>
      <div class="dsr-decide-row"><span class="dsr-decide-lbl">Even 1-1 → decider (${p.even})</span>
        <span class="dsr-decide-val">won it <b>${p.even_won}</b> · <span class="l">lost it <b>${p.even_lost}</b></span></span></div>
      <div class="dsr-decide-row"><span class="dsr-decide-lbl">Went down 0-2 (${p.down02})</span>
        <span class="dsr-decide-val">swept 0-3 <b>${p.down02_swept}</b> · won the dead rubber <b>${p.down02_rubber}</b></span></div>
    </div>
    ${p.up20 ? `<div class="dsr-note">When up 2-0, they close it out <b>${Math.round(p.up20_closed / p.up20 * 100)}%</b> of the time and drop the meaningless game 3 <b>${Math.round(p.up20_dropped / p.up20 * 100)}%</b> of the time.</div>` : ''}
  ` : `<div class="history-empty">No complete practice sets tracked since fetching began.</div>`;

  return `<div class="reco-section" style="margin-top:22px;border-top:2px solid var(--border);padding-top:16px;">
    <div class="reco-heading">Duel-Set Record <span class="reco-heading-note">supplementary — how their SETS end (win rate + 2-0 vs 2-1), split by format</span></div>
    <div class="dsr-cols">
      <div class="dsr-col"><div class="dsr-col-head dsr-crl">Official CRL <span>first-to-2, stops</span></div>${crlBody}</div>
      <div class="dsr-col"><div class="dsr-col-head dsr-prac">Practice <span>full Bo3 always played</span></div>${pracBody}</div>
    </div>
  </div>`;
}

// Duel-history mode: 'crl' (Official CRL) or 'practice'. Toggled inside the modal.
let _histMode = 'crl';
function renderHistoryList(g, mode) {
  const isCrl = (mode || _histMode) === 'crl';
  const src = (isCrl ? DATA.group_a_history : DATA.group_a_practice_history) || {};
  const games = src[g.tag] || [];
  if (!games.length) {
    return `<div class="history-empty">No tracked ${isCrl ? 'Official CRL' : 'practice'} games for ${g.name} yet.</div>`;
  }
  return games.map(historyRow).join('');
}
function renderHistoryToggle(g) {
  const crlN = ((DATA.group_a_history || {})[g.tag] || []).length;
  const pracN = ((DATA.group_a_practice_history || {})[g.tag] || []).length;
  return `<div class="hist-toggle">
    <button class="hist-tab ${_histMode === 'crl' ? 'active' : ''}" data-mode="crl">Official CRL <span class="hist-count">${crlN}</span></button>
    <button class="hist-tab ${_histMode === 'practice' ? 'active' : ''}" data-mode="practice">Practice <span class="hist-count">${pracN}</span></button>
  </div>`;
}
function openHistoryModal(g) {
  _histMode = 'crl';
  const statusTag = g.status === 'on_deck'
    ? '<span class="scouted-tag" title="Not yet a confirmed Group A member -- scouted ahead of time in case disqualifications reshuffle the group before Day 2.">on deck</span> '
    : g.status === 'reference'
    ? '<span class="shadow-tag" title="Not actually in your group -- kept here only for your own reference.">reference only</span> '
    : '';
  historyModalTitle.innerHTML = `${statusTag}${g.name} -- Duel History`;
  historyModalSubtitle.textContent =
    "Their decks on the left, opponents' on the right, most recent first. Toggle Official CRL vs practice below.";
  const notes = (DATA.player_notes || {})[g.tag] || [];
  const noteHtml = notes.length
    ? `<div class="player-note">📌 <b>Watch note${notes.length > 1 ? 's' : ''}:</b>` +
      (notes.length > 1
        ? `<ul class="player-note-list">${notes.map(n => `<li>${n}</li>`).join('')}</ul>`
        : ` ${notes[0]}`) + `</div>`
    : '';
  // Pre-render BOTH history lists ONCE (they're the only thing the toggle changes).
  // Switching then just swaps the cached list instead of rebuilding the whole modal --
  // the old code re-rendered every section (thousands of card icons) on each click,
  // which caused multi-second/​minute lag for players with long histories.
  const histHtml = {
    crl: renderHistoryList(g, 'crl'),
    practice: renderHistoryList(g, 'practice'),
  };
  historyModalBody.innerHTML =
    noteHtml +
    renderRecentPracticeSection(g) +
    renderRecommendationSection(g) +
    renderMatchupPrepSection(g) +
    renderSequencingSection(g) +
    '<div class="reco-heading" style="margin-top:18px;">Duel History</div>' +
    renderHistoryToggle(g) +
    `<div id="histList">
       <div class="hist-pane" data-pane="crl"${_histMode === 'crl' ? '' : ' style="display:none"'}>${histHtml.crl}</div>
       <div class="hist-pane" data-pane="practice"${_histMode === 'practice' ? '' : ' style="display:none"'}>${histHtml.practice}</div>
     </div>` +
    renderDuelSetRecordSection(g);
  // Both lists are already in the DOM; the toggle just flips which pane is visible -- a pure
  // CSS show/hide, so switching is instant no matter how many games a player has.
  historyModalBody.querySelectorAll('.hist-tab').forEach(btn => {
    btn.onclick = () => {
      if (btn.dataset.mode === _histMode) return;
      _histMode = btn.dataset.mode;
      historyModalBody.querySelectorAll('#histList .hist-pane').forEach(p =>
        p.style.display = (p.dataset.pane === _histMode) ? '' : 'none');
      historyModalBody.querySelectorAll('.hist-tab').forEach(b =>
        b.classList.toggle('active', b.dataset.mode === _histMode));
    };
  });
  historyModal.classList.add('show');
}

// ---------- Search ----------
const searchInput = document.getElementById('searchInput');
const suggestionsEl = document.getElementById('suggestions');
let activeIdx = -1;

// Matches by player NAME first, then falls back to player TAG (e.g. #80ULUJLYY or the bare
// tag digits without '#') -- added 2026-07-18 so a renamed/mismatched account (tracked under
// one name, but the API now reports a different in-game name -- see the ElMollejas/MH Axel
// case) can still be found by searching the tag you originally recorded for them.
function playerTagsOf(p) {
  return (p['Player Tag(s)'] || '').split(',').map(t => t.trim()).filter(Boolean);
}
function renderSuggestions(query) {
  const q = query.trim().toLowerCase();
  if (!q) { suggestionsEl.classList.remove('show'); suggestionsEl.innerHTML = ''; return; }
  const qTag = q.replace(/^#/, '');
  const nameMatches = players.filter(p => p['Player'].toLowerCase().includes(q));
  const nameMatchSet = new Set(nameMatches.map(p => p['Player']));
  const tagMatches = players.filter(p =>
    !nameMatchSet.has(p['Player']) &&
    playerTagsOf(p).some(t => t.toLowerCase().replace(/^#/, '').includes(qTag))
  );
  const matches = nameMatches.concat(tagMatches).slice(0, 8);
  if (!matches.length) { suggestionsEl.classList.remove('show'); suggestionsEl.innerHTML = ''; return; }
  suggestionsEl.innerHTML = '';
  matches.forEach((p, i) => {
    const btn = document.createElement('button');
    const matchedByTag = tagMatches.includes(p);
    const tagSuffix = matchedByTag ? '  \\u00b7  matched tag ' + playerTagsOf(p).find(t => t.toLowerCase().replace(/^#/, '').includes(qTag)) : '';
    btn.textContent = p['Player'] + (p['_is_scouted'] ? ' (scouted)' : p['_is_extended'] ? ' (extended roster)' : p['_is_shadow'] ? ' (opponent)' : '') + '  \\u00b7  ' + p['Total Games'] + ' games, ' + p['Win Rate'] + ' win rate' + tagSuffix;
    btn.onmousedown = (e) => { e.preventDefault(); selectPlayer(p['Player']); };
    suggestionsEl.appendChild(btn);
  });
  suggestionsEl.classList.add('show');
  activeIdx = -1;
}

searchInput.addEventListener('input', () => renderSuggestions(searchInput.value));
searchInput.addEventListener('focus', () => renderSuggestions(searchInput.value));
searchInput.addEventListener('keydown', (e) => {
  const items = Array.from(suggestionsEl.querySelectorAll('button'));
  if (e.key === 'ArrowDown') { e.preventDefault(); activeIdx = Math.min(activeIdx + 1, items.length - 1); }
  else if (e.key === 'ArrowUp') { e.preventDefault(); activeIdx = Math.max(activeIdx - 1, 0); }
  else if (e.key === 'Enter') {
    e.preventDefault();
    if (activeIdx >= 0 && items[activeIdx]) { items[activeIdx].dispatchEvent(new Event('mousedown')); }
    else {
      const q = searchInput.value.trim().toLowerCase();
      const qTag = q.replace(/^#/, '');
      const match = players.find(p => p['Player'].toLowerCase() === q) ||
                    players.find(p => p['Player'].toLowerCase().includes(q)) ||
                    players.find(p => playerTagsOf(p).some(t => t.toLowerCase().replace(/^#/, '').includes(qTag)));
      if (match) selectPlayer(match['Player']);
    }
    return;
  } else { return; }
  items.forEach((it, i) => it.classList.toggle('active', i === activeIdx));
});
document.addEventListener('click', (e) => {
  if (!e.target.closest('.search-box')) suggestionsEl.classList.remove('show');
});

function deckCard(text) {
  if (!text) return '<div class="deck-card" style="color:var(--text-muted);">-</div>';
  const m = text.match(/^(.*)\\s\\(([^)]+)\\)$/);
  const cardsText = m ? m[1] : text;
  const meta = m ? m[2] : '';
  const cardNames = parseCardList(text);
  return '<div class="deck-card">' + cardIconStrip(cardNames, 34) +
    '<div class="deck-meta">' + meta + '</div></div>';
}
function winconBadge(text) {
  if (!text) return '';
  const m = text.match(/^(.*)\\s\\(([^)]+)\\)$/);
  const cardName = m ? m[1] : text;
  const meta = m ? m[2] : '';
  return '<div class="wincon-badge">' + cardIconStrip([cardName], 28) +
    '<div><div>' + cardName + '</div><div class="wincon-meta">' + meta + '</div></div></div>';
}

function briefDeckRow(entry) {
  if (!entry) return '';
  const cardNames = entry.deck.split(', ');
  const pctText = (entry.win_rate * 100).toFixed(0) + '%';
  return '<div class="brief-deck-row">' + cardIconStrip(cardNames, 22) +
    '<div class="meta">' + pctText + ' (' + entry.games + ' game' + (entry.games === 1 ? '' : 's') + ')</div></div>';
}
function briefWinconRow(text) {
  if (!text) return '';
  const m = text.match(/^(.*)\\s\\(([^)]+)\\)$/);
  const cardName = m ? m[1] : text;
  const meta = m ? m[2] : '';
  return '<div class="brief-wincon-row">' + cardIconStrip([cardName], 22) +
    '<div>' + cardName + ' <span class="n">(' + meta + ')</span></div></div>';
}
// Elixir-tendency labels are a rough heuristic bucketed from average deck elixir
// cost (real elixirCost values from the API) -- not an official Supercell category,
// just a quick read on cycle-speed tendency.
function elixirTag(avg) {
  if (avg == null) return '';
  if (avg < 3.4) return 'Fast cycle';
  if (avg < 4.0) return 'Balanced';
  return 'Heavy / beatdown';
}
function renderBrief(name, p) {
  const brief = playerBriefs[name];
  const wincons = [p['Top Win Condition #1'], p['Top Win Condition #2'], p['Top Win Condition #3']].filter(Boolean);
  const bestDecks = brief ? brief.best_decks : [];
  const worstDecks = brief ? brief.worst_decks : [];
  const elixirVal = brief ? brief.avg_elixir : null;

  let matchupHtml;
  if (bestDecks.length && worstDecks.length && bestDecks.length === 1 && bestDecks[0].deck === worstDecks[0].deck) {
    matchupHtml = '<div class="brief-empty" style="margin-bottom:2px;">Only one deck with enough games tracked:</div>' + briefDeckRow(bestDecks[0]);
  } else {
    matchupHtml = (bestDecks[0]
        ? '<div class="brief-empty" style="margin-bottom:2px;">Best:</div>' + briefDeckRow(bestDecks[0])
        : '<div class="brief-empty">No deck with enough games yet.</div>')
      + (worstDecks[0]
        ? '<div class="brief-empty" style="margin:6px 0 2px;">Worst:</div>' + briefDeckRow(worstDecks[0])
        : '');
  }

  document.getElementById('briefGrid').innerHTML = `
    <div>
      <div class="brief-col-label">Top Win Conditions</div>
      ${wincons.length ? wincons.map(briefWinconRow).join('') : '<div class="brief-empty">None identified yet.</div>'}
    </div>
    <div>
      <div class="brief-col-label">Best / Worst Matchup Decks</div>
      ${matchupHtml}
    </div>
    <div>
      <div class="brief-col-label">Elixir Tendency</div>
      ${elixirVal != null
        ? `<div class="brief-elixir">${elixirVal.toFixed(1)}<span class="unit">avg elixir</span></div>`
          + `<div class="brief-elixir-tag" title="Rough heuristic bucketed from average deck elixir cost across this player's tracked decks">${elixirTag(elixirVal)}</div>`
        : '<div class="brief-empty">Not enough deck data yet.</div>'}
    </div>
  `;
}

// ---------- Deck Explorer ----------
// JS port of the Python _deck_overlap / _cluster_decks_by_overlap union-find helpers
// (build_dashboard.py / build_duel_workbook.py) -- groups decks sharing at least
// `threshold` of their 8 cards into one "family" so 1-2 card swaps of what's really
// the same deck still show up aggregated together, instead of as separate rows.
function deckOverlap(deckKeyA, deckKeyB) {
  const a = new Set(deckKeyA.split(', '));
  let n = 0;
  deckKeyB.split(', ').forEach(c => { if (a.has(c)) n++; });
  return n;
}
function clusterDecksByOverlap(deckKeys, threshold) {
  const parent = {};
  deckKeys.forEach(d => { parent[d] = d; });
  function find(x) {
    while (parent[x] !== x) { parent[x] = parent[parent[x]]; x = parent[x]; }
    return x;
  }
  function union(a, b) {
    const ra = find(a), rb = find(b);
    if (ra !== rb) parent[ra] = rb;
  }
  const deckList = Array.from(deckKeys);
  for (let i = 0; i < deckList.length; i++) {
    for (let j = i + 1; j < deckList.length; j++) {
      if (deckOverlap(deckList[i], deckList[j]) >= threshold) union(deckList[i], deckList[j]);
    }
  }
  const families = {};
  deckList.forEach(d => {
    const root = find(d);
    (families[root] = families[root] || []).push(d);
  });
  return families;
}
// Builds aggregated "archetype" rows for a player's tracked games: filters by
// category ('all' = Practice + Official CRL, 'crl' = Official CRL only), clusters
// the distinct decks played into families at the given card-overlap threshold, and
// rolls up games/wins per family. Representative deck shown is the most-played exact
// deck within that family.
function computeDeckFamilies(tag, category, threshold) {
  const games = playerDecks[tag] || [];
  const filtered = category === 'crl' ? games.filter(g => g.c === 'crl') : games;
  if (!filtered.length) return [];

  const exactGames = {};
  const exactWins = {};
  filtered.forEach(g => {
    exactGames[g.d] = (exactGames[g.d] || 0) + 1;
    if (g.w) exactWins[g.d] = (exactWins[g.d] || 0) + 1;
  });
  const deckKeys = Object.keys(exactGames);
  const families = clusterDecksByOverlap(deckKeys, threshold);

  const rows = [];
  Object.values(families).forEach(members => {
    let totalGames = 0, totalWins = 0, repDeck = null, repGames = -1;
    members.forEach(dk => {
      const g = exactGames[dk] || 0;
      const w = exactWins[dk] || 0;
      totalGames += g;
      totalWins += w;
      if (g > repGames) { repGames = g; repDeck = dk; }
    });
    rows.push({
      deck: repDeck,
      games: totalGames,
      wins: totalWins,
      winRate: totalGames ? totalWins / totalGames : 0,
      variants: members.length,
    });
  });
  rows.sort((a, b) => b.games - a.games || b.winRate - a.winRate);
  return rows;
}
function deckExplorerRow(row) {
  const cardNames = row.deck.split(', ');
  const pct = (row.winRate * 100).toFixed(0) + '%';
  const variantNote = row.variants > 1 ? row.variants + ' variants clustered' : 'exact match only';
  return '<div class="deck-explorer-row">' + cardIconStrip(cardNames, 30) +
    '<div class="der-meta">' +
      '<div class="der-games">' + row.games + ' game' + (row.games === 1 ? '' : 's') + '</div>' +
      '<div class="der-wr">' + pct + ' win rate (' + row.wins + '-' + (row.games - row.wins) + ')</div>' +
      '<div class="der-variants">' + variantNote + '</div>' +
    '</div></div>';
}
let currentProfileTag = null;
function renderDeckExplorer(tag) {
  currentProfileTag = tag;
  const resultsEl = document.getElementById('deckExplorerResults');
  if (!tag || !playerDecks[tag] || !playerDecks[tag].length) {
    resultsEl.innerHTML = '<div class="history-empty">No full 8-card deck games recorded for this player yet.</div>';
    return;
  }
  const category = document.getElementById('deckExplorerCategory').value;
  const threshold = parseInt(document.getElementById('deckExplorerThreshold').value, 10);
  const rows = computeDeckFamilies(tag, category, threshold);
  resultsEl.innerHTML = rows.length
    ? rows.map(deckExplorerRow).join('')
    : '<div class="history-empty">No games in this category yet -- try "Practice + Official CRL" or a lower card-match threshold.</div>';
}
document.getElementById('deckExplorerCategory').addEventListener('change', () => {
  if (currentProfileTag) renderDeckExplorer(currentProfileTag);
});
document.getElementById('deckExplorerThreshold').addEventListener('change', () => {
  if (currentProfileTag) renderDeckExplorer(currentProfileTag);
});

function selectPlayer(name) {
  const p = playerByName[name];
  document.getElementById('noResult').style.display = 'none';
  if (!p) {
    document.getElementById('profile').classList.remove('show');
    document.getElementById('noResult').style.display = 'block';
    return;
  }
  searchInput.value = name;
  suggestionsEl.classList.remove('show');

  const tagsForProfile = playerTagsOf(p);
  document.getElementById('profName').innerHTML = name + (p['_is_scouted']
    ? ' <span class="scouted-tag" title="One-off scouted player -- not on the tracked roster. Profile built directly from their own recent battle log (~25-30 most recent battles), not from paired duels against your roster.">scouted</span>'
    : p['_is_extended']
    ? ' <span class="extended-tag" title="Extended roster -- a permanently-tracked opponent (encountered in an Official CRL game), not one of the original 48. Fetched and archived like the main roster, but kept out of Official CRL opponent-detection so games vs. the original 48 still classify correctly.">extended roster</span>'
    : p['_is_shadow']
    ? ' <span class="shadow-tag" title="Not one of the ~50 tracked players -- this profile is built only from games where a tracked player faced them, so coverage may be thin.">seen as opponent</span>'
    : '') + (tagsForProfile.length
      ? ` <span class="table-note" style="font-size:12px;">${tagsForProfile.join(', ')}</span>`
      : '');
  const pct = parsePct(p['Win Rate']);
  const wrClass = pct === null ? '' : winRateClass(pct);
  const officialCount = p['Official CRL Games'] || 0;
  document.getElementById('statRow').innerHTML = `
    <div class="stat-tile"><div class="num">${p['Total Games']}</div><div class="label">Total Games</div></div>
    <div class="stat-tile"><div class="num">${p['Total Wins']}</div><div class="label">Total Wins</div></div>
    <div class="stat-tile"><div class="num ${wrClass}">${p['Win Rate']}</div><div class="label">Win Rate</div></div>
    ${matchCategoryFilter === 'all' ? `<div class="stat-tile"><div class="num">${p['Practice Games'] || 0} / ${officialCount}</div><div class="label">Practice / Official CRL</div></div>` : ''}
  `;

  renderBrief(name, p);

  document.getElementById('mostPlayedGrid').innerHTML =
    [p['Most-Played Deck #1'], p['Most-Played Deck #2'], p['Most-Played Deck #3']].map(deckCard).join('');
  document.getElementById('bestWinRateGrid').innerHTML =
    [p['Best Win-Rate Deck #1'], p['Best Win-Rate Deck #2'], p['Best Win-Rate Deck #3']].map(deckCard).join('');
  document.getElementById('winconBadges').innerHTML =
    [p['Top Win Condition #1'], p['Top Win Condition #2'], p['Top Win Condition #3']]
      .filter(Boolean).map(winconBadge).join('') || '<span style="color:var(--text-muted); font-size:13px;">No win conditions identified yet.</span>';

  renderDeckExplorer(tagsForProfile[0] || null);

  document.getElementById('profile').classList.add('show');
}

// ---------- Match Category filter + weight toggle ----------
const weightToggleEl = document.getElementById('weightToggle');
const weightToggleLabelEl = document.getElementById('weightToggleLabel');

function onPlayerPoolChanged() {
  const currentName = searchInput.value.trim();
  rebuildPlayerPool();
  renderQuickRoster();
  if (currentName && playerByName[currentName]) {
    selectPlayer(currentName);
  } else if (currentName) {
    // the previously-selected player has no games in the newly-selected category
    document.getElementById('profile').classList.remove('show');
    document.getElementById('noResult').style.display = 'block';
  }
}

document.querySelectorAll('#matchCategoryTabs .category-tab').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('#matchCategoryTabs .category-tab').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    matchCategoryFilter = btn.dataset.category;
    weightToggleLabelEl.classList.toggle('disabled', matchCategoryFilter !== 'all');
    onPlayerPoolChanged();
  });
});
weightToggleEl.addEventListener('change', () => {
  weightOfficialGames = weightToggleEl.checked;
  onPlayerPoolChanged();
});

// ---------- Generic sortable/filterable table ----------
// opts (added 2026-07-19): { filterInputId, filterFn, minGamesKey, minGamesSelectId,
// showCountSelectId } -- minGamesKey is the row field to compare against the "Min games
// played" dropdown (hides small-sample rows, e.g. a 1-game deck showing a misleading
// 100% win rate), showCountSelectId caps how many rows render at once so long tables
// (Deck Stats, Popular Win-Con Sets) don't dominate the page.
function buildTable(tableId, rows, rowRenderer, opts) {
  opts = opts || {};
  const { filterInputId, filterFn, minGamesKey, minGamesSelectId, showCountSelectId } = opts;
  const table = document.getElementById(tableId);
  const tbody = table.querySelector('tbody');
  const ths = Array.from(table.querySelectorAll('th'));
  let sortKey = null, sortDir = 1;
  let currentRows = rows;

  function computeBase() {
    let base = rows;
    if (filterInputId) {
      const q = document.getElementById(filterInputId).value.trim().toLowerCase();
      if (q) base = base.filter(r => filterFn(r, q));
    }
    if (minGamesKey && minGamesSelectId) {
      const minVal = parseInt(document.getElementById(minGamesSelectId).value, 10) || 0;
      if (minVal > 0) base = base.filter(r => (r[minGamesKey] || 0) >= minVal);
    }
    return base;
  }
  function render() {
    let shown = currentRows;
    if (showCountSelectId) {
      const limVal = document.getElementById(showCountSelectId).value;
      if (limVal !== 'all') shown = shown.slice(0, parseInt(limVal, 10));
    }
    tbody.innerHTML = shown.length
      ? shown.map(rowRenderer).join('')
      : `<tr><td colspan="${ths.length}" class="empty-hint" style="text-align:center;">No rows match these filters -- try lowering "Min games played".</td></tr>`;
  }
  function applySort() {
    if (!sortKey) return;
    const th = ths.find(t => t.dataset.key === sortKey);
    const type = th.dataset.type;
    currentRows = currentRows.slice().sort((a, b) => {
      let av = a[sortKey], bv = b[sortKey];
      if (type === 'num') {
        av = typeof av === 'number' ? av : parsePct(av) ?? 0;
        bv = typeof bv === 'number' ? bv : parsePct(bv) ?? 0;
        return (av - bv) * sortDir;
      }
      av = (av || '').toString().toLowerCase();
      bv = (bv || '').toString().toLowerCase();
      return av.localeCompare(bv) * sortDir;
    });
  }
  function refresh() {
    currentRows = computeBase();
    applySort();
    render();
  }
  ths.forEach(th => {
    th.addEventListener('click', () => {
      const key = th.dataset.key;
      if (sortKey === key) sortDir *= -1; else { sortKey = key; sortDir = th.dataset.type === 'num' ? -1 : 1; }
      ths.forEach(t => t.querySelector('.arrow')?.remove());
      const arrow = document.createElement('span');
      arrow.className = 'arrow';
      arrow.textContent = sortDir === 1 ? '\\u25B2' : '\\u25BC';
      th.appendChild(arrow);
      applySort();
      render();
    });
  });
  if (filterInputId) document.getElementById(filterInputId).addEventListener('input', refresh);
  if (minGamesSelectId) document.getElementById(minGamesSelectId).addEventListener('change', refresh);
  if (showCountSelectId) document.getElementById(showCountSelectId).addEventListener('change', render);
  refresh();
}

function wrBar(pctRaw) {
  const pct = typeof pctRaw === 'number' ? pctRaw * 100 : parsePct(pctRaw);
  if (pct === null) return '<span style="color:var(--text-muted);">-</span>';
  const clamped = Math.max(0, Math.min(100, pct));
  return `<div class="wr-bar-wrap"><span>${pct.toFixed(0)}%</span><div class="wr-bar-track"><div class="wr-bar-fill" style="width:${clamped}%;"></div></div></div>`;
}

buildTable('winconTable', winconSets, r => `
  <tr>
    <td>${cardIconStrip(r['Win-Con Set'].split('+'), 26)}<div class="row-label">${r['Win-Con Set']}</div></td>
    <td class="num">${r['Times Played (Duels)']}</td>
    <td class="num">${r['Games Played']}</td>
    <td class="num">${wrBar(r['Win Rate'])}</td>
    <td>${r['Players Who Used This']}</td>
  </tr>
`, {
  filterInputId: 'winconFilter',
  filterFn: (r, q) => r['Win-Con Set'].toLowerCase().includes(q) || (r['Players Who Used This']||'').toLowerCase().includes(q),
  minGamesKey: 'Games Played', minGamesSelectId: 'winconMinGames', showCountSelectId: 'winconShowCount',
});

buildTable('deckTable', deckStats, r => `
  <tr>
    <td>${cardIconStrip(r['Deck (sorted)'].split(', '), 26)}<div class="row-label">${r['Deck (sorted)']}</div></td>
    <td>${r['Used By']}</td>
    <td class="num">${r['Games Played']}</td>
    <td class="num">${wrBar(r['Win Rate'])}</td>
  </tr>
`, {
  filterInputId: 'deckFilter',
  filterFn: (r, q) => r['Deck (sorted)'].toLowerCase().includes(q) || (r['Used By']||'').toLowerCase().includes(q),
  minGamesKey: 'Games Played', minGamesSelectId: 'deckMinGames', showCountSelectId: 'deckShowCount',
});

// ---------- Best Picks (game-day quick reference, added 2026-07-18) ----------
const BEST_PICKS = DATA.best_picks || {};
let bestPicksCategory = 'all';   // 'all' | 'practice' | 'official'
let bestPicksView = 'decks';     // 'decks' | 'duel_sets' | 'wincon_sets'
const bestPicksResultsEl = document.getElementById('bestPicksResults');
const bestPicksNoteEl = document.getElementById('bestPicksNote');

function bestPicksDeckRow(row, i) {
  const cards = row.deck.split(', ');
  return `<tr>
    <td><span class="best-picks-rank">${i + 1}</span>${cardIconStrip(cards, 26)}</td>
    <td class="num">${row.games}</td>
    <td class="num">${wrBar(row.win_rate)}</td>
  </tr>`;
}

function bestPicksWinconSetRow(row, i) {
  const cards = row.wincon_set.split('+');
  return `<tr>
    <td><span class="best-picks-rank">${i + 1}</span>${cardIconStrip(cards, 26)}<div class="row-label">${row.wincon_set}</div></td>
    <td class="num">${row.duels}</td>
    <td class="num">${row.games}</td>
    <td class="num">${wrBar(row.win_rate)}</td>
  </tr>`;
}

function bestPicksDuelSetRow(row, i) {
  const decksHtml = row.example_decks.map((deckKey, di) => {
    const cards = deckKey.split(', ');
    const famSize = row.family_sizes[di];
    const famNote = famSize > 1 ? ` <span class="table-note" style="font-size:10.5px;">+${famSize - 1} similar deck${famSize - 1 === 1 ? '' : 's'}</span>` : '';
    return `<div>${cardIconStrip(cards, 22)}${famNote}</div>`;
  }).join('');
  return `<tr>
    <td><span class="best-picks-rank">${i + 1}</span><div class="duelset-decks">${decksHtml}</div></td>
    <td class="num">${row.duels}</td>
    <td class="num">${row.games}</td>
    <td class="num">${wrBar(row.win_rate)}</td>
  </tr>`;
}

// "Min times played" / "Show top N" controls (added 2026-07-19): the backend now bakes
// in every deck/win-con-set/duel-set regardless of sample size (min 1), so a 1-game
// deck with a lucky win no longer silently dominates a win-rate sort -- these two
// dropdowns filter and cap the list live, no rebuild needed.
const bestPicksMinPlayedEl = document.getElementById('bestPicksMinPlayed');
const bestPicksShowCountEl = document.getElementById('bestPicksShowCount');

function limitRows(list) {
  const v = bestPicksShowCountEl.value;
  return v === 'all' ? list : list.slice(0, parseInt(v, 10));
}

function renderBestPicks() {
  const bucket = BEST_PICKS[bestPicksCategory] || { decks: [], wincon_sets: [], duel_sets: [], duel_sets_threshold: null };
  const minPlayed = parseInt(bestPicksMinPlayedEl.value, 10) || 1;
  if (bestPicksView === 'decks') {
    const filtered = bucket.decks.filter(r => r.games >= minPlayed);
    const shown = limitRows(filtered);
    bestPicksNoteEl.textContent = filtered.length
      ? `Showing ${shown.length} of ${filtered.length} decks with >= ${minPlayed} game${minPlayed === 1 ? '' : 's'} played (${bucket.decks.length} total tracked), ranked by win rate.`
      : `No decks with >= ${minPlayed} games played yet in this category -- try lowering "Min times played".`;
    bestPicksResultsEl.innerHTML = shown.length ? `
      <table><thead><tr><th>Deck</th><th class="num">Games</th><th class="num">Win Rate</th></tr></thead>
      <tbody>${shown.map(bestPicksDeckRow).join('')}</tbody></table>
    ` : '<div class="empty-hint">Nothing to show yet.</div>';
  } else if (bestPicksView === 'wincon_sets') {
    const filtered = bucket.wincon_sets.filter(r => r.duels >= minPlayed);
    const shown = limitRows(filtered);
    bestPicksNoteEl.textContent = filtered.length
      ? `Showing ${shown.length} of ${filtered.length} win-con sets with >= ${minPlayed} duel${minPlayed === 1 ? '' : 's'} played (${bucket.wincon_sets.length} total tracked; each duel needs all 3 games captured), ranked by win rate.`
      : `No win-con sets with >= ${minPlayed} duels yet in this category -- try lowering "Min times played".`;
    bestPicksResultsEl.innerHTML = shown.length ? `
      <table><thead><tr><th>Win-Con Set</th><th class="num">Duels</th><th class="num">Games</th><th class="num">Win Rate</th></tr></thead>
      <tbody>${shown.map(bestPicksWinconSetRow).join('')}</tbody></table>
    ` : '<div class="empty-hint">Nothing to show yet.</div>';
  } else {
    const t = bucket.duel_sets_threshold;
    const filtered = bucket.duel_sets.filter(r => r.duels >= minPlayed);
    const shown = limitRows(filtered);
    bestPicksNoteEl.textContent = filtered.length
      ? `Showing ${shown.length} of ${filtered.length} duel sets with >= ${minPlayed} duel${minPlayed === 1 ? '' : 's'} played (${bucket.duel_sets.length} total tracked). Decks compared as a match if they share at least ${t}/8 cards -- ${t === 8 ? 'exact decks only' : `loosened from an exact match down to ${t}/8 shared cards to find enough repeats`}.`
      : `No duel-set combination has >= ${minPlayed} duels yet in this category, even at the loosest comparison -- try lowering "Min times played".`;
    bestPicksResultsEl.innerHTML = shown.length ? `
      <table><thead><tr><th>Decks Used Together</th><th class="num">Duels</th><th class="num">Games</th><th class="num">Win Rate</th></tr></thead>
      <tbody>${shown.map(bestPicksDuelSetRow).join('')}</tbody></table>
    ` : '<div class="empty-hint">Nothing to show yet.</div>';
  }
}

document.querySelectorAll('#bestPicksCategoryTabs .category-tab').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('#bestPicksCategoryTabs .category-tab').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    bestPicksCategory = btn.dataset.category;
    renderBestPicks();
  });
});
document.querySelectorAll('#bestPicksViewTabs .view-tab').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('#bestPicksViewTabs .view-tab').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    bestPicksView = btn.dataset.view;
    renderBestPicks();
  });
});
bestPicksMinPlayedEl.addEventListener('change', renderBestPicks);
bestPicksShowCountEl.addEventListener('change', renderBestPicks);
renderBestPicks();

// ---------- Page nav (Scout Tools vs. Best Picks & Stats, added 2026-07-19) ----------
document.querySelectorAll('.page-nav-btn').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.page-nav-btn').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    document.querySelectorAll('.page').forEach(p => { p.style.display = 'none'; });
    document.getElementById(btn.dataset.page).style.display = '';
  });
});

// ---------- What Might Follow? predictor ----------
// transitions.wincon[fromWincon] = [[toWincon, count], ...]  (single-card fallback, top 8, roster-wide)
// transitions.wincon_by_player[player][fromWincon] = [[toWincon, count], ...]
// transitions.wincon_set["A::B"] = [[toWincon, count], ...]  -- exact-match on the FULL SET of win
// cons seen together in the earlier game (a deck can run more than one at once)
// transitions.wincon_set_by_player[player]["A::B"] = [[toWincon, count], ...]
// transitions.wincon3 / wincon3_by_player: same idea but single-card, keyed "g1|||g2"
// transitions.wincon3_set / wincon3_set_by_player: exact-set version, keyed "A::B|||C::D"
// transitions.set_sep ("::") joins multiple win cons selected for ONE game;
// transitions.triple_sep ("|||") separates the G1 set from the G2 set in wincon3(_set) keys.
const SET_SEP = transitions.set_sep || '::';
const TRIPLE_SEP = transitions.triple_sep || '|||';
let predictorMode = 'wincon';
let g1Selected = new Set();
let g2Selected = new Set();
const predictorScopeEl = document.getElementById('predictorScope');
const predictorPickersEl = document.getElementById('predictorPickers');
const predictorResultsEl = document.getElementById('predictorResults');
const predictorSubEl = document.getElementById('predictorSub');
const predictorDqNoteEl = document.getElementById('predictorDqNote');
const allWincons = (transitions.all_wincons || []).slice().sort();
if (predictorDqNoteEl) {
  const excluded = transitions.excluded_uncertain_duels || 0;
  predictorDqNoteEl.textContent = `Data quality: ${excluded} "uncertain-start" duel(s) -- the first duel found for a player/opponent pair, which may really be a continuation of an earlier duel outside the API's fetch window -- are excluded here since their game order can't be verified. Incomplete duels (<3 games) are excluded too. See the workbook's Data Quality sheet for details.`;
}

const PREDICTOR_SUB_TEXT = {
  wincon: "Check every win condition your opponent played in an earlier game of this duel (a deck can run more than one at once -- e.g. Goblin Barrel + Wall Breakers + Miner together). Shows what most commonly showed up in a LATER game of the same duel, aggregated across the tracked roster's duel history.",
  wincon3: "Check the win condition(s) your opponent played in game 1, AND the win condition(s) they played in game 2 of this duel. Shows what most commonly showed up in game 3+ of the same duel. Needs duels with at least 3 tracked games, so coverage is thinner than the game 2 predictor.",
};

players.forEach(p => {
  const opt = document.createElement('option');
  opt.value = p['Player'];
  opt.textContent = p['Player'] + ' (their own tendency)';
  predictorScopeEl.appendChild(opt);
});

function pickerColumn(label, selectedSet, targetKey) {
  const chips = allWincons.map(w => {
    const active = selectedSet.has(w);
    return `<span class="picker-chip ${active ? 'active' : ''}" data-wc="${w.replace(/"/g,'&quot;')}">${w}</span>`;
  }).join('');
  return `<div class="picker-col">
    <div class="picker-col-label">${label}</div>
    <div class="picker-search-row">
      <input type="text" class="picker-search-input" data-target="${targetKey}" placeholder="Time-crunch search (Enter to add) -- wb, gob, skelly, hog...">
    </div>
    <div class="picker-search-status" data-target-status="${targetKey}"></div>
    <div class="picker-chips" data-target="${targetKey}">${chips || '<span class="empty-hint" style="margin:0;">No win conditions identified yet.</span>'}</div>
  </div>`;
}

// Fast, typo/abbreviation-tolerant search: type "wb", "gob barrel", "skelly", "dart",
// etc. and Enter adds the matched win condition to that column without having to
// scan/click through the full chip grid -- built for the between-games time crunch.
function wirePickerSearchInputs() {
  predictorPickersEl.querySelectorAll('.picker-search-input').forEach(input => {
    input.addEventListener('keydown', e => {
      if (e.key !== 'Enter') return;
      e.preventDefault();
      const raw = input.value.trim();
      const targetKey = input.dataset.target;
      const statusEl = predictorPickersEl.querySelector(`.picker-search-status[data-target-status="${targetKey}"]`);
      if (!raw) return;
      const target = targetKey === 'g2' ? g2Selected : g1Selected;
      const match = fuzzyMatchName(raw, allWincons);
      if (match) {
        target.add(match);
        renderPickers();
        renderPredictor();
        const refocused = predictorPickersEl.querySelector(`.picker-search-input[data-target="${targetKey}"]`);
        if (refocused) refocused.focus();
      } else if (statusEl) {
        statusEl.textContent = `couldn't match "${raw}" to a win condition`;
        statusEl.className = 'picker-search-status warn';
      }
    });
  });
}

function renderPickers() {
  predictorSubEl.textContent = PREDICTOR_SUB_TEXT[predictorMode] || '';
  if (predictorMode === 'wincon3') {
    predictorPickersEl.innerHTML =
      pickerColumn('Game 1 win condition(s)', g1Selected, 'g1') +
      pickerColumn('Game 2 win condition(s)', g2Selected, 'g2');
  } else {
    predictorPickersEl.innerHTML = pickerColumn('Earlier-game win condition(s)', g1Selected, 'g1');
  }
  predictorPickersEl.querySelectorAll('.picker-chip').forEach(chip => {
    chip.addEventListener('click', () => {
      const target = chip.parentElement.dataset.target === 'g2' ? g2Selected : g1Selected;
      const wc = chip.dataset.wc;
      if (target.has(wc)) target.delete(wc); else target.add(wc);
      chip.classList.toggle('active');
      renderPredictor();
    });
  });
  wirePickerSearchInputs();
}

// Looks up exact-set data first (transitions.*_set*); if that combo has no history yet,
// falls back to aggregating the single-card data across every card in the selection (a
// looser but still useful approximation) and flags the result as approximate.
function lookupSetOrFallback(setDict, setDictByPlayer, singleDict, singleDictByPlayer, selectedCards, scope) {
  if (!selectedCards.length) return { results: [], usedScope: '', approx: false };
  const setKey = selectedCards.slice().sort().join(SET_SEP);
  if (scope !== '__all__' && setDictByPlayer[scope] && setDictByPlayer[scope][setKey] && setDictByPlayer[scope][setKey].length) {
    return { results: setDictByPlayer[scope][setKey], usedScope: scope + "'s own duels", approx: false };
  }
  if (setDict[setKey] && setDict[setKey].length) {
    return { results: setDict[setKey], usedScope: 'roster-wide', approx: false };
  }
  // Fallback: aggregate each selected card's individual single-card data.
  const agg = {};
  let usedPlayerData = false;
  selectedCards.forEach(card => {
    let entries;
    if (scope !== '__all__' && singleDictByPlayer[scope] && singleDictByPlayer[scope][card] && singleDictByPlayer[scope][card].length) {
      entries = singleDictByPlayer[scope][card];
      usedPlayerData = true;
    } else {
      entries = singleDict[card] || [];
    }
    entries.forEach(([name, count]) => { agg[name] = (agg[name] || 0) + count; });
  });
  const results = Object.entries(agg).sort((a, b) => b[1] - a[1]).slice(0, 8);
  const usedScope = usedPlayerData
    ? `approximate -- aggregated from ${scope}'s own duels per win condition individually (no exact combo history yet)`
    : `approximate -- aggregated roster-wide per win condition individually (no exact combo history yet)`;
  return { results, usedScope, approx: true };
}

function renderPredictor() {
  const scope = predictorScopeEl.value;
  const g1 = Array.from(g1Selected);

  let lookup;
  if (predictorMode === 'wincon') {
    if (!g1.length) { predictorResultsEl.innerHTML = '<div class="empty-hint">Check at least one win condition above.</div>'; return; }
    lookup = lookupSetOrFallback(transitions.wincon_set, transitions.wincon_set_by_player, transitions.wincon, transitions.wincon_by_player, g1, scope);
  } else {
    const g2 = Array.from(g2Selected);
    if (!g1.length || !g2.length) { predictorResultsEl.innerHTML = '<div class="empty-hint">Check at least one win condition for both Game 1 and Game 2.</div>'; return; }
    const setKey = g1.slice().sort().join(SET_SEP) + TRIPLE_SEP + g2.slice().sort().join(SET_SEP);
    if (scope !== '__all__' && transitions.wincon3_set_by_player[scope] && transitions.wincon3_set_by_player[scope][setKey] && transitions.wincon3_set_by_player[scope][setKey].length) {
      lookup = { results: transitions.wincon3_set_by_player[scope][setKey], usedScope: scope + "'s own duels", approx: false };
    } else if (transitions.wincon3_set[setKey] && transitions.wincon3_set[setKey].length) {
      lookup = { results: transitions.wincon3_set[setKey], usedScope: 'roster-wide', approx: false };
    } else {
      // Fallback: aggregate every (g1 card, g2 card) single-pair combo.
      const agg = {};
      let usedPlayerData = false;
      g1.forEach(w1 => {
        g2.forEach(w2 => {
          const tripleKey = w1 + TRIPLE_SEP + w2;
          let entries;
          if (scope !== '__all__' && transitions.wincon3_by_player[scope] && transitions.wincon3_by_player[scope][tripleKey] && transitions.wincon3_by_player[scope][tripleKey].length) {
            entries = transitions.wincon3_by_player[scope][tripleKey];
            usedPlayerData = true;
          } else {
            entries = transitions.wincon3[tripleKey] || [];
          }
          entries.forEach(([name, count]) => { agg[name] = (agg[name] || 0) + count; });
        });
      });
      const results = Object.entries(agg).sort((a, b) => b[1] - a[1]).slice(0, 8);
      const usedScope = usedPlayerData
        ? `approximate -- aggregated from ${scope}'s own duels per G1+G2 card pair individually (no exact combo history yet)`
        : `approximate -- aggregated roster-wide per G1+G2 card pair individually (no exact combo history yet)`;
      lookup = { results, usedScope, approx: true };
    }
  }

  const { results, usedScope } = lookup;
  if (!results.length) {
    const emptyMsg = predictorMode === 'wincon3'
      ? 'No game 3 data found for this G1 + G2 combo yet -- needs duels that reached at least 3 tracked games with this opening sequence.'
      : 'No later-game data found for this combo yet.';
    predictorResultsEl.innerHTML = `<div class="empty-hint">${emptyMsg}</div>`;
    return;
  }
  const maxCount = Math.max(...results.map(r => r[1]));
  const total = results.reduce((s, r) => s + r[1], 0);
  const scopeNote = `<div style="font-size:12px; color:var(--text-muted); margin-bottom:10px;">Scope: ${usedScope}</div>`;
  predictorResultsEl.innerHTML = scopeNote + results.map((r, i) => {
    const [name, count] = r;
    const pct = maxCount ? (count / maxCount * 100) : 0;
    const shareOfTotal = total ? Math.round(count / total * 100) : 0;
    const icons = cardIconStrip([name], 28);
    return `<div class="predictor-row">
      <span class="predictor-rank">#${i+1}</span>
      ${icons}
      <div style="flex:1; min-width:120px;">
        <strong style="font-size:13px;">${name}</strong>
        <div class="predictor-bar-track"><div class="predictor-bar-fill" style="width:${pct}%;"></div></div>
      </div>
      <span class="predictor-count">${count}x (${shareOfTotal}%)</span>
    </div>`;
  }).join('');
}

document.querySelectorAll('.predictor-tab').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.predictor-tab').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    predictorMode = btn.dataset.mode;
    renderPickers();
    renderPredictor();
  });
});
predictorScopeEl.addEventListener('change', () => {
  // keep the search box in sync when the plain dropdown is used directly
  predictorScopeSearchEl.value = predictorScopeEl.value === '__all__' ? '' : predictorScopeEl.value;
  renderPredictor();
});

// ---------- Predictor scope search (searchable alternative to the plain dropdown) ----------
const predictorScopeSearchEl = document.getElementById('predictorScopeSearch');
const predictorScopeSuggestionsEl = document.getElementById('predictorScopeSuggestions');
let predictorScopeActiveIdx = -1;

function selectPredictorScope(name) {
  predictorScopeEl.value = name;
  predictorScopeSearchEl.value = name;
  predictorScopeSuggestionsEl.classList.remove('show');
  renderPredictor();
}

function renderPredictorScopeSuggestions(query) {
  const q = query.trim().toLowerCase();
  if (!q) { predictorScopeSuggestionsEl.classList.remove('show'); predictorScopeSuggestionsEl.innerHTML = ''; return; }
  const matches = players.filter(p => p['Player'].toLowerCase().includes(q)).slice(0, 8);
  if (!matches.length) { predictorScopeSuggestionsEl.classList.remove('show'); predictorScopeSuggestionsEl.innerHTML = ''; return; }
  predictorScopeSuggestionsEl.innerHTML = '';
  matches.forEach(p => {
    const btn = document.createElement('button');
    btn.textContent = p['Player'] + (p['_is_scouted'] ? ' (scouted)' : p['_is_extended'] ? ' (extended roster)' : p['_is_shadow'] ? ' (opponent)' : '') + '  \\u00b7  ' + p['Total Games'] + ' games';
    btn.onmousedown = (e) => { e.preventDefault(); selectPredictorScope(p['Player']); };
    predictorScopeSuggestionsEl.appendChild(btn);
  });
  predictorScopeSuggestionsEl.classList.add('show');
  predictorScopeActiveIdx = -1;
}

predictorScopeSearchEl.addEventListener('input', () => {
  if (!predictorScopeSearchEl.value.trim()) {
    // cleared search -- reset scope back to "all"
    predictorScopeEl.value = '__all__';
    renderPredictor();
  }
  renderPredictorScopeSuggestions(predictorScopeSearchEl.value);
});
predictorScopeSearchEl.addEventListener('focus', () => renderPredictorScopeSuggestions(predictorScopeSearchEl.value));
predictorScopeSearchEl.addEventListener('keydown', (e) => {
  const items = Array.from(predictorScopeSuggestionsEl.querySelectorAll('button'));
  if (e.key === 'ArrowDown') { e.preventDefault(); predictorScopeActiveIdx = Math.min(predictorScopeActiveIdx + 1, items.length - 1); }
  else if (e.key === 'ArrowUp') { e.preventDefault(); predictorScopeActiveIdx = Math.max(predictorScopeActiveIdx - 1, 0); }
  else if (e.key === 'Enter') {
    e.preventDefault();
    if (predictorScopeActiveIdx >= 0 && items[predictorScopeActiveIdx]) { items[predictorScopeActiveIdx].dispatchEvent(new Event('mousedown')); }
    else {
      const q = predictorScopeSearchEl.value.trim().toLowerCase();
      const match = players.find(p => p['Player'].toLowerCase() === q) ||
                    players.find(p => p['Player'].toLowerCase().includes(q));
      if (match) selectPredictorScope(match['Player']);
    }
    return;
  } else { return; }
  items.forEach((it, i) => it.classList.toggle('active', i === predictorScopeActiveIdx));
});
document.addEventListener('click', (e) => {
  if (!e.target.closest('.predictor-scope-search')) predictorScopeSuggestionsEl.classList.remove('show');
});

renderPickers();
renderPredictor();

// ---------- Fuzzy name matching (shared: card pool + win-con pool) ----------
// Handles, in order: exact name, alias-expanded multi-word match (order independent),
// acronym-of-initials match, prefix match, plain substring match (e.g. "dart" ->
// "Dart Goblin"), raw multi-word substring match, and finally edit-distance typo
// tolerance -- each tier only fires if it resolves to exactly ONE unambiguous card, so a
// vague query never silently picks the wrong thing. This is a starting alias list, not
// exhaustive -- easy to extend if a common shorthand is missing.
const WORD_ALIASES = {
  'gob': 'goblin', 'gobs': 'goblin', 'goblins': 'goblins',
  'skelly': 'skeleton', 'skel': 'skeleton', 'skellies': 'skeleton', 'skele': 'skeleton',
  'barb': 'barbarian', 'barbs': 'barbarian',
  'musk': 'musketeer', 'musk.': 'musketeer',
  'wb': 'wall breakers',
  'fb': 'fireball',
  'is': 'ice spirit',
  'ig': 'ice golem',
  'gy': 'graveyard',
  'eq': 'earthquake',
  'bt': 'bomb tower',
  'aq': 'archer queen',
  'mk': 'mega knight',
  'gk': 'golden knight',
  'sk': 'skeleton king',
  'lj': 'lumberjack',
  'mm': 'mighty miner',
  'pekka': 'p e k k a',
  'log': 'the log',
  'xbow': 'x-bow',
  'rg': 'royal giant',
  'rh': 'royal hogs',
  'gb': 'goblin barrel',
  'sb': 'skeleton barrel',
  'br': 'battle ram',
  'mp': 'mini pekka',
};

// Standalone shorthand for a specific card (checked against the WHOLE query, not
// token-by-token) -- covers cases where the generic singular alias above would be
// ambiguous on its own (e.g. "skelly" alone should mean the "Skeletons" swarm card, but
// WORD_ALIASES maps it to the singular "skeleton" so compound queries like "skelly
// barrel" still resolve to "Skeleton Barrel").
const PHRASE_ALIASES = {
  'skelly': 'skeletons', 'skellies': 'skeletons', 'skel': 'skeletons',
  'barb': 'barbarians', 'barbs': 'barbarians',
  'gob': 'goblins', 'gobs': 'goblins',
};

function normalizeForMatch(s) {
  return s.toLowerCase().replace(/[.\-']/g, '').replace(/\s+/g, ' ').trim();
}

function expandAliasTokens(raw) {
  const tokens = normalizeForMatch(raw).split(' ').filter(Boolean);
  return tokens.map(t => WORD_ALIASES[t] || t).join(' ').split(' ').filter(Boolean);
}

function levenshtein(a, b) {
  const m = a.length, n = b.length;
  const dp = [];
  for (let i = 0; i <= m; i++) { dp.push(new Array(n + 1).fill(0)); dp[i][0] = i; }
  for (let j = 0; j <= n; j++) dp[0][j] = j;
  for (let i = 1; i <= m; i++) {
    for (let j = 1; j <= n; j++) {
      dp[i][j] = a[i - 1] === b[j - 1] ? dp[i - 1][j - 1] : 1 + Math.min(dp[i - 1][j - 1], dp[i - 1][j], dp[i][j - 1]);
    }
  }
  return dp[m][n];
}

function fuzzyMatchName(raw, pool) {
  const qRaw = (raw || '').trim();
  if (!qRaw || !pool.length) return null;
  const qNorm = normalizeForMatch(qRaw);
  const qCompact = qNorm.replace(/\s+/g, '');

  const exact = pool.find(n => normalizeForMatch(n) === qNorm);
  if (exact) return exact;

  if (PHRASE_ALIASES[qNorm]) {
    const phraseTarget = PHRASE_ALIASES[qNorm];
    const phraseMatch = pool.find(n => normalizeForMatch(n) === phraseTarget);
    if (phraseMatch) return phraseMatch;
  }

  const expandedTokens = expandAliasTokens(qRaw);
  if (expandedTokens.length) {
    const allTokMatches = pool.filter(n => {
      const nNorm = normalizeForMatch(n);
      return expandedTokens.every(t => nNorm.includes(t));
    });
    if (allTokMatches.length === 1) return allTokMatches[0];
  }

  const acronymMatches = pool.filter(n => {
    const initials = n.split(/\s+/).map(w => (w.replace(/[^A-Za-z]/g, '')[0] || '')).join('').toLowerCase();
    return initials && initials === qCompact;
  });
  if (acronymMatches.length === 1) return acronymMatches[0];

  const starts = pool.filter(n => normalizeForMatch(n).startsWith(qNorm));
  if (starts.length === 1) return starts[0];

  const contains = pool.filter(n => normalizeForMatch(n).includes(qNorm));
  if (contains.length === 1) return contains[0];

  const rawTokens = qNorm.split(' ').filter(Boolean);
  if (rawTokens.length > 1) {
    const tokMatches = pool.filter(n => {
      const nNorm = normalizeForMatch(n);
      return rawTokens.every(t => nNorm.includes(t));
    });
    if (tokMatches.length === 1) return tokMatches[0];
  }

  if (qNorm.length >= 3) {
    const scored = pool.map(n => ({ n, d: levenshtein(qNorm, normalizeForMatch(n)) })).sort((a, b) => a.d - b.d);
    const maxDist = qNorm.length <= 5 ? 1 : 2;
    if (scored[0] && scored[0].d <= maxDist && (!scored[1] || scored[1].d > scored[0].d)) {
      return scored[0].n;
    }
  }

  return null;
}

// ---------- Deck Predictor (pick 8 cards -> likely full decks for the next game) ----------
// Two modes sharing the same "slot" machinery: deck2 predicts game 2 from one 8-card
// deck (game 1); deck3 predicts game 3 from TWO 8-card decks (game 1 AND game 2).
const MAX_DECK_CARDS = 8;
let deckMode = 'deck2';
const deckSlotG1El = document.getElementById('deckSlot_g1');
const deckSlotG2El = document.getElementById('deckSlot_g2');
const deckResultsEl = document.getElementById('deckPredictorResults');
const deckPredictorSubEl = document.getElementById('deckPredictorSub');
const deckPredictorDqNoteEl = document.getElementById('deckPredictorDqNote');
if (deckPredictorDqNoteEl) {
  const excludedDeck = transitions.excluded_uncertain_duels || 0;
  deckPredictorDqNoteEl.textContent = `Data quality: ${excludedDeck} "uncertain-start" duel(s) are excluded from these predictions for the same reason (unverifiable game order). See the workbook's Data Quality sheet for details.`;
}

const DECK_PREDICTOR_SUB = {
  deck2: "Click the cards your opponent played in an earlier game (up to 8) from the full card pool below to rebuild their deck -- or type in the filter box and hit Enter to add the top match fast. Card matching is fuzzy: partial names, common abbreviations (wb, gob, skelly, barb...), acronyms (initials of each word), and minor typos are all recognized, not just exact spelling.",
  deck3: "Build BOTH the Game 1 deck and the Game 2 deck (same fast fuzzy entry works in each box), then see the closest-matching FULL decks for game 3. Needs duels with 3+ tracked games; an exact (deck1, deck2) match is rare given how many decks exist, so this usually falls back to a card-overlap search blended from both decks (Game 2 weighted slightly higher as the more recent signal).",
};

function deckTileIcon(name) {
  const url = cardIcons[name];
  return url ? `<img src="${url}" alt="">` : '';
}

function makeDeckSlot(suffix) {
  return {
    suffix,
    selected: [],
    poolEl: document.getElementById('deckPredictorPool_' + suffix),
    selectedEl: document.getElementById('deckPredictorSelected_' + suffix),
    countEl: document.getElementById('deckPredictorCount_' + suffix),
    filterEl: document.getElementById('deckPredictorFilter_' + suffix),
    clearBtn: document.getElementById('deckPredictorClear_' + suffix),
    bulkEl: document.getElementById('deckPredictorBulk_' + suffix),
    bulkAddBtn: document.getElementById('deckPredictorBulkAdd_' + suffix),
    bulkStatusEl: document.getElementById('deckPredictorBulkStatus_' + suffix),
  };
}

function renderSlotPool(slot) {
  const q = (slot.filterEl.value || '').toLowerCase();
  const list = allCards.filter(n => n.toLowerCase().includes(q));
  slot.poolEl.innerHTML = list.map(name => {
    const isSel = slot.selected.includes(name);
    const disabled = !isSel && slot.selected.length >= MAX_DECK_CARDS;
    return `<div class="deck-tile ${isSel ? 'selected' : ''} ${disabled ? 'disabled' : ''}" data-card="${name.replace(/"/g,'&quot;')}">${deckTileIcon(name)}<span>${name}</span></div>`;
  }).join('') || '<div class="empty-hint">No cards match that filter.</div>';
}

function renderSlotSelected(slot) {
  slot.countEl.textContent = `${slot.selected.length} / ${MAX_DECK_CARDS} selected`;
  if (!slot.selected.length) {
    slot.selectedEl.innerHTML = '<div class="empty-hint">Click cards below to build the deck you saw...</div>';
    return;
  }
  slot.selectedEl.innerHTML = slot.selected.map(name =>
    `<div class="deck-tile selected" data-card="${name.replace(/"/g,'&quot;')}">${deckTileIcon(name)}<span>${name}</span> &times;</div>`
  ).join('');
}

function toggleSlotCard(slot, name) {
  const idx = slot.selected.indexOf(name);
  if (idx >= 0) slot.selected.splice(idx, 1);
  else if (slot.selected.length < MAX_DECK_CARDS) slot.selected.push(name);
  renderSlotPool(slot);
  renderSlotSelected(slot);
  renderDeckPredictor();
}

// Adds a card by exact name if it isn't already selected and there's room. Returns true if added.
function addSlotCard(slot, name) {
  if (!name || slot.selected.includes(name) || slot.selected.length >= MAX_DECK_CARDS) return false;
  slot.selected.push(name);
  return true;
}

function wireSlot(slot) {
  slot.poolEl.addEventListener('click', e => {
    const tile = e.target.closest('.deck-tile');
    if (!tile || tile.classList.contains('disabled')) return;
    toggleSlotCard(slot, tile.dataset.card);
  });
  slot.selectedEl.addEventListener('click', e => {
    const tile = e.target.closest('.deck-tile');
    if (!tile) return;
    toggleSlotCard(slot, tile.dataset.card);
  });
  slot.filterEl.addEventListener('input', () => renderSlotPool(slot));
  slot.filterEl.addEventListener('keydown', e => {
    if (e.key !== 'Enter') return;
    e.preventDefault();
    const raw = slot.filterEl.value.trim();
    if (!raw) return;
    const available = allCards.filter(n => !slot.selected.includes(n));
    const match = fuzzyMatchName(raw, available);
    if (match && addSlotCard(slot, match)) {
      slot.filterEl.value = '';
      renderSlotPool(slot);
      renderSlotSelected(slot);
      renderDeckPredictor();
    }
  });
  slot.clearBtn.addEventListener('click', () => {
    slot.selected = [];
    renderSlotPool(slot);
    renderSlotSelected(slot);
    renderDeckPredictor();
  });

  function runSlotBulkAdd() {
    const raw = slot.bulkEl.value;
    if (!raw.trim()) { slot.bulkStatusEl.textContent = ''; slot.bulkStatusEl.className = 'deck-predictor-bulk-status'; return; }
    const parts = raw.split(/[,\\n]/).map(s => s.trim()).filter(Boolean);
    const matched = [];
    const unmatched = [];
    parts.forEach(part => {
      const available = allCards.filter(n => !matched.includes(n));
      const name = fuzzyMatchName(part, available);
      if (name) matched.push(name); else unmatched.push(part);
    });
    let added = 0;
    matched.forEach(name => { if (addSlotCard(slot, name)) added++; });
    renderSlotPool(slot);
    renderSlotSelected(slot);
    renderDeckPredictor();

    const bits = [];
    if (added) bits.push(`added ${added} card${added !== 1 ? 's' : ''}`);
    const skippedAlready = matched.length - added;
    if (skippedAlready > 0) bits.push(`${skippedAlready} already selected/deck full`);
    if (unmatched.length) bits.push(`couldn't match: ${unmatched.join(', ')}`);
    slot.bulkStatusEl.textContent = bits.join(' -- ') || 'Nothing to add.';
    slot.bulkStatusEl.className = 'deck-predictor-bulk-status ' + (unmatched.length ? 'warn' : 'ok');
    if (!unmatched.length) slot.bulkEl.value = '';
  }

  slot.bulkAddBtn.addEventListener('click', runSlotBulkAdd);
  slot.bulkEl.addEventListener('keydown', e => {
    if (e.key === 'Enter' && !e.shiftKey) {
      e.preventDefault();
      runSlotBulkAdd();
    }
  });
}

const deckSlotG1 = makeDeckSlot('g1');
const deckSlotG2 = makeDeckSlot('g2');
wireSlot(deckSlotG1);
wireSlot(deckSlotG2);

document.querySelectorAll('.predictor-tab[data-deckmode]').forEach(btn => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.predictor-tab[data-deckmode]').forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    deckMode = btn.dataset.deckmode;
    deckSlotG2El.style.display = deckMode === 'deck3' ? '' : 'none';
    deckPredictorSubEl.textContent = DECK_PREDICTOR_SUB[deckMode] || '';
    document.getElementById('deckSlotLabel_g1').textContent = deckMode === 'deck3' ? 'Game 1 deck' : 'Earlier-game deck';
    renderDeckPredictor();
  });
});

function overlapCandidates(selectedCards) {
  const selSet = new Set(selectedCards);
  const scoreByTo = {};
  Object.keys(transitions.deck).forEach(fromKey => {
    const fromCards = fromKey.split(', ');
    const overlap = fromCards.filter(c => selSet.has(c)).length;
    if (overlap >= 4) {  // at least half the deck shared
      transitions.deck[fromKey].forEach(([toKey, count]) => {
        const weighted = count * (overlap / MAX_DECK_CARDS);
        scoreByTo[toKey] = (scoreByTo[toKey] || 0) + weighted;
      });
    }
  });
  return scoreByTo;
}

function renderDeckResultsList(results, isExact, note) {
  if (!results.length) {
    deckResultsEl.innerHTML = '<div class="empty-hint">No historical decks found that resemble this one closely enough yet.</div>';
    return;
  }
  const maxCount = Math.max(...results.map(r => r[1]));
  const total = results.reduce((s, r) => s + r[1], 0);
  const noteHtml = note ? `<div style="font-size:12px; color:var(--text-muted); margin-bottom:10px;">${note}</div>` : '';
  deckResultsEl.innerHTML = noteHtml + results.map((r, i) => {
    const [name, count] = r;
    const pct = maxCount ? (count / maxCount * 100) : 0;
    const shareOfTotal = total ? Math.round(count / total * 100) : 0;
    const icons = cardIconStrip(name.split(', '), 22);
    return `<div class="predictor-row">
      <span class="predictor-rank">#${i+1}</span>
      ${icons}
      <div style="flex:1; min-width:120px;">
        <div class="row-label" style="margin-top:0;">${name}</div>
        <div class="predictor-bar-track"><div class="predictor-bar-fill" style="width:${pct}%;"></div></div>
      </div>
      <span class="predictor-count">${isExact ? count + 'x' : '~' + count} (${shareOfTotal}%)</span>
    </div>`;
  }).join('');
}

function renderDeckPredictor() {
  if (deckMode === 'deck2') {
    if (deckSlotG1.selected.length !== MAX_DECK_CARDS) {
      deckResultsEl.innerHTML = `<div class="empty-hint">Select all 8 cards to see predicted decks (${deckSlotG1.selected.length}/8 so far).</div>`;
      return;
    }
    const deckKey = deckSlotG1.selected.slice().sort().join(', ');
    const exact = transitions.deck[deckKey];
    if (exact && exact.length) {
      renderDeckResultsList(exact, true, 'Exact match -- this precise 8-card deck has been seen before in the tracked duel history.');
      return;
    }
    const scoreByTo = overlapCandidates(deckSlotG1.selected);
    const ranked = Object.entries(scoreByTo).sort((a, b) => b[1] - a[1]).slice(0, 8);
    const results = ranked.map(([name, score]) => [name, Math.round(score * 10) / 10]);
    renderDeckResultsList(results, false, results.length
      ? 'No exact match for this deck yet -- these are the closest historical decks (sharing at least half the cards), weighted by card overlap. Treat as approximate.'
      : '');
    return;
  }

  // deck3 mode: needs both slots full
  if (deckSlotG1.selected.length !== MAX_DECK_CARDS || deckSlotG2.selected.length !== MAX_DECK_CARDS) {
    deckResultsEl.innerHTML = `<div class="empty-hint">Select all 8 cards for BOTH the Game 1 deck (${deckSlotG1.selected.length}/8) and the Game 2 deck (${deckSlotG2.selected.length}/8).</div>`;
    return;
  }
  const deck1Key = deckSlotG1.selected.slice().sort().join(', ');
  const deck2Key = deckSlotG2.selected.slice().sort().join(', ');
  const tripleSep = transitions.triple_sep || '|||';
  const tripleKey = `${deck1Key}${tripleSep}${deck2Key}`;
  const exact3 = (transitions.deck3 || {})[tripleKey];
  if (exact3 && exact3.length) {
    renderDeckResultsList(exact3, true, 'Exact match -- this precise Game 1 + Game 2 deck pair has been seen before in the tracked duel history.');
    return;
  }
  // Fallback: blend overlap-weighted candidates from BOTH selected decks -- Game 2 is
  // weighted slightly higher since it's the more recent, more predictive signal.
  const scores1 = overlapCandidates(deckSlotG1.selected);
  const scores2 = overlapCandidates(deckSlotG2.selected);
  const combined = {};
  Object.entries(scores1).forEach(([k, v]) => { combined[k] = (combined[k] || 0) + v * 0.75; });
  Object.entries(scores2).forEach(([k, v]) => { combined[k] = (combined[k] || 0) + v * 1.0; });
  const ranked = Object.entries(combined).sort((a, b) => b[1] - a[1]).slice(0, 8);
  const results = ranked.map(([name, score]) => [name, Math.round(score * 10) / 10]);
  renderDeckResultsList(results, false, results.length
    ? 'No exact match for this Game 1 + Game 2 combo yet -- these are the closest historical decks (blended card-overlap signal from both selected decks, Game 2 weighted slightly higher). Treat as approximate.'
    : '');
}

renderSlotPool(deckSlotG1);
renderSlotSelected(deckSlotG1);
renderSlotPool(deckSlotG2);
renderSlotSelected(deckSlotG2);
renderDeckPredictor();
</script>
</body>
</html>
"""

html = html.replace("__DATA_JSON__", data_json)

with open("crl_opponent_scout.html", "w", encoding="utf-8") as f:
    f.write(html)

print("wrote", len(html), "bytes")
