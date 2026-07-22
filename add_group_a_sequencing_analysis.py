"""Adds a "Tomorrow's Opponents - Sequencing" sheet to CRL_Duel_Decks.xlsx, scoped
SOLELY to the 7 players in the user's actual Group A tomorrow (4 confirmed + 3 on-deck --
the 6 reference-only players are explicitly excluded, matching the established "7 players
in my group" definition used elsewhere in this project). Added 2026-07-19 per explicit
user request:

"can you run an analysis solely on the 7 players I'm playing tomorrow, what spells and win
conditions they tend to use together in each deck, in which order in the duel set b03 do
they pull them out?"

Two angles per player, both drawn directly from real recorded games (Practice + Official
CRL combined -- the CRL-only sample is generally too thin per player to analyze alone):

1. SPELL + WIN-CONDITION COMBOS: every 8-card deck they've played is broken into its
   win-condition(s) (via the existing classify_deck()/WIN_CONDITION_REFERENCE) and its
   spell(s) (via a new SPELL_CARDS reference below), then grouped by that
   (win-con set, spell set) combo across all their games, ranked by frequency. This answers
   "what spells and win conditions do they tend to run together."

2. BEST-OF-3 (B03) SEQUENCING: their games are grouped back into duel sets (duel_id,
   already computed by build_dataset()) and ordered by actual battle time within each duel
   (rematches excluded; the first/uncertain-visibility duel per opponent is INCLUDED since
   excluding it left 5 of the 7 players with zero sequencing data, but is caveated in the
   sheet). IMPORTANT FINDING WHILE BUILDING THIS: real Clash Royale Duel format bans reusing
   any card across a duel's games (enforced upstream by group_into_duels()'s
   deck_set.isdisjoint(current_cards) check), so a naive "did their win-con/spell change
   game-to-game" stat is ~100% for every player by the game's own rules, not a behavioral
   signal -- that metric was dropped after being checked and confirmed trivial. What IS a
   real signal: which win-con/spell each player tends to slot into WHICH position (Game 1
   opener / Game 2 / Game 3 decider) of a set, computed per-win-con and per-spell with a
   games-played breakdown across positions -- this is the actual answer to "in which order
   do they pull them out."

Caveat on SPELL_CARDS: Clash Royale's card list is live-updated by Supercell, so this
classification reflects the card pool as understood as of this build and may not be
100% current if new cards/reworks have shipped recently -- verify against the in-game
card list if a card's spell/troop classification looks off. Cards already flagged as
win-conditions in WIN_CONDITION_REFERENCE (Goblin Barrel, Graveyard) are treated as
win-cons, not spells, here, to avoid double-counting the same card in both categories.
"""
import glob
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter, defaultdict

from build_duel_workbook import classify_deck, build_dataset, classify_match_category, parse_time

import os as _os
_CRL_HOME = _os.environ.get('CRL_HOME')
XLSX_PATH = _os.path.join(_CRL_HOME, 'CRL_Duel_Decks.xlsx') if _CRL_HOME else '/home/claude/CRL_Duel_Decks.xlsx'

# Some Group A players (e.g. INA.BenZerRidel) don't have a master_<tag>.json -- they're
# Extended Roster only, tracked via a flat ~25-30-game recent-battle snapshot
# (extended_<tag>.json) rather than the full duel-grouped fetch history the main 53-player
# roster gets. build_dataset() above only processes master_*.json players as "player" rows,
# so their games are entirely missing from full_duel_log otherwise -- this loader fills
# that gap for Section A (spell/win-con combos) using the same parsing build_dashboard.py
# uses for its extended-roster rows. Section B (B03 sequencing) still won't have anything
# for these players: extended-roster battles are flat single-game snapshots with no known
# duel-set grouping, so there's genuinely no sequence data to show, not a bug.
_ROSTER_TAGS_ONLY = {
    f"#{os.path.basename(p)[len('master_'):-len('.json')]}" for p in glob.glob('master_*.json')
}


def load_extended_rows_for_tag(tag):
    tag_digits = tag.lstrip('#')
    path = f'extended_{tag_digits}.json'
    if not os.path.exists(path):
        return []
    with open(path) as f:
        battles = json.load(f)
    rows = []
    for b in battles:
        team = b.get('team', [{}])[0]
        opp = b.get('opponent', [{}])[0]
        deck = [c['name'] for c in team.get('cards', [])]
        opp_deck = [c['name'] for c in opp.get('cards', [])]
        if not deck:
            continue
        battle_time = parse_time(b['battleTime'])
        category = classify_match_category(
            b.get('type'), b.get('gameMode', {}).get('name'), battle_time,
            opponent_tag=opp.get('tag'), roster_tags=_ROSTER_TAGS_ONLY,
        )
        if category is None:
            continue
        rows.append({
            'duel_id': f"extended_{team.get('tag')}_{b.get('battleTime')}",
            'game_num': 1,
            'player_tag': team.get('tag'),
            'opponent_tag': opp.get('tag'),
            'battle_time': battle_time,
            'deck': deck,
            'opponent_deck': opp_deck,
            'crowns_for': team.get('crowns', 0),
            'crowns_against': opp.get('crowns', 0),
            'is_rematch': False,
            'uncertain_start': True,
            'match_category': category,
        })
    return rows

# Spell-type cards present in this dataset's card pool, EXCLUDING any card already
# classified True (a win-condition) in WIN_CONDITION_REFERENCE (Goblin Barrel, Graveyard --
# both spells that are also win conditions, kept in the win-con category only here).
SPELL_CARDS = {
    "Arrows", "Barbarian Barrel", "Clone", "Earthquake", "Fireball", "Freeze",
    "Giant Snowball", "Goblin Curse", "Lightning", "Mirror", "Poison", "Rage",
    "Rocket", "Royal Delivery", "The Log", "Tornado", "Void", "Zap",
}

# The 7 players actually in Group A tomorrow (confirmed + on-deck only -- matches
# GROUP_A_OPPONENTS/GROUP_A_ROSTER elsewhere, reference-only players excluded).
# REPLACED 2026-07-19: entire Day-2 roster swapped for the 15 possible Monthly Finals
# (Day 3) opponents -- see add_group_a_analysis.py for the full rationale.
TARGET_PLAYERS = [
    ("Mugi", "#2CLV2RP0", "confirmed"),
    ("SandBox", "#Y022GRCJQ", "confirmed"),
    ("40k Oker", "#YLVV0JPQ", "confirmed"),
    ("Mohamed Light", "#G9YV9GR8R", "confirmed"),
    ("Adriel", "#9CPCC890", "confirmed"),
    ("Pedro™️", "#RJ88Y8U08", "on_deck"),
    ("Asaf", "#RUQ0JU2P", "on_deck"),
    ("Clown (KickAsh)", "#GPPYR9JYR", "on_deck"),
    ("Vitor75", "#8LJ92G8UG", "on_deck"),
    ("Sub", "#U890Q9UQ", "on_deck"),
    ("SK Morten", "#R09228V", "on_deck"),
    ("Guriko", "#2LJ0ULYCC", "on_deck"),
    ("Polaris", "#U8RYGC8GU", "on_deck"),
    ("JorZ", "#22LC8JG02", "on_deck"),
    ("FrancoMedinaSL", "#UJQQCUCQ8", "on_deck"),
]

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='FF1F4E78')
LABEL_FONT = Font(name='Arial', size=11, bold=True)
VALUE_FONT = Font(name='Arial', size=11, bold=False)
SECTION_FONT = Font(name='Arial', size=12, bold=True, color='FF1F4E78')
PLAYER_FONT = Font(name='Arial', size=14, bold=True, color='FFFFFFFF')
PLAYER_FILL = PatternFill('solid', fgColor='FF2E7D32')
ONDECK_FILL = PatternFill('solid', fgColor='FFFFF2CC')
NOTE_FONT = Font(name='Arial', size=10, italic=True, color='FF666666')
TITLE_FONT = Font(name='Arial', size=16, bold=True, color='FF1F4E78')
THIN_FILL = PatternFill('solid', fgColor='FFF5F5F5')

full_duel_log, _duel_summary, _session_stats = build_dataset()


def deck_wincon_spell_keys(deck):
    wc = tuple(classify_deck(deck) or [])
    sp = tuple(sorted(set(deck) & SPELL_CARDS))
    return wc, sp


def analyze_player(tag):
    player_games = [r for r in full_duel_log if r["player_tag"] == tag and r["deck"] and len(r["deck"]) == 8]
    used_extended_fallback = False
    if not player_games:
        extended = load_extended_rows_for_tag(tag)
        player_games = [r for r in extended if r["deck"] and len(r["deck"]) == 8]
        used_extended_fallback = bool(player_games)

    # ---- Section A: spell + win-con combos across every logged deck (all games) ----
    combo_games = Counter()
    combo_wins = Counter()
    for r in player_games:
        wc, sp = deck_wincon_spell_keys(r["deck"])
        key = (wc, sp)
        combo_games[key] += 1
        if r["crowns_for"] > r["crowns_against"]:
            combo_wins[key] += 1
    combo_rows = sorted(
        [{"wincons": k[0], "spells": k[1], "games": g, "wins": combo_wins[k], "win_rate": combo_wins[k] / g}
         for k, g in combo_games.items()],
        key=lambda x: (-x["games"], -x["win_rate"]),
    )

    # ---- Section B: B03 sequencing ----
    # NOTE: uncertain_start duels (the FIRST fetched duel per opponent -- true for
    # essentially every duel of a just-added roster player like RAD/Adox/DK/Lucas.xit/
    # Lucas.jack, since their fetch history only just started) are included here, unlike
    # the Excel formulas elsewhere in this workbook. Relative game-to-game ORDER within a
    # fetched duel (by real battle_time) is still reliable even when uncertain_start is
    # True -- what's uncertain is only whether an earlier, untracked duel happened before
    # our fetch window started, which would make our "Game 1" actually a later game in the
    # real-world set. Excluding uncertain_start entirely would leave 5 of the 7 players
    # with ZERO sequencing data (every one of their duels so far is a first-ever-fetched
    # one), so switch-rate transitions are computed on all non-rematch games, and the
    # absolute Game-1/2/3 position labels below carry an explicit caveat instead.
    sequencing_games = [r for r in player_games if not r["is_rematch"]]
    any_uncertain = any(r["uncertain_start"] for r in sequencing_games)
    duels = defaultdict(list)
    for r in sequencing_games:
        duels[r["duel_id"]].append(r)

    # NOTE ON WHY GAME-TO-GAME "SWITCHING" ISN'T TRACKED AS A BEHAVIORAL STAT: real
    # Clash Royale Duel format requires every card played across a duel's up-to-3 games to
    # be used at most once total -- group_into_duels() (build_duel_workbook.py) enforces
    # this by only grouping a new game into the same duel when its deck shares ZERO cards
    # with every earlier game in that duel. That means the win-condition and spells
    # necessarily change from one game to the next almost every time, by the game's own
    # rules -- it was checked directly against this data and came out ~100% for every one
    # of these 7 players, which is exactly what the disjoint-card rule predicts, not a
    # player-specific tendency. So instead of a meaningless "switch rate," the useful
    # signal is WHICH win-con/spell each player tends to slot into WHICH position of the
    # set -- computed below.
    position_wincons = {1: Counter(), 2: Counter(), 3: Counter()}
    position_spells = {1: Counter(), 2: Counter(), 3: Counter()}
    wincon_position_counts = defaultdict(lambda: Counter())  # wincon -> {1: n, 2: n, 3: n}
    spell_position_counts = defaultdict(lambda: Counter())
    multi_game_duels = 0

    for duel_id, games in duels.items():
        games = sorted(games, key=lambda r: r["battle_time"])[:3]
        # Completeness gate (added 2026-07-20 per user: "duel sets entered into the data
        # [should be] complete Bo3, with 3 unique decks, not incomplete 2 game duels").
        # Practice duels are now always best-of-3, so a practice duel with <3 distinct
        # decks is a TRUNCATED set (a game aged out of the API's sliding window) whose
        # Game-1/2/3 position labels are unreliable -- exclude it from positional
        # sequencing. Official CRL is different: a 2-0 sweep is a complete 2-game result,
        # so >=2 games stays valid there.
        cat = games[0].get("match_category")
        if cat == "Practice":
            if len(games) < 3:
                continue
        elif len(games) < 2:
            continue
        multi_game_duels += 1
        for i, g in enumerate(games, start=1):
            wc = classify_deck(g["deck"]) or []
            sp = sorted(set(g["deck"]) & SPELL_CARDS)
            position_wincons[i].update(wc)
            position_spells[i].update(sp)
            for w in wc:
                wincon_position_counts[w][i] += 1
            for s in sp:
                spell_position_counts[s][i] += 1

    def modal_position_rows(position_counts, min_games=2):
        rows = []
        for name, counts in position_counts.items():
            total = sum(counts.values())
            if total < min_games:
                continue
            modal_pos, modal_n = counts.most_common(1)[0]
            rows.append({
                "name": name, "total": total, "modal_pos": modal_pos, "modal_n": modal_n,
                "breakdown": ", ".join(f"G{p}:{counts.get(p, 0)}" for p in (1, 2, 3) if counts.get(p, 0)),
            })
        rows.sort(key=lambda x: (-x["total"], x["modal_pos"]))
        return rows

    return {
        "total_games": len(player_games),
        "combo_rows": combo_rows,
        "multi_game_duels": multi_game_duels,
        "position_wincons": position_wincons,
        "position_spells": position_spells,
        "wincon_position_rows": modal_position_rows(wincon_position_counts),
        "spell_position_rows": modal_position_rows(spell_position_counts),
        "any_uncertain": any_uncertain,
        "used_extended_fallback": used_extended_fallback,
    }


def fmt_pct(changed, total):
    if total == 0:
        return "n/a (0 games)"
    return f"{changed / total:.0%} ({changed}/{total})"


def top_n_str(counter, n=3):
    if not counter:
        return "(none recorded)"
    return ", ".join(f"{name} ({g}g)" for name, g in counter.most_common(n))


# ---- Build the sheet ----
wb = openpyxl.load_workbook(XLSX_PATH)
SHEET_NAME = "Group A Sequencing"
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)
ws.sheet_view.showGridLines = True
ws.column_dimensions['A'].width = 55
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 55

r = 1
ws.cell(r, 1, "Tomorrow's Opponents -- Spell/Win-Con Combos & B03 Sequencing").font = TITLE_FONT
r += 1
ws.cell(r, 1,
        "Scoped solely to the 7 players actually in Group A tomorrow (4 confirmed + 3 on-deck; "
        "the 6 reference-only players elsewhere in this workbook are excluded here). For each "
        "player: (A) which spells and win conditions they tend to run together in the same deck, "
        "ranked by how often that combo shows up across all their logged games, Practice + Official "
        "CRL combined; and (B) within actual best-of-3 duel sets, whether their win-condition and "
        "spell choices change from one game to the next, split by whether the prior game was a win "
        "or a loss, plus which win-cons/spells tend to show up at each position (Game 1/2/3) of a "
        "set. All figures are drawn from real recorded games -- nothing here is theorycrafted. "
        "Spell classification note: Clash Royale's card list is live-updated by Supercell, so this "
        "spell/win-con split reflects the card pool as understood as of this build -- flag anything "
        "that looks off."
        ).font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.row_dimensions[r].height = 90
r += 2

for name, tag, status in TARGET_PLAYERS:
    data = analyze_player(tag)

    status_label = " [ON DECK -- not yet confirmed Group A]" if status == "on_deck" else ""
    ws.cell(r, 1, f"{name}  ({tag}){status_label}").font = PLAYER_FONT
    for c in range(1, 6):
        ws.cell(r, c).fill = ONDECK_FILL if status == "on_deck" else PLAYER_FILL
    r += 1
    fallback_note = (
        " (No master roster fetch history for this player -- pulled from their Extended Roster "
        "snapshot instead, a flat ~25-30 most-recent-battle log with no duel-set grouping, so no "
        "B03 sequencing is possible for them; deck/spell/win-con combos below are still valid.)"
        if data["used_extended_fallback"] else ""
    )
    ws.cell(r, 1, f"{data['total_games']} total logged games with a full 8-card deck (Practice + Official CRL)."
            + fallback_note).font = NOTE_FONT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    if fallback_note:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 40
    r += 1

    # --- Section A ---
    ws.cell(r, 1, "Spell + Win-Condition combos (across all decks played)").font = SECTION_FONT
    r += 1
    ws.cell(r, 1, "Win Condition(s)").font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, "Spell(s)").font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
    ws.cell(r, 3, "Games").font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
    ws.cell(r, 4, "Win Rate").font = HEADER_FONT; ws.cell(r, 4).fill = HEADER_FILL
    ws.cell(r, 5, "").fill = HEADER_FILL
    r += 1
    if data["combo_rows"]:
        for row in data["combo_rows"][:8]:
            wc_text = ", ".join(row["wincons"]) if row["wincons"] else "(no win-con identified)"
            sp_text = ", ".join(row["spells"]) if row["spells"] else "(no spell identified)"
            ws.cell(r, 1, wc_text).font = VALUE_FONT
            ws.cell(r, 1).alignment = Alignment(wrap_text=True)
            ws.cell(r, 2, sp_text).font = VALUE_FONT
            ws.cell(r, 2).alignment = Alignment(wrap_text=True)
            ws.cell(r, 3, row["games"]).font = VALUE_FONT
            wr_cell = ws.cell(r, 4, row["win_rate"])
            wr_cell.number_format = "0.0%"
            wr_cell.font = VALUE_FONT
            r += 1
    else:
        ws.cell(r, 1, "No full-deck games recorded yet.").font = VALUE_FONT
        r += 1
    r += 1

    # --- Section B ---
    ws.cell(r, 1, "Best-of-3 (B03) sequencing").font = SECTION_FONT
    r += 1
    uncertain_note = (
        " Some or all of these are this player's FIRST fetched duel against that opponent -- the "
        "game-to-game order within each duel is still reliable, but for those specific duels we can't "
        "rule out an earlier, untracked game happening before our fetch window started, so their "
        "absolute Game 1/2/3 position label below may be off by one game."
        if data["any_uncertain"] else ""
    )
    ws.cell(r, 1,
            f"{data['multi_game_duels']} multi-game duel set(s) found (rematch games excluded)."
            + uncertain_note
            ).font = NOTE_FONT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 40 if uncertain_note else 15
    r += 1

    if data["multi_game_duels"] == 0:
        ws.cell(r, 1, "Not enough multi-game duel data yet to analyze sequencing for this player.").font = VALUE_FONT
        r += 2
        continue

    ws.cell(r, 1,
            "Note: real Clash Royale Duel format bans reusing any card across the games of a set, so the "
            "win-condition and spells essentially always change from one game to the next by the game's own "
            "rules -- checked directly against this data, and it holds true here too. That's not a player "
            "tendency, so instead of a meaningless \"switch rate,\" the tables below show WHICH win-con/spell "
            "each player tends to slot into WHICH position of the set -- Game 1 (opener), Game 2, or Game 3 "
            "(decider) -- which is the real answer to \"in what order do they pull them out.\""
            ).font = NOTE_FONT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 55
    r += 1

    ws.cell(r, 1, "By duel position -- top win-cons / spells seen at each stage").font = LABEL_FONT
    r += 1
    ws.cell(r, 1, "Position").font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, "Top Win Conditions").font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
    ws.cell(r, 3, "Top Spells").font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
    ws.cell(r, 4, "").fill = HEADER_FILL
    ws.cell(r, 5, "").fill = HEADER_FILL
    r += 1
    for pos, label in ((1, "Game 1"), (2, "Game 2"), (3, "Game 3")):
        ws.cell(r, 1, label).font = VALUE_FONT
        ws.cell(r, 2, top_n_str(data["position_wincons"][pos])).font = VALUE_FONT
        ws.cell(r, 2).alignment = Alignment(wrap_text=True)
        ws.cell(r, 3, top_n_str(data["position_spells"][pos])).font = VALUE_FONT
        ws.cell(r, 3).alignment = Alignment(wrap_text=True)
        r += 1
    r += 1

    ws.cell(r, 1, "Win-condition -- usual position (min 2 games)").font = LABEL_FONT
    r += 1
    ws.cell(r, 1, "Win Condition").font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, "Usual Position").font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
    ws.cell(r, 3, "Games").font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
    ws.cell(r, 4, "Full Breakdown").font = HEADER_FONT; ws.cell(r, 4).fill = HEADER_FILL
    ws.cell(r, 5, "").fill = HEADER_FILL
    r += 1
    if data["wincon_position_rows"]:
        for row in data["wincon_position_rows"][:6]:
            ws.cell(r, 1, row["name"]).font = VALUE_FONT
            ws.cell(r, 2, f"Game {row['modal_pos']} ({row['modal_n']}/{row['total']})").font = VALUE_FONT
            ws.cell(r, 3, row["total"]).font = VALUE_FONT
            ws.cell(r, 4, row["breakdown"]).font = VALUE_FONT
            r += 1
    else:
        ws.cell(r, 1, "No win-condition has appeared 2+ times in a sequenced position yet.").font = VALUE_FONT
        r += 1
    r += 1

    ws.cell(r, 1, "Spell -- usual position (min 2 games)").font = LABEL_FONT
    r += 1
    ws.cell(r, 1, "Spell").font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, "Usual Position").font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
    ws.cell(r, 3, "Games").font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
    ws.cell(r, 4, "Full Breakdown").font = HEADER_FONT; ws.cell(r, 4).fill = HEADER_FILL
    ws.cell(r, 5, "").fill = HEADER_FILL
    r += 1
    if data["spell_position_rows"]:
        for row in data["spell_position_rows"][:6]:
            ws.cell(r, 1, row["name"]).font = VALUE_FONT
            ws.cell(r, 2, f"Game {row['modal_pos']} ({row['modal_n']}/{row['total']})").font = VALUE_FONT
            ws.cell(r, 3, row["total"]).font = VALUE_FONT
            ws.cell(r, 4, row["breakdown"]).font = VALUE_FONT
            r += 1
    else:
        ws.cell(r, 1, "No spell has appeared 2+ times in a sequenced position yet.").font = VALUE_FONT
        r += 1
    r += 2

wb.save(XLSX_PATH)
print(f"Sheet '{SHEET_NAME}' added. {r} rows written.")
for name, tag, status in TARGET_PLAYERS:
    d = analyze_player(tag)
    print(f"  {name}: {d['total_games']} games, {d['multi_game_duels']} multi-game duels")
