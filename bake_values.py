"""bake_values.py -- compute the Win-Con Sets and Deck Stats aggregate columns in Python
and write them as LITERAL values, replacing the COUNTIF/SUMIF formulas that previously
required a slow LibreOffice recalc pass.

WHY: build_dashboard.py reads 'Win-Con Sets' and 'Deck Stats' via openpyxl data_only=True,
which only returns numbers if the workbook's formulas have cached values. LibreOffice recalc
(recalc.py) baked those values but grew slow enough as the archive expanded that the whole
GitHub 'update' job blew past the 60-min ceiling and was killed. This replicates the exact
formula semantics in Python (validated cell-for-cell against a real recalced workbook:
Deck Stats 0/6141 mismatches; Win-Con Sets differences were only stale-cache rows the recalc
had left behind), so no LibreOffice is needed at all. Columns are located by HEADER NAME so
this keeps working if column positions shift.

Formulas replicated (from the workbook build):
  Duel Log:      Result = Win/Loss/Draw from Crowns For vs Crowns Against
                 'Stats Eligible' == 'Yes' gates Deck Stats counts
  Duel Summary:  Games Played  = # Duel Log rows for that Duel ID
                 Games Won     = # of those rows whose Result == 'Win'
  Win-Con Sets:  Times Played (Duels) = COUNTIF(Duel Summary Win-Con Set == this set)
                 Games Played/Won     = SUMIF over those duels' Games Played/Won
                 Win Rate             = Games Won / Games Played
  Deck Stats:    Games/Wins/Losses/Draws = COUNTIFS(Own Deck Key == this deck, AJ=Yes[, Result])
                 Win Rate                = Wins / Games
"""
import sys
from collections import defaultdict
import openpyxl


def hdr_index(ws):
    return {c.value: i for i, c in enumerate(ws[1]) if c.value is not None}


def hdr_col(ws, name):
    for c in ws[1]:
        if c.value == name:
            return c.column  # 1-based
    raise KeyError(f"header {name!r} not found in {ws.title!r}")


def main(path):
    wb = openpyxl.load_workbook(path, data_only=False)

    # ---- Duel Log: per-duel game/win counts; per-deck (Stats Eligible) records ----
    dl = wb["Duel Log"]
    H = hdr_index(dl)
    iA, iY, iZ, iAD, iAJ = (H["Duel ID"], H["Crowns For"], H["Crowns Against"],
                            H["Own Deck Key"], H["Stats Eligible"])
    duel_games = defaultdict(int)
    duel_wins = defaultdict(int)
    deck = defaultdict(lambda: [0, 0, 0, 0])  # games, wins, losses, draws
    for r in dl.iter_rows(min_row=2, values_only=True):
        did = r[iA]
        if did is None:
            continue
        y = r[iY] or 0
        z = r[iZ] or 0
        res = "Win" if y > z else ("Loss" if y < z else "Draw")
        duel_games[did] += 1
        if res == "Win":
            duel_wins[did] += 1
        if r[iAJ] == "Yes":
            dk = r[iAD]
            if dk:
                d = deck[dk]
                d[0] += 1
                d[1] += res == "Win"
                d[2] += res == "Loss"
                d[3] += res == "Draw"

    # ---- Duel Summary: bake Games Played/Won; accumulate per Win-Con Set ----
    ds = wb["Duel Summary"]
    jA = hdr_col(ds, "Duel ID")
    jM = hdr_col(ds, "Win-Con Set (Duel)")
    jG = hdr_col(ds, "Games Played")
    jH = hdr_col(ds, "Games Won")
    set_duels = defaultdict(int)
    set_games = defaultdict(int)
    set_wins = defaultdict(int)
    for rr in range(2, ds.max_row + 1):
        did = ds.cell(rr, jA).value
        if did is None:
            continue
        g = duel_games.get(did, 0)
        w = duel_wins.get(did, 0)
        ds.cell(rr, jG).value = g   # bake literal Games Played
        ds.cell(rr, jH).value = w   # bake literal Games Won
        m = ds.cell(rr, jM).value
        if m:
            set_duels[m] += 1
            set_games[m] += g
            set_wins[m] += w

    # ---- Win-Con Sets ----
    ws = wb["Win-Con Sets"]
    kKey = hdr_col(ws, "Win-Con Set")
    kC = hdr_col(ws, "Times Played (Duels)")
    kD = hdr_col(ws, "Games Played")
    kE = hdr_col(ws, "Games Won")
    kF = hdr_col(ws, "Win Rate")
    n = 0
    for rr in range(2, ws.max_row + 1):
        a = ws.cell(rr, kKey).value
        if not a:
            continue
        d = set_games.get(a, 0)
        e = set_wins.get(a, 0)
        ws.cell(rr, kC).value = set_duels.get(a, 0)
        ws.cell(rr, kD).value = d
        ws.cell(rr, kE).value = e
        ws.cell(rr, kF).value = (e / d) if d else ""
        n += 1
    wcs_rows = n

    # ---- Deck Stats ----
    ws = wb["Deck Stats"]
    kKey = hdr_col(ws, "Deck (sorted)")
    kC = hdr_col(ws, "Games Played")
    kD = hdr_col(ws, "Wins")
    kE = hdr_col(ws, "Losses")
    kF = hdr_col(ws, "Draws")
    kG = hdr_col(ws, "Win Rate")
    n = 0
    for rr in range(2, ws.max_row + 1):
        a = ws.cell(rr, kKey).value
        if not a:
            continue
        g, w, l, dr = deck.get(a, [0, 0, 0, 0])
        ws.cell(rr, kC).value = g
        ws.cell(rr, kD).value = w
        ws.cell(rr, kE).value = l
        ws.cell(rr, kF).value = dr
        ws.cell(rr, kG).value = (w / g) if g else ""
        n += 1
    ds_rows = n

    wb.save(path)
    print(f"baked Win-Con Sets ({wcs_rows} rows) + Deck Stats ({ds_rows} rows); no LibreOffice used")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "CRL_Duel_Decks.xlsx")
