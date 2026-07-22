"""Adds a "Group A Scouting" sheet to CRL_Duel_Decks.xlsx: for each opponent in Batan's
Day-2 group-stage group (snake-seeded, seeds 1/16/17/32/33/48/49/64), a data-driven
breakdown of every deck they've played, their most-played / best-win-rate / worst-win-rate
decks, their top win-condition sets, "decks that have beaten them" pulled directly from
their own recorded battle history (their own logged losses) -- NOT a theorycrafted
card-counter list, since we don't have a verified source for general counter matchups --
PLUS (added 2026-07-19, per explicit user request) a "Recommended for Tomorrow" section:
the best-performing deck / win-con set / duel-set pair FROM OUR ENTIRE TRACKED DATA POOL
(every roster/extended-roster player's Practice + Official CRL games combined), filtered
down to games where the opposing side played a deck matching this specific player's own
top win conditions. Also empirical, not theorycrafted -- every recommendation keeps its
games/win-rate so a thin sample stays visibly thin.
Added 2026-07-18, per explicit user request (user = 老板 Ι Batan'宙斯, seed 1, Group A).
Extended 2026-07-19: group may still change due to possible disqualifications, so 3 more
players the user asked to have ready regardless of final seeding -- Adox, Lucas✨杰克 (a
different Lucas from Lucas.xit✨之安神), and DK -- added to the same analysis, clearly
marked "on deck / not yet confirmed Group A" rather than folded in as if their group
membership were final.

Corrected 2026-07-19 (same day, user follow-up): the user's actual confirmed Group A is
INA.BenZerRidel, Lucas.xit✨之安神, RAD, and SandBox -- Wyze❤️Ultimo, fluffypotato99, and
LF丨张✨Ink❤️llb were NOT actually in the user's group and have been removed. JL Viiper is
explicitly NOT in the user's group either, but stays in this analysis marked "reference
only" per the user's explicit request ("keep viiper there since i want to see his decks
for my reference").

Data sources: master_<tag>.json for anyone on the main roster (JL Viiper, Adox, RAD, DK,
SandBox, Lucas✨杰克 -- all on full roster), extended_<tag>.json as fallback for
INA.BenZerRidel (still Extended-Roster-only, no master_ file yet)."""
import glob
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

from build_duel_workbook import (
    classify_deck, build_dataset, compute_counter_recommendations,
)

import os as _os
_CRL_HOME = _os.environ.get('CRL_HOME')
XLSX_PATH = _os.path.join(_CRL_HOME, 'CRL_Duel_Decks.xlsx') if _CRL_HOME else '/home/claude/CRL_Duel_Decks.xlsx'
CRL_DIR = _CRL_HOME if _CRL_HOME else '/mnt/user-data/uploads/CRL'

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='FF1F4E78')
NAME_FONT = Font(name='Arial', size=13, bold=True, color='FF1F4E78')
LABEL_FONT = Font(name='Arial', size=11, bold=True)
VALUE_FONT = Font(name='Arial', size=11, bold=False)
NOTE_FONT = Font(name='Arial', size=10, italic=True, color='FF666666')
RECO_FONT = Font(name='Arial', size=11, bold=False, color='FF1F6E3E')
PENDING_FILL = PatternFill('solid', fgColor='FFFCE8E6')
ONDECK_FILL = PatternFill('solid', fgColor='FFFFF7E0')

# (Display name, source filename under CRL_DIR (preferring master_ over extended_ if a
# player has both), tag, status)
# status is one of:
#   "confirmed" -- an actual Group A opponent per the user's own confirmed roster
#   "on_deck"   -- NOT yet confirmed Group A, but scouted ahead of time in case
#                  disqualifications reshuffle the group before Day 2
#   "reference" -- explicitly NOT in the user's group -- kept only because the user wants
#                  to see this player's decks for their own reference
#
# Updated 2026-07-19 per user correction: "remove wyze, fullly potatoe, ink from the day 2
# analysis feature, and add sandbox, once sandbox is there that's the entirety of my group
# except viiper... keep viiper there since i want to see his decks for my reference."
# Removed: Wyze❤️Ultimo, fluffypotato99, LF丨张✨Ink❤️llb (none of these are actually in the
# user's group). Added: SandBox (confirmed). JL Viiper's status changed confirmed ->
# reference (user clarified Viiper is NOT actually in their group, just wanted for
# reference). Adox/Lucas✨杰克/DK (added earlier this session as on-deck) are unaffected by
# this correction -- still on_deck.
# REPLACED 2026-07-19: per explicit user request ("replace the current day 2 opponents
# with the 15 opponents who I am up against in the monthly final, give me the same
# stats/info and dashboard features on those players as you did for the day 2 players"),
# the entire prior Day-2 Group A roster (7 opponents + 6 reference-only players) is
# replaced with the 15 possible Monthly Finals (Day 3) opponents. "status" is reused as:
#   "confirmed" -- one of the user's 5 "projected" (most-likely) Day 3 opponents
#   "on_deck"   -- one of the 10 other possible Day 3 opponents
# (no more "reference" category this round -- fully replaced, not appended to)
GROUP_A_OPPONENTS = [
    ("Mugi", "master_2CLV2RP0.json", "#2CLV2RP0", "confirmed"),
    ("SandBox", "master_Y022GRCJQ.json", "#Y022GRCJQ", "confirmed"),
    ("40k Oker", None, "#YLVV0JPQ", "confirmed"),
    ("Mohamed Light", "master_G9YV9GR8R.json", "#G9YV9GR8R", "confirmed"),
    ("Adriel", "master_9CPCC890.json", "#9CPCC890", "confirmed"),
    ("Pedro™️", "master_RJ88Y8U08.json", "#RJ88Y8U08", "on_deck"),
    ("Asaf", "master_RUQ0JU2P.json", "#RUQ0JU2P", "on_deck"),
    ("Clown (KickAsh)", "master_GPPYR9JYR.json", "#GPPYR9JYR", "on_deck"),
    ("Vitor75", "master_8LJ92G8UG.json", "#8LJ92G8UG", "on_deck"),
    ("Sub", "master_U890Q9UQ.json", "#U890Q9UQ", "on_deck"),
    ("SK Morten", "master_R09228V.json", "#R09228V", "on_deck"),
    ("Guriko", "master_2LJ0ULYCC.json", "#2LJ0ULYCC", "on_deck"),
    ("Polaris", "master_U8RYGC8GU.json", "#U8RYGC8GU", "on_deck"),
    ("JorZ", "master_22LC8JG02.json", "#22LC8JG02", "on_deck"),
    ("FrancoMedinaSL", "master_UJQQCUCQ8.json", "#UJQQCUCQ8", "on_deck"),
]


def deck_key(cards):
    return ", ".join(sorted(cards))


def analyze(path):
    with open(path) as f:
        battles = json.load(f)
    deck_games = Counter()
    deck_wins = Counter()
    wincon_games = Counter()
    wincon_wins = Counter()
    counters = Counter()
    for b in battles:
        team = b["team"][0]
        opp = b["opponent"][0]
        deck = team.get("cards", [])
        if not deck:
            continue
        dk = deck_key([c["name"] for c in deck])
        won = team.get("crowns", 0) > opp.get("crowns", 0)
        deck_games[dk] += 1
        if won:
            deck_wins[dk] += 1
        wincons = classify_deck([c["name"] for c in deck]) or ["(none classified)"]
        for wc in wincons:
            wincon_games[wc] += 1
            if won:
                wincon_wins[wc] += 1
        if not won:
            opp_dk = deck_key([c["name"] for c in opp.get("cards", [])])
            if opp_dk:
                counters[opp_dk] += 1
    return {
        "total_games": sum(deck_games.values()),
        "deck_games": deck_games, "deck_wins": deck_wins,
        "wincon_games": wincon_games, "wincon_wins": wincon_wins,
        "counters": counters,
    }


# ---- Full tracked data pool for counter-recommendations (added 2026-07-19) ----
# Same duel_log the main workbook is built from -- every tracked/extended-roster player's
# Practice + Official CRL games combined, per explicit user instruction to use "all
# games, practice and crl as a whole" as the data pool for recommendations.
_full_duel_log, _duel_summary, _session_stats = build_dataset()
print(f"Counter-recommendation data pool: {len(_full_duel_log)} games loaded from the "
      f"main tracked roster's Duel Log (Practice + Official CRL combined).")

wb = openpyxl.load_workbook(XLSX_PATH)
if 'Group A Scouting' in wb.sheetnames:
    del wb['Group A Scouting']
ws = wb.create_sheet('Group A Scouting')
ws.column_dimensions['A'].width = 90
ws.sheet_view.showGridLines = True

r = 1
ws.cell(r, 1, "Group A Scouting Report -- vs 老板 Ι Batan'宙斯 (Day 2 group stage)").font = \
    Font(name='Arial', size=15, bold=True, color='FF1F4E78')
r += 1
ws.cell(r, 1,
        "Snake-seeded from the Day 1 Swiss final standings per the official rulebook (4.1.3.8.3): "
        "seeds 1, 16, 17, 32, 33, 48, 49, 64 -> Group A. Each section below is built directly from "
        "that player's own recorded battle history (their master/extended-roster archive) -- deck "
        "frequency, win rate, top win-con sets, 'decks that beat them' (drawn only from games "
        "they actually lost, not a theorycrafted counter list), and a data-driven 'Recommended "
        "for Tomorrow' pick (see below) -- flag any small-sample rows as directional only, per "
        "usual caveat. Players marked ON DECK are NOT yet confirmed Group A members -- scouted "
        "ahead of time in case disqualifications reshuffle the group before Day 2."
        ).font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 2

for name, fname, tag, status in GROUP_A_OPPONENTS:
    if status == "on_deck":
        header_text = f"{name}  [ON DECK -- not yet confirmed Group A]"
    elif status == "reference":
        header_text = f"{name}  [REFERENCE ONLY -- not in your group]"
    else:
        header_text = name
    ws.cell(r, 1, header_text).font = NAME_FONT
    if status in ("on_deck", "reference"):
        ws.cell(r, 1).fill = ONDECK_FILL
    r += 1
    if fname is None:
        ws.cell(r, 1,
                f"No data yet -- {name} ({tag}) has not played any tracked or extended-roster "
                "player so far, so there's no battle log for them in this archive. Run "
                "fetch_scout_player.py against this tag (or add to extended_roster_tags.json and "
                "re-run fetch_extended_roster.py) to populate this section."
                ).font = VALUE_FONT
        ws.cell(r, 1).fill = PENDING_FILL
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        r += 2
        continue

    data = analyze(os.path.join(CRL_DIR, fname))
    ws.cell(r, 1, f"{data['total_games']} games in their recent battle log ({fname}).").font = NOTE_FONT
    r += 1

    ws.cell(r, 1, "Most-played decks").font = LABEL_FONT
    r += 1
    for dk, g in data["deck_games"].most_common(5):
        w = data["deck_wins"][dk]
        ws.cell(r, 1, f"  {g}g, {w}w ({w/g:.0%}) -- {dk}").font = VALUE_FONT
        r += 1

    wr_ranked = sorted(
        [(dk, g, data["deck_wins"][dk]) for dk, g in data["deck_games"].items() if g >= 2],
        key=lambda x: -(x[2] / x[1])
    )
    if wr_ranked:
        best = wr_ranked[0]
        worst = wr_ranked[-1]
        ws.cell(r, 1, f"Highest win-rate deck (min 2 games): {best[2]}/{best[1]} "
                       f"({best[2]/best[1]:.0%}) -- {best[0]}").font = VALUE_FONT
        r += 1
        ws.cell(r, 1, f"Lowest win-rate deck (min 2 games): {worst[2]}/{worst[1]} "
                       f"({worst[2]/worst[1]:.0%}) -- {worst[0]}").font = VALUE_FONT
        r += 1
    r += 1

    ws.cell(r, 1, "Win-condition sets (by games played)").font = LABEL_FONT
    r += 1
    top_wincons_ranked = data["wincon_games"].most_common(6)
    for wc, g in top_wincons_ranked:
        w = data["wincon_wins"][wc]
        ws.cell(r, 1, f"  {g}g, {w}w ({w/g:.0%}) -- {wc}").font = VALUE_FONT
        r += 1
    r += 1

    ws.cell(r, 1, "Decks that have beaten them (from their own logged losses)").font = LABEL_FONT
    r += 1
    if data["counters"]:
        for dk, c in data["counters"].most_common(5):
            ws.cell(r, 1, f"  beat them {c}x -- {dk}").font = VALUE_FONT
            r += 1
    else:
        ws.cell(r, 1, "  No logged losses in this snapshot -- no counter data available yet.").font = VALUE_FONT
        r += 1
    r += 1

    # ---- Recommended for Tomorrow (added 2026-07-19) ----
    ws.cell(r, 1, "Recommended for Tomorrow (data-driven, from our ENTIRE tracked pool)").font = LABEL_FONT
    r += 1
    top_wincon_names = [wc for wc, _g in top_wincons_ranked[:2] if wc != "(none classified)"]
    if not top_wincon_names:
        ws.cell(r, 1, "  Not enough classified win-con data on this player yet to build a "
                       "recommendation.").font = VALUE_FONT
        r += 1
    else:
        reco = compute_counter_recommendations(_full_duel_log, top_wincon_names, min_games=3, top_n=3)
        ws.cell(r, 1, f"  Based on their top win-con(s) -- {', '.join(top_wincon_names)} -- "
                       f"filtered across our whole tracked Duel Log: {reco['sample_size']} of our "
                       "own games (Practice + Official CRL combined) faced an opponent playing one "
                       "of those win conditions.").font = NOTE_FONT
        r += 1
        if reco["sample_size"] == 0:
            ws.cell(r, 1, "  No games in our data pool have faced this win-con archetype yet -- "
                           "no recommendation possible.").font = VALUE_FONT
            r += 1
        else:
            ws.cell(r, 1, "  Best deck to bring (our win rate vs this archetype, min 3 games):").font = VALUE_FONT
            r += 1
            if reco["best_decks"]:
                for row in reco["best_decks"]:
                    ws.cell(r, 1, f"    {row['wins']}/{row['games']} ({row['win_rate']:.0%}) -- "
                                   f"{row['deck']}").font = RECO_FONT
                    r += 1
            else:
                ws.cell(r, 1, "    No single deck reaches the 3-game minimum yet.").font = VALUE_FONT
                r += 1
            ws.cell(r, 1, "  Best win-con set to bring (min 3 games):").font = VALUE_FONT
            r += 1
            if reco["best_wincon_sets"]:
                for row in reco["best_wincon_sets"]:
                    ws.cell(r, 1, f"    {row['wins']}/{row['games']} ({row['win_rate']:.0%}) -- "
                                   f"{row['wincon_set']}").font = RECO_FONT
                    r += 1
            else:
                ws.cell(r, 1, "    No win-con set reaches the 3-game minimum yet.").font = VALUE_FONT
                r += 1
            ws.cell(r, 1, "  Best duel-set (decks used together across a duel vs this archetype, min 3 games):").font = VALUE_FONT
            r += 1
            if reco["best_duel_sets"]:
                for row in reco["best_duel_sets"]:
                    ws.cell(r, 1, f"    {row['wins']}/{row['games']} ({row['win_rate']:.0%}) -- "
                                   f"{row['duel_set']}").font = RECO_FONT
                    r += 1
            else:
                ws.cell(r, 1, "    No duel-set combination reaches the 3-game minimum yet.").font = VALUE_FONT
                r += 1
    r += 2

wb.save(XLSX_PATH)
print(f"Group A Scouting sheet added. {r} rows written.")
