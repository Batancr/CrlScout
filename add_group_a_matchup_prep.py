"""Adds a "Group A Matchup Prep" sheet to CRL_Duel_Decks.xlsx -- the same "My Matchup
Analysis" treatment built for the user (add_batan_matchup_analysis.py), generalized to run
for ALL 7 Day 2 opponents (4 confirmed + 3 on-deck; the 6 reference-only players are
excluded, same scope as Group A Sequencing). Added 2026-07-19 per explicit user request:

"For the my matchup analysis page, can you make such a page in the excel, and a feature as
well in the day 2, for all my day 2 opponents..."

For EACH of the 7 players, the same structure/logic as My Matchup Analysis, generalized
from "the user" to "this specific opponent":

1. Their own most-played decks / win-conditions (from their own game history).
2. EMPIRICAL -- what has actually beaten THEM personally: pulled from every game where
   they were the player and lost, ranking the beating deck by its win rate against them
   specifically (min 3 games faced).
3. PREDICTIVE / COMMUNITY-WIDE -- their top 3 win-conditions filtered across the ENTIRE
   tracked pool (every roster/extended-roster player's games), ranking the OPPONENT side in
   that slice by win rate -- what the wider dataset says beats decks like theirs, backed by
   a larger sample than their own history alone.
4. Double-confirmed cross-reference: decks in BOTH lists above.

This is the flip side of "Recommended for Tomorrow" in the Group A Scouting sheet (which
answers "what should WE bring against them" using only OUR own results) -- this sheet
answers "what's most likely to beat THEM, and what already has," prep material for planning
actual picks against each specific opponent.

Reuses the same Extended-Roster fallback loader as add_group_a_sequencing_analysis.py for
INA.BenZerRidel (no master_<tag>.json -- Extended Roster only).
"""
import glob
import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

from build_duel_workbook import classify_deck, build_dataset, classify_match_category, parse_time

import os as _os
_CRL_HOME = _os.environ.get('CRL_HOME')
XLSX_PATH = _os.path.join(_CRL_HOME, 'CRL_Duel_Decks.xlsx') if _CRL_HOME else '/home/claude/CRL_Duel_Decks.xlsx'

# REPLACED 2026-07-19: entire Day-2 roster swapped for the 15 possible Monthly Finals
# (Day 3) opponents -- see add_group_a_analysis.py for the full rationale.
TARGET_PLAYERS = [
    ("Mugi", "#2CLV2RP0", "confirmed"),
    ("SandBox", "#Y022GRCJQ", "confirmed"),
    ("40k Oker", "#YLVV0JPQ", "confirmed"),
    ("Mohamed Light", "#G9YV9GR8R", "confirmed"),
    ("Adriel", "#9CPCC890", "confirmed"),
    ("Pedro™️", "#RJ88Y8U08", "on_deck"),
    ("Asaf", "#RUQ0JU2P", "on_deck"),
    ("Clown (KickAsh)", "#GPPYR9JYR", "on_deck"),
    ("Vitor75", "#8LJ92G8UG", "on_deck"),
    ("Sub", "#U890Q9UQ", "on_deck"),
    ("SK Morten", "#R09228V", "on_deck"),
    ("Guriko", "#2LJ0ULYCC", "on_deck"),
    ("Polaris", "#U8RYGC8GU", "on_deck"),
    ("JorZ", "#22LC8JG02", "on_deck"),
    ("FrancoMedinaSL", "#UJQQCUCQ8", "on_deck"),
]

MIN_GAMES_DECK = 3
MIN_GAMES_WINCON = 5

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='FF1F4E78')
LABEL_FONT = Font(name='Arial', size=11, bold=True)
VALUE_FONT = Font(name='Arial', size=11, bold=False)
SECTION_FONT = Font(name='Arial', size=12, bold=True, color='FF1F4E78')
PLAYER_FONT = Font(name='Arial', size=14, bold=True, color='FFFFFFFF')
PLAYER_FILL = PatternFill('solid', fgColor='FF2E7D32')
ONDECK_FILL = PatternFill('solid', fgColor='FFFFF2CC')
NOTE_FONT = Font(name='Arial', size=10, italic=True, color='FF666666')
TITLE_FONT = Font(name='Arial', size=16, bold=True, color='FF1F4E78')
THREAT_FILL = PatternFill('solid', fgColor='FFFCE8E6')

_ROSTER_TAGS_ONLY = {
    f"#{os.path.basename(p)[len('master_'):-len('.json')]}" for p in glob.glob('master_*.json')
}


def load_extended_rows_for_tag(tag):
    tag_digits = tag.lstrip('#')
    path = f'extended_{tag_digits}.json'
    if not os.path.exists(path):
        return []
    with open(path) as f:
        battles = json.load(f)
    rows = []
    for b in battles:
        team = b.get('team', [{}])[0]
        opp = b.get('opponent', [{}])[0]
        deck = [c['name'] for c in team.get('cards', [])]
        opp_deck = [c['name'] for c in opp.get('cards', [])]
        if not deck:
            continue
        battle_time = parse_time(b['battleTime'])
        category = classify_match_category(
            b.get('type'), b.get('gameMode', {}).get('name'), battle_time,
            opponent_tag=opp.get('tag'), roster_tags=_ROSTER_TAGS_ONLY,
        )
        if category is None:
            continue
        rows.append({
            'player_tag': team.get('tag'),
            'opponent_tag': opp.get('tag'),
            'battle_time': battle_time,
            'deck': deck,
            'opponent_deck': opp_deck,
            'crowns_for': team.get('crowns', 0),
            'crowns_against': opp.get('crowns', 0),
        })
    return rows


full_duel_log, _duel_summary, _session_stats = build_dataset()


def analyze_target(tag):
    target_games = [r for r in full_duel_log if r['player_tag'] == tag and r['deck'] and len(r['deck']) == 8]
    used_extended_fallback = False
    if not target_games:
        extended = load_extended_rows_for_tag(tag)
        target_games = [r for r in extended if r['deck'] and len(r['deck']) == 8]
        used_extended_fallback = bool(target_games)

    # ---- 1) their own most-played decks / win-cons ----
    deck_games = Counter()
    deck_wins = Counter()
    wincon_games = Counter()
    wincon_wins = Counter()
    for r in target_games:
        dk = ', '.join(sorted(r['deck']))
        won = r['crowns_for'] > r['crowns_against']
        deck_games[dk] += 1
        if won:
            deck_wins[dk] += 1
        for wc in (classify_deck(r['deck']) or []):
            wincon_games[wc] += 1
            if won:
                wincon_wins[wc] += 1
    top_wincons = [wc for wc, _g in wincon_games.most_common(3)]

    # ---- 2) EMPIRICAL -- decks that have actually beaten them ----
    opp_deck_games_vs_target = Counter()
    opp_deck_wins_vs_target = Counter()  # wins FOR THE OPPONENT, i.e. target's losses
    for r in target_games:
        if not r.get('opponent_deck') or len(r['opponent_deck']) != 8:
            continue
        dk = ', '.join(sorted(r['opponent_deck']))
        opp_deck_games_vs_target[dk] += 1
        if r['crowns_for'] < r['crowns_against']:
            opp_deck_wins_vs_target[dk] += 1

    # ---- 3) PREDICTIVE -- across the ENTIRE tracked pool, what beats their win-con pool ----
    target_set = set(top_wincons)
    predicted = {'sample_size': 0, 'top_decks': [], 'top_wincons': []}
    if target_set:
        filtered = [
            r for r in full_duel_log
            if r.get('deck') and set(classify_deck(r['deck']) or []) & target_set
        ]
        opp_deck_games = Counter()
        opp_deck_wins = Counter()
        opp_wincon_games = Counter()
        opp_wincon_wins = Counter()
        for r in filtered:
            if not r.get('opponent_deck') or len(r['opponent_deck']) != 8:
                continue
            dk = ', '.join(sorted(r['opponent_deck']))
            opp_deck_games[dk] += 1
            opp_won = r['crowns_for'] < r['crowns_against']
            if opp_won:
                opp_deck_wins[dk] += 1
            for wc in (classify_deck(r['opponent_deck']) or []):
                opp_wincon_games[wc] += 1
                if opp_won:
                    opp_wincon_wins[wc] += 1
        top_decks = sorted(
            [{'deck': dk, 'games': g, 'wins': opp_deck_wins[dk], 'win_rate': opp_deck_wins[dk] / g}
             for dk, g in opp_deck_games.items() if g >= MIN_GAMES_DECK],
            key=lambda x: (-x['win_rate'], -x['games']),
        )[:6]
        top_wc = sorted(
            [{'wincon': wc, 'games': g, 'wins': opp_wincon_wins[wc], 'win_rate': opp_wincon_wins[wc] / g}
             for wc, g in opp_wincon_games.items() if g >= MIN_GAMES_WINCON],
            key=lambda x: (-x['win_rate'], -x['games']),
        )[:6]
        predicted = {'sample_size': len(filtered), 'top_decks': top_decks, 'top_wincons': top_wc}

    # ---- 4) double-confirmed cross-reference ----
    personal_threat_decks = {
        dk for dk, g in opp_deck_games_vs_target.items()
        if g >= 2 and opp_deck_wins_vs_target[dk] / g >= 0.5
    }
    predicted_deck_keys = {row['deck'] for row in predicted['top_decks']}
    double_confirmed = sorted(personal_threat_decks & predicted_deck_keys)

    return {
        'total_games': len(target_games),
        'used_extended_fallback': used_extended_fallback,
        'deck_games': deck_games, 'deck_wins': deck_wins,
        'wincon_games': wincon_games, 'wincon_wins': wincon_wins,
        'top_wincons': top_wincons,
        'opp_deck_games_vs_target': opp_deck_games_vs_target,
        'opp_deck_wins_vs_target': opp_deck_wins_vs_target,
        'predicted': predicted,
        'double_confirmed': double_confirmed,
    }


# ---- Build the sheet ----
wb = openpyxl.load_workbook(XLSX_PATH)
SHEET_NAME = 'Group A Matchup Prep'
if SHEET_NAME in wb.sheetnames:
    del wb[SHEET_NAME]
ws = wb.create_sheet(SHEET_NAME)
ws.sheet_view.showGridLines = True
ws.column_dimensions['A'].width = 70
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 55

r = 1
ws.cell(r, 1, 'Group A Matchup Prep -- What Beats Each Day 2 Opponent').font = TITLE_FONT
r += 1
ws.cell(r, 1,
        'Scoped solely to the 7 players actually in Group A tomorrow (4 confirmed + 3 on-deck). '
        'For each: their own most-played decks/win-cons, what has actually beaten them personally '
        '(empirical), and what the wider tracked pool says beats decks like theirs (predictive, '
        'much larger sample). This is the flip side of "Recommended for Tomorrow" in Group A '
        'Scouting (which only uses OUR own results) -- here the goal is picking the strongest '
        'answer against each specific opponent. All figures are drawn from real recorded games; '
        'nothing here is theorycrafted.'
        ).font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.row_dimensions[r].height = 70
r += 2

for name, tag, status in TARGET_PLAYERS:
    data = analyze_target(tag)

    status_label = ' [ON DECK -- not yet confirmed Group A]' if status == 'on_deck' else ''
    ws.cell(r, 1, f'{name}  ({tag}){status_label}').font = PLAYER_FONT
    for c in range(1, 5):
        ws.cell(r, c).fill = ONDECK_FILL if status == 'on_deck' else PLAYER_FILL
    r += 1
    fallback_note = (
        ' (No master roster fetch history -- pulled from their Extended Roster snapshot instead, '
        'a flat recent-battle log; still valid for the sections below.)'
        if data['used_extended_fallback'] else ''
    )
    ws.cell(r, 1, f"{data['total_games']} total logged games with a full 8-card deck." + fallback_note).font = NOTE_FONT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    r += 1

    if data['total_games'] == 0:
        ws.cell(r, 1, 'No games recorded for this player yet.').font = VALUE_FONT
        r += 2
        continue

    # --- Section 1 ---
    ws.cell(r, 1, 'Their most-played win conditions').font = SECTION_FONT
    r += 1
    ws.cell(r, 1, 'Win Condition').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, 'Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
    ws.cell(r, 3, 'Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
    ws.cell(r, 4, '').fill = HEADER_FILL
    r += 1
    for wc, g in data['wincon_games'].most_common(6):
        w = data['wincon_wins'][wc]
        is_top3 = wc in data['top_wincons']
        ws.cell(r, 1, wc + (' *' if is_top3 else '')).font = LABEL_FONT if is_top3 else VALUE_FONT
        ws.cell(r, 2, g).font = VALUE_FONT
        wr_cell = ws.cell(r, 3, w / g if g else 0)
        wr_cell.number_format = '0.0%'
        wr_cell.font = VALUE_FONT
        r += 1
    ws.cell(r, 1, '* = top 3 win condition, used below to find likely counters').font = NOTE_FONT
    r += 1

    # --- Section 2: empirical ---
    ws.cell(r, 1, f'What has actually beaten them (min {MIN_GAMES_DECK} games faced)').font = SECTION_FONT
    r += 1
    ws.cell(r, 1, 'Opponent Deck').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, 'Games vs Them').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
    ws.cell(r, 3, 'Their Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
    ws.cell(r, 4, '').fill = HEADER_FILL
    r += 1
    ranked_threats = sorted(
        [(dk, g, data['opp_deck_wins_vs_target'][dk]) for dk, g in data['opp_deck_games_vs_target'].items()
         if g >= MIN_GAMES_DECK],
        key=lambda x: (-(x[2] / x[1]), -x[1]),
    )
    if ranked_threats:
        for dk, g, w in ranked_threats[:8]:
            ws.cell(r, 1, dk).font = VALUE_FONT
            ws.cell(r, 1).alignment = Alignment(wrap_text=True)
            ws.cell(r, 2, g).font = VALUE_FONT
            wr_cell = ws.cell(r, 3, w / g)
            wr_cell.number_format = '0.0%'
            wr_cell.font = VALUE_FONT
            if w / g >= 0.6:
                for c in (1, 2, 3):
                    ws.cell(r, c).fill = THREAT_FILL
            r += 1
    else:
        ws.cell(r, 1, f'No single opponent deck has faced them {MIN_GAMES_DECK}+ times yet.').font = VALUE_FONT
        r += 1
    r += 1

    # --- Section 3: predictive ---
    predicted = data['predicted']
    ws.cell(r, 1, 'What the wider tracked pool says beats their win-con archetype').font = SECTION_FONT
    r += 1
    ws.cell(r, 1,
            f"Their top win-cons ({', '.join(data['top_wincons']) or 'none identified'}) filtered across the "
            f"whole tracked pool -- {predicted['sample_size']} games. Ranked by how often the OPPONENT side "
            "won in those games."
            ).font = NOTE_FONT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    r += 1
    ws.cell(r, 1, 'Deck').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
    ws.cell(r, 2, 'Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
    ws.cell(r, 3, 'Their Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
    ws.cell(r, 4, '').fill = HEADER_FILL
    r += 1
    if predicted['top_decks']:
        for row in predicted['top_decks']:
            ws.cell(r, 1, row['deck']).font = VALUE_FONT
            ws.cell(r, 1).alignment = Alignment(wrap_text=True)
            ws.cell(r, 2, row['games']).font = VALUE_FONT
            wr_cell = ws.cell(r, 3, row['win_rate'])
            wr_cell.number_format = '0.0%'
            wr_cell.font = VALUE_FONT
            if row['deck'] in data['double_confirmed']:
                ws.cell(r, 4, '⚠ also beat them personally').font = VALUE_FONT
                for c in (1, 2, 3, 4):
                    ws.cell(r, c).fill = THREAT_FILL
            r += 1
    else:
        ws.cell(r, 1, f"No single deck reaches the {MIN_GAMES_DECK}-game minimum against this win-con pool yet.").font = VALUE_FONT
        r += 1
    r += 1

    # --- Section 4: double-confirmed ---
    ws.cell(r, 1, 'Double-confirmed threats (beat them personally AND beat their win-con archetype pool-wide)').font = LABEL_FONT
    r += 1
    if data['double_confirmed']:
        for dk in data['double_confirmed']:
            ws.cell(r, 1, dk).font = VALUE_FONT
            ws.cell(r, 1).alignment = Alignment(wrap_text=True)
            ws.cell(r, 1).fill = THREAT_FILL
            r += 1
    else:
        ws.cell(r, 1, 'None yet -- samples still small in places.').font = VALUE_FONT
        r += 1
    r += 2

wb.save(XLSX_PATH)
print(f"Sheet '{SHEET_NAME}' added. {r} rows written.")
for name, tag, status in TARGET_PLAYERS:
    d = analyze_target(tag)
    print(f"  {name}: {d['total_games']} games, top wincons={d['top_wincons']}, "
          f"predicted sample={d['predicted']['sample_size']}, double-confirmed={d['double_confirmed']}")
