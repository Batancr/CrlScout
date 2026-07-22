"""Adds a "My Matchup Analysis" sheet to CRL_Duel_Decks.xlsx, scoped specifically to
老板 Ι Batan'宙斯 (#9RQ8YRYQL) -- the user. Added 2026-07-19 per explicit user request:
"run an analysis on what decks do best vs me, historically from all my games, as well as
according to what decks and win conditions I play most often (what decks would counter the
majority of my win condition and deck pool that I play most often), this way I can predict
what my opponents might be thinking to play against me and counter them back."

Two distinct, both data-driven (not theorycrafted) angles, covering both halves of the
request:

1. EMPIRICAL -- what has actually beaten me: pulled directly from the user's own logged
   games (every game where Player Tag = the user's tag, all match categories combined,
   Practice + Official CRL), ranking opponent decks by their real win rate against the
   user specifically (min games threshold so thin samples stay visibly thin).

2. PREDICTIVE / COMMUNITY-WIDE -- what's likely to be brought against the user: the
   user's own most-played decks/win-conditions are identified first, then the ENTIRE
   tracked data pool (every roster/extended-roster player's games, not just the user's
   own) is filtered down to games where ANY tracked player used one of those same win
   conditions, and the OPPONENT side's decks/win-cons in that filtered slice are ranked by
   how often THEY won. This answers "what does the community's data say beats decks like
   mine" with a much larger sample than the user's own history alone provides -- exactly
   the "predict what opponents might be thinking to counter me" angle the user asked for.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter

from build_duel_workbook import classify_deck, build_dataset

import os as _os
_CRL_HOME = _os.environ.get('CRL_HOME')
XLSX_PATH = _os.path.join(_CRL_HOME, 'CRL_Duel_Decks.xlsx') if _CRL_HOME else '/home/claude/CRL_Duel_Decks.xlsx'
BATAN_TAG = '#9RQ8YRYQL'
BATAN_NAME = "老板 Ι Batan'宙斯"

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='FF1F4E78')
LABEL_FONT = Font(name='Arial', size=11, bold=True)
VALUE_FONT = Font(name='Arial', size=11, bold=False)
SECTION_FONT = Font(name='Arial', size=12, bold=True, color='FF1F4E78')
NOTE_FONT = Font(name='Arial', size=10, italic=True, color='FF666666')
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='FF1F4E78')
THREAT_FILL = PatternFill('solid', fgColor='FFFCE8E6')

MIN_GAMES_DECK = 3
MIN_GAMES_WINCON = 5

full_duel_log, _duel_summary, _session_stats = build_dataset()
my_games = [r for r in full_duel_log if r['player_tag'] == BATAN_TAG]

# ---------------------------------------------------------------------------
# 1) My most-played decks / win conditions
# ---------------------------------------------------------------------------
my_deck_games = Counter()
my_deck_wins = Counter()
my_wincon_games = Counter()
my_wincon_wins = Counter()
for r in my_games:
    if not r['deck'] or len(r['deck']) != 8:
        continue
    dk = ', '.join(sorted(r['deck']))
    won = r['crowns_for'] > r['crowns_against']
    my_deck_games[dk] += 1
    if won:
        my_deck_wins[dk] += 1
    for wc in (classify_deck(r['deck']) or []):
        my_wincon_games[wc] += 1
        if won:
            my_wincon_wins[wc] += 1

my_top_wincons = [wc for wc, _g in my_wincon_games.most_common(3)]

# ---------------------------------------------------------------------------
# 2) EMPIRICAL -- decks that have actually beaten me (and their overall record vs me)
# ---------------------------------------------------------------------------
opp_deck_games_vs_me = Counter()
opp_deck_wins_vs_me = Counter()  # wins FOR THE OPPONENT, i.e. my losses
for r in my_games:
    if not r['opponent_deck'] or len(r['opponent_deck']) != 8:
        continue
    dk = ', '.join(sorted(r['opponent_deck']))
    opp_deck_games_vs_me[dk] += 1
    if r['crowns_for'] < r['crowns_against']:
        opp_deck_wins_vs_me[dk] += 1

# ---------------------------------------------------------------------------
# 3) PREDICTIVE -- across the ENTIRE tracked pool, what beats players using MY win-cons
# ---------------------------------------------------------------------------
def compute_predicted_counters(duel_log, target_wincons, min_games_deck=3, min_games_wincon=5, top_n=6):
    target = set(target_wincons)
    if not target:
        return {'sample_size': 0, 'top_decks': [], 'top_wincons': []}
    filtered = [
        r for r in duel_log
        if r.get('deck') and set(classify_deck(r['deck']) or []) & target
    ]
    opp_deck_games = Counter()
    opp_deck_wins = Counter()
    opp_wincon_games = Counter()
    opp_wincon_wins = Counter()
    for r in filtered:
        if not r.get('opponent_deck') or len(r['opponent_deck']) != 8:
            continue
        dk = ', '.join(sorted(r['opponent_deck']))
        opp_deck_games[dk] += 1
        opp_won = r['crowns_for'] < r['crowns_against']
        if opp_won:
            opp_deck_wins[dk] += 1
        for wc in (classify_deck(r['opponent_deck']) or []):
            opp_wincon_games[wc] += 1
            if opp_won:
                opp_wincon_wins[wc] += 1
    top_decks = sorted(
        [{'deck': dk, 'games': g, 'wins': opp_deck_wins[dk], 'win_rate': opp_deck_wins[dk] / g}
         for dk, g in opp_deck_games.items() if g >= min_games_deck],
        key=lambda x: (-x['win_rate'], -x['games']),
    )[:top_n]
    top_wincons = sorted(
        [{'wincon': wc, 'games': g, 'wins': opp_wincon_wins[wc], 'win_rate': opp_wincon_wins[wc] / g}
         for wc, g in opp_wincon_games.items() if g >= min_games_wincon],
        key=lambda x: (-x['win_rate'], -x['games']),
    )[:top_n]
    return {'sample_size': len(filtered), 'top_decks': top_decks, 'top_wincons': top_wincons}


predicted = compute_predicted_counters(full_duel_log, my_top_wincons, MIN_GAMES_DECK, MIN_GAMES_WINCON)

# ---------------------------------------------------------------------------
# 4) Cross-reference -- decks appearing in BOTH the "beat me personally" list AND the
#    "beats my win-con archetype broadly" list are the double-confirmed threats.
# ---------------------------------------------------------------------------
personal_threat_decks = {
    dk for dk, g in opp_deck_games_vs_me.items()
    if g >= 2 and opp_deck_wins_vs_me[dk] / g >= 0.5
}
predicted_deck_keys = {row['deck'] for row in predicted['top_decks']}
double_confirmed = sorted(personal_threat_decks & predicted_deck_keys)

# ---- Build the sheet ----
wb = openpyxl.load_workbook(XLSX_PATH)
if 'My Matchup Analysis' in wb.sheetnames:
    del wb['My Matchup Analysis']
ws = wb.create_sheet('My Matchup Analysis')
ws.sheet_view.showGridLines = True
ws.column_dimensions['A'].width = 85
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 60

r = 1
ws.cell(r, 1, f'My Matchup Analysis -- {BATAN_NAME}').font = TITLE_FONT
r += 1
ws.cell(r, 1,
        'What historically beats me, what I play most, and what the community\'s own data says '
        'counters players who run decks/win-cons like mine -- so I can anticipate what opponents '
        'are likely to bring and prepare an answer. All figures are drawn from real recorded games '
        '(Practice + Official CRL combined unless noted); nothing here is theorycrafted.'
        ).font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 2

# ---- Section 1: my most-played decks / win-cons ----
ws.cell(r, 1, 'My most-played decks').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'Deck').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, '').fill = HEADER_FILL
r += 1
for dk, g in my_deck_games.most_common(8):
    w = my_deck_wins[dk]
    ws.cell(r, 1, dk).font = VALUE_FONT
    ws.cell(r, 1).alignment = Alignment(wrap_text=True)
    ws.cell(r, 2, g).font = VALUE_FONT
    wr_cell = ws.cell(r, 3, w / g if g else 0)
    wr_cell.number_format = '0.0%'
    wr_cell.font = VALUE_FONT
    r += 1
r += 1

ws.cell(r, 1, 'My most-played win conditions').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'Win Condition').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, '').fill = HEADER_FILL
r += 1
for wc, g in my_wincon_games.most_common(8):
    w = my_wincon_wins[wc]
    is_top3 = wc in my_top_wincons
    ws.cell(r, 1, wc + (' *' if is_top3 else '')).font = LABEL_FONT if is_top3 else VALUE_FONT
    ws.cell(r, 2, g).font = VALUE_FONT
    wr_cell = ws.cell(r, 3, w / g if g else 0)
    wr_cell.number_format = '0.0%'
    wr_cell.font = VALUE_FONT
    r += 1
ws.cell(r, 1, '* = top 3 win condition, used below to find likely counters').font = NOTE_FONT
r += 2

# ---- Section 2: empirical -- decks that beat me ----
ws.cell(r, 1, 'What has actually beaten me (from my own game history)').font = SECTION_FONT
r += 1
ws.cell(r, 1, f'Ranked by their win rate against me specifically, minimum {MIN_GAMES_DECK} games faced. '
               'This is the direct, personal answer to "what beats me" -- drawn only from games I actually '
               'played.').font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 1
ws.cell(r, 1, 'Opponent Deck').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Games vs Me').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Their Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, '').fill = HEADER_FILL
r += 1
ranked_threats = sorted(
    [(dk, g, opp_deck_wins_vs_me[dk]) for dk, g in opp_deck_games_vs_me.items() if g >= MIN_GAMES_DECK],
    key=lambda x: (-(x[2] / x[1]), -x[1]),
)
if ranked_threats:
    for dk, g, w in ranked_threats[:10]:
        ws.cell(r, 1, dk).font = VALUE_FONT
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        ws.cell(r, 2, g).font = VALUE_FONT
        wr_cell = ws.cell(r, 3, w / g)
        wr_cell.number_format = '0.0%'
        wr_cell.font = VALUE_FONT
        if w / g >= 0.6:
            for c in (1, 2, 3):
                ws.cell(r, c).fill = THREAT_FILL
        r += 1
else:
    ws.cell(r, 1, f'No single opponent deck has faced me {MIN_GAMES_DECK}+ times yet.').font = VALUE_FONT
    r += 1
r += 1
ws.cell(r, 1, '(Red-highlighted rows = 60%+ win rate against me -- your most proven personal threats.)').font = NOTE_FONT
r += 2

# ---- Section 3: predictive -- community-wide counters to my win-con pool ----
ws.cell(r, 1, "Predicted counters -- what beats players using MY top win conditions, across the whole tracked pool").font = SECTION_FONT
r += 1
ws.cell(r, 1,
        f"My top 3 win conditions ({', '.join(my_top_wincons)}) filtered across every tracked/extended-roster "
        f"player's games (not just mine) -- {predicted['sample_size']} games in our whole pool had a tracked "
        "player using one of these. Ranked by how often the OPPONENT side won in those games -- this is the "
        "broader, community-data answer to \"what might my opponents be thinking of bringing against me,\" "
        "since it's not limited to my own (smaller) game history."
        ).font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 1

ws.cell(r, 1, f'Top decks that beat my win-con archetype (min {MIN_GAMES_DECK} games)').font = LABEL_FONT
r += 1
ws.cell(r, 1, 'Deck').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Their Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, '').fill = HEADER_FILL
r += 1
if predicted['top_decks']:
    for row in predicted['top_decks']:
        ws.cell(r, 1, row['deck']).font = VALUE_FONT
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        ws.cell(r, 2, row['games']).font = VALUE_FONT
        wr_cell = ws.cell(r, 3, row['win_rate'])
        wr_cell.number_format = '0.0%'
        wr_cell.font = VALUE_FONT
        if row['deck'] in double_confirmed:
            ws.cell(r, 4, '⚠ also beat me personally (see section above)').font = VALUE_FONT
            for c in (1, 2, 3, 4):
                ws.cell(r, c).fill = THREAT_FILL
        r += 1
else:
    ws.cell(r, 1, f'No single deck reaches the {MIN_GAMES_DECK}-game minimum against this win-con pool yet.').font = VALUE_FONT
    r += 1
r += 1

ws.cell(r, 1, f'Top win conditions that beat my win-con archetype (min {MIN_GAMES_WINCON} games)').font = LABEL_FONT
r += 1
ws.cell(r, 1, 'Win Condition').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Their Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, '').fill = HEADER_FILL
r += 1
if predicted['top_wincons']:
    for row in predicted['top_wincons']:
        ws.cell(r, 1, row['wincon']).font = VALUE_FONT
        ws.cell(r, 2, row['games']).font = VALUE_FONT
        wr_cell = ws.cell(r, 3, row['win_rate'])
        wr_cell.number_format = '0.0%'
        wr_cell.font = VALUE_FONT
        r += 1
else:
    ws.cell(r, 1, f'No single win condition reaches the {MIN_GAMES_WINCON}-game minimum yet.').font = VALUE_FONT
    r += 1
r += 2

# ---- Section 4: double-confirmed threats ----
ws.cell(r, 1, 'Double-confirmed threats (beat me personally AND beat my win-con archetype community-wide)').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'These decks show up in BOTH lists above -- proven against me directly, and statistically strong '
               'against decks like mine across the whole pool. These are the ones most worth pre-planning an '
               'answer for.').font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 1
if double_confirmed:
    for dk in double_confirmed:
        ws.cell(r, 1, dk).font = VALUE_FONT
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        ws.cell(r, 1).fill = THREAT_FILL
        r += 1
else:
    ws.cell(r, 1, 'None yet -- no deck currently appears in both lists (samples are still small in places).').font = VALUE_FONT
    r += 1
r += 2

ws.cell(r, 1, 'Read').font = SECTION_FONT
r += 1
top3_wr = [(wc, my_wincon_wins[wc] / my_wincon_games[wc]) for wc in my_top_wincons if my_wincon_games[wc]]
read_lines = (
    f"My top 3 win conditions are {', '.join(f'{wc} ({wr:.0%})' for wc, wr in top3_wr)} -- this is the "
    "archetype pool an opponent scouting me would most likely prepare for. The 'Predicted counters' section "
    "above shows what the wider player pool has found effective against that same archetype, independent of "
    "my own personal results, so it's a good proxy for what a well-prepared opponent brings. Cross-referencing "
    "against what has actually beaten me personally narrows that down to the decks worth taking most seriously "
    "-- see the double-confirmed list. Small-sample rows (below the games-minimum threshold, or just barely "
    "above it) should be treated as directional, not conclusive -- re-run this after more games accumulate, "
    "especially more Official CRL games specifically."
)
ws.cell(r, 1, read_lines).font = VALUE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.row_dimensions[r].height = 110

wb.save(XLSX_PATH)
print(f"My Matchup Analysis sheet added. {r} rows written.")
print(f"My top 3 win-cons: {my_top_wincons}")
print(f"Predicted-counter sample size: {predicted['sample_size']} games")
print(f"Double-confirmed threats: {double_confirmed}")
