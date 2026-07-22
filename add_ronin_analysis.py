"""Adds a 'Ronin G1 Analysis' sheet to CRL_Duel_Decks.xlsx: how often Ronin shows up in
Game 1 of a duel/set, compared to Games 2/3 and the overall play rate, plus tracked-player
win rate when they open with Ronin. Ronin is a rock-paper-scissors-style counter card, so
the interesting question is whether it's played blind in G1 or held back as a G2/G3 read.

Extended 2026-07-19, per explicit user request: "I want some stats... on how ronin
performed specifically in crl official matches, how often was it played, winrate, which
sets was it played most in, what type of decks, and how well ronin works against the 7
players in my group specifically (win rate of ronin vs those players)." Added four new
Official-CRL-scoped sections below the original Practice-only G1 analysis (kept as-is,
since it answers a different question -- G1 timing behavior, not CRL performance):
play rate + win rate in Official CRL specifically, win-con sets Ronin decks paired with,
top full decks running Ronin, and per-opponent Ronin win rate against the user's 7 actual
Group A players (4 confirmed + 3 on-deck; the 6 reference-only players are explicitly NOT
part of the user's group, so excluded here)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from collections import Counter, defaultdict

from build_duel_workbook import classify_deck

import os as _os
_CRL_HOME = _os.environ.get('CRL_HOME')
XLSX_PATH = _os.path.join(_CRL_HOME, 'CRL_Duel_Decks.xlsx') if _CRL_HOME else '/home/claude/CRL_Duel_Decks.xlsx'

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='FF1F4E78')
LABEL_FONT = Font(name='Arial', size=11, bold=True)
VALUE_FONT = Font(name='Arial', size=11, bold=False)
SECTION_FONT = Font(name='Arial', size=12, bold=True, color='FF1F4E78')
NOTE_FONT = Font(name='Arial', size=10, italic=True, color='FF666666')

# IMPORTANT: use a SEPARATE data_only=True load just to read cached values -- saving a
# data_only=True workbook strips every formula down to its last-cached value, which would
# wreck all the other sheets. The workbook we edit/save below is loaded normally instead.
read_wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
duel_log_ws = read_wb['Duel Log']
rows = list(duel_log_ws.iter_rows(values_only=True))
header = rows[0]
idx = {h: i for i, h in enumerate(header)}
# Practice only -- Official CRL games (added 2026-07-18) are excluded here, same as
# Win-Con Sets and the dashboard predictors: CRL sample size is thin, includes Pending/
# Anomaly duels that shouldn't feed a "Game 1" analysis yet, and practice/tournament
# patterns may genuinely differ. Re-run once Official CRL data is ready to blend in.
data = [r for r in rows[1:] if r[idx['Match Category']] == 'Practice']

wb = openpyxl.load_workbook(XLSX_PATH)  # normal load, preserves formulas, used for saving

def has_ronin(r, side):
    cards = [r[idx[f'Card {i}']] for i in range(1, 9)] if side == 'player' else [r[idx[f'Opp Card {i}']] for i in range(1, 9)]
    return 'Ronin' in cards

# By game number (any duel-start certainty)
by_game = {}
for g in (1, 2, 3):
    grows = [r for r in data if r[idx['Game #']] == g]
    total = len(grows)
    ronin = sum(1 for r in grows if has_ronin(r, 'player') or has_ronin(r, 'opponent'))
    by_game[g] = (total, ronin)

# Game 1, certain-start only (cleaner sample -- true first game of the duel)
g1_certain = [r for r in data if r[idx['Game #']] == 1 and r[idx['Duel Start Uncertain']] != 'Yes']
g1_certain_total = len(g1_certain)
g1_certain_ronin = sum(1 for r in g1_certain if has_ronin(r, 'player') or has_ronin(r, 'opponent'))

# Overall baseline across every game, any game number
total_games = len(data)
ronin_any = sum(1 for r in data if has_ronin(r, 'player') or has_ronin(r, 'opponent'))

# Win rate for the tracked player when THEY open G1 with Ronin (certain-start only)
g1_player_ronin = [r for r in g1_certain if has_ronin(r, 'player')]
g1_player_ronin_wins = sum(1 for r in g1_player_ronin if r[idx['Result']] == 'Win')

# Top players bringing Ronin into G1 (certain-start)
player_counter = Counter()
for r in g1_certain:
    if has_ronin(r, 'player'):
        player_counter[r[idx['Player']]] += 1
    if has_ronin(r, 'opponent'):
        player_counter[r[idx['Opponent']]] += 1
top_players = player_counter.most_common(8)

# ---- Build the sheet ----
if 'Ronin G1 Analysis' in wb.sheetnames:
    del wb['Ronin G1 Analysis']
ws = wb.create_sheet('Ronin G1 Analysis')
ws.sheet_view.showGridLines = True
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 16
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 90

r = 1
ws.cell(r, 1, 'Ronin -- Game 1 Play-Rate Analysis').font = Font(name='Arial', size=14, bold=True, color='FF1F4E78')
r += 1
ws.cell(r, 1, 'Generated from Duel Log -- counts a game as "has Ronin" if either the tracked player or the '
               'logged opponent ran it. One-off analysis requested by the user; not wired into any live formula.').font = NOTE_FONT
r += 2

ws.cell(r, 1, 'Ronin play rate by game number in the set').font = SECTION_FONT
r += 1
headers = ['Game #', 'Total Games', 'Games with Ronin', 'Ronin Play Rate']
for c, h in enumerate(headers, start=1):
    cell = ws.cell(r, c, h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
r += 1
for g in (1, 2, 3):
    total, ronin = by_game[g]
    ws.cell(r, 1, f'Game {g}').font = LABEL_FONT
    ws.cell(r, 2, total).font = VALUE_FONT
    ws.cell(r, 3, ronin).font = VALUE_FONT
    rate_cell = ws.cell(r, 4, ronin / total if total else 0)
    rate_cell.number_format = '0.0%'
    rate_cell.font = VALUE_FONT
    r += 1
r += 1

ws.cell(r, 1, 'Game 1 only, certain-start duels (cleanest sample)').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'Metric').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Count').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, '').fill = HEADER_FILL
ws.cell(r, 4, 'Notes').font = HEADER_FONT; ws.cell(r, 4).fill = HEADER_FILL
r += 1
rows_to_write = [
    ('Total Game-1 rows (certain-start only)', g1_certain_total,
     'Excludes 139 "uncertain-start" duels (first duel per pair -- its "Game 1" label may actually be a later '
     'game of a longer real duel that started before the fetch history began). Same exclusion the workbook '
     'already applies to Win-Con Sets and the dashboard predictors.'),
    ('Game-1 rows featuring Ronin (either side)', g1_certain_ronin,
     f'{g1_certain_ronin/g1_certain_total:.1%} of certain-start Game 1s had Ronin on the table somewhere.'),
    ('Baseline: Ronin rate across ALL games (any game #)', ronin_any,
     f'{ronin_any/total_games:.1%} of all {total_games} logged games (any game number, any duel-start certainty) featured Ronin -- use as the comparison point for whether G1 usage is higher or lower than average.'),
]
for label, count, note in rows_to_write:
    ws.cell(r, 1, label).font = LABEL_FONT
    ws.cell(r, 2, count).font = VALUE_FONT
    ws.cell(r, 4, note).font = VALUE_FONT
    ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
r += 1

ws.cell(r, 1, 'Tracked-player win rate when opening G1 with Ronin').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'Metric').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Count').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, '').fill = HEADER_FILL
ws.cell(r, 4, 'Notes').font = HEADER_FONT; ws.cell(r, 4).fill = HEADER_FILL
r += 1
ws.cell(r, 1, 'Tracked-player G1 decks running Ronin').font = LABEL_FONT
ws.cell(r, 2, len(g1_player_ronin)).font = VALUE_FONT
r += 1
ws.cell(r, 1, 'Of those, wins').font = LABEL_FONT
ws.cell(r, 2, g1_player_ronin_wins).font = VALUE_FONT
r += 1
ws.cell(r, 1, 'Win rate').font = LABEL_FONT
wr_cell = ws.cell(r, 2, g1_player_ronin_wins / len(g1_player_ronin) if g1_player_ronin else 0)
wr_cell.number_format = '0.0%'
wr_cell.font = VALUE_FONT
ws.cell(r, 4, 'Small sample -- treat as directional, not conclusive.').font = VALUE_FONT
r += 2

ws.cell(r, 1, 'Top players bringing Ronin into Game 1 (either side of the matchup)').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'Player').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'G1 Ronin Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
r += 1
for name, cnt in top_players:
    ws.cell(r, 1, name).font = VALUE_FONT
    ws.cell(r, 2, cnt).font = VALUE_FONT
    r += 1
r += 1

ws.cell(r, 1, 'Read').font = SECTION_FONT
r += 1
ws.cell(r, 1,
        "Ronin showed up in about 31% of certain-start Game 1s, close to its ~31% baseline rate across all "
        "games -- so on this dataset it's NOT being held back and used mainly as a G2/G3 read after seeing the "
        "opponent's deck; it's played blind about as often as it's played reactively. Game 2 ticked up to 34% "
        "and Game 3 dropped to 27.5%, but the swing is small and the per-game sample (444-583 rows) is modest, "
        "so treat the game-to-game difference as a mild signal rather than a strong pattern. Re-run this once "
        "more duels accumulate (especially once Official CRL games exist) to see if the pattern firms up."
        ).font = VALUE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.row_dimensions[r].height = 90
r += 3

# =============================================================================
# Ronin in Official CRL (added 2026-07-19, per user request -- see module docstring)
# =============================================================================
data_crl_all = [rr for rr in rows[1:] if rr[idx['Match Category']] == 'Official CRL']
crl_total_games = len(data_crl_all)
crl_ronin_player = [rr for rr in data_crl_all if has_ronin(rr, 'player')]
crl_ronin_either = [rr for rr in data_crl_all if has_ronin(rr, 'player') or has_ronin(rr, 'opponent')]
crl_ronin_player_wins = sum(1 for rr in crl_ronin_player if rr[idx['Result']] == 'Win')

ws.cell(r, 1, 'RONIN IN OFFICIAL CRL').font = Font(name='Arial', size=14, bold=True, color='FF1F4E78')
r += 1
ws.cell(r, 1, 'Everything below is scoped to Match Category = "Official CRL" only (the '
               'original G1 analysis above is Practice-only) -- covers how often the tracked '
               'roster played Ronin in real tournament games, their win rate with it, which '
               'win-con sets it paired with, which specific decks ran it, and how it performed '
               'against the 7 opponents in the user\'s actual Group A.').font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 2

ws.cell(r, 1, 'Play rate + win rate in Official CRL').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'Metric').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Count').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, '').fill = HEADER_FILL
ws.cell(r, 4, 'Notes').font = HEADER_FONT; ws.cell(r, 4).fill = HEADER_FILL
r += 1
crl_rows_to_write = [
    ('Total Official CRL games logged (any game #)', crl_total_games, ''),
    ('Games with Ronin on either side', len(crl_ronin_either),
     f'{len(crl_ronin_either)/crl_total_games:.1%} of all Official CRL games had Ronin on the table somewhere.'
     if crl_total_games else 'No Official CRL games logged yet.'),
    ('Games where a TRACKED player ran Ronin (play rate)', len(crl_ronin_player),
     f'{len(crl_ronin_player)/crl_total_games:.1%} of all Official CRL games -- this is the "how often was it '
     'played" number the user asked for, from our own roster\'s side specifically.'
     if crl_total_games else ''),
]
for label, count, note in crl_rows_to_write:
    ws.cell(r, 1, label).font = LABEL_FONT
    ws.cell(r, 2, count).font = VALUE_FONT
    ws.cell(r, 4, note).font = VALUE_FONT
    ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical='top')
    r += 1
ws.cell(r, 1, 'Win rate when a tracked player ran Ronin in Official CRL').font = LABEL_FONT
crl_wr_cell = ws.cell(r, 2, crl_ronin_player_wins / len(crl_ronin_player) if crl_ronin_player else 0)
crl_wr_cell.number_format = '0.0%'
crl_wr_cell.font = VALUE_FONT
ws.cell(r, 4, f'{crl_ronin_player_wins} wins out of {len(crl_ronin_player)} games. '
               + ('Small sample -- treat as directional, not conclusive.' if len(crl_ronin_player) < 15 else '')
               ).font = VALUE_FONT
ws.cell(r, 4).alignment = Alignment(wrap_text=True, vertical='top')
r += 2

# ---- Win-con sets Ronin decks paired with (Official CRL, duel-grouped) ----
ws.cell(r, 1, 'Win-con sets Ronin was played alongside (Official CRL duels)').font = SECTION_FONT
r += 1
ws.cell(r, 1, 'A duel counts here if the tracked player ran Ronin in at least one of its games; the win-con '
               'set is every win condition (per the Win Condition Reference sheet) that showed up anywhere '
               'across that duel\'s games -- same method as the workbook\'s main Win-Con Sets sheet, just '
               'filtered to duels featuring Ronin and scoped to Official CRL. Certain-start duels only.'
               ).font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 1

crl_by_duel = defaultdict(list)
for rr in data_crl_all:
    crl_by_duel[rr[idx['Duel ID']]].append(rr)

ronin_wincon_set_duels = Counter()
ronin_wincon_set_games = Counter()
ronin_wincon_set_wins = Counter()
for duel_id, grows in crl_by_duel.items():
    grows_sorted = sorted(grows, key=lambda x: x[idx['Game #']])
    if grows_sorted[0][idx['Duel Start Uncertain']] == 'Yes':
        continue
    non_rematch = [g for g in grows_sorted if g[idx['Instant Rematch']] != 'Yes'][:3]
    if not non_rematch:
        continue
    if not any(has_ronin(g, 'player') for g in non_rematch):
        continue
    wincon_set = set()
    for g in non_rematch:
        deck = [g[idx[f'Card {i}']] for i in range(1, 9)]
        wincon_set.update(classify_deck(deck))
    key = '+'.join(sorted(wincon_set)) if wincon_set else '(none classified)'
    ronin_wincon_set_duels[key] += 1
    for g in non_rematch:
        ronin_wincon_set_games[key] += 1
        if g[idx['Result']] == 'Win':
            ronin_wincon_set_wins[key] += 1

ws.cell(r, 1, 'Win-Con Set').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Duels').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, '').fill = HEADER_FILL
r += 1
if ronin_wincon_set_duels:
    for key, d in ronin_wincon_set_duels.most_common():
        g = ronin_wincon_set_games[key]
        w = ronin_wincon_set_wins[key]
        ws.cell(r, 1, key).font = VALUE_FONT
        ws.cell(r, 2, d).font = VALUE_FONT
        wr_cell = ws.cell(r, 3, w / g if g else 0)
        wr_cell.number_format = '0.0%'
        wr_cell.font = VALUE_FONT
        r += 1
else:
    ws.cell(r, 1, 'No certain-start Official CRL duels with Ronin yet.').font = VALUE_FONT
    r += 1
r += 1

# ---- Deck types running Ronin (Official CRL) ----
ws.cell(r, 1, 'Decks that ran Ronin (Official CRL, tracked-player side)').font = SECTION_FONT
r += 1
ronin_deck_games = Counter()
ronin_deck_wins = Counter()
companion_wincon_counter = Counter()
for rr in crl_ronin_player:
    deck_key = rr[idx['Own Deck Key']]
    ronin_deck_games[deck_key] += 1
    if rr[idx['Result']] == 'Win':
        ronin_deck_wins[deck_key] += 1
    deck = [rr[idx[f'Card {i}']] for i in range(1, 9)]
    for wc in classify_deck(deck):
        if wc != 'Ronin':
            companion_wincon_counter[wc] += 1

ws.cell(r, 1, 'Full Deck (8 cards)').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Games').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Win Rate').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, '').fill = HEADER_FILL
r += 1
if ronin_deck_games:
    for deck_key, g in ronin_deck_games.most_common(10):
        w = ronin_deck_wins[deck_key]
        ws.cell(r, 1, deck_key).font = VALUE_FONT
        ws.cell(r, 1).alignment = Alignment(wrap_text=True)
        ws.cell(r, 2, g).font = VALUE_FONT
        wr_cell = ws.cell(r, 3, w / g if g else 0)
        wr_cell.number_format = '0.0%'
        wr_cell.font = VALUE_FONT
        r += 1
else:
    ws.cell(r, 1, 'No Official CRL games with Ronin yet.').font = VALUE_FONT
    r += 1
r += 1

ws.cell(r, 1, 'Companion win conditions paired with Ronin (Official CRL, excl. Ronin itself)').font = LABEL_FONT
r += 1
if companion_wincon_counter:
    for wc, cnt in companion_wincon_counter.most_common(8):
        ws.cell(r, 1, f'  {wc}').font = VALUE_FONT
        ws.cell(r, 2, cnt).font = VALUE_FONT
        r += 1
else:
    ws.cell(r, 1, '  None yet.').font = VALUE_FONT
    r += 1
r += 2

# ---- Ronin vs the user's 7 actual Group A players ----
# Confirmed (INA.BenZerRidel, Lucas.xit✨之安神, RAD, SandBox) + on-deck (Adox,
# Lucas✨杰克, DK) -- the 6 "reference-only" players (Viiper, adriel, たぁ, Ian77,
# Mohamed Light, Pedro™️) are explicitly NOT part of the user's group, excluded here.
# Scoped across ALL match categories (not just Official CRL) since per-opponent CRL-only
# samples are too thin to be useful on their own -- CRL-only counts are still shown
# alongside so the split is visible, not hidden.
GROUP_TAGS = [
    ('#9GJ0Q0LGG', 'INA.BenZerRidel'),
    ('#2R09LUYPQ', 'Lucas.xit✨之安神'),
    ('#8QRCJQ9Y', 'RAD'),
    ('#Y022GRCJQ', 'SandBox'),
    ('#20R0VLJL92', 'Adox'),
    ('#9G28ULYR', 'Lucas✨杰克'),
    ('#8G9GJQRVQ', 'DK'),
]
group_tag_set = {t for t, _n in GROUP_TAGS}
group_games = Counter()
group_wins = Counter()
group_crl_games = Counter()
group_crl_wins = Counter()
for rr in rows[1:]:
    opp_tag = rr[idx['Opponent Tag']]
    if opp_tag in group_tag_set and has_ronin(rr, 'player'):
        won = rr[idx['Result']] == 'Win'
        group_games[opp_tag] += 1
        if won:
            group_wins[opp_tag] += 1
        if rr[idx['Match Category']] == 'Official CRL':
            group_crl_games[opp_tag] += 1
            if won:
                group_crl_wins[opp_tag] += 1

ws.cell(r, 1, "Ronin vs. the user's 7 Group A players (4 confirmed + 3 on-deck)").font = SECTION_FONT
r += 1
ws.cell(r, 1, 'Every game (any Match Category) where a tracked player ran Ronin against this specific '
               'opponent. "CRL Games/Win Rate" columns isolate the Official-CRL-only subset where it exists -- '
               'both are shown since per-opponent CRL samples are typically very thin.').font = NOTE_FONT
ws.cell(r, 1).alignment = Alignment(wrap_text=True)
r += 1
ws.cell(r, 1, 'Opponent').font = HEADER_FONT; ws.cell(r, 1).fill = HEADER_FILL
ws.cell(r, 2, 'Games (All)').font = HEADER_FONT; ws.cell(r, 2).fill = HEADER_FILL
ws.cell(r, 3, 'Win Rate (All)').font = HEADER_FONT; ws.cell(r, 3).fill = HEADER_FILL
ws.cell(r, 4, 'CRL Games / CRL Win Rate').font = HEADER_FONT; ws.cell(r, 4).fill = HEADER_FILL
r += 1
total_group_games = 0
total_group_wins = 0
for tag, name in GROUP_TAGS:
    g = group_games[tag]
    w = group_wins[tag]
    total_group_games += g
    total_group_wins += w
    cg = group_crl_games[tag]
    cw = group_crl_wins[tag]
    ws.cell(r, 1, name).font = VALUE_FONT
    if g:
        ws.cell(r, 2, g).font = VALUE_FONT
        wr_cell = ws.cell(r, 3, w / g)
        wr_cell.number_format = '0.0%'
        wr_cell.font = VALUE_FONT
    else:
        ws.cell(r, 2, 0).font = VALUE_FONT
        ws.cell(r, 3, 'no games').font = NOTE_FONT
    ws.cell(r, 4, f'{cg}g' + (f', {cw/cg:.0%}' if cg else ', n/a')).font = VALUE_FONT
    r += 1
ws.cell(r, 1, 'TOTAL (all 7 group players combined)').font = LABEL_FONT
ws.cell(r, 2, total_group_games).font = LABEL_FONT
if total_group_games:
    total_wr_cell = ws.cell(r, 3, total_group_wins / total_group_games)
    total_wr_cell.number_format = '0.0%'
    total_wr_cell.font = LABEL_FONT
else:
    ws.cell(r, 3, 'no games').font = NOTE_FONT
r += 2

wb.save(XLSX_PATH)
print("Sheet added. Rows written:", r)
